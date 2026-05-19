#!/usr/bin/env python3
"""
Rebuild SK_Delivery_System.xlsx — single-pass, no xlwings needed.

Creates a BRAND NEW file with openpyxl. Since this file never had Power Query,
there's no metadata for openpyxl to strip → no corruption.

Power Query will be re-added by the user through Excel's GUI later.
"""
import openpyxl, os, sys

CORRUPT_FILE = '/Users/pabloherrera/Documents/Claude/Projects/route optimization/sk_optimizer/data/SK_Delivery_System.xlsx'
CLEAN_FILE   = '/Users/pabloherrera/Documents/Claude/Projects/route optimization/sk_optimizer/data/SK_Delivery_System_clean.xlsx'

# ═══ Step 1: Read all data (values) from corrupt file ═══
print("Step 1: Reading data values...")
wb_old = openpyxl.load_workbook(CORRUPT_FILE, data_only=True, read_only=True)
sheet_data = {}
for name in wb_old.sheetnames:
    ws = wb_old[name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    sheet_data[name] = rows
    print(f"  {name}: {len(rows)} rows")
wb_old.close()

# ═══ Step 2: Read all formulas from corrupt file ═══
print("\nStep 2: Reading formulas...")
wb_f = openpyxl.load_workbook(CORRUPT_FILE, read_only=True)
formula_map = {}
for sname in ['Client_List', 'Optimizer_Input']:
    ws = wb_f[sname]
    formulas = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formulas[cell.coordinate] = cell.value
    formula_map[sname] = formulas
    print(f"  {sname}: {len(formulas)} formulas")
wb_f.close()

# ═══ Step 3: Create brand new workbook ═══
print("\nStep 3: Creating new workbook with all data + formulas...")
wb = openpyxl.Workbook()

target_sheets = ['Delivery_Log', 'Client_List', 'Optimizer_Input', 'Trucks',
                 'Depot', 'Client_Time_Windows', 'Client_Closures', 'Query', 'Anova_Live']

for i, name in enumerate(target_sheets):
    if i == 0:
        ws = wb.active
        ws.title = name
    else:
        ws = wb.create_sheet(name)

    # Write data values
    if name in sheet_data and sheet_data[name]:
        for r_idx, row in enumerate(sheet_data[name], 1):
            for c_idx, val in enumerate(row, 1):
                if val is not None:
                    ws.cell(r_idx, c_idx).value = val
        print(f"  {name}: {len(sheet_data[name])} rows of data")

# ═══ Step 4: Restore Client_List formulas H-Q ═══
print("\nStep 4: Restoring Client_List H-Q formulas...")
ws_cl = wb['Client_List']
count = 0
for coord, formula in formula_map.get('Client_List', {}).items():
    col = ''.join(c for c in coord if c.isalpha())
    if col in ['H','I','J','K','L','M','N','O','P','Q']:
        ws_cl[coord].value = formula
        count += 1
print(f"  {count} formulas")

# ═══ Step 5: Add Client_List R-X Anova formulas ═══
print("\nStep 5: Adding Client_List R-X Anova formulas...")
# Headers
for i, h in enumerate(['RTU ID', 'Anova Level (lbs)', 'Anova % Full',
                        'Anova Last Reading', 'Anova Deliverable',
                        'Anova Age (hrs)', 'Anova Source']):
    ws_cl.cell(3, 18 + i).value = h

for row in range(4, 176):
    r = str(row)
    ws_cl[f'R{r}'].value = f'=IFERROR(_xlfn.XLOOKUP($A{r},Anova_Live!$B$2:$B$60,Anova_Live!$A$2:$A$60,""),"")'
    ws_cl[f'S{r}'].value = f'=IFERROR(_xlfn.XLOOKUP($A{r},Anova_Live!$B$2:$B$60,Anova_Live!$E$2:$E$60,""),"")'
    ws_cl[f'T{r}'].value = f'=IFERROR(_xlfn.XLOOKUP($A{r},Anova_Live!$B$2:$B$60,Anova_Live!$K$2:$K$60,""),"")'
    ws_cl[f'U{r}'].value = f'=IFERROR(_xlfn.XLOOKUP($A{r},Anova_Live!$B$2:$B$60,Anova_Live!$H$2:$H$60,""),"")'
    ws_cl[f'V{r}'].value = f'=IFERROR(_xlfn.XLOOKUP($A{r},Anova_Live!$B$2:$B$60,Anova_Live!$L$2:$L$60,""),"")'
    ws_cl[f'W{r}'].value = f'=IFERROR(_xlfn.XLOOKUP($A{r},Anova_Live!$B$2:$B$60,Anova_Live!$M$2:$M$60,""),"")'
    ws_cl[f'X{r}'].value = f'=IFERROR(_xlfn.XLOOKUP($A{r},Anova_Live!$B$2:$B$60,Anova_Live!$N$2:$N$60,""),"")'
print(f"  172 rows × 7 cols = 1204 formulas")

# ═══ Step 6: Restore Optimizer_Input formulas C-R ═══
print("\nStep 6: Restoring Optimizer_Input C-R formulas...")
ws_oi = wb['Optimizer_Input']
count = 0
for coord, formula in formula_map.get('Optimizer_Input', {}).items():
    col = ''.join(c for c in coord if c.isalpha())
    if col not in ['S', 'T', 'U', 'A', 'B']:
        ws_oi[coord].value = formula
        count += 1
print(f"  {count} formulas")

# ═══ Step 7: Add Optimizer_Input S-U Anova formulas ═══
print("\nStep 7: Adding Optimizer_Input S-U Anova formulas...")
ws_oi.cell(5, 19).value = 'Anova Level (lbs)'
ws_oi.cell(5, 20).value = 'Anova Updated'
ws_oi.cell(5, 21).value = 'Anova Status'

for row in range(6, 176):
    r = str(row)
    ws_oi[f'S{r}'].value = f'=IFERROR(_xlfn.XLOOKUP(B{r},Client_List!$A$4:$A$200,Client_List!$S$4:$S$200,""),"")'
    ws_oi[f'T{r}'].value = f'=IFERROR(_xlfn.XLOOKUP(B{r},Client_List!$A$4:$A$200,Client_List!$U$4:$U$200,""),"")'
    ws_oi[f'U{r}'].value = f'=IFERROR(IF(S{r}="","",IF(_xlfn.XLOOKUP(B{r},Client_List!$A$4:$A$200,Client_List!$W$4:$W$200,999)<=24,"LIVE",IF(_xlfn.XLOOKUP(B{r},Client_List!$A$4:$A$200,Client_List!$W$4:$W$200,999)<=72,"STALE","OLD"))),"")'
print(f"  170 rows × 3 cols = 510 formulas")

# ═══ Step 8: Build Anova_Live formulas ═══
print("\nStep 8: Building Anova_Live formulas...")
ws_al = wb['Anova_Live']

# Headers
al_headers = ['device_id', 'client_id', 'client_name', 'customer_ref',
              'display_value', 'display_unit', 'product', 'timestamp',
              'received_at', 'tank_capacity', 'pct_full', 'deliverable_lbs',
              'age_hours', 'source']
for c, h in enumerate(al_headers, 1):
    ws_al.cell(1, c).value = h

# Known devices in rows 2-4
devices = [(124697513, 1054), (121381187, 19117), (124677515, '')]
for i, (dev, cid) in enumerate(devices):
    r = i + 2
    ws_al.cell(r, 1).value = dev
    ws_al.cell(r, 2).value = cid
    ws_al.cell(r, 14).value = 'push'

# Formulas for all rows 2-54
for row in range(2, 55):
    r = str(row)
    ws_al[f'C{r}'].value = f'=IFERROR(_xlfn.XLOOKUP(B{r},Client_List!$A$4:$A$200,Client_List!$B$4:$B$200,""),"")'
    ws_al[f'D{r}'].value = f'=IFERROR(_xlfn.XLOOKUP(TEXT(A{r},"0"),Query!$A$2:$A$100,Query!$B$2:$B$100,""),"")'
    ws_al[f'E{r}'].value = f'=IFERROR(_xlfn.XLOOKUP(TEXT(A{r},"0"),Query!$A$2:$A$100,Query!$C$2:$C$100,""),"")'
    ws_al[f'F{r}'].value = f'=IFERROR(_xlfn.XLOOKUP(TEXT(A{r},"0"),Query!$A$2:$A$100,Query!$D$2:$D$100,""),"")'
    ws_al[f'G{r}'].value = f'=IFERROR(_xlfn.XLOOKUP(TEXT(A{r},"0"),Query!$A$2:$A$100,Query!$G$2:$G$100,""),"")'
    ws_al[f'H{r}'].value = f'=IFERROR(_xlfn.XLOOKUP(TEXT(A{r},"0"),Query!$A$2:$A$100,Query!$H$2:$H$100,""),"")'
    ws_al[f'I{r}'].value = f'=IFERROR(_xlfn.XLOOKUP(TEXT(A{r},"0"),Query!$A$2:$A$100,Query!$I$2:$I$100,""),"")'
    ws_al[f'J{r}'].value = f'=IFERROR(_xlfn.XLOOKUP(B{r},Client_List!$A$4:$A$200,Client_List!$J$4:$J$200,""),"")'
    ws_al[f'K{r}'].value = f'=IFERROR(IF(AND(E{r}<>"",J{r}<>""),ROUND(E{r}/J{r}*100,1),""),"")'
    ws_al[f'L{r}'].value = f'=IFERROR(IF(AND(E{r}<>"",J{r}<>""),ROUND(J{r}-E{r},0),""),"")'
    ws_al[f'M{r}'].value = f'=IFERROR(IF(H{r}<>"",ROUND((NOW()-H{r})*24,1),""),"")'
print(f"  53 rows × 11 cols = 583 formulas")

# ═══ Step 9: Save ═══
print("\nStep 9: Saving...")
wb.save(CLEAN_FILE)
size = os.path.getsize(CLEAN_FILE)
print(f"  Saved: {CLEAN_FILE}")
print(f"  Size: {size:,} bytes")

# ═══ Summary ═══
print(f"""
═══ REBUILD COMPLETE ═══
File: {CLEAN_FILE}
Size: {size:,} bytes
Sheets: {[ws.title for ws in wb.worksheets]}

Data chain:
  Query (Power Query target) → Anova_Live (XLOOKUP) → Client_List R-X (XLOOKUP) → Optimizer_Input S-U (XLOOKUP)

Next steps:
  1. Open in Excel — should open with NO repair dialog
  2. Rename to SK_Delivery_System.xlsx
  3. Set up Power Query: Data → Get Data → From Web → paste M code → Close & Load To → Query sheet cell A1
""")
