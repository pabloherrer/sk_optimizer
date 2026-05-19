#!/usr/bin/env python3
"""Phase 1: Create brand new data-only file with openpyxl."""
import openpyxl, os

CORRUPT_FILE = '/Users/pabloherrera/Documents/Claude/Projects/route optimization/sk_optimizer/data/SK_Delivery_System.xlsx'
CLEAN_FILE   = '/Users/pabloherrera/Documents/Claude/Projects/route optimization/sk_optimizer/data/SK_Delivery_System_clean.xlsx'

print("Reading data from existing file...")
wb_old = openpyxl.load_workbook(CORRUPT_FILE, data_only=True, read_only=True)

sheet_data = {}
for name in wb_old.sheetnames:
    ws = wb_old[name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    sheet_data[name] = rows
    print(f"  {name}: {len(rows)} rows")
wb_old.close()

target_sheets = ['Delivery_Log', 'Client_List', 'Optimizer_Input', 'Trucks',
                 'Depot', 'Client_Time_Windows', 'Client_Closures', 'Query', 'Anova_Live']

print("\nWriting new workbook...")
wb = openpyxl.Workbook()

for i, name in enumerate(target_sheets):
    if i == 0:
        ws = wb.active
        ws.title = name
    else:
        ws = wb.create_sheet(name)

    if name in sheet_data and sheet_data[name]:
        for r_idx, row in enumerate(sheet_data[name], 1):
            for c_idx, val in enumerate(row, 1):
                if val is not None:
                    ws.cell(r_idx, c_idx).value = val
        print(f"  {name}: {len(sheet_data[name])} rows")

# Anova_Live headers + device data
ws_al = wb['Anova_Live']
al_headers = ['device_id', 'client_id', 'client_name', 'customer_ref',
              'display_value', 'display_unit', 'product', 'timestamp',
              'received_at', 'tank_capacity', 'pct_full', 'deliverable_lbs',
              'age_hours', 'source']
for c, h in enumerate(al_headers, 1):
    ws_al.cell(1, c).value = h

devices = [(124697513, 1054), (121381187, 19117), (124677515, '')]
for i, (dev, cid) in enumerate(devices):
    r = i + 2
    ws_al.cell(r, 1).value = dev
    ws_al.cell(r, 2).value = cid
    ws_al.cell(r, 14).value = 'push'

# Client_List Anova headers
ws_cl = wb['Client_List']
for i, h in enumerate(['RTU ID', 'Anova Level (lbs)', 'Anova % Full',
                        'Anova Last Reading', 'Anova Deliverable',
                        'Anova Age (hrs)', 'Anova Source']):
    ws_cl.cell(3, 18 + i).value = h

# Optimizer_Input Anova headers
ws_oi = wb['Optimizer_Input']
ws_oi.cell(5, 19).value = 'Anova Level (lbs)'
ws_oi.cell(5, 20).value = 'Anova Updated'
ws_oi.cell(5, 21).value = 'Anova Status'

wb.save(CLEAN_FILE)
print(f"\nSaved: {CLEAN_FILE} ({os.path.getsize(CLEAN_FILE):,} bytes)")
print("Phase 1 DONE.")
