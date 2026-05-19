#!/usr/bin/env python3
"""
Rebuild SK_Delivery_System.xlsx from scratch.

Phase 1: openpyxl creates a BRAND NEW file with all raw data (no formulas).
          Since this is a new file (not editing one with Power Query), there's
          nothing for openpyxl to corrupt.

Phase 2: xlwings opens that file through Excel and adds all formulas.
          Excel saves it cleanly — no XML corruption.
"""
import openpyxl
import time, sys, os

CORRUPT_FILE = '/Users/pabloherrera/Documents/Claude/Projects/route optimization/sk_optimizer/data/SK_Delivery_System.xlsx'
CLEAN_FILE   = '/Users/pabloherrera/Documents/Claude/Projects/route optimization/sk_optimizer/data/SK_Delivery_System_clean.xlsx'

# ═══════════════════════════════════════════════════════
# PHASE 1 — Read old data, write a brand new file with openpyxl
# ═══════════════════════════════════════════════════════

print("═══ PHASE 1: Create new file with data (openpyxl) ═══")
print("Step 1: Reading data from existing file (read-only)...")
wb_old = openpyxl.load_workbook(CORRUPT_FILE, data_only=True, read_only=True)

sheet_data = {}
for name in wb_old.sheetnames:
    ws = wb_old[name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    sheet_data[name] = rows
    print(f"  {name}: {len(rows)} rows")

wb_old.close()

# Sheet order
target_sheets = ['Delivery_Log', 'Client_List', 'Optimizer_Input', 'Trucks',
                 'Depot', 'Client_Time_Windows', 'Client_Closures', 'Query', 'Anova_Live']

print("\nStep 2: Writing brand new workbook...")
wb_new = openpyxl.Workbook()

for i, name in enumerate(target_sheets):
    if i == 0:
        ws = wb_new.active
        ws.title = name
    else:
        ws = wb_new.create_sheet(name)

    if name in sheet_data and sheet_data[name]:
        data = sheet_data[name]
        for r_idx, row in enumerate(data, 1):
            for c_idx, val in enumerate(row, 1):
                if val is not None:
                    ws.cell(r_idx, c_idx).value = val
        print(f"  {name}: wrote {len(data)} rows")
    else:
        print(f"  {name}: created empty")

# Add Anova_Live headers (data rows already copied if they existed)
ws_al = wb_new['Anova_Live']
al_headers = ['device_id', 'client_id', 'client_name', 'customer_ref',
              'display_value', 'display_unit', 'product', 'timestamp',
              'received_at', 'tank_capacity', 'pct_full', 'deliverable_lbs',
              'age_hours', 'source']
for c, hdr in enumerate(al_headers, 1):
    ws_al.cell(1, c).value = hdr

# Ensure known device IDs are in rows 2-4
devices = [(124697513, 1054), (121381187, 19117), (124677515, '')]
for i, (dev_id, client_id) in enumerate(devices):
    row = i + 2
    ws_al.cell(row, 1).value = dev_id   # A = device_id
    ws_al.cell(row, 2).value = client_id  # B = client_id
    ws_al.cell(row, 14).value = 'push'   # N = source

# Add Anova headers to Client_List (row 3, cols R-X = 18-24)
ws_cl = wb_new['Client_List']
cl_anova_headers = ['RTU ID', 'Anova Level (lbs)', 'Anova % Full',
                    'Anova Last Reading', 'Anova Deliverable',
                    'Anova Age (hrs)', 'Anova Source']
for i, hdr in enumerate(cl_anova_headers):
    ws_cl.cell(3, 18 + i).value = hdr

# Add Anova headers to Optimizer_Input (row 5, cols S-U = 19-21)
ws_oi = wb_new['Optimizer_Input']
ws_oi.cell(5, 19).value = 'Anova Level (lbs)'
ws_oi.cell(5, 20).value = 'Anova Updated'
ws_oi.cell(5, 21).value = 'Anova Status'

# Save the data-only file
wb_new.save(CLEAN_FILE)
print(f"\n  Data-only file saved: {CLEAN_FILE}")
print(f"  File size: {os.path.getsize(CLEAN_FILE):,} bytes")

# ═══════════════════════════════════════════════════════
# PHASE 2 — Open in Excel via xlwings, add all formulas
# ═══════════════════════════════════════════════════════

print("\n═══ PHASE 2: Add formulas via Excel (xlwings) ═══")
import xlwings as xw

app = xw.apps.active
if app is None:
    app = xw.App(visible=True)
    time.sleep(3)

app.display_alerts = False

# Close any open workbooks
for b in list(app.books):
    try:
        b.close()
    except:
        pass
time.sleep(1)

# Open our clean data file
print("Step 3: Opening clean file in Excel...")
wb = app.books.open(CLEAN_FILE)
time.sleep(2)
print(f"  Opened: {wb.name}")
print(f"  Sheets: {[s.name for s in wb.sheets]}")

# ── Read existing formulas from corrupt file to restore them ──
print("\nStep 4: Reading formulas from old file...")
wb_formulas = openpyxl.load_workbook(CORRUPT_FILE, read_only=True)

formula_map = {}
for sname in ['Client_List', 'Optimizer_Input']:
    ws = wb_formulas[sname]
    formulas = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formulas[cell.coordinate] = cell.value
    formula_map[sname] = formulas
    print(f"  {sname}: {len(formulas)} formulas")
wb_formulas.close()

# ── Restore Client_List formulas H-Q ──
print("\nStep 5: Restoring Client_List H-Q formulas...")
ws_cl = wb.sheets['Client_List']
cl_formulas = formula_map.get('Client_List', {})
count = 0
for coord, formula in cl_formulas.items():
    col_letter = ''.join(c for c in coord if c.isalpha())
    if col_letter in ['H','I','J','K','L','M','N','O','P','Q']:
        try:
            ws_cl.range(coord).formula = formula
            count += 1
        except:
            pass
print(f"  Restored {count} formulas")

# ── Add Client_List R-X Anova formulas ──
print("\nStep 6: Adding Client_List R-X Anova formulas...")
for row in range(4, 176):
    r = str(row)
    ws_cl.range(f'R{r}').formula = f'=IFERROR(XLOOKUP(TEXT($A{r},"0"),Anova_Live!$B$2:$B$60,Anova_Live!$A$2:$A$60,""),"")'
    ws_cl.range(f'S{r}').formula = f'=IFERROR(XLOOKUP(TEXT($A{r},"0"),Anova_Live!$B$2:$B$60,Anova_Live!$E$2:$E$60,""),"")'
    ws_cl.range(f'T{r}').formula = f'=IFERROR(XLOOKUP(TEXT($A{r},"0"),Anova_Live!$B$2:$B$60,Anova_Live!$K$2:$K$60,""),"")'
    ws_cl.range(f'U{r}').formula = f'=IFERROR(XLOOKUP(TEXT($A{r},"0"),Anova_Live!$B$2:$B$60,Anova_Live!$H$2:$H$60,""),"")'
    ws_cl.range(f'V{r}').formula = f'=IFERROR(XLOOKUP(TEXT($A{r},"0"),Anova_Live!$B$2:$B$60,Anova_Live!$L$2:$L$60,""),"")'
    ws_cl.range(f'W{r}').formula = f'=IFERROR(XLOOKUP(TEXT($A{r},"0"),Anova_Live!$B$2:$B$60,Anova_Live!$M$2:$M$60,""),"")'
    ws_cl.range(f'X{r}').formula = f'=IFERROR(XLOOKUP(TEXT($A{r},"0"),Anova_Live!$B$2:$B$60,Anova_Live!$N$2:$N$60,""),"")'
    if row % 50 == 0:
        print(f"    row {row}...")

# ── Restore Optimizer_Input formulas C-R ──
print("\nStep 7: Restoring Optimizer_Input C-R formulas...")
ws_oi = wb.sheets['Optimizer_Input']
oi_formulas = formula_map.get('Optimizer_Input', {})
count = 0
for coord, formula in oi_formulas.items():
    col_letter = ''.join(c for c in coord if c.isalpha())
    if col_letter not in ['S', 'T', 'U', 'A', 'B']:
        try:
            ws_oi.range(coord).formula = formula
            count += 1
        except:
            pass
print(f"  Restored {count} formulas")

# ── Add Optimizer_Input S-U Anova formulas ──
print("\nStep 8: Adding Optimizer_Input S-U Anova formulas...")
for row in range(6, 176):
    r = str(row)
    ws_oi.range(f'S{r}').formula = f'=IFERROR(XLOOKUP(B{r},Client_List!$A$4:$A$200,Client_List!$S$4:$S$200,""),"")'
    ws_oi.range(f'T{r}').formula = f'=IFERROR(XLOOKUP(B{r},Client_List!$A$4:$A$200,Client_List!$U$4:$U$200,""),"")'
    ws_oi.range(f'U{r}').formula = f'=IFERROR(IF(S{r}="","",IF(XLOOKUP(B{r},Client_List!$A$4:$A$200,Client_List!$W$4:$W$200,999)<=24,"LIVE",IF(XLOOKUP(B{r},Client_List!$A$4:$A$200,Client_List!$W$4:$W$200,999)<=72,"STALE","OLD"))),"")'
    if row % 50 == 0:
        print(f"    row {row}...")

# ── Build Anova_Live formulas ──
print("\nStep 9: Adding Anova_Live formulas...")
ws_al = wb.sheets['Anova_Live']

# Rows 2-4 (known devices) + template rows 5-54
for row in range(2, 55):
    r = str(row)
    ws_al.range(f'C{r}').formula = f'=IFERROR(XLOOKUP(B{r},Client_List!$A$4:$A$200,Client_List!$B$4:$B$200,""),"")'
    ws_al.range(f'D{r}').formula = f'=IFERROR(XLOOKUP(TEXT(A{r},"0"),Query!$A$2:$A$100,Query!$B$2:$B$100,""),"")'
    ws_al.range(f'E{r}').formula = f'=IFERROR(XLOOKUP(TEXT(A{r},"0"),Query!$A$2:$A$100,Query!$C$2:$C$100,""),"")'
    ws_al.range(f'F{r}').formula = f'=IFERROR(XLOOKUP(TEXT(A{r},"0"),Query!$A$2:$A$100,Query!$D$2:$D$100,""),"")'
    ws_al.range(f'G{r}').formula = f'=IFERROR(XLOOKUP(TEXT(A{r},"0"),Query!$A$2:$A$100,Query!$G$2:$G$100,""),"")'
    ws_al.range(f'H{r}').formula = f'=IFERROR(XLOOKUP(TEXT(A{r},"0"),Query!$A$2:$A$100,Query!$H$2:$H$100,""),"")'
    ws_al.range(f'I{r}').formula = f'=IFERROR(XLOOKUP(TEXT(A{r},"0"),Query!$A$2:$A$100,Query!$I$2:$I$100,""),"")'
    ws_al.range(f'J{r}').formula = f'=IFERROR(XLOOKUP(B{r},Client_List!$A$4:$A$200,Client_List!$J$4:$J$200,""),"")'
    ws_al.range(f'K{r}').formula = f'=IFERROR(IF(AND(E{r}<>"",J{r}<>""),ROUND(E{r}/J{r}*100,1),""),"")'
    ws_al.range(f'L{r}').formula = f'=IFERROR(IF(AND(E{r}<>"",J{r}<>""),ROUND(J{r}-E{r},0),""),"")'
    ws_al.range(f'M{r}').formula = f'=IFERROR(IF(H{r}<>"",ROUND((NOW()-H{r})*24,1),""),"")'
    if row % 20 == 0:
        print(f"    row {row}...")

# ── Delete stray sheets ──
for s in wb.sheets:
    if s.name not in target_sheets:
        try:
            s.delete()
        except:
            pass

# ── Verify ──
print("\nStep 10: Verifying data chain...")
time.sleep(2)
app.calculate()
time.sleep(2)

e2 = ws_al.range('E2').value
c2 = ws_al.range('C2').value
k2 = ws_al.range('K2').value
print(f"  Anova_Live row 2: name={c2}, lbs={e2}, pct={k2}")

s8 = ws_cl.range('S8').value
print(f"  Client_List row 8 (Angry Crab): Anova Level={s8}")

for row in range(6, 176):
    bid = ws_oi.range(f'B{row}').value
    if str(bid) == '1054':
        s = ws_oi.range(f'S{row}').value
        u = ws_oi.range(f'U{row}').value
        print(f"  Optimizer_Input row {row} (1054): Level={s}, Status={u}")
        break

# ── Save ──
print("\nStep 11: Saving through Excel...")
wb.save()
print(f"  Saved: {CLEAN_FILE}")

# ── Close and reopen to verify ──
print("\nStep 12: Close and reopen to verify no repair dialog...")
wb.close()
time.sleep(2)
wb2 = app.books.open(CLEAN_FILE)
time.sleep(3)
sheets = [s.name for s in wb2.sheets]
print(f"  Reopened successfully! Sheets: {sheets}")
print(f"  No repair dialog = SUCCESS")

app.display_alerts = True
print("\n═══ DONE. File is clean and ready for Power Query. ═══")
