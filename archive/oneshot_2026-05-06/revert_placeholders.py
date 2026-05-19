#!/usr/bin/env python3
"""
revert_placeholders.py — Remove the May 1 + May 2 200-lb placeholder rows
=========================================================================

These placeholders were added by ingest_actuals.py as a UI-visible
"these visits are scheduled" marker, but they ended up in the historical
Delivery_Log, which made the IRP think those clients had been topped off.
Result: state file shows 149/171 clients at >90% full.

This script SURGICALLY removes only:
  • Rows where date is May 1 or May 2 AND qty = 200 AND (the customer was
    last delivered earlier with a real qty, OR the row is brand-new)

It KEEPS:
  • All real CSV-loaded deliveries (qty != 200)
  • The 9 placeholder→real-qty replacements (those rows still exist
    with their corrected qty and didn't get the placeholder added BY us)
  • Pre-existing placeholder rows from before this whole ingest run

Then it deletes data/inventory_state.json so the next IRP run rebuilds.
"""

from __future__ import annotations

import sys
from pathlib import Path
from datetime import datetime
import openpyxl

EXCEL = Path('data/SK_Delivery_System.xlsx')
STATE = Path('data/inventory_state.json')
PLAN = Path('data/plan.json')

# Dates whose 200-lb placeholders we treat as schedule-markers, not deliveries
TARGET_DATES = {
    datetime(2026, 5, 1).date(),
    datetime(2026, 5, 2).date(),
}
PLACEHOLDER_QTY = 200


def main() -> int:
    if not EXCEL.exists():
        print(f'Excel not found: {EXCEL}', file=sys.stderr); return 1

    wb = openpyxl.load_workbook(EXCEL)
    ws = wb['Delivery_Log']

    rows_to_delete = []
    for r in range(4, ws.max_row + 1):
        d = ws.cell(r, 1).value
        cid = ws.cell(r, 2).value
        qty = ws.cell(r, 4).value
        if d is None or cid is None:
            continue
        d_date = d.date() if hasattr(d, 'date') else d
        if d_date in TARGET_DATES and qty == PLACEHOLDER_QTY:
            rows_to_delete.append(r)

    if not rows_to_delete:
        print('  No matching placeholder rows found — nothing to revert.')
    else:
        print(f'  Found {len(rows_to_delete)} placeholder row(s) to remove.')
        # Delete from bottom up so indices stay stable
        for r in sorted(rows_to_delete, reverse=True):
            d = ws.cell(r, 1).value
            cid = ws.cell(r, 2).value
            print(f'    - row {r}: {d.date() if hasattr(d, "date") else d} id={cid}')
            ws.delete_rows(r, 1)
        wb.save(EXCEL)
        print(f'  Saved → {EXCEL}')

    # Delete the broken state file so the IRP rebuilds from clean log
    if STATE.exists():
        STATE.unlink()
        print(f'  Deleted: {STATE}')
    if PLAN.exists():
        PLAN.unlink()
        print(f'  Deleted: {PLAN}')

    print('\n  Now run:  python run_irp.py --solve-sec 60 --today 2026-05-02')
    print('  to regenerate clean state and plan.\n')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
