"""
v2.ingest.overrides — Operator overrides from the optional Overrides sheet.

Sheet schema (one row per override):
    type         — pin | forbid | lock | reading | consumption
    client_id    — required for pin/forbid/reading/consumption
    date         — required for pin/lock/reading; first day for forbid/consumption
    date_end     — optional — last day for forbid/consumption (date inclusive)
    lbs          — manual tank reading (lbs) for type=reading
                   manual consumption rate (lbs/day) for type=consumption
    truck        — truck_id for type=lock
    stop_n       — locked_through_stop for type=lock (0 = entire day)
    reason       — required free-text justification
    operator     — operator initials/name

If the sheet doesn't exist, returns an empty Overrides().

If the sheet exists but contains a malformed row, raises
OverrideValidationError with the spreadsheet row number — overrides are
high-stakes operator inputs and silent skipping is dangerous.
"""
from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

import pandas as pd
from openpyxl import load_workbook

from v2.domain.overrides import (
    Forbid,
    Lock,
    ManualConsumption,
    ManualReading,
    Overrides,
    Pin,
)


class OverrideValidationError(ValueError):
    """Raised when a row in the Overrides sheet cannot be parsed."""


_HEADER_ALIASES: dict[str, str] = {
    'type': 'type',
    'kind': 'type',
    'client_id': 'client_id',
    'client': 'client_id',
    'date': 'date',
    'date_start': 'date',
    'start_date': 'date',
    'date_end': 'date_end',
    'end_date': 'date_end',
    'lbs': 'lbs',
    'qty': 'lbs',
    'truck': 'truck',
    'truck_id': 'truck',
    'stop_n': 'stop_n',
    'locked_through_stop': 'stop_n',
    'reason': 'reason',
    'operator': 'operator',
}

_VALID_TYPES: frozenset[str] = frozenset(
    {'pin', 'forbid', 'lock', 'reading', 'consumption'}
)


def load_overrides(input_file: Path) -> Overrides:
    """
    Read the Overrides sheet (if present). Returns Overrides() if the
    sheet doesn't exist.

    Raises OverrideValidationError with the offending row number if any
    row is malformed.
    """
    input_file = Path(input_file)
    if not input_file.exists():
        return Overrides()

    try:
        wb = load_workbook(str(input_file), data_only=True)
    except Exception:
        return Overrides()

    if 'Overrides' not in wb.sheetnames:
        return Overrides()

    ws = wb['Overrides']
    # Header row may be row 1, 2, or 3 — find the first row whose first
    # cell contains a recognisable header name.
    header_row_idx = _find_header_row(ws)
    if header_row_idx is None:
        return Overrides()

    raw_headers = [
        ws.cell(header_row_idx, c).value
        for c in range(1, ws.max_column + 1)
    ]
    col_map: dict[str, int] = {}
    for i, h in enumerate(raw_headers):
        if h is None:
            continue
        key = _HEADER_ALIASES.get(str(h).strip().lower())
        if key:
            col_map[key] = i

    if 'type' not in col_map:
        raise OverrideValidationError(
            "Overrides sheet present but no 'type' column found "
            f"(header row {header_row_idx})."
        )

    pins: list[Pin] = []
    forbids: list[Forbid] = []
    locks: list[Lock] = []
    readings: list[ManualReading] = []
    consumptions: list[ManualConsumption] = []

    for r in range(header_row_idx + 1, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if not any(v is not None and str(v).strip() != '' for v in row):
            continue

        try:
            otype = _str(row, col_map.get('type')).lower()
            if not otype:
                continue
            if otype not in _VALID_TYPES:
                raise OverrideValidationError(
                    f"Row {r}: unknown override type '{otype}' "
                    f"(expected one of {sorted(_VALID_TYPES)})"
                )

            client_id = _str(row, col_map.get('client_id'))
            d_start = _date(row, col_map.get('date'))
            d_end = _date(row, col_map.get('date_end'))
            lbs = _float(row, col_map.get('lbs'))
            truck = _str(row, col_map.get('truck'))
            stop_n = _int(row, col_map.get('stop_n'))
            reason = _str(row, col_map.get('reason'))
            operator = _str(row, col_map.get('operator'))

            if not reason:
                raise OverrideValidationError(
                    f"Row {r} ({otype}): 'reason' is required."
                )

            if otype == 'pin':
                if not client_id or d_start is None:
                    raise OverrideValidationError(
                        f"Row {r} (pin): need client_id and date."
                    )
                pins.append(Pin(
                    client_id=client_id,
                    date=d_start,
                    reason=reason,
                    operator=operator,
                ))

            elif otype == 'forbid':
                if not client_id or d_start is None:
                    raise OverrideValidationError(
                        f"Row {r} (forbid): need client_id and date."
                    )
                if d_end is None:
                    dates = (d_start,)
                else:
                    if d_end < d_start:
                        raise OverrideValidationError(
                            f"Row {r} (forbid): date_end < date."
                        )
                    dates = _date_range(d_start, d_end)
                forbids.append(Forbid(
                    client_id=client_id,
                    dates=dates,
                    reason=reason,
                    operator=operator,
                ))

            elif otype == 'lock':
                if d_start is None or not truck:
                    raise OverrideValidationError(
                        f"Row {r} (lock): need date and truck."
                    )
                locks.append(Lock(
                    date=d_start,
                    truck_id=truck,
                    locked_through_stop=int(stop_n or 0),
                    reason=reason,
                    operator=operator,
                ))

            elif otype == 'reading':
                if not client_id or d_start is None or lbs is None:
                    raise OverrideValidationError(
                        f"Row {r} (reading): need client_id, date, lbs."
                    )
                # Promote date → datetime at noon for the as_of timestamp
                # (operator typically reports the reading taken "today",
                # not at a specific minute).
                as_of = datetime.combine(d_start, datetime.min.time()).replace(
                    hour=12,
                )
                readings.append(ManualReading(
                    client_id=client_id,
                    current_lbs=float(lbs),
                    as_of=as_of,
                    reason=reason,
                    operator=operator,
                ))

            elif otype == 'consumption':
                if not client_id or d_start is None or lbs is None:
                    raise OverrideValidationError(
                        f"Row {r} (consumption): need client_id, date, "
                        "and lbs (= rate lbs/day)."
                    )
                consumptions.append(ManualConsumption(
                    client_id=client_id,
                    rate_lbs_per_day=float(lbs),
                    effective_from=d_start,
                    effective_to=d_end,
                    reason=reason,
                    operator=operator,
                ))

        except OverrideValidationError:
            raise
        except Exception as e:
            raise OverrideValidationError(
                f"Row {r}: malformed override — {e}"
            ) from e

    return Overrides(
        pins=tuple(pins),
        forbids=tuple(forbids),
        locks=tuple(locks),
        readings=tuple(readings),
        consumptions=tuple(consumptions),
    )


# ── Internal helpers ─────────────────────────────────────────────────────────

def _find_header_row(ws) -> Optional[int]:
    for r in range(1, min(5, ws.max_row + 1)):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            if str(v).strip().lower() in _HEADER_ALIASES:
                return r
    return None


def _str(row: list[Any], idx: Optional[int]) -> str:
    if idx is None or idx >= len(row) or row[idx] is None:
        return ''
    return str(row[idx]).strip()


def _float(row: list[Any], idx: Optional[int]) -> Optional[float]:
    if idx is None or idx >= len(row) or row[idx] is None:
        return None
    try:
        return float(row[idx])
    except (TypeError, ValueError):
        return None


def _int(row: list[Any], idx: Optional[int]) -> Optional[int]:
    f = _float(row, idx)
    if f is None:
        return None
    return int(round(f))


def _date(row: list[Any], idx: Optional[int]) -> Optional[date]:
    if idx is None or idx >= len(row) or row[idx] is None:
        return None
    v = row[idx]
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    try:
        ts = pd.Timestamp(v)
        if pd.isna(ts):
            return None
        return ts.date()
    except Exception:
        return None


def _date_range(start: date, end: date) -> tuple[date, ...]:
    from datetime import timedelta
    out: list[date] = []
    cursor = start
    while cursor <= end:
        out.append(cursor)
        cursor += timedelta(days=1)
    return tuple(out)
