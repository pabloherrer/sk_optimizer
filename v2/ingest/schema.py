"""
v2.ingest.schema — operator-edited tables from SK_Delivery_System.xlsx.

  • Client_Time_Windows  → dict[client_id, (open_min, close_min)]
                           Times are minutes from shift start (not midnight).
                           When the sheet expresses HH:MM clock times, we
                           subtract the depot shift start (6:00 = 360 min)
                           per the v2 domain convention.

  • Client_Closures      → dict[client_id, tuple[date, ...]]
                           Recurring-DOW closures get expanded across a
                           ±2-year window so the solver can simply look up
                           "is X in the closure set for date D?".

  • Excluded IDs         → frozenset[str]
                           Hardcoded list of Tucson / Flagstaff far-cluster
                           clients (Saturday far-runs, not metro).
"""
from __future__ import annotations

import warnings
from datetime import date, timedelta
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook

warnings.filterwarnings('ignore')


_WORKDAYS_DEFAULT: tuple[str, ...] = ('Tue', 'Wed', 'Thu', 'Fri', 'Sat')
_WEEKDAY_NAMES: tuple[str, ...] = (
    'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun',
)


def load_depot_config(input_file: Path) -> dict:
    """
    Read the operator-editable Depot sheet (key-value pairs).

    Returns a dict with normalized keys:
      shift_start_min   — minutes since midnight (e.g., 360 for 06:00)
      shift_end_min     — minutes since midnight (e.g., 960 for 16:00)
      morning_load_min  — depot loading time before truck leaves
      evening_unload_min — depot unloading time on return
      depot_lat / depot_lon — depot coordinates
      work_days         — tuple of weekday abbreviations

    These OVERRIDE corresponding values in fleet.yaml. The Excel is operator
    truth; YAML is engineering default.
    """
    wb = load_workbook(str(input_file), data_only=True)
    if 'Depot' not in wb.sheetnames:
        return {}
    ws = wb['Depot']

    def _parse_hhmm(v) -> int:
        """'06:00' or '6:00 AM' or '16:00:00' or datetime.time or 6 → minutes since midnight."""
        if v is None:
            return 360  # 6 AM fallback
        # datetime.time / datetime.datetime
        if hasattr(v, 'hour') and hasattr(v, 'minute'):
            return int(v.hour) * 60 + int(v.minute)
        if isinstance(v, (int, float)):
            return int(v) * 60 if v < 24 else int(v)
        s = str(v).strip().upper()
        # Strip AM/PM if present
        is_pm = 'PM' in s
        is_am = 'AM' in s
        s = s.replace('AM', '').replace('PM', '').strip()
        if ':' in s:
            parts = s.split(':')
            h = int(parts[0])
            m = int(parts[1]) if len(parts) > 1 else 0
            # parts[2] (seconds) ignored
            if is_pm and h < 12:
                h += 12
            if is_am and h == 12:
                h = 0
            return h * 60 + m
        return int(s) * 60

    cfg: dict = {}
    for r in range(2, ws.max_row + 1):
        key = ws.cell(r, 1).value
        val = ws.cell(r, 2).value
        if key is None:
            continue
        key = str(key).strip()
        if key in ('Key', 'Field'):
            continue
        if key == 'Depot_Lat':
            cfg['depot_lat'] = float(val)
        elif key == 'Depot_Lon':
            cfg['depot_lon'] = float(val)
        elif key in ('Shift_Start_HHMM', 'Shift_Start'):
            cfg['shift_start_min'] = _parse_hhmm(val)
        elif key in ('Shift_End_HHMM', 'Shift_End'):
            cfg['shift_end_min'] = _parse_hhmm(val)
        elif key == 'Morning_Load_Min':
            cfg['morning_load_min'] = int(val or 0)
        elif key == 'Evening_Unload_Min':
            cfg['evening_unload_min'] = int(val or 0)
        elif key == 'Work_Days':
            cfg['work_days'] = tuple(s.strip() for s in str(val).split(',') if s.strip())
    return cfg


# Hardcoded set of clients on Tucson/Flagstaff far-cluster Saturday routes.
# These are NOT routed by the metro optimizer.
_EXCLUDED_FAR_CLUSTER_IDS: frozenset[str] = frozenset({
    # Flagstaff
    '11005', '12021', '15032', '15004',
    # Prescott (en route to Flagstaff)
    '16052', '20089',
    # New River
    '18036',
    # Tucson / Casa Grande
    '1057', '10012', '15033', '15028', '15021', '16027',
})


def load_time_windows(
    input_file: Path,
    shift_start_min: int = 360,
) -> dict[str, tuple[int, int]]:
    """
    Read Client_Time_Windows sheet → dict[client_id, (open, close)].

    Times in the returned tuple are minutes-from-shift-start
    (NOT minutes-from-midnight). Default shift_start = 6:00 AM (360 min).

    The sheet may carry per-DOW windows; for v2 we collapse to a single
    earliest-open / latest-close envelope per client. This matches how
    `Client.time_window_min` is defined (one window per client; daily
    DOW handling is delegated to a future enhancement).
    """
    input_file = Path(input_file)
    if not input_file.exists():
        return {}
    try:
        wb = load_workbook(str(input_file), data_only=True)
    except Exception:
        return {}
    if 'Client_Time_Windows' not in wb.sheetnames:
        return {}

    ws = wb['Client_Time_Windows']
    by_client: dict[str, list[tuple[int, int]]] = {}

    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or not row[0]:
            continue
        cid = _normalize_id(row[0])
        if not cid or cid.startswith('EXAMPLE_'):
            continue

        # New schema: col 2 = customer, col 3 = DOW, col 4 = open, col 5 = close
        # Legacy schema: col 2 = DOW, col 3 = open, col 4 = close
        col3 = row[2] if len(row) > 2 else None
        col3_str = str(col3).strip() if col3 is not None else ''
        if ':' in col3_str and len(row) > 3 and row[3] and ':' in str(row[3]):
            # legacy
            open_raw = str(row[2]).strip() if row[2] else ''
            close_raw = str(row[3]).strip() if len(row) > 3 and row[3] else ''
        else:
            open_raw = (str(row[3]).strip()
                        if len(row) > 3 and row[3] else '')
            close_raw = (str(row[4]).strip()
                         if len(row) > 4 and row[4] else '')
        if not (open_raw and close_raw):
            continue
        try:
            open_min = _time_to_min(open_raw)
            close_min = _time_to_min(close_raw)
        except ValueError:
            continue

        by_client.setdefault(cid, []).append((open_min, close_min))

    out: dict[str, tuple[int, int]] = {}
    for cid, windows in by_client.items():
        # Envelope: earliest open, latest close (relative to midnight).
        earliest = min(o for o, _ in windows)
        latest = max(c for _, c in windows)
        # Shift to "from shift start". Clamp at 0 (a client opening BEFORE
        # the shift starts means "available from minute 0").
        open_rel = max(0, earliest - shift_start_min)
        close_rel = max(0, latest - shift_start_min)
        if close_rel <= open_rel:
            continue
        out[cid] = (open_rel, close_rel)

    return out


def load_closures(
    input_file: Path,
    horizon_window_years: int = 2,
) -> dict[str, tuple[date, ...]]:
    """
    Read Client_Closures sheet → dict[client_id, tuple[date, ...]].

    Date-range closures (Start_Date..End_Date) expand to one date per day.
    Recurring-weekday closures (e.g. 'Tue' every week) are expanded across
    ±horizon_window_years from today.

    For each client_id, the returned tuple is sorted, deduplicated.
    """
    input_file = Path(input_file)
    if not input_file.exists():
        return {}
    try:
        wb = load_workbook(str(input_file), data_only=True)
    except Exception:
        return {}
    if 'Client_Closures' not in wb.sheetnames:
        return {}

    ws = wb['Client_Closures']
    today = date.today()
    window_start = today - timedelta(days=365 * horizon_window_years)
    window_end = today + timedelta(days=365 * horizon_window_years)

    by_client: dict[str, set[date]] = {}

    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or not row[0]:
            continue
        cid = _normalize_id(row[0])
        if not cid or cid.startswith('EXAMPLE_'):
            continue

        # Detect schema
        col2 = row[1] if len(row) > 1 else None
        new_schema = isinstance(col2, str) or col2 is None
        if new_schema:
            rec_dow_raw = (str(row[2]).strip()
                           if len(row) > 2 and row[2] else '')
            start_raw = row[3] if len(row) > 3 else None
            end_raw = row[4] if len(row) > 4 else None
        else:
            rec_dow_raw = ''
            start_raw = row[1]
            end_raw = row[2]

        rec_dow = rec_dow_raw[:3].title() if rec_dow_raw else ''
        if rec_dow not in ('', *_WEEKDAY_NAMES):
            rec_dow = ''

        target = by_client.setdefault(cid, set())

        if rec_dow:
            dow_idx = _WEEKDAY_NAMES.index(rec_dow)
            cursor = window_start
            # Walk forward day by day. Cheap enough at ±2 years (~1500 days).
            while cursor <= window_end:
                if cursor.weekday() == dow_idx:
                    target.add(cursor)
                cursor += timedelta(days=1)
            continue

        # Date-range closure
        start_d = _to_date(start_raw)
        end_d = _to_date(end_raw)
        if start_d is None or end_d is None:
            continue
        if end_d < start_d:
            continue
        cursor = start_d
        while cursor <= end_d:
            target.add(cursor)
            cursor += timedelta(days=1)

    return {
        cid: tuple(sorted(dates))
        for cid, dates in by_client.items()
        if dates
    }


def load_excluded_ids(input_file: Path) -> frozenset[str]:
    """
    Return the immutable set of client_ids on the alternating Saturday
    far-cluster routes (Tucson / Flagstaff / Prescott / New River).

    Currently this is a fixed list maintained in code — it changes only
    when SK adds or drops a far-cluster account, which is rare and
    requires a code change anyway because the Saturday rotation logic
    is hardcoded elsewhere.

    The `input_file` argument is accepted for future compatibility (we
    may move this list into the workbook), but is not used today.
    """
    _ = input_file  # noqa: F841  reserved for future sheet-driven version
    return _EXCLUDED_FAR_CLUSTER_IDS


# ── Internal helpers ─────────────────────────────────────────────────────────

def _normalize_id(raw) -> str:
    """Mirror of v2.ingest.excel._normalize_id."""
    if raw is None:
        return ''
    if isinstance(raw, float):
        if raw != raw:
            return ''
        if raw.is_integer():
            return str(int(raw))
        return str(raw).strip().upper()
    if isinstance(raw, int):
        return str(raw)
    s = str(raw).strip()
    if not s:
        return ''
    try:
        f = float(s)
        if f.is_integer():
            return str(int(f))
    except (ValueError, TypeError):
        pass
    return s.upper()


def _time_to_min(time_str: str) -> int:
    parts = time_str.split(':')
    if len(parts) != 2:
        raise ValueError(f"Invalid time format: {time_str}")
    h = int(parts[0])
    m = int(parts[1])
    return h * 60 + m


def _to_date(v) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, date) and not hasattr(v, 'hour'):
        return v
    try:
        ts = pd.Timestamp(v)
        if pd.isna(ts):
            return None
        return ts.date()
    except Exception:
        return None
