"""
v2.ingest.excel — Excel readers for Client_List and Delivery_Log.

Reads SK_Delivery_System.xlsx and produces:
  • A frozen tuple[Client, ...] from Client_List
  • A pandas DataFrame of deliveries (forecaster input)

ID normalisation rule:
  • Pure-numeric IDs are stored as their canonical integer string ("4015")
  • Alphanumeric IDs are uppercased verbatim ("4015a" → "4015A")
This must match across Client_List, Delivery_Log, Query (Anova), schemas,
and overrides — otherwise rows silently fail to join.
"""
from __future__ import annotations

import warnings
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook

from v2.domain.client import Client

warnings.filterwarnings('ignore')


# ── Canonical product mapping (hardcoded — does not depend on legacy config) ─

_PRODUCTS_CANONICAL: tuple[str, ...] = ('CANOLA', 'FRYERS CHOICE')

_PRODUCT_ALIASES: dict[str, str] = {
    'CANOLA OIL':          'CANOLA',
    'CANOLA':              'CANOLA',
    '100% CANOLA':         'CANOLA',
    'FRYERS CHOICE BLEND': 'FRYERS CHOICE',
    'FRYERS CHOICE':       'FRYERS CHOICE',
    "FRYER'S CHOICE":      'FRYERS CHOICE',
    'SOYBEAN OIL':         'CANOLA',
    'VEGETABLE OIL':       'FRYERS CHOICE',
}

_DO_NOT_SCHEDULE_TRUE: frozenset[str] = frozenset({'Y', 'YES', 'TRUE', '1'})


# ── Public API ───────────────────────────────────────────────────────────────

def load_clients(input_file: Path) -> tuple[Client, ...]:
    """
    Read the Client_List sheet starting at row 4 and return a tuple of
    frozen Client records.

    Filters out:
      • Header rows where col A == 'ID'
      • Rows with no Customer name
      • Rows missing Lat or Lon (unroutable)

    Prints summary:  "Clients loaded: N (routable: M)"
    where routable = clients that have valid Lat AND Lon.
    """
    input_file = Path(input_file)
    wb = load_workbook(str(input_file), data_only=True)
    ws = wb['Client_List']

    seen_ids: set[str] = set()
    out: list[Client] = []
    total = 0
    routable = 0

    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or row[0] is None:
            continue
        # Skip header / spacer rows
        if str(row[0]).strip().upper() == 'ID':
            continue
        if len(row) < 2 or not row[1]:
            continue

        cid = _normalize_id(row[0])
        if not cid:
            continue
        if cid in seen_ids:
            # Duplicate: keep first occurrence
            continue
        seen_ids.add(cid)

        customer = str(row[1]).strip()
        lat = _to_float(row[7] if len(row) > 7 else None)
        lon = _to_float(row[8] if len(row) > 8 else None)
        tank = _to_float(row[9] if len(row) > 9 else None)
        product_raw = (str(row[10]).strip().upper()
                       if len(row) > 10 and row[10] else '')

        # Skip rows with no GPS — not routable, can't be in problem
        if lat is None or lon is None:
            total += 1
            continue
        if tank is None or tank <= 0:
            total += 1
            continue

        product = _PRODUCT_ALIASES.get(product_raw, _PRODUCTS_CANONICAL[0])

        # Map operator/notes columns
        service_override = _to_int(row[11]) if len(row) > 11 else None
        address = _build_address(
            row[4] if len(row) > 4 else None,
            row[5] if len(row) > 5 else None,
        )
        phone = (str(row[13]).strip()
                 if len(row) > 13 and row[13] else '')
        notes = (str(row[15]).strip()
                 if len(row) > 15 and row[15] else '')
        do_not_schedule = _parse_bool(
            row[16] if len(row) > 16 else None
        )

        client = Client(
            id=cid,
            customer=customer,
            lat=float(lat),
            lon=float(lon),
            tank_capacity_lbs=int(round(tank)),
            product=product,
            do_not_schedule=do_not_schedule,
            excluded=False,                 # set later by build_problem
            address=address,
            phone=phone,
            notes=notes,
            service_min_override=service_override,
            time_window_min=None,           # set later from Client_Time_Windows
            closed_dates=(),                # set later from Client_Closures
        )
        out.append(client)
        total += 1
        routable += 1

    print(f"  Clients loaded: {total} (routable: {routable})")
    return tuple(out)


def load_deliveries(input_file: Path) -> pd.DataFrame:
    """
    Read the Delivery_Log sheet. Returns a DataFrame with columns:
        Date, Customer, Tank_lbs, Qty_lbs, Is_Placeholder

    Rows are dropped if:
      • Date is missing or unparseable
      • Qty ≤ 0 or non-numeric
      • Customer ID is missing

    Customer column holds the resolved name (from Client_List). If a delivery
    row references an ID not in Client_List, the raw ID string is used.
    """
    input_file = Path(input_file)
    wb = load_workbook(str(input_file), data_only=True)

    # Build ID → Customer-name lookup. Use the same normalisation as
    # load_clients so '4015A' on both sheets joins.
    id_to_name: dict[str, str] = {}
    cl = wb['Client_List']
    for r in cl.iter_rows(min_row=4, values_only=True):
        if not r or r[0] is None or len(r) < 2 or not r[1]:
            continue
        if str(r[0]).strip().upper() == 'ID':
            continue
        cid = _normalize_id(r[0])
        if cid:
            id_to_name[cid] = str(r[1]).strip()

    ws = wb['Delivery_Log']
    records: list[dict] = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or row[0] is None:
            continue
        qty = row[3] if len(row) > 3 else None
        if not isinstance(qty, (int, float)) or qty <= 0:
            continue

        raw_date = row[0]
        try:
            d = pd.Timestamp(
                raw_date.date() if hasattr(raw_date, 'date') else raw_date
            )
        except Exception:
            continue
        if pd.isna(d):
            continue

        if len(row) < 2 or row[1] is None:
            continue
        cid = _normalize_id(row[1])
        if not cid:
            continue
        customer = id_to_name.get(cid, cid)

        tank_lbs = _to_float(row[7]) if len(row) > 7 else None

        records.append({
            'Date':           d,
            'Customer':       customer,
            'Tank_lbs':       tank_lbs,
            'Qty_lbs':        float(qty),
            # 200-lb is the human-entered "I delivered something but don't
            # know the amount yet" placeholder. Forecaster must exclude
            # these from rate computation.
            'Is_Placeholder': float(qty) == 200.0,
        })

    df = pd.DataFrame(records,
                      columns=['Date', 'Customer', 'Tank_lbs',
                               'Qty_lbs', 'Is_Placeholder'])
    if not df.empty:
        df = df[df['Date'] >= '2020-01-01'].copy()
        df = df.sort_values(['Customer', 'Date']).reset_index(drop=True)
    return df


# ── Internal helpers ─────────────────────────────────────────────────────────

def _normalize_id(raw) -> str:
    """
    Canonical ID form:
      • Pure-numeric (1234, 1234.0)  → "1234"
      • Alphanumeric ("4015a")      → "4015A"  (uppercased, stripped)
      • None / empty                 → ""
    """
    if raw is None:
        return ''
    if isinstance(raw, float):
        if raw != raw:  # NaN
            return ''
        if raw.is_integer():
            return str(int(raw))
        return str(raw).strip().upper()
    if isinstance(raw, int):
        return str(raw)
    s = str(raw).strip()
    if not s:
        return ''
    # If it looks like a float string ("4015.0"), strip the trailing .0
    try:
        f = float(s)
        if f.is_integer():
            return str(int(f))
    except (ValueError, TypeError):
        pass
    return s.upper()


def _to_float(v) -> Optional[float]:
    if v is None:
        return None
    try:
        f = float(v)
        if f != f:  # NaN
            return None
        return f
    except (ValueError, TypeError):
        return None


def _to_int(v) -> Optional[int]:
    f = _to_float(v)
    if f is None:
        return None
    return int(round(f))


def _build_address(street, city) -> str:
    parts = [str(x).strip() for x in (street, city) if x]
    return ', '.join(parts) + ' AZ' if parts else ''


def _parse_bool(v) -> bool:
    if v is None:
        return False
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v == 1
    s = str(v).strip().upper()
    return s in _DO_NOT_SCHEDULE_TRUE
