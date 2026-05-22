"""
v2.ingest.anova — Anova sensor readings from the Query sheet.

The Query sheet is populated by Power Query from the Azure API. Customer
names look like 'HAR - 8031 - HAROLDS' — the SK client ID is embedded in
the middle. We extract it with a regex.

Output dict keyed by client_id (canonical normalised form), value is a
plain dict with these fields:
    level_lbs : float           — most recent observation
    timestamp : pd.Timestamp    — when the sensor read it
    age_hours : float           — hours between timestamp and now
    source    : 'sensor' | 'sensor-projected' | 'stale'
                  ≤ 24h  → sensor
                  24-72h → sensor-projected   (consumer projects forward)
                  > 72h  → stale              (consumer falls back to estimate)
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

_FRESH_HOURS = 24.0
_STALE_HOURS = 72.0
_ID_RE = re.compile(r'\b(\d{1,6})\b')


def load_anova_readings(input_file: Path) -> dict[str, dict[str, Any]]:
    """
    Read the Query sheet from the input Excel and return per-client
    sensor readings.

    Returns
    -------
    {
      client_id: {
        'level_lbs': float,
        'timestamp': pd.Timestamp,   # naive (no tz)
        'age_hours': float,
        'source':    'sensor' | 'sensor-projected' | 'stale',
      }
    }

    Empty dict if the Query sheet does not exist or has no parseable rows.
    Rows with level ≤ 0 or missing client_id are silently skipped.
    """
    input_file = Path(input_file)
    if not input_file.exists():
        return {}

    try:
        wb = load_workbook(str(input_file), data_only=True)
    except Exception:
        return {}

    if 'Query' not in wb.sheetnames:
        return {}

    ws = wb['Query']
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if not headers:
        return {}
    col_idx = {h: i for i, h in enumerate(headers) if h is not None}

    # Required columns
    if 'customer' not in col_idx or 'display_value' not in col_idx:
        return {}

    cust_i = col_idx['customer']
    val_i = col_idx['display_value']
    ts_i = col_idx.get('timestamp', None)
    if ts_i is None:
        ts_i = col_idx.get('received_at', None)

    now = pd.Timestamp.now(tz='UTC').tz_localize(None)
    readings: dict[str, dict[str, Any]] = {}

    for r in range(2, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if not any(v is not None for v in row):
            continue

        customer = row[cust_i] if cust_i < len(row) else None
        level_raw = row[val_i] if val_i < len(row) else None
        ts_raw = row[ts_i] if (ts_i is not None and ts_i < len(row)) else None

        if not customer:
            continue
        try:
            level = float(level_raw)
        except (TypeError, ValueError):
            continue
        if level <= 0:
            continue

        m = _ID_RE.search(str(customer))
        if not m:
            continue
        cid = m.group(1)

        # Parse timestamp; missing → treat as now (age 0)
        if ts_raw is None:
            ts_naive = now
        else:
            try:
                ts = pd.Timestamp(ts_raw)
                ts_naive = ts.tz_localize(None) if ts.tz else ts
            except Exception:
                ts_naive = now

        age_h = max((now - ts_naive).total_seconds() / 3600.0, 0.0)
        if age_h <= _FRESH_HOURS:
            source = 'sensor'
        elif age_h <= _STALE_HOURS:
            source = 'sensor-projected'
        else:
            source = 'stale'

        prev = readings.get(cid)
        if prev is None or ts_naive > prev['timestamp']:
            readings[cid] = {
                'level_lbs': round(float(level), 1),
                'timestamp': ts_naive,
                'age_hours': age_h,
                'source':    source,
            }

    return readings
