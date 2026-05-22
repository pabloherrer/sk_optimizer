"""
v2.reporting.excel — Excel workbook writer (rebuilt).

The dispatcher's main artifact. Built around feedback that the old layout
hid the math (Tank_Capacity was missing so a 945-lb refill into a 1050-lb
tank looked like "100% fill"), lacked driver-friendly mileage between
stops, and dropped the v1 stockout-risk section at the bottom of each
PRINT page.

Sheet order:
    Summary
    Today's_Plan          (the committed days — driver targets)
    Week_Outlook          (all horizon, color-coded by status)
    At_Risk               (DTE ≤ 5 anywhere in horizon, including deferred)
    Deferred              (clients not scheduled + why)
    PRINT <Day1> ...      (one landscape sheet per delivery day)
    Diagnostics

Phones come from `problem.clients` when a ProblemInstance is supplied;
otherwise the phone column is blank — drivers still get everything else.
"""
from __future__ import annotations
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.worksheet import Worksheet


# ─────────────────────────────────────────────────────────────────────────────
# Visual discipline
# ─────────────────────────────────────────────────────────────────────────────

# Truck colors — must match map.py so printouts and map agree at a glance.
TRUCK_HEX: Dict[str, str] = {
    'Truck2': '1A6FAF',  # blue
    'Truck9': 'C0392B',  # red
}

# Urgency tints — very light so they read after photocopying.
URGENCY_FILL: Dict[str, str] = {
    'stockout': 'F4B5B5',
    'critical': 'F9CFA6',
    'urgent':   'FCE9A8',
    'normal':   '',
}

THIN = Side(style='thin', color='B0B0B0')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADER_FILL = PatternFill('solid', fgColor='2C3E50')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
SECTION_FONT = Font(bold=True, size=13, color='2C3E50')
WARNING_FONT = Font(bold=True, color='C0392B')
KPI_LABEL_FONT = Font(bold=True, size=9, color='7F8C8D')
KPI_VALUE_FONT = Font(bold=True, size=18, color='2C3E50')

PANEL_FILL = PatternFill('solid', fgColor='F7F9FB')
PANEL_BORDER = Border(
    left=Side(style='thin', color='D5DBDF'),
    right=Side(style='thin', color='D5DBDF'),
    top=Side(style='thin', color='D5DBDF'),
    bottom=Side(style='thin', color='D5DBDF'),
)

# Stockout-risk section header background (v1 amber)
CAUTION_HEX = '993300'


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _minutes_to_hhmm(shift_start_min: int, minute_offset: int) -> str:
    total = shift_start_min + minute_offset
    h = (total // 60) % 24
    m = total % 60
    suffix = 'AM' if h < 12 else 'PM'
    h12 = h % 12 or 12
    return f'{h12}:{m:02d} {suffix}'


def _safe_truck_color(truck_id: str) -> str:
    return TRUCK_HEX.get(truck_id, '7F8C8D')


def _write_header_row(
    ws: Worksheet,
    row: int,
    headers: List[str],
    start_col: int = 1,
    fill_hex: str = '2C3E50',
) -> None:
    fill = PatternFill('solid', fgColor=fill_hex)
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=start_col + i, value=h)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER


def _autosize_columns(
    ws: Worksheet,
    min_width: int = 9,
    max_width: int = 48,
    overrides: Optional[Dict[int, int]] = None,
) -> None:
    overrides = overrides or {}
    widths: Dict[int, int] = {}
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if cell.value is None:
                continue
            text = str(cell.value)
            # Use the longest single line — wrap_text handles the rest.
            longest = max((len(s) for s in text.splitlines()), default=0)
            widths[cell.column] = max(widths.get(cell.column, min_width), longest + 2)
    for col, w in widths.items():
        if col in overrides:
            ws.column_dimensions[get_column_letter(col)].width = overrides[col]
        else:
            ws.column_dimensions[get_column_letter(col)].width = min(w, max_width)


def _committed_dates(plan) -> List[date]:
    if not plan.horizon_dates:
        return []
    n = min(plan.commit_days, len(plan.horizon_dates))
    return list(plan.horizon_dates[:n])


def _routes_on(plan, d: date) -> List:
    return [r for (rd, _), r in sorted(plan.routes.items()) if rd == d]


def _shift_start_min(plan) -> int:
    return getattr(plan, 'shift_start_min', 360)


def _phone_lookup(problem) -> Dict[str, str]:
    """client_id → phone (empty string if unknown / no problem given)."""
    if problem is None:
        return {}
    out: Dict[str, str] = {}
    for c in getattr(problem, 'clients', ()):
        out[c.id] = c.phone or ''
    return out


def _fill_pct(stop) -> float:
    """% of tank that THIS delivery represents.
    delivery / tank_capacity × 100. A 70% fill = we delivered 70% of the
    tank's capacity (driver arrived with ~30% remaining).
    NOT level_after / cap (which would always be 100% when we fill to full).
    """
    return (stop.delivery_lbs / stop.tank_capacity_lbs * 100
            if stop.tank_capacity_lbs else 0)


# ─────────────────────────────────────────────────────────────────────────────
# Summary — operator dashboard
# ─────────────────────────────────────────────────────────────────────────────

def _kpi_card(
    ws: Worksheet,
    row: int,
    col: int,
    label: str,
    value: str,
    color: str = '2C3E50',
) -> None:
    """One 2-row tall KPI tile."""
    label_cell = ws.cell(row=row, column=col, value=label.upper())
    label_cell.font = KPI_LABEL_FONT
    label_cell.alignment = Alignment(horizontal='center')
    label_cell.fill = PANEL_FILL
    label_cell.border = PANEL_BORDER

    val_cell = ws.cell(row=row + 1, column=col, value=value)
    val_cell.font = Font(bold=True, size=18, color=color)
    val_cell.alignment = Alignment(horizontal='center', vertical='center')
    val_cell.fill = PANEL_FILL
    val_cell.border = PANEL_BORDER
    ws.row_dimensions[row + 1].height = 28


def _build_summary(ws: Worksheet, plan, problem=None) -> None:
    ws.title = 'Summary'
    ws.sheet_view.showGridLines = False

    # Title + run metadata
    ws.cell(row=1, column=1, value='SK ROUTE PLAN').font = Font(
        bold=True, size=20, color='2C3E50'
    )
    ws.cell(row=2, column=1, value=f'{plan.today.strftime("%A, %B %d, %Y")} · run {plan.run_id}').font = Font(
        size=10, color='7F8C8D'
    )
    ws.cell(row=3, column=1, value=(
        f'Horizon {plan.horizon_dates[0].strftime("%b %d")} – '
        f'{plan.horizon_dates[-1].strftime("%b %d")}  '
        f'·  {plan.commit_days} committed days  '
        f'·  generated {plan.generated_at.strftime("%Y-%m-%d %H:%M")}'
    )).font = Font(size=10, color='7F8C8D')

    # ── KPI cards ───────────────────────────────────────────────────────
    ot_days = sum(1 for r in plan.routes.values() if r.overtime_minutes > 0)
    deferred_n = len(plan.deferred)
    n_truckdays = sum(1 for r in plan.routes.values() if r.stops)
    kpis = [
        ('Stops',       f'{plan.total_stops:,}',                   '2C3E50'),
        ('Lbs',         f'{round(plan.total_lbs_delivered):,}',    '1A6FAF'),
        ('Miles',       f'{round(plan.total_miles):,.0f}',         '2C3E50'),
        ('Avg Fill',    f'{round(plan.avg_fill_pct, 1)}%',         '27AE60'),
        ('OT Days',     str(ot_days),                              ('C0392B' if ot_days else '7F8C8D')),
        ('Deferred',    str(deferred_n),                           ('C0392B' if deferred_n else '7F8C8D')),
        ('Truck-Days',  str(n_truckdays),                          '2C3E50'),
    ]
    kpi_row = 5
    for i, (label, val, color) in enumerate(kpis):
        col = 1 + i * 2
        _kpi_card(ws, kpi_row, col, label, val, color=color)
        ws.column_dimensions[get_column_letter(col)].width = 12
        ws.column_dimensions[get_column_letter(col + 1)].width = 2

    # ── Per truck-day table ─────────────────────────────────────────────
    table_row = kpi_row + 4
    ws.cell(row=table_row - 1, column=1, value='Per-Truck-Day Detail').font = SECTION_FONT
    headers = [
        'Date', 'Truck', 'Stops', 'Lbs', 'Cap %', 'Shift %', 'OT (min)',
        'Avg Fill %', 'Miles', 'Depart', 'Return',
    ]
    _write_header_row(ws, table_row, headers)
    target_minutes = getattr(plan, 'shift_target_min', 480)
    shift_start = _shift_start_min(plan)

    r = table_row + 1
    for (d, truck_id), route in sorted(plan.routes.items()):
        if not route.stops:
            continue
        avg_fill = sum(_fill_pct(s) for s in route.stops) / len(route.stops)
        cap_pct = route.cap_pct or (route.total_load_lbs / 10000.0 * 100)
        shift_pct = route.total_minutes / target_minutes * 100 if target_minutes else 0
        depart = _minutes_to_hhmm(shift_start, route.depart_depot_min)
        ret = _minutes_to_hhmm(shift_start, route.return_depot_min)
        ws.cell(row=r, column=1, value=d.strftime('%a %b %d'))
        tcell = ws.cell(row=r, column=2, value=truck_id)
        tcell.fill = PatternFill('solid', fgColor=_safe_truck_color(truck_id))
        tcell.font = Font(bold=True, color='FFFFFF')
        ws.cell(row=r, column=3, value=len(route.stops))
        ws.cell(row=r, column=4, value=round(route.total_load_lbs)).number_format = '#,##0'
        ws.cell(row=r, column=5, value=round(cap_pct, 1)).number_format = '0.0"%"'
        ws.cell(row=r, column=6, value=round(shift_pct, 1)).number_format = '0.0"%"'
        ot_cell = ws.cell(row=r, column=7, value=route.overtime_minutes or '')
        if route.overtime_minutes:
            ot_cell.font = WARNING_FONT
        ws.cell(row=r, column=8, value=round(avg_fill, 1)).number_format = '0.0"%"'
        ws.cell(row=r, column=9, value=round(route.total_miles, 1)).number_format = '0.0'
        ws.cell(row=r, column=10, value=depart)
        ws.cell(row=r, column=11, value=ret)
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).alignment = Alignment(horizontal='center')
        r += 1

    # ── Capacity warnings ───────────────────────────────────────────────
    r += 2
    ws.cell(row=r, column=1, value='Capacity Warnings').font = SECTION_FONT
    r += 1
    if plan.capacity_warnings:
        for w in plan.capacity_warnings:
            cell = ws.cell(row=r, column=1, value=str(w))
            cell.font = WARNING_FONT
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
            r += 1
    else:
        ws.cell(row=r, column=1, value='(none — fleet has headroom)').font = Font(italic=True, color='808080')
        r += 1

    # ── Solver status ───────────────────────────────────────────────────
    r += 2
    ws.cell(row=r, column=1, value='Solver').font = SECTION_FONT
    r += 1
    _write_header_row(ws, r, ['Status', 'Solve Seconds', 'Objective Cost'])
    r += 1
    ws.cell(row=r, column=1, value=plan.solver_status)
    ws.cell(row=r, column=2, value=round(plan.solve_seconds, 2)).number_format = '0.00'
    ws.cell(row=r, column=3, value=round(plan.objective_cost_dollars, 2)).number_format = '$#,##0.00'
    for c in range(1, 4):
        ws.cell(row=r, column=c).border = BORDER
        ws.cell(row=r, column=c).alignment = Alignment(horizontal='center')

    # Column widths — wider Date / Truck / numeric stack
    _autosize_columns(ws, overrides={1: 14, 2: 9, 10: 11, 11: 11})


# ─────────────────────────────────────────────────────────────────────────────
# Today's_Plan / Week_Outlook — per-stop tables
# ─────────────────────────────────────────────────────────────────────────────

PER_STOP_HEADERS = [
    'Date', 'Truck', 'Seq', 'Client_ID', 'Customer', 'Address', 'Phone',
    'ETA', 'Dist_to_mi', 'Refill_Lbs',
    'Tank_Capacity',
    # TODAY's actual level (from Anova/state at plan time) — what the
    # dispatcher would see if they checked the tank RIGHT NOW.
    'Today_Lbs', 'Today_%',
    # Days from today until truck arrives at this stop.
    'Days_Out',
    # Projected level AT ARRIVAL (= Today_Lbs − Days_Out × rate, clamped).
    'Tank_Before', 'Tank_After', 'Fill_%',
    'DTE', 'Urgency', 'Notes',
]


def _write_per_stop_row(
    ws: Worksheet,
    r: int,
    plan,
    d: date,
    route,
    stop,
    phones: Dict[str, str],
    include_status: bool,
    committed: set,
) -> None:
    shift_start = _shift_start_min(plan)
    eta = _minutes_to_hhmm(shift_start, stop.arrival_min)
    fill = _fill_pct(stop)
    phone = phones.get(stop.client_id, '')
    today_lbs = getattr(stop, 'current_lbs_today', stop.level_at_arrival_lbs)
    today_pct = (today_lbs / stop.tank_capacity_lbs * 100
                 if stop.tank_capacity_lbs else 0)
    days_out = getattr(stop, 'days_to_arrival', 0)
    vals = [
        d.strftime('%a %b %d'),
        route.truck_id,
        stop.sequence,
        stop.client_id,
        stop.customer,
        stop.address,
        phone,
        eta,
        round(stop.travel_miles, 1),
        round(stop.delivery_lbs),
        stop.tank_capacity_lbs,
        round(today_lbs),
        round(today_pct, 1),
        int(days_out),
        round(stop.level_at_arrival_lbs),
        round(stop.level_after_lbs),
        round(fill, 1),
        round(stop.days_until_stockout_at_arrival, 1),
        stop.urgency_tier,
        stop.notes or '',
    ]
    if include_status:
        vals.append('COMMITTED' if d in committed else 'TENTATIVE')

    for i, v in enumerate(vals):
        cell = ws.cell(row=r, column=i + 1, value=v)
        cell.border = BORDER
        cell.alignment = Alignment(vertical='center', wrap_text=(i in (4, 5, 16)))

    ws.cell(row=r, column=5).font = Font(bold=True)
    tcell = ws.cell(row=r, column=2)
    tcell.fill = PatternFill('solid', fgColor=_safe_truck_color(route.truck_id))
    tcell.font = Font(bold=True, color='FFFFFF')
    tcell.alignment = Alignment(horizontal='center')

    # Urgency tint (skip truck column so we don't blow away its color)
    fill_color = URGENCY_FILL.get(stop.urgency_tier, '')
    if fill_color:
        for c in range(1, len(vals) + 1):
            if c == 2:
                continue
            ws.cell(row=r, column=c).fill = PatternFill('solid', fgColor=fill_color)

    # Number formats (1-based column indexes)
    ws.cell(row=r, column=9).number_format = '0.0'      # dist
    ws.cell(row=r, column=10).number_format = '#,##0'   # refill
    ws.cell(row=r, column=11).number_format = '#,##0'   # capacity
    ws.cell(row=r, column=12).number_format = '#,##0'   # tank before
    ws.cell(row=r, column=13).number_format = '#,##0'   # tank after
    ws.cell(row=r, column=14).number_format = '0.0"%"'  # fill
    ws.cell(row=r, column=15).number_format = '0.0'     # DTE

    if include_status:
        status_col = len(vals)
        s_cell = ws.cell(row=r, column=status_col)
        if s_cell.value == 'TENTATIVE':
            s_cell.font = Font(italic=True, color='7F8C8D')


def _build_per_stop_table(
    ws: Worksheet,
    plan,
    dates: List[date],
    phones: Dict[str, str],
    include_status: bool = False,
) -> None:
    headers = list(PER_STOP_HEADERS)
    if include_status:
        headers.append('Status')
    _write_header_row(ws, 1, headers)

    committed = set(_committed_dates(plan))
    r = 2
    for d in dates:
        for route in _routes_on(plan, d):
            for stop in route.stops:
                _write_per_stop_row(
                    ws, r, plan, d, route, stop, phones,
                    include_status, committed,
                )
                r += 1

    if r == 2:
        ws.cell(row=2, column=1, value='(no routes scheduled in this window)').font = (
            Font(italic=True, color='808080')
        )

    ws.freeze_panes = 'A2'
    _autosize_columns(
        ws,
        overrides={
            1: 12,   # Date
            2: 8,    # Truck
            3: 5,    # Seq
            4: 10,   # Client_ID
            5: 32,   # Customer
            6: 36,   # Address
            7: 14,   # Phone
            8: 9,    # ETA
            9: 8,    # Dist_to
            10: 10,  # Refill
            11: 12,  # Capacity
            12: 11,  # Before
            13: 11,  # After
            14: 8,   # Fill %
            15: 6,   # DTE
            16: 9,   # Urgency
            17: 32,  # Notes
        },
    )


def _build_todays_plan(ws: Worksheet, plan, phones: Dict[str, str]) -> None:
    ws.title = "Today's_Plan"
    _build_per_stop_table(ws, plan, _committed_dates(plan), phones, include_status=False)


def _build_week_outlook(ws: Worksheet, plan, phones: Dict[str, str]) -> None:
    ws.title = 'Week_Outlook'
    _build_per_stop_table(ws, plan, list(plan.horizon_dates), phones, include_status=True)


# ─────────────────────────────────────────────────────────────────────────────
# At_Risk — every client at risk anywhere in horizon
# ─────────────────────────────────────────────────────────────────────────────

DTE_AT_RISK_THRESHOLD = 5.0


def _build_at_risk(ws: Worksheet, plan, phones: Dict[str, str], problem=None) -> None:
    ws.title = 'At_Risk'
    headers = [
        'Client_ID', 'Customer', 'Phone', 'DTE_at_Visit',
        'Scheduled_Date', 'Truck', 'Status', 'Reason',
    ]
    _write_header_row(ws, 1, headers)

    # Map client_id → earliest (date, truck, stop) visit
    earliest: Dict[str, Tuple[date, str, object]] = {}
    for (d, truck_id), route in sorted(plan.routes.items()):
        for stop in route.stops:
            if stop.client_id not in earliest:
                earliest[stop.client_id] = (d, truck_id, stop)

    customer_lookup: Dict[str, str] = {}
    if problem is not None:
        for c in problem.clients:
            customer_lookup[c.id] = c.customer

    rows: List[Tuple] = []

    # 1) Scheduled clients with DTE-at-visit <= threshold or urgency != normal
    seen: set = set()
    for cid, (d, tid, stop) in earliest.items():
        risky = (
            stop.urgency_tier in ('stockout', 'critical', 'urgent')
            or stop.days_until_stockout_at_arrival <= DTE_AT_RISK_THRESHOLD
        )
        if risky:
            seen.add(cid)
            rows.append((
                cid, stop.customer, phones.get(cid, ''),
                round(stop.days_until_stockout_at_arrival, 1),
                d.strftime('%a %b %d'), tid, 'SCHEDULED',
                f'urgency={stop.urgency_tier}',
            ))

    # 2) Deferred clients — always include (operator needs to know)
    for cid, reason in sorted(plan.deferred.items()):
        if cid in seen:
            continue
        customer = customer_lookup.get(cid, '(deferred)')
        rows.append((
            cid, customer, phones.get(cid, ''), '',
            '', '', 'DEFERRED', reason,
        ))

    rows.sort(key=lambda row: (
        row[6] == 'DEFERRED',   # deferred sinks to bottom
        9999 if not isinstance(row[3], (int, float)) else row[3],
        row[0],
    ))

    r = 2
    for row in rows:
        for i, v in enumerate(row):
            cell = ws.cell(row=r, column=i + 1, value=v)
            cell.border = BORDER
            cell.alignment = Alignment(vertical='center', wrap_text=(i in (1, 7)))
        ws.cell(row=r, column=2).font = Font(bold=True)
        # Deferred = red, else amber by DTE
        if row[6] == 'DEFERRED':
            tint = URGENCY_FILL['stockout']
        elif isinstance(row[3], (int, float)) and row[3] < 0:
            tint = URGENCY_FILL['stockout']
        elif isinstance(row[3], (int, float)) and row[3] <= 2:
            tint = URGENCY_FILL['critical']
        else:
            tint = URGENCY_FILL['urgent']
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).fill = PatternFill('solid', fgColor=tint)
        r += 1

    if r == 2:
        ws.cell(row=2, column=1, value='(no at-risk clients)').font = Font(italic=True, color='808080')

    ws.freeze_panes = 'A2'
    _autosize_columns(ws, overrides={2: 32, 3: 14, 8: 40})


# ─────────────────────────────────────────────────────────────────────────────
# Deferred
# ─────────────────────────────────────────────────────────────────────────────

def _build_deferred(ws: Worksheet, plan, problem=None) -> None:
    ws.title = 'Deferred'
    headers = ['Client_ID', 'Customer', 'Reason']
    _write_header_row(ws, 1, headers)

    name_lookup: Dict[str, str] = {}
    if problem is not None:
        for c in problem.clients:
            name_lookup[c.id] = c.customer

    r = 2
    for cid, reason in sorted(plan.deferred.items()):
        ws.cell(row=r, column=1, value=cid).border = BORDER
        ws.cell(row=r, column=2, value=name_lookup.get(cid, '')).border = BORDER
        ws.cell(row=r, column=3, value=reason).border = BORDER
        r += 1
    if r == 2:
        ws.cell(row=2, column=1, value='(no deferred clients)').font = Font(italic=True, color='808080')

    ws.freeze_panes = 'A2'
    _autosize_columns(ws, overrides={2: 30, 3: 60})


# ─────────────────────────────────────────────────────────────────────────────
# PRINT sheets — one per delivery day, landscape, with stockout block at bottom
# ─────────────────────────────────────────────────────────────────────────────

PRINT_HEADERS = [
    'Stop#', 'Customer', 'Address', 'Phone', 'ETA',
    'Dist (mi)', 'Refill Lbs', 'Tank Cap', 'Tank Before',
    'Tank After', 'Fill %', 'DTE', 'Notes',
]
PRINT_COL_WIDTHS = {
    1: 6, 2: 30, 3: 30, 4: 14, 5: 9, 6: 8, 7: 10,
    8: 9, 9: 11, 10: 11, 11: 7, 12: 6, 13: 24,
}


def _stockout_block(
    ws: Worksheet,
    row_start: int,
    plan,
    d: date,
    phones: Dict[str, str],
    problem=None,
) -> int:
    """Write the per-day "stockout risk" block at the bottom of a PRINT sheet.

    Only shows clients who are AT RISK ON THIS SPECIFIC DAY (d) — i.e. whose
    tank will be at/near zero on day `d` but who are NOT on day `d`'s route.
    Sorted by current-tank-DTE ascending, capped to a manageable list.

    Earlier version listed every client whose *first scheduled visit* had low
    DTE-at-arrival — which dumped ~25 clients scheduled across the whole
    horizon on every print sheet (some scheduled 10+ days out). The dispatcher
    can't act on those today, so they were noise.
    """
    # Cap: never overwhelm the print sheet. Top N most urgent only.
    MAX_RISKS_SHOWN = 10
    # Day-d DTE threshold to be "at risk right now"
    CURRENT_DTE_THRESHOLD = 3.0

    visited_today: set = set()
    for route in _routes_on(plan, d):
        for stop in route.stops:
            visited_today.add(stop.client_id)

    # Earliest scheduled visit per client (so we can show "Scheduled Wed Jun 03"
    # next to their phone — the dispatcher decides whether to bring it forward).
    earliest: Dict[str, Tuple[date, str, object]] = {}
    for (rd, tid), route in sorted(plan.routes.items()):
        for stop in route.stops:
            if stop.client_id not in earliest:
                earliest[stop.client_id] = (rd, tid, stop)

    # Compute each client's PROJECTED DTE on day `d`. Needs problem state.
    # If problem isn't passed, fall back to the legacy at-arrival DTE so
    # we don't crash — but the per-day risk view requires `problem`.
    today_real = problem.today if problem is not None else d
    days_from_today_to_d = max(0, (d - today_real).days)

    # Map client_id → list of (delivery_date, delivery_lbs) for the WHOLE plan.
    # We use this to credit deliveries that happen on/before day d when we
    # project a client's tank level forward. Without this we'd flag a client
    # that already got a truck visit on Friday as "empty on Wednesday".
    deliveries_by_cid: Dict[str, List[Tuple[date, float]]] = {}
    for (rd, _), route in plan.routes.items():
        for stop in route.stops:
            deliveries_by_cid.setdefault(stop.client_id, []).append(
                (rd, float(stop.delivery_lbs))
            )

    # Clients deferred for non-actionable reasons (far-cluster, DNS, no data,
    # not in matrix) should NOT show up in the at-risk block — the dispatcher
    # can't do anything about them this run. Only flag if the deferral reason
    # would be actionable (e.g. "NOT_NEEDED_THIS_HORIZON" — those CAN be
    # promoted to today's route).
    UNACTIONABLE_DEFER_REASONS = (
        'EXCLUDED', 'DO_NOT_SCHEDULE', 'NOT_IN_MATRIX',
        'INSUFFICIENT_CONSUMPTION_DATA',
    )
    deferred_unactionable: set = {
        cid for cid, reason in plan.deferred.items()
        if any(r in (reason or '') for r in UNACTIONABLE_DEFER_REASONS)
    }

    risks: List[Tuple] = []
    if problem is not None:
        for cid, ts in problem.initial_tanks.items():
            if cid in visited_today:
                continue
            if cid in deferred_unactionable:
                continue
            rate = float(ts.rate_lbs_per_day or 0.0)
            if rate <= 0:
                continue

            # Project tank level forward day-by-day from today to day d.
            # Apply any scheduled delivery on or before each step.
            tank_cap = 0.0
            for c in problem.clients:
                if c.id == cid:
                    tank_cap = float(c.tank_capacity_lbs)
                    break
            level = float(ts.current_lbs)
            cursor = today_real
            crossed_zero = False
            min_level_at_d = level
            client_deliveries = sorted(deliveries_by_cid.get(cid, []))
            while cursor <= d:
                # Apply any delivery that happens on `cursor`
                for dd, qty in client_deliveries:
                    if dd == cursor:
                        level = min(tank_cap or (level + qty), level + qty)
                # Drain one day
                if cursor < d:
                    level = level - rate
                    if level < 0:
                        crossed_zero = True
                cursor += timedelta(days=1)
            level_at_d = max(0.0, level)
            dte_at_d = level_at_d / rate if rate > 0 else 999.0

            # "At risk on day d" criterion: tank actually low ON DAY d.
            # Negative DTE means it would have crossed zero before d, but we
            # clamp display to 0 minimum.
            if dte_at_d > CURRENT_DTE_THRESHOLD:
                continue

            # Find the NEXT scheduled delivery strictly AFTER day d.
            next_visit: Optional[Tuple[date, str]] = None
            prev_visit: Optional[Tuple[date, str]] = None
            for (dd, tk), rte in plan.routes.items():
                for s in rte.stops:
                    if s.client_id == cid:
                        if dd > d:
                            if next_visit is None or dd < next_visit[0]:
                                next_visit = (dd, tk)
                        elif dd <= d:
                            if prev_visit is None or dd > prev_visit[0]:
                                prev_visit = (dd, tk)

            if next_visit is not None:
                rd, tid = next_visit
                days_dry_before_visit = (rd - d).days - 1   # full days w/o oil
                # If next visit is tomorrow — driver is coming, not actionable
                if days_dry_before_visit < 1:
                    continue
                sched_label = rd.strftime('%a %b %d')
                truck_label = tid
                status = 'SCHEDULED'
            elif prev_visit is not None:
                # Client was served earlier in the plan but goes dry again
                # before horizon end with no 2nd visit. Real alarm — they
                # need a follow-up in next horizon.
                rd, tid = prev_visit
                sched_label = f'last {rd.strftime("%a %b %d")}'
                truck_label = tid
                status = 'NEEDS 2ND VISIT'
            else:
                # No visit anywhere in plan and not deferred — truly unscheduled.
                sched_label = '—'
                truck_label = '—'
                status = 'NOT SCHEDULED'

            # Customer name lookup
            customer = ''
            for c in problem.clients:
                if c.id == cid:
                    customer = c.customer
                    break

            risks.append((
                cid, customer, phones.get(cid, ''),
                round(dte_at_d, 1),
                sched_label, truck_label, status,
            ))
    else:
        # Legacy fallback: at-arrival DTE on first scheduled visit
        for cid, (rd, tid, stop) in earliest.items():
            if cid in visited_today: continue
            if stop.days_until_stockout_at_arrival > CURRENT_DTE_THRESHOLD: continue
            if rd <= d: continue
            risks.append((
                cid, stop.customer, phones.get(cid, ''),
                round(stop.days_until_stockout_at_arrival, 1),
                rd.strftime('%a %b %d'), tid, 'SCHEDULED',
            ))

    if not risks:
        return row_start

    # Sort by urgency (lowest DTE first), then ID. Cap.
    risks.sort(key=lambda r_: (
        9999 if not isinstance(r_[3], (int, float)) else r_[3],
        r_[0],
    ))
    truncated = max(0, len(risks) - MAX_RISKS_SHOWN)
    risks = risks[:MAX_RISKS_SHOWN]

    # Section title
    cur = row_start + 1
    title = ws.cell(row=cur, column=1, value=(
        f'AT-RISK ON {d.strftime("%a %b %d").upper()} — CALL THESE CLIENTS IF POSSIBLE'
    ))
    title.font = Font(bold=True, size=13, color='FFFFFF')
    title.fill = PatternFill('solid', fgColor=CAUTION_HEX)
    title.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=13)
    ws.row_dimensions[cur].height = 24
    cur += 1

    subtitle_text = (
        f"Clients projected to be at/near empty on {d.strftime('%a %b %d')} but "
        f"NOT on today's route. Sorted most urgent first."
    )
    if truncated:
        subtitle_text += f' ({truncated} more not shown.)'
    subtitle = ws.cell(row=cur, column=1, value=subtitle_text)
    subtitle.font = Font(italic=True, size=10, color='7F8C8D')
    ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=13)
    cur += 1

    headers = ['Client_ID', 'Customer', 'Phone',
               f'DTE on {d.strftime("%a")}',
               'Scheduled', 'Truck', 'Status']
    _write_header_row(ws, cur, headers, fill_hex=CAUTION_HEX)
    cur += 1

    for row in risks:
        for i, v in enumerate(row):
            cell = ws.cell(row=cur, column=i + 1, value=v)
            cell.border = BORDER
            cell.font = Font(size=11)
            cell.alignment = Alignment(vertical='center')
        ws.cell(row=cur, column=2).font = Font(bold=True, size=11)
        dte = row[3]
        if row[6] == 'NOT SCHEDULED':
            tint = URGENCY_FILL['stockout']
        elif isinstance(dte, (int, float)) and dte <= 0:
            tint = URGENCY_FILL['stockout']
        elif isinstance(dte, (int, float)) and dte <= 1.5:
            tint = URGENCY_FILL['critical']
        else:
            tint = URGENCY_FILL['urgent']
        for c in range(1, len(headers) + 1):
            ws.cell(row=cur, column=c).fill = PatternFill('solid', fgColor=tint)
        ws.row_dimensions[cur].height = 20
        cur += 1
    return cur


def _build_print_day(
    ws: Worksheet,
    plan,
    d: date,
    phones: Dict[str, str],
    problem=None,
) -> None:
    """Driver-friendly per-day printable. Landscape. Stockout block at bottom."""
    ws.title = f'PRINT {d.strftime("%a %b %d")}'[:31]
    ws.sheet_view.showGridLines = False

    # Print setup — landscape Letter, fit to width
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = 1  # Letter
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(
        left=0.4, right=0.4, top=0.5, bottom=0.5, header=0.2, footer=0.2,
    )
    ws.oddHeader.center.text = (
        f'&"Arial,Bold"&14SK Oil Sales — Route Schedule    '
        f'&"Arial"&11{d.strftime("%A %b %d, %Y")}'
    )
    ws.oddFooter.right.text = 'Page &P of &N'
    ws.oddFooter.left.text = f'Run {plan.run_id[:8]}  ·  Generated {plan.generated_at.strftime("%Y-%m-%d %H:%M")}'

    shift_start = _shift_start_min(plan)
    routes = _routes_on(plan, d)
    row = 1
    if not routes:
        cell = ws.cell(row=1, column=1, value=(
            f'{d.strftime("%A %b %d, %Y")} — no trucks dispatched'
        ))
        cell.font = SECTION_FONT
        for col, w in PRINT_COL_WIDTHS.items():
            ws.column_dimensions[get_column_letter(col)].width = w
        return

    for route in routes:
        depart = _minutes_to_hhmm(shift_start, route.depart_depot_min)
        ret = _minutes_to_hhmm(shift_start, route.return_depot_min)
        hrs, mins = divmod(route.total_minutes, 60)
        ot = f' · OT {route.overtime_minutes} min' if route.overtime_minutes else ''
        title = (
            f'{route.truck_id}  ·  {len(route.stops)} stops  ·  '
            f'{round(route.total_load_lbs):,} lbs ({round(route.cap_pct or 0)}% cap)  ·  '
            f'{round(route.total_miles, 1)} mi  ·  '
            f'Depart {depart}  ·  Return ~{ret}  ·  {hrs}h {mins:02d}m{ot}'
        )
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = Font(bold=True, color='FFFFFF', size=14)
        cell.fill = PatternFill('solid', fgColor=_safe_truck_color(route.truck_id))
        cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
        ws.row_dimensions[row].height = 28
        row += 1

        # Compartment manifest (if present)
        comp_lines = []
        if getattr(route, 'compartment_a_product', None) and route.compartment_a_lbs:
            comp_lines.append(
                f'A: {route.compartment_a_product}  {round(route.compartment_a_lbs):,} lbs'
            )
        if getattr(route, 'compartment_b_product', None) and route.compartment_b_lbs:
            comp_lines.append(
                f'B: {route.compartment_b_product}  {round(route.compartment_b_lbs):,} lbs'
            )
        if comp_lines:
            cell = ws.cell(row=row, column=1, value='LOAD: ' + '   ·   '.join(comp_lines))
            cell.font = Font(size=11, bold=True)
            cell.fill = PatternFill('solid', fgColor='ECF0F1')
            cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
            ws.row_dimensions[row].height = 20
            row += 1

        # Header row
        _write_header_row(ws, row, PRINT_HEADERS)
        ws.row_dimensions[row].height = 28
        row += 1

        for stop in route.stops:
            eta = _minutes_to_hhmm(shift_start, stop.arrival_min)
            fill = _fill_pct(stop)
            phone = phones.get(stop.client_id, '')
            vals = [
                stop.sequence,
                stop.customer,
                stop.address,
                phone,
                eta,
                round(stop.travel_miles, 1),
                round(stop.delivery_lbs),
                stop.tank_capacity_lbs,
                round(stop.level_at_arrival_lbs),
                round(stop.level_after_lbs),
                round(fill, 1),
                round(stop.days_until_stockout_at_arrival, 1),
                stop.notes or '',
            ]
            for i, v in enumerate(vals):
                c = ws.cell(row=row, column=i + 1, value=v)
                c.border = BORDER
                c.font = Font(size=11)
                left_align_cols = (2, 3, 13)
                c.alignment = Alignment(
                    horizontal='left' if (i + 1) in left_align_cols else 'center',
                    vertical='center',
                    wrap_text=(i + 1) in (2, 3, 13),
                )
            ws.cell(row=row, column=1).font = Font(bold=True, size=14)
            ws.cell(row=row, column=2).font = Font(bold=True, size=12)
            ws.cell(row=row, column=6).number_format = '0.0'
            ws.cell(row=row, column=7).number_format = '#,##0'
            ws.cell(row=row, column=7).font = Font(bold=True, size=12)
            ws.cell(row=row, column=8).number_format = '#,##0'
            ws.cell(row=row, column=9).number_format = '#,##0'
            ws.cell(row=row, column=10).number_format = '#,##0'
            ws.cell(row=row, column=11).number_format = '0.0"%"'
            ws.cell(row=row, column=12).number_format = '0.0'
            # Urgency row tint
            fill_color = URGENCY_FILL.get(stop.urgency_tier, '')
            if fill_color:
                for c_idx in range(1, len(vals) + 1):
                    ws.cell(row=row, column=c_idx).fill = PatternFill('solid', fgColor=fill_color)
            ws.row_dimensions[row].height = 26
            row += 1

        row += 2  # spacer between trucks

    # Stockout-risk section
    row = _stockout_block(ws, row, plan, d, phones, problem=problem)

    # Driver signature
    row += 2
    sig_cell = ws.cell(row=row, column=1, value='Driver signature: ____________________________________________     Date: ____________')
    sig_cell.font = Font(size=11, italic=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)

    # Column widths
    for col, w in PRINT_COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # Print area covers everything written
    ws.print_area = f'A1:{get_column_letter(13)}{row}'


# ─────────────────────────────────────────────────────────────────────────────
# Diagnostics
# ─────────────────────────────────────────────────────────────────────────────

def _build_diagnostics(
    ws: Worksheet,
    plan,
    invariant_results: Optional[List[Tuple[str, bool, str]]] = None,
) -> None:
    ws.title = 'Diagnostics'
    ws.cell(row=1, column=1, value='Solver').font = SECTION_FONT
    _write_header_row(ws, 2, ['Field', 'Value'])
    rows = [
        ('Status', plan.solver_status),
        ('Solve seconds', round(plan.solve_seconds, 2)),
        ('Objective ($)', round(plan.objective_cost_dollars, 2)),
        ('Run ID', plan.run_id),
        ('Generated', plan.generated_at.strftime('%Y-%m-%d %H:%M:%S')),
        ('Today', str(plan.today)),
        ('Horizon dates', f'{len(plan.horizon_dates)} days'),
        ('Commit days', plan.commit_days),
        ('Shift start (min)', getattr(plan, 'shift_start_min', 360)),
        ('Shift target (min)', getattr(plan, 'shift_target_min', 480)),
        ('Total stops', plan.total_stops),
        ('Total lbs', round(plan.total_lbs_delivered)),
        ('Total miles', round(plan.total_miles, 1)),
        ('Total minutes', plan.total_minutes),
        ('Avg fill %', round(plan.avg_fill_pct, 1)),
        ('% stops under target fill', round(plan.pct_stops_under_target_fill, 1)),
    ]
    r = 3
    for k, v in rows:
        ws.cell(row=r, column=1, value=k).border = BORDER
        ws.cell(row=r, column=2, value=v).border = BORDER
        r += 1

    r += 2
    ws.cell(row=r, column=1, value='Invariant Checks').font = SECTION_FONT
    r += 1
    _write_header_row(ws, r, ['Check', 'Passed', 'Detail'])
    r += 1
    if invariant_results:
        for name, passed, detail in invariant_results:
            ws.cell(row=r, column=1, value=name).border = BORDER
            pcell = ws.cell(row=r, column=2, value='PASS' if passed else 'FAIL')
            pcell.font = Font(bold=True, color=('27AE60' if passed else 'C0392B'))
            pcell.border = BORDER
            ws.cell(row=r, column=3, value=detail).border = BORDER
            r += 1
    else:
        ws.cell(row=r, column=1, value='(invariants not run at write time)').font = Font(italic=True, color='808080')

    _autosize_columns(ws, overrides={3: 60})


# ─────────────────────────────────────────────────────────────────────────────
# Public entrypoint
# ─────────────────────────────────────────────────────────────────────────────

def write_plan_excel(
    plan,
    output_path: Path,
    invariant_results: Optional[List[Tuple[str, bool, str]]] = None,
    problem=None,
) -> None:
    """Render `plan` to a multi-sheet Excel workbook at `output_path`.

    Pass `problem` (a ProblemInstance) to enrich sheets with client
    metadata (phone, customer names for deferred). If omitted, those
    columns are blank — everything else works.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    phones = _phone_lookup(problem)

    wb = Workbook()
    _build_summary(wb.active, plan, problem=problem)
    _build_todays_plan(wb.create_sheet(), plan, phones)
    _build_week_outlook(wb.create_sheet(), plan, phones)
    _build_at_risk(wb.create_sheet(), plan, phones, problem=problem)
    _build_deferred(wb.create_sheet(), plan, problem=problem)

    # Per-day printables — one sheet per delivery day that has any route
    days_with_routes = sorted({d for (d, _) in plan.routes.keys()})
    for d in days_with_routes:
        _build_print_day(wb.create_sheet(), plan, d, phones, problem=problem)

    _build_diagnostics(wb.create_sheet(), plan, invariant_results)

    wb.save(str(output_path))
