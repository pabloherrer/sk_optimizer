"""
v2 ingest layer smoke tests.

Run: ./sk_venv/bin/python3 -m pytest v2/tests/test_ingest.py -v
"""
from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

import numpy as np
import pandas as pd
import pytest
from openpyxl import Workbook

# Make v2 imports work whether tests run from project root or v2/tests
sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

PROJECT_ROOT = Path(__file__).resolve().parents[2]
DATA_DIR = PROJECT_ROOT / 'data'
INPUT_XLSX = DATA_DIR / 'SK_Delivery_System.xlsx'
MATRIX_NPZ = DATA_DIR / 'osrm_full_matrix_with_ids.npz'
V2_CONFIG = PROJECT_ROOT / 'v2' / 'config'


# ── Helpers for synthetic Excel files ────────────────────────────────────────

def _write_minimal_client_list(
    path: Path,
    rows: list[tuple],
    sheet_name: str = 'Client_List',
) -> None:
    """
    Write a workbook with a Client_List sheet whose data starts at row 4
    (matching the production sheet layout). `rows` is a list of tuples
    (id, customer, zone, zone_code, street, city, state, lat, lon, tank,
     product, service_min, notes, phone, anova, notes2, do_not_schedule).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    # Row 1: arbitrary title. Rows 2-3: blank. Row 4+: data.
    ws.cell(1, 1, 'Client List')
    for i, row in enumerate(rows, start=4):
        for j, val in enumerate(row, start=1):
            ws.cell(i, j, val)
    # Empty Delivery_Log to satisfy load_deliveries
    dl = wb.create_sheet('Delivery_Log')
    dl.cell(1, 1, 'Delivery Log')
    wb.save(str(path))


# ── 1. Real workbook: clients load and are immutable ─────────────────────────

@pytest.mark.skipif(not INPUT_XLSX.exists(),
                    reason='SK_Delivery_System.xlsx not present')
def test_load_clients_from_excel():
    from v2.ingest.excel import load_clients
    from v2.domain.client import Client

    clients = load_clients(INPUT_XLSX)
    assert isinstance(clients, tuple)
    assert len(clients) >= 150, (
        f'Expected ≥150 routable clients, got {len(clients)}'
    )

    # Every client is the frozen domain type
    for c in clients:
        assert isinstance(c, Client)
        # Frozen — assignment raises
        with pytest.raises(Exception):
            c.customer = 'CHANGED'  # type: ignore[misc]

    # IDs are unique
    ids = [c.id for c in clients]
    assert len(ids) == len(set(ids)), 'Duplicate client IDs returned'

    # Lat/lon are populated (we filter unroutable rows)
    for c in clients:
        assert c.lat != 0.0 or c.lon != 0.0
        assert c.tank_capacity_lbs > 0
        assert c.product in ('CANOLA', 'FRYERS CHOICE')


# ── 2. ID normalisation: 4015A / 4015a / 4015 ───────────────────────────────

def test_id_normalization(tmp_path: Path):
    from v2.ingest.excel import load_clients

    xlsx = tmp_path / 'syn.xlsx'
    # Three rows: alphanumeric upper, alphanumeric lower, pure-int
    rows = [
        # id, customer, zone, zone_code, street, city, state, lat, lon,
        # tank, product, service_min, notes, phone
        ('4015A', 'Restaurant A', '1', '1A', '100 Main', 'Phx', 'AZ',
         33.5, -112.1, 350, 'CANOLA', None, None, None),
        ('4015a', 'Restaurant A Dup', '1', '1A', '100 Main', 'Phx', 'AZ',
         33.5, -112.1, 350, 'CANOLA', None, None, None),
        (4015, 'Restaurant Int', '1', '1A', '101 Main', 'Phx', 'AZ',
         33.6, -112.2, 350, 'CANOLA', None, None, None),
    ]
    _write_minimal_client_list(xlsx, rows)
    clients = load_clients(xlsx)

    ids = [c.id for c in clients]
    # All three normalise to either "4015A" or "4015". '4015A' and '4015a'
    # collapse (dedup keeps first); pure-int "4015" is a distinct ID.
    assert '4015A' in ids
    assert '4015' in ids
    # No duplicates after normalisation
    assert len(ids) == len(set(ids))


# ── 3. Matrix loads and shape matches client count ───────────────────────────

@pytest.mark.skipif(not MATRIX_NPZ.exists(),
                    reason='osrm_full_matrix_with_ids.npz not present')
def test_load_matrix():
    from v2.ingest.matrix import load_matrix

    dm, tm, idx = load_matrix(MATRIX_NPZ)
    assert dm.ndim == 2 and dm.shape[0] == dm.shape[1]
    assert tm.shape == dm.shape
    assert dm.dtype == np.int64
    assert tm.dtype == np.int64
    assert len(idx) == dm.shape[0]
    # Truck-speed factor (1.25 default) makes times larger than raw OSRM.
    # Raw cells contain seconds; we don't have raw here, but we can at least
    # check that the time matrix is non-negative and diag is zero.
    assert (tm >= 0).all()
    assert (np.diag(tm) == 0).all()
    assert (np.diag(dm) == 0).all()


# ── 4. Anova readings load (real Query sheet) ────────────────────────────────

@pytest.mark.skipif(not INPUT_XLSX.exists(),
                    reason='SK_Delivery_System.xlsx not present')
def test_anova_readings():
    from v2.ingest.anova import load_anova_readings

    readings = load_anova_readings(INPUT_XLSX)
    # Typical install has 50–100 readings — if the file is missing the
    # Query sheet, that's a different test and we skip.
    if not readings:
        pytest.skip('Query sheet absent in this workbook variant')
    assert len(readings) >= 50, (
        f'Expected ≥50 Anova readings, got {len(readings)}'
    )
    for cid, rec in readings.items():
        assert isinstance(cid, str)
        assert rec['level_lbs'] > 0
        assert rec['source'] in ('sensor', 'sensor-projected', 'stale')
        assert isinstance(rec['timestamp'], pd.Timestamp)
        assert rec['age_hours'] >= 0


# ── 5. Overrides: empty when sheet missing ───────────────────────────────────

def test_overrides_empty_when_sheet_missing(tmp_path: Path):
    from v2.ingest.overrides import load_overrides

    xlsx = tmp_path / 'noov.xlsx'
    wb = Workbook()
    wb.active.title = 'OnlySomeOtherSheet'
    wb.save(str(xlsx))

    ov = load_overrides(xlsx)
    assert ov.is_empty()


def test_overrides_nonexistent_file(tmp_path: Path):
    from v2.ingest.overrides import load_overrides
    ov = load_overrides(tmp_path / 'nope.xlsx')
    assert ov.is_empty()


def test_overrides_malformed_row_raises(tmp_path: Path):
    """A malformed Overrides row must NOT be silently skipped."""
    from v2.ingest.overrides import load_overrides, OverrideValidationError

    xlsx = tmp_path / 'badov.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = 'Overrides'
    # Header
    headers = ['type', 'client_id', 'date', 'date_end', 'lbs',
               'truck', 'stop_n', 'reason', 'operator']
    for j, h in enumerate(headers, start=1):
        ws.cell(1, j, h)
    # Bad row: type=pin but no date
    ws.cell(2, 1, 'pin')
    ws.cell(2, 2, 'C001')
    ws.cell(2, 8, 'Customer call')
    wb.save(str(xlsx))

    with pytest.raises(OverrideValidationError):
        load_overrides(xlsx)


def test_overrides_round_trip(tmp_path: Path):
    """A well-formed Overrides sheet round-trips into Pin/Forbid/Lock/etc."""
    from v2.ingest.overrides import load_overrides

    xlsx = tmp_path / 'goodov.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = 'Overrides'
    headers = ['type', 'client_id', 'date', 'date_end', 'lbs',
               'truck', 'stop_n', 'reason', 'operator']
    for j, h in enumerate(headers, start=1):
        ws.cell(1, j, h)
    # pin
    ws.cell(2, 1, 'pin')
    ws.cell(2, 2, 'C001')
    ws.cell(2, 3, date(2026, 5, 22))
    ws.cell(2, 8, 'must visit')
    ws.cell(2, 9, 'pmh')
    # reading
    ws.cell(3, 1, 'reading')
    ws.cell(3, 2, 'C002')
    ws.cell(3, 3, date(2026, 5, 22))
    ws.cell(3, 5, 250.5)
    ws.cell(3, 8, 'manual gauge')
    wb.save(str(xlsx))

    ov = load_overrides(xlsx)
    assert len(ov.pins) == 1
    assert ov.pins[0].client_id == 'C001'
    assert len(ov.readings) == 1
    assert ov.readings[0].current_lbs == 250.5


# ── 6. Deliveries load and use canonical IDs ─────────────────────────────────

@pytest.mark.skipif(not INPUT_XLSX.exists(),
                    reason='SK_Delivery_System.xlsx not present')
def test_load_deliveries():
    from v2.ingest.excel import load_deliveries
    df = load_deliveries(INPUT_XLSX)
    assert set(['Date', 'Customer', 'Tank_lbs', 'Qty_lbs',
                'Is_Placeholder']).issubset(df.columns)
    assert (df['Qty_lbs'] > 0).all()
    assert df['Date'].notna().all()


# ── 7. Schemas: time windows + closures + excluded ───────────────────────────

@pytest.mark.skipif(not INPUT_XLSX.exists(),
                    reason='SK_Delivery_System.xlsx not present')
def test_schema_loaders_run():
    from v2.ingest.schema import (
        load_closures,
        load_excluded_ids,
        load_time_windows,
    )
    tw = load_time_windows(INPUT_XLSX)
    cl = load_closures(INPUT_XLSX)
    ex = load_excluded_ids(INPUT_XLSX)
    assert isinstance(tw, dict)
    assert isinstance(cl, dict)
    assert isinstance(ex, frozenset)
    # Far-cluster guard set is hardcoded and non-empty
    assert len(ex) > 0
    # If time windows present, they are (int, int)
    for cid, w in tw.items():
        assert isinstance(cid, str)
        assert isinstance(w, tuple) and len(w) == 2
        assert w[0] < w[1]


# ── 8. Full build_problem_instance integration ──────────────────────────────

@pytest.mark.skipif(
    not (INPUT_XLSX.exists() and MATRIX_NPZ.exists()),
    reason='Need both real Excel and matrix',
)
def test_build_problem_instance_smoke():
    from v2.ingest.build_problem import build_problem_instance
    from v2.domain.problem import ProblemInstance

    prob = build_problem_instance(
        config_dir=V2_CONFIG,
        input_file=INPUT_XLSX,
        matrix_file=MATRIX_NPZ,
        today=date(2026, 5, 21),
        run_id='test-run-001',
    )
    assert isinstance(prob, ProblemInstance)
    assert len(prob.clients) >= 150
    assert len(prob.trucks) == 2
    assert {t.id for t in prob.trucks} == {'Truck2', 'Truck9'}
    assert len(prob.horizon_dates) > 0
    # Every client has a TankState
    for c in prob.clients:
        assert c.id in prob.initial_tanks
        ts = prob.initial_tanks[c.id]
        assert ts.current_lbs >= 0
        assert ts.current_lbs <= c.tank_capacity_lbs
        assert ts.source in (
            'manual', 'sensor', 'sensor-projected', 'estimated',
        )
    # Saturday truck rule: at least one (date, truck) is False when Saturday
    saturdays = [d for d in prob.horizon_dates if d.weekday() == 5]
    if saturdays:
        sat = saturdays[0]
        availability = {
            tid: prob.truck_available[(sat, tid)] for tid in
            (t.id for t in prob.trucks)
        }
        # Truck9 is NOT in saturday_trucks per fleet.yaml
        assert availability.get('Truck9') is False
        assert availability.get('Truck2') is True
