"""
Tests for the v2 reporting layer.

Each test constructs a small synthetic Plan, runs one writer, and asserts the
output file exists with reasonable content. Edge cases covered:
  • Empty Plan (no routes)
  • Plan with only Truck2 dispatched (no Truck9)
  • Plan with deferred clients
"""
from __future__ import annotations
import csv
import json
import sys
from datetime import date, datetime
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from v2.domain.plan import Plan, Route, Stop  # noqa: E402
from v2.reporting import (  # noqa: E402
    write_all_outputs,
    write_plan_archive,
    write_plan_excel,
    write_route_map,
    write_smartservice_csv,
)


# ─────────────────────────────────────────────────────────────────────────────
# Synthetic fixtures
# ─────────────────────────────────────────────────────────────────────────────

def _stop(seq: int, cid: str, customer: str, arrival_min: int,
          delivery: float = 600, urgency: str = 'normal',
          lat: float = 33.5, lon: float = -112.1) -> Stop:
    return Stop(
        sequence=seq,
        client_id=cid,
        customer=customer,
        address=f'{100 + seq} Main St',
        lat=lat, lon=lon,
        product='CANOLA',
        tank_capacity_lbs=1000,
        level_at_arrival_lbs=200,
        delivery_lbs=delivery,
        level_after_lbs=200 + delivery,
        arrival_min=arrival_min,
        setup_min=15,
        pump_min=20,
        depart_min=arrival_min + 35,
        travel_miles=5.0,
        cumulative_miles=5.0 * seq,
        days_until_stockout_at_arrival=3.0 if urgency == 'normal' else 0.5,
        urgency_tier=urgency,
    )


def _route(d: date, truck_id: str, stops: tuple) -> Route:
    return Route(
        date=d, truck_id=truck_id, territory_label='NE',
        stops=stops,
        compartment_a_product='CANOLA', compartment_a_lbs=5000,
        compartment_b_product='FRYERS CHOICE', compartment_b_lbs=5000,
        depart_depot_min=0,
        return_depot_min=stops[-1].depart_min + 30 if stops else 0,
        total_minutes=(stops[-1].depart_min + 30) if stops else 0,
        overtime_minutes=0,
        total_miles=sum(s.travel_miles for s in stops),
        cost_miles_dollars=10.0, cost_labor_dollars=50.0,
        cost_overtime_dollars=0.0, cost_dispatch_dollars=50.0,
        cost_total_dollars=110.0,
        total_load_lbs=sum(s.delivery_lbs for s in stops),
        cap_pct=sum(s.delivery_lbs for s in stops) / 10000 * 100,
    )


def _make_plan(
    n_stops_t2: int = 3,
    n_stops_t9: int = 2,
    deferred: dict = None,
    extra_day: bool = False,
) -> Plan:
    today = date(2026, 5, 21)
    horizon = (today, date(2026, 5, 22)) if extra_day else (today,)

    routes = {}
    if n_stops_t2 > 0:
        stops_t2 = tuple(
            _stop(i + 1, f'T2-{i}', f'Customer T2-{i}', arrival_min=60 * (i + 1),
                  urgency='critical' if i == 0 else 'normal')
            for i in range(n_stops_t2)
        )
        routes[(today, 'Truck2')] = _route(today, 'Truck2', stops_t2)
    if n_stops_t9 > 0:
        stops_t9 = tuple(
            _stop(i + 1, f'T9-{i}', f'Customer T9-{i}', arrival_min=60 * (i + 1),
                  urgency='stockout' if i == 0 else 'urgent',
                  lat=33.6 + 0.01 * i, lon=-112.2 - 0.01 * i)
            for i in range(n_stops_t9)
        )
        routes[(today, 'Truck9')] = _route(today, 'Truck9', stops_t9)

    total_stops = sum(len(r.stops) for r in routes.values())
    total_lbs = sum(r.total_load_lbs for r in routes.values())
    total_miles = sum(r.total_miles for r in routes.values())
    total_min = sum(r.total_minutes for r in routes.values())

    return Plan(
        run_id='test-run-001',
        generated_at=datetime(2026, 5, 20, 22, 0),
        today=today,
        horizon_dates=horizon,
        commit_days=1,
        routes=routes,
        deferred=deferred or {},
        solve_seconds=1.5,
        objective_cost_dollars=110.0,
        solver_status='OPTIMAL',
        total_stops=total_stops,
        total_lbs_delivered=total_lbs,
        total_miles=total_miles,
        total_minutes=total_min,
        avg_fill_pct=80.0,
        pct_stops_under_target_fill=10.0,
        capacity_warnings=('Tue May 26: 94% cap, consider extra truck',),
    )


def _empty_plan() -> Plan:
    today = date(2026, 5, 21)
    return Plan(
        run_id='empty-run',
        generated_at=datetime(2026, 5, 20, 22, 0),
        today=today,
        horizon_dates=(today,),
        commit_days=1,
        routes={},
        deferred={},
        solve_seconds=0.1,
        objective_cost_dollars=0.0,
        solver_status='OPTIMAL',
        total_stops=0,
        total_lbs_delivered=0.0,
        total_miles=0.0,
        total_minutes=0,
        avg_fill_pct=0.0,
        pct_stops_under_target_fill=0.0,
    )


# ─────────────────────────────────────────────────────────────────────────────
# Excel
# ─────────────────────────────────────────────────────────────────────────────

def test_excel_writes_expected_sheets(tmp_path: Path):
    plan = _make_plan(deferred={'CLI-99': 'Closed for renovation'})
    path = tmp_path / 'plan.xlsx'
    write_plan_excel(plan, path)
    assert path.exists() and path.stat().st_size > 0

    from openpyxl import load_workbook
    wb = load_workbook(path)
    names = wb.sheetnames
    assert 'Summary' in names
    assert "Today's_Plan" in names
    assert 'Week_Outlook' in names
    assert 'At_Risk' in names
    assert 'Deferred' in names
    assert 'Diagnostics' in names
    # One PRINT sheet for the one day with routes
    assert any(n.startswith('PRINT') for n in names)


def test_excel_empty_plan(tmp_path: Path):
    plan = _empty_plan()
    path = tmp_path / 'empty.xlsx'
    write_plan_excel(plan, path)
    assert path.exists()
    from openpyxl import load_workbook
    wb = load_workbook(path)
    # Even with no routes, the framework sheets exist
    assert 'Summary' in wb.sheetnames
    assert 'Deferred' in wb.sheetnames


def test_excel_deferred_populates(tmp_path: Path):
    deferred = {'CLI-1': 'Tank still 70% full', 'CLI-2': 'Closed Tuesday'}
    plan = _make_plan(deferred=deferred)
    path = tmp_path / 'plan.xlsx'
    write_plan_excel(plan, path)

    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb['Deferred']
    # Header + two data rows
    rows = list(ws.iter_rows(values_only=True))
    data_rows = [r for r in rows[1:] if r[0]]
    assert len(data_rows) == 2
    ids = {r[0] for r in data_rows}
    assert ids == {'CLI-1', 'CLI-2'}


# ─────────────────────────────────────────────────────────────────────────────
# SmartService CSV
# ─────────────────────────────────────────────────────────────────────────────

def test_smartservice_csv_rows_match_stops(tmp_path: Path):
    plan = _make_plan(n_stops_t2=3, n_stops_t9=2)
    path = tmp_path / 'ss.csv'
    n = write_smartservice_csv(plan, plan.today, path)
    assert n == 5

    with open(path, newline='') as f:
        reader = csv.DictReader(f)
        rows = list(reader)
    assert len(rows) == 5
    # Both drivers represented
    drivers = {r['Employee Full Name'] for r in rows}
    assert 'JOE' in drivers and 'JASON B' in drivers
    # Stops numbered correctly
    t2 = [int(r['Stop']) for r in rows if r['Crew'] == 'Truck2']
    assert sorted(t2) == [1, 2, 3]


def test_smartservice_csv_empty_day(tmp_path: Path):
    plan = _empty_plan()
    path = tmp_path / 'ss.csv'
    n = write_smartservice_csv(plan, plan.today, path)
    assert n == 0
    # Header is still written
    assert path.exists()
    with open(path) as f:
        first = f.readline()
    assert 'Customer' in first


# ─────────────────────────────────────────────────────────────────────────────
# Map
# ─────────────────────────────────────────────────────────────────────────────

def test_route_map_writes_html(tmp_path: Path):
    plan = _make_plan()
    path = tmp_path / 'map.html'
    result = write_route_map(plan, path)
    assert result == path
    assert path.exists()
    content = path.read_text()
    assert '<html' in content.lower() or 'leaflet' in content.lower()
    # Both trucks should appear as layers
    assert 'Truck2' in content
    assert 'Truck9' in content


def test_route_map_empty_plan(tmp_path: Path):
    plan = _empty_plan()
    path = tmp_path / 'map.html'
    write_route_map(plan, path)
    assert path.exists() and path.stat().st_size > 0


# ─────────────────────────────────────────────────────────────────────────────
# Archive
# ─────────────────────────────────────────────────────────────────────────────

def test_archive_roundtrip(tmp_path: Path):
    plan = _make_plan(deferred={'CLI-9': 'closed'})
    path = write_plan_archive(plan, tmp_path)
    assert path.exists()
    assert path.parent.name == 'archive'

    data = json.loads(path.read_text())
    assert data['run_id'] == 'test-run-001'
    assert data['solver_status'] == 'OPTIMAL'
    # routes flattened to a list (tuple-key serialization)
    assert isinstance(data['routes'], list)
    assert len(data['routes']) == 2


# ─────────────────────────────────────────────────────────────────────────────
# write_all_outputs
# ─────────────────────────────────────────────────────────────────────────────

def test_write_all_outputs_full_bundle(tmp_path: Path):
    plan = _make_plan()
    artifacts = write_all_outputs(plan, tmp_path)

    assert artifacts['excel'].exists()
    assert artifacts['csv'].exists()
    assert artifacts['map'].exists()
    assert artifacts['archive'].exists()
    # PDFs are no longer produced — PRINT sheets in the Excel replace them.
    assert 'pdfs' not in artifacts


def test_write_all_outputs_only_truck2(tmp_path: Path):
    plan = _make_plan(n_stops_t9=0)
    artifacts = write_all_outputs(plan, tmp_path)
    assert artifacts['excel'].exists()
    assert artifacts['map'].exists()


def test_write_all_outputs_empty_plan(tmp_path: Path):
    plan = _empty_plan()
    artifacts = write_all_outputs(plan, tmp_path)
    assert artifacts['excel'].exists()
    assert artifacts['csv'].exists()
    assert artifacts['map'].exists()
    assert artifacts['archive'].exists()
