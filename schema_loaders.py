"""
schema_loaders.py
=================
Load configuration and constraints from SK_Delivery_System.xlsx sheets.

Functions expose Client_Time_Windows, Client_Closures, Trucks, and Depot
configuration as Python objects, with graceful fallbacks to config.py defaults
when sheets are missing or incomplete.
"""

import warnings
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
from config import INPUT_FILE, TRUCKS as DEFAULT_TRUCKS

warnings.filterwarnings('ignore')


# ── Public API ────────────────────────────────────────────────────────────────

def load_time_windows(input_file: str | Path = INPUT_FILE) -> pd.DataFrame:
    """
    Load Client_Time_Windows sheet.

    Schema (one row per rule, human-friendly):
      Col 1: Client_ID
      Col 2: Customer (name; for human readability — ignored by code)
      Col 3: Day_of_Week — 'Tue' / 'Wed' / ... / 'Sat'  OR  'All' / '*' /
             empty to mean every workday
      Col 4: Open_HHMM    e.g. '9:00'
      Col 5: Close_HHMM   e.g. '10:00'
      Col 6: Notes (optional, free text)

    Backwards-compat: also accepts the legacy 4-column format
    (Client_ID, Day_of_Week, Open, Close) by detecting whether col 4
    looks like an HH:MM string vs a customer name.

    Loader EXPANDS 'All'/'*'/empty to all workdays, returning one row
    per (client, day) pair as before.
    """
    workdays = ['Tue', 'Wed', 'Thu', 'Fri', 'Sat']
    try:
        wb = load_workbook(str(input_file), data_only=True)
        if 'Client_Time_Windows' not in wb.sheetnames:
            return pd.DataFrame(columns=['Client_ID', 'Day_of_Week', 'Open_Min', 'Close_Min'])

        ws = wb['Client_Time_Windows']
        records = []

        for row in ws.iter_rows(min_row=4, values_only=True):
            if not row[0]:
                continue
            client_id = str(row[0]).strip()
            if client_id.startswith('EXAMPLE_'):
                continue

            # Detect schema version. New schema: col 2 = customer name (free text);
            # col 3 = day; col 4 = open; col 5 = close. Legacy: col 2 = day;
            # col 3 = open; col 4 = close. Distinguish by checking whether col 3
            # looks like an HH:MM time.
            new_schema = False
            col3 = row[2] if len(row) > 2 else None
            if col3 is not None:
                col3_str = str(col3).strip()
                # If col 3 is HH:MM-ish and col 4 is too → legacy schema
                if ':' in col3_str and len(row) > 3 and row[3] and ':' in str(row[3]):
                    new_schema = False
                else:
                    new_schema = True

            if new_schema:
                day_raw   = str(row[2]).strip() if row[2] else ''
                open_raw  = str(row[3]).strip() if len(row) > 3 and row[3] else ''
                close_raw = str(row[4]).strip() if len(row) > 4 and row[4] else ''
            else:
                day_raw   = str(row[1]).strip() if row[1] else ''
                open_raw  = str(row[2]).strip() if row[2] else ''
                close_raw = str(row[3]).strip() if row[3] else ''

            if not (open_raw and close_raw):
                continue

            try:
                open_min  = _time_to_min(open_raw)
                close_min = _time_to_min(close_raw)
            except ValueError:
                continue

            day_upper = day_raw.upper()
            if day_upper in ('', 'ALL', '*', 'ANY', 'EVERY', 'WEEKDAY', 'WEEKDAYS'):
                days_to_apply = workdays
            else:
                days_to_apply = [day_raw]

            for d in days_to_apply:
                records.append({
                    'Client_ID': client_id,
                    'Day_of_Week': d,
                    'Open_Min': open_min,
                    'Close_Min': close_min,
                })

        return pd.DataFrame(records)
    except Exception:
        return pd.DataFrame(columns=['Client_ID', 'Day_of_Week', 'Open_Min', 'Close_Min'])


def load_closures(input_file: str | Path = INPUT_FILE) -> pd.DataFrame:
    """
    Load Client_Closures sheet.

    Schema (one row per rule, human-friendly):
      Col 1: Client_ID
      Col 2: Customer (name; for human readability — ignored by code)
      Col 3: Recurring_DOW — 'Tue' / 'Wed' / ... / 'Sat'  OR empty.
             If set: applies every week on that weekday (no need to list
             individual dates).
      Col 4: Start_Date  (only used if Recurring_DOW is empty)
      Col 5: End_Date    (only used if Recurring_DOW is empty)
      Col 6: Reason

    Backwards-compat: legacy 4-col format
      (Client_ID, Start_Date, End_Date, Reason)
    is also accepted — detected by whether col 2 is a date or a name.

    Returns DataFrame columns:
      Client_ID, Start_Date, End_Date, Recurring_DOW, Reason
    """
    try:
        wb = load_workbook(str(input_file), data_only=True)
        if 'Client_Closures' not in wb.sheetnames:
            return pd.DataFrame(columns=['Client_ID', 'Start_Date', 'End_Date',
                                         'Recurring_DOW', 'Reason'])

        ws = wb['Client_Closures']
        records = []

        for row in ws.iter_rows(min_row=4, values_only=True):
            if not row[0]:
                continue
            client_id = str(row[0]).strip()
            if client_id.startswith('EXAMPLE_'):
                continue

            # Detect schema. New: col 2 = customer (text). Legacy: col 2 = date.
            col2 = row[1] if len(row) > 1 else None
            new_schema = isinstance(col2, str) or col2 is None

            if new_schema:
                rec_dow_raw = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                start_date  = row[3] if len(row) > 3 else None
                end_date    = row[4] if len(row) > 4 else None
                reason      = str(row[5]).strip() if len(row) > 5 and row[5] else ''
            else:
                rec_dow_raw = ''
                start_date  = row[1]
                end_date    = row[2]
                reason      = str(row[3]).strip() if row[3] else ''

            rec_dow_norm = rec_dow_raw[:3].title() if rec_dow_raw else ''
            if rec_dow_norm not in ('', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun'):
                rec_dow_norm = ''

            # If recurring DOW is set, dates are optional (use sentinels)
            if rec_dow_norm:
                start_date_p = pd.Timestamp('1970-01-01')
                end_date_p   = pd.Timestamp('2099-12-31')
            else:
                # Date range required for non-recurring closures
                if start_date is None or end_date is None:
                    continue
                try:
                    start_date_p = pd.Timestamp(start_date)
                    end_date_p   = pd.Timestamp(end_date)
                except (ValueError, AttributeError, TypeError):
                    continue

            records.append({
                'Client_ID':     client_id,
                'Start_Date':    start_date_p,
                'End_Date':      end_date_p,
                'Recurring_DOW': rec_dow_norm,
                'Reason':        reason,
            })

        return pd.DataFrame(records)
    except Exception:
        return pd.DataFrame(columns=['Client_ID', 'Start_Date', 'End_Date',
                                     'Recurring_DOW', 'Reason'])


def load_trucks(input_file: str | Path = INPUT_FILE) -> dict:
    """
    Load Trucks sheet and return as dict matching config.TRUCKS format.

    Dict structure:
      {truck_id: {
          'capacity_lbs': int,
          'pump_rate_lbs_per_min': float,
          'fixed_setup_min': int,
          'shift_min': int,
          'cost_per_mile': float,
          'compartment_a_lbs': int,
          'compartment_b_lbs': int,
          'active': bool
      }}

    Only includes rows where Active='yes' (case-insensitive).
    Falls back to config.TRUCKS if sheet is missing or all rows inactive.
    """
    try:
        wb = load_workbook(str(input_file), data_only=True)
        if 'Trucks' not in wb.sheetnames:
            return DEFAULT_TRUCKS

        ws = wb['Trucks']
        trucks_dict = {}

        for row in ws.iter_rows(min_row=4, values_only=True):
            if not row[0]:
                continue

            truck_id = str(row[0]).strip()
            if truck_id.startswith('EXAMPLE_'):
                continue

            # Check if active
            active = str(row[7]).strip().lower() if row[7] else 'no'
            if active != 'yes':
                continue

            try:
                comp_a = float(row[1])
                comp_b = float(row[2])
                pump_rate = float(row[3])
                setup_min = int(float(row[4]))
                shift_min = int(float(row[5]))
                cost_per_mile = float(row[6])

                trucks_dict[truck_id] = {
                    'capacity_lbs': comp_a + comp_b,
                    'pump_rate_lbs_per_min': pump_rate,
                    'fixed_setup_min': setup_min,
                    'shift_min': shift_min,
                    'cost_per_mile': cost_per_mile,
                    'compartment_a_lbs': comp_a,
                    'compartment_b_lbs': comp_b,
                    'active': True,
                }
            except (ValueError, TypeError, IndexError):
                # Skip malformed rows
                continue

        # Fall back to defaults if no valid trucks found
        return trucks_dict if trucks_dict else DEFAULT_TRUCKS
    except Exception:
        return DEFAULT_TRUCKS


def load_depot_config(input_file: str | Path = INPUT_FILE) -> dict:
    """
    Load Depot sheet as key-value pairs.

    Returns dict with keys:
      - depot_lat (float)
      - depot_lon (float)
      - shift_start_min (int: minutes since midnight)
      - shift_end_min (int: minutes since midnight)
      - morning_load_min (int)
      - evening_unload_min (int)
      - work_days (list of str: ['Tue', 'Wed', ...])

    Falls back to config.py defaults if sheet is missing or incomplete.
    """
    from config import (DEPOT_LAT, DEPOT_LON, SHIFT_MIN, DAYS,
                        SHIFT_MIN as DEFAULT_SHIFT_MIN)

    defaults = {
        'depot_lat': DEPOT_LAT,
        'depot_lon': DEPOT_LON,
        'shift_start_min': 6 * 60,  # 06:00
        'shift_end_min': 16 * 60,   # 16:00
        'morning_load_min': 30,
        'evening_unload_min': 15,
        'work_days': DAYS,
    }

    try:
        wb = load_workbook(str(input_file), data_only=True)
        if 'Depot' not in wb.sheetnames:
            return defaults

        ws = wb['Depot']
        config_dict = dict(defaults)

        for row in ws.iter_rows(min_row=4, values_only=True):
            if not row[0]:
                continue

            param = str(row[0]).strip()
            value = row[1]

            try:
                if param == 'Depot_Lat':
                    config_dict['depot_lat'] = float(value)
                elif param == 'Depot_Lon':
                    config_dict['depot_lon'] = float(value)
                elif param == 'Shift_Start_HHMM':
                    config_dict['shift_start_min'] = _time_to_min(str(value).strip())
                elif param == 'Shift_End_HHMM':
                    config_dict['shift_end_min'] = _time_to_min(str(value).strip())
                elif param == 'Morning_Load_Min':
                    config_dict['morning_load_min'] = int(float(value))
                elif param == 'Evening_Unload_Min':
                    config_dict['evening_unload_min'] = int(float(value))
                elif param == 'Work_Days':
                    days_str = str(value).strip()
                    config_dict['work_days'] = [d.strip() for d in days_str.split(',')]
            except (ValueError, TypeError):
                # Keep default for this parameter
                pass

        return config_dict
    except Exception:
        return defaults


def is_client_open(client_id: str, day_name: str, arrival_min: int,
                   time_windows_df: pd.DataFrame) -> bool:
    """
    Check if a client is open on a given day at a given arrival time.

    Parameters
    ----------
    client_id : str
        Client ID (e.g., 'C001')
    day_name : str
        Day of week (e.g., 'Tue')
    arrival_min : int
        Arrival time in minutes since midnight (e.g., 540 for 9:00 AM)
    time_windows_df : pd.DataFrame
        Result from load_time_windows()

    Returns
    -------
    bool
        True if client is open at arrival_min.
        True if no window defined for (client_id, day_name) = no restriction.
    """
    if time_windows_df.empty:
        return True

    match = time_windows_df[
        (time_windows_df['Client_ID'] == client_id) &
        (time_windows_df['Day_of_Week'] == day_name)
    ]

    if match.empty:
        # No restriction defined
        return True

    row = match.iloc[0]
    return row['Open_Min'] <= arrival_min < row['Close_Min']


_WEEKDAY_NAMES = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']


def is_client_closed_on(client_id: str, date, closures_df: pd.DataFrame) -> bool:
    """
    Check if a client is closed on a given date. Handles BOTH:
      • Date-range closures (Start_Date ≤ date ≤ End_Date)
      • Recurring weekday closures (Recurring_DOW matches date.weekday())

    Returns True if any closure rule covers (client_id, date).
    """
    if closures_df.empty:
        return False

    if isinstance(date, str):
        date = pd.Timestamp(date)
    elif hasattr(date, 'date'):
        date = pd.Timestamp(date.date())
    else:
        date = pd.Timestamp(date)
    weekday_short = _WEEKDAY_NAMES[date.weekday()]

    # Match this client's rows
    rows = closures_df[closures_df['Client_ID'] == client_id]
    if rows.empty:
        return False

    # Recurring weekday rule
    if 'Recurring_DOW' in rows.columns:
        rec = rows[rows['Recurring_DOW'] == weekday_short]
        if not rec.empty:
            return True

    # Date-range rule
    in_range = rows[
        (rows['Start_Date'] <= date) & (date <= rows['End_Date'])
    ]
    # If the row has a Recurring_DOW, the date range is sentinel — skip those
    if 'Recurring_DOW' in in_range.columns:
        in_range = in_range[in_range['Recurring_DOW'] == '']
    return len(in_range) > 0


# ── Helpers ───────────────────────────────────────────────────────────────────

def _time_to_min(time_str: str) -> int:
    """
    Convert HH:MM string to minutes since midnight.

    Raises ValueError if format is invalid.
    """
    parts = time_str.split(':')
    if len(parts) != 2:
        raise ValueError(f"Invalid time format: {time_str}")
    h, m = int(parts[0]), int(parts[1])
    return h * 60 + m
