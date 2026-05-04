"""
anova_history.py — Historical Anova readings report
====================================================

Reads data/anova_history.csv and writes data/anova_tank_report.xlsx with:
  • Sheet 'Current'  — latest reading per client, color-coded by % full
  • Sheet 'History'  — all readings sorted by client → timestamp
  • Sheet 'Trends'   — daily average level per client

Run:  python anova_history.py
"""

import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

sys.path.insert(0, str(Path(__file__).parent))
from config import DATA_DIR

HISTORY_FILE = DATA_DIR / 'anova_history.csv'
REPORT_FILE = DATA_DIR / 'anova_tank_report.xlsx'

GREEN = PatternFill('solid', fgColor='FFCCFFCC')   # ≥ 40 % full
YELLOW = PatternFill('solid', fgColor='FFFFFFAA')  # 20–40 %
RED = PatternFill('solid', fgColor='FFFF9999')     # < 20 %
HEADER_FILL = PatternFill('solid', fgColor='FFD9E1F2')


def _color_for_pct(pct: float) -> PatternFill:
    if pct < 20:
        return RED
    if pct < 40:
        return YELLOW
    return GREEN


def _write_headers(ws, headers):
    bold = Font(bold=True)
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = bold
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal='left')


def build_report() -> Path:
    if not HISTORY_FILE.exists():
        print(f'  ⚠ No history file at {HISTORY_FILE} — run anova_fetch.py first.')
        return None

    df = pd.read_csv(HISTORY_FILE)
    if df.empty:
        print('  ⚠ History file is empty.')
        return None

    df['timestamp'] = pd.to_datetime(df['timestamp'], errors='coerce')
    df = df.dropna(subset=['timestamp'])
    df['date'] = df['timestamp'].dt.date

    wb = Workbook()

    # ── Sheet: Current ──────────────────────────────────────────────────
    ws_cur = wb.active
    ws_cur.title = 'Current'
    headers = ['client_id', 'client_name', 'rtu_id', 'level_lbs',
               'pct_full', 'tank_capacity_lbs', 'product', 'last_reading']
    _write_headers(ws_cur, headers)

    latest = df.sort_values('timestamp').groupby('client_id').tail(1)
    latest = latest.sort_values('pct_full')
    for r_idx, (_, r) in enumerate(latest.iterrows(), start=2):
        pct = float(r.get('pct_full', 0))
        fill = _color_for_pct(pct)
        vals = [
            str(r.get('client_id', '')),
            str(r.get('client_name', '')),
            str(r.get('rtu_id', '')),
            float(r.get('level_lbs', 0)),
            pct,
            float(r.get('tank_capacity_lbs', 0)),
            str(r.get('product', '')),
            str(r['timestamp']),
        ]
        for i, v in enumerate(vals, start=1):
            c = ws_cur.cell(row=r_idx, column=i, value=v)
            c.fill = fill
    ws_cur.freeze_panes = 'A2'
    for col_letter, w in zip('ABCDEFGH', [10, 38, 12, 11, 9, 14, 16, 22]):
        ws_cur.column_dimensions[col_letter].width = w

    # ── Sheet: History ──────────────────────────────────────────────────
    ws_hist = wb.create_sheet('History')
    hist_headers = ['client_id', 'client_name', 'rtu_id', 'timestamp',
                    'level_lbs', 'pct_full', 'product']
    _write_headers(ws_hist, hist_headers)

    hist = df.sort_values(['client_id', 'timestamp'])
    for r_idx, (_, r) in enumerate(hist.iterrows(), start=2):
        ws_hist.cell(row=r_idx, column=1, value=str(r.get('client_id', '')))
        ws_hist.cell(row=r_idx, column=2, value=str(r.get('client_name', '')))
        ws_hist.cell(row=r_idx, column=3, value=str(r.get('rtu_id', '')))
        ws_hist.cell(row=r_idx, column=4, value=str(r['timestamp']))
        ws_hist.cell(row=r_idx, column=5, value=float(r.get('level_lbs', 0)))
        ws_hist.cell(row=r_idx, column=6, value=float(r.get('pct_full', 0)))
        ws_hist.cell(row=r_idx, column=7, value=str(r.get('product', '')))
    ws_hist.freeze_panes = 'A2'
    for col_letter, w in zip('ABCDEFG', [10, 38, 12, 22, 11, 9, 16]):
        ws_hist.column_dimensions[col_letter].width = w

    # ── Sheet: Trends (daily avg per client) ────────────────────────────
    ws_tr = wb.create_sheet('Trends')
    trend_headers = ['client_id', 'client_name', 'date', 'avg_level_lbs',
                     'avg_pct_full', 'n_readings']
    _write_headers(ws_tr, trend_headers)

    daily = (
        df.groupby(['client_id', 'client_name', 'date'])
          .agg(avg_level_lbs=('level_lbs', 'mean'),
               avg_pct_full=('pct_full', 'mean'),
               n_readings=('level_lbs', 'count'))
          .reset_index()
          .sort_values(['client_id', 'date'])
    )
    for r_idx, (_, r) in enumerate(daily.iterrows(), start=2):
        ws_tr.cell(row=r_idx, column=1, value=str(r['client_id']))
        ws_tr.cell(row=r_idx, column=2, value=str(r['client_name']))
        ws_tr.cell(row=r_idx, column=3, value=str(r['date']))
        ws_tr.cell(row=r_idx, column=4, value=round(float(r['avg_level_lbs']), 1))
        ws_tr.cell(row=r_idx, column=5, value=round(float(r['avg_pct_full']), 1))
        ws_tr.cell(row=r_idx, column=6, value=int(r['n_readings']))
    ws_tr.freeze_panes = 'A2'
    for col_letter, w in zip('ABCDEF', [10, 38, 12, 14, 12, 11]):
        ws_tr.column_dimensions[col_letter].width = w

    wb.save(str(REPORT_FILE))
    print(f'✓ Report written → {REPORT_FILE}')
    print(f'  Current: {len(latest)} clients | History: {len(hist)} rows | Trends: {len(daily)} day-rows')
    return REPORT_FILE


if __name__ == '__main__':
    build_report()
