#!/usr/bin/env python3
"""
backtest_rolling.py — Rolling 2-week simulation vs actual driver deliveries.

Unlike backtest.py (which runs each day independently), this script carries
optimizer state forward: deliveries the optimizer "makes" on Day 1 update
inventory before planning Day 2.  This prevents double-counting clients
and gives a fair apples-to-apples comparison against what drivers did.

For each delivery day:
  1. Use historical data up to the simulation start date
  2. Run optimizer → commit Day 0 routes
  3. "Execute" optimizer deliveries (update inventory)
  4. Deplete all tanks by one day of consumption
  5. Advance and repeat

Then compare 2-week totals: optimizer vs actual drivers.
"""

import sys, json, io, re
from pathlib import Path
from contextlib import redirect_stdout, redirect_stderr
from collections import defaultdict

import pandas as pd
import numpy as np

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from config import (
    INPUT_FILE, MATRIX_FILE, DAYS, NUM_DAYS,
    EXCLUDED_CLIENT_IDS, PRODUCTS, METERS_PER_MILE,
)
from load_data import load_all
from forecast_consumption import estimate_consumption_rates
from inventory import enrich_snapshot
from router import load_matrix
from state import initialise_state_from_snapshot
from schema_loaders import load_time_windows, load_closures, load_depot_config
from unified_solver import solve_week
from run_unified import compute_plan_dates


_WEEKDAY_SHORT = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']


def load_actual_deliveries(input_file):
    """Load actual deliveries with qty and tank size from the log."""
    import openpyxl
    wb = openpyxl.load_workbook(input_file, data_only=True)
    ws = wb['Delivery_Log']

    ws_cl = wb['Client_List']
    tank_by_name = {}
    for row in ws_cl.iter_rows(min_row=4, values_only=True):
        cust_name = row[1]
        if cust_name:
            tank_by_name[str(cust_name).strip()] = float(row[9] or 0)

    by_date = defaultdict(list)
    for row in ws.iter_rows(min_row=4, values_only=True):
        dt, cust = row[0], row[1]
        if dt and cust:
            cust_s = str(cust).strip()
            date_str = pd.Timestamp(dt).strftime('%Y-%m-%d')
            nid = extract_numeric_id(cust_s)
            qty = float(row[6] or 0)
            tank = float(row[5] or 0)
            if tank == 0:
                tank = tank_by_name.get(cust_s, 0)
            by_date[date_str].append({
                'id': nid,
                'customer': cust_s,
                'qty': qty,
                'tank': tank,
            })
    return by_date


def extract_numeric_id(customer_str):
    parts = customer_str.replace('-', ' ').split()
    for p in parts:
        m = re.match(r'^(\d{3,})[A-Za-z]?$', p)
        if m:
            return int(m.group(1))
    return None


def estimate_driver_miles(client_ids, node_index_map, dist_matrix):
    """
    Estimate miles for a set of client IDs using nearest-neighbor TSP.
    This gives drivers a BEST CASE (optimal-ish path) — real miles are higher.
    Returns miles (converted from meters via METERS_PER_MILE).
    """
    # Map client IDs to matrix indices
    indices = []
    for cid in client_ids:
        # Try matching by numeric ID prefix
        for key, idx in node_index_map.items():
            m = re.match(r'^(\d+)', str(key))
            if m and int(m.group(1)) == cid:
                indices.append(idx)
                break

    if not indices:
        return 0.0

    # Depot is index 0
    depot = 0
    unvisited = set(indices)
    route = [depot]
    current = depot
    total_m = 0.0

    while unvisited:
        # Find nearest unvisited
        best_dist = float('inf')
        best_node = None
        for n in unvisited:
            d = dist_matrix[current][n]
            if d < best_dist:
                best_dist = d
                best_node = n
        if best_node is None:
            break
        total_m += best_dist
        route.append(best_node)
        unvisited.remove(best_node)
        current = best_node

    # Return to depot
    total_m += dist_matrix[current][depot]

    return total_m / METERS_PER_MILE


def run_rolling_comparison(start_plan_date: str, n_days: int = 10,
                           solve_sec: int = 8, output_file: str = None):
    """
    Run rolling simulation starting from start_plan_date, compare against
    actual driver deliveries for the same period.

    start_plan_date: the evening we "plan". E.g. '2026-03-09' means we plan
                     Sunday evening, first delivery day is Mon→Tue (Mar 10).
    n_days: number of delivery days to simulate (10 = 2 weeks Tue-Sat).
    """
    plan_start = pd.Timestamp(start_plan_date).normalize()

    # ── Load static data ────────────────────────────────────────────────
    clients_raw, deliveries_full = load_all(INPUT_FILE)
    dm, tm, nim = load_matrix(MATRIX_FILE)
    tw = load_time_windows(INPUT_FILE)
    cl = load_closures(INPUT_FILE)
    dc = load_depot_config(INPUT_FILE)

    # ── Truncate deliveries to plan start date ──────────────────────────
    cutoff = plan_start + pd.Timedelta(days=1)
    deliveries_cut = deliveries_full[deliveries_full['Date'] < cutoff].copy()
    print(f"  Using {len(deliveries_cut)} deliveries up to {plan_start.date()}")

    # ── Estimate consumption rates from truncated data ──────────────────
    clients_df = estimate_consumption_rates(deliveries_cut, clients_raw, today=plan_start)
    state = initialise_state_from_snapshot(clients_df)
    snapshot = enrich_snapshot(clients_df, state)

    # ── Build mutable inventory tracker ─────────────────────────────────
    inventory = {}
    rates = {}
    tanks = {}
    for _, row in snapshot.iterrows():
        cid = row['ID']
        inventory[cid] = float(row.get('Current_lbs', 0))
        rates[cid] = float(row.get('Avg_LbsPerDay', 0))
        tanks[cid] = float(row.get('Tank_lbs', 0))

    # ── Load actual driver deliveries ───────────────────────────────────
    actual_by_date = load_actual_deliveries(INPUT_FILE)

    workday_set = set(DAYS)

    # ── Simulate ────────────────────────────────────────────────────────
    opt_days = []
    current_date = plan_start + pd.Timedelta(days=1)
    delivery_day_count = 0

    print(f"\n{'═' * 100}")
    print(f"  Rolling comparison: plan from {plan_start.date()}, {n_days} delivery days")
    print(f"{'═' * 100}")
    print(f"  {'Date':<12} {'Day':<4} {'OptStp':>6} {'DrvStp':>6} {'OptLbs':>8} {'DrvLbs':>8}"
          f"  {'OptMi':>6} {'DrvMi':>6}  {'Opt/s':>6} {'Drv/s':>6}  {'Opt%':>5} {'Drv%':>5}")
    print(f"  {'-' * 100}")

    while delivery_day_count < n_days:
        day_name = _WEEKDAY_SHORT[current_date.weekday()]
        if day_name not in workday_set:
            for cid in inventory:
                inventory[cid] = max(0, inventory[cid] - rates.get(cid, 0))
            current_date += pd.Timedelta(days=1)
            continue

        delivery_day_count += 1
        date_str = current_date.strftime('%Y-%m-%d')

        # ── Update snapshot with current inventory ──────────────────
        snap = snapshot.copy()
        for i, row in snap.iterrows():
            cid = row['ID']
            if cid in inventory:
                cur = max(0, inventory[cid])
                tank = tanks.get(cid, 1)
                rate = rates.get(cid, 0)
                snap.at[i, 'Current_lbs'] = cur
                snap.at[i, 'Refill_lbs'] = max(0, tank - cur)
                snap.at[i, 'Days_Until_Stockout'] = cur / rate if rate > 0 else 999
                from inventory import urgency_tier
                snap.at[i, 'Urgency'] = urgency_tier(cur / rate if rate > 0 else 999)

        # ── Run solver ──────────────────────────────────────────────
        plan_evening = current_date - pd.Timedelta(days=1)
        plan_dates = compute_plan_dates(plan_evening)

        try:
            buf = io.StringIO()
            with redirect_stdout(buf), redirect_stderr(buf):
                routes, deferred = solve_week(
                    snap, dm, tm, nim, solve_seconds=solve_sec,
                    time_windows_df=tw, closures_df=cl,
                    today=plan_evening,
                    depot_config=dc, plan_dates=plan_dates,
                )
        except Exception as e:
            print(f"  {date_str}  {day_name}  ERROR: {e}")
            for cid in inventory:
                inventory[cid] = max(0, inventory[cid] - rates.get(cid, 0))
            current_date += pd.Timedelta(days=1)
            continue

        # ── Extract Day 0 committed deliveries ──────────────────────
        day0 = routes.get(0, pd.DataFrame())
        opt_stops = 0
        opt_lbs = 0
        opt_miles = 0.0
        opt_fill_ratios = []

        if not day0.empty:
            opt_stops = len(day0)
            # Miles: sum per-truck route distances (each stop has the route total)
            for truck in day0['Truck'].unique():
                sub = day0[day0['Truck'] == truck]
                opt_miles += float(sub['Route_Dist_mi'].iloc[0])

            for _, stop in day0.iterrows():
                cid = stop['ID']
                refill = float(stop.get('Refill_lbs', 0))
                tank = float(stop.get('Tank_lbs', 0)) or tanks.get(cid, 0)
                opt_lbs += refill
                if tank > 0:
                    opt_fill_ratios.append(refill / tank * 100)

                # Execute: refill this client's tank
                inventory[cid] = min(tanks.get(cid, 10000),
                                     inventory.get(cid, 0) + refill)

        opt_avg_fill = sum(opt_fill_ratios) / len(opt_fill_ratios) if opt_fill_ratios else 0

        # ── Actual driver deliveries for this date ──────────────────
        actual = actual_by_date.get(date_str, [])
        drv_stops = len(actual)
        drv_lbs = sum(r['qty'] for r in actual)
        drv_fill_ratios = [r['qty'] / r['tank'] * 100 for r in actual if r['tank'] > 0]
        drv_avg_fill = sum(drv_fill_ratios) / len(drv_fill_ratios) if drv_fill_ratios else 0

        # Estimate driver miles (nearest-neighbor through their actual clients)
        drv_client_ids = [r['id'] for r in actual if r['id'] is not None]
        drv_miles = estimate_driver_miles(drv_client_ids, nim, dm)

        opt_per_stop = opt_lbs / opt_stops if opt_stops else 0
        drv_per_stop = drv_lbs / drv_stops if drv_stops else 0

        opt_days.append({
            'date': date_str,
            'day': day_name,
            'opt_stops': opt_stops,
            'drv_stops': drv_stops,
            'opt_lbs': int(opt_lbs),
            'drv_lbs': int(drv_lbs),
            'opt_lbs_per_stop': round(opt_per_stop),
            'drv_lbs_per_stop': round(drv_per_stop),
            'opt_fill_pct': round(opt_avg_fill, 1),
            'drv_fill_pct': round(drv_avg_fill, 1),
            'opt_miles': round(opt_miles, 1),
            'drv_miles': round(drv_miles, 1),
        })

        print(f"  {date_str:<12} {day_name:<4} {opt_stops:>6} {drv_stops:>6}"
              f" {int(opt_lbs):>8,} {int(drv_lbs):>8,}"
              f"  {opt_miles:>6.0f} {drv_miles:>6.0f}"
              f"  {opt_per_stop:>6.0f} {drv_per_stop:>6.0f}"
              f"  {opt_avg_fill:>5.0f}% {drv_avg_fill:>5.0f}%")

        # ── Deplete ALL tanks ───────────────────────────────────────
        for cid in inventory:
            inventory[cid] = max(0, inventory[cid] - rates.get(cid, 0))

        current_date += pd.Timedelta(days=1)

    # ── Summary ─────────────────────────────────────────────────────────
    print(f"  {'-' * 100}")

    tot_opt_stops = sum(d['opt_stops'] for d in opt_days)
    tot_drv_stops = sum(d['drv_stops'] for d in opt_days)
    tot_opt_lbs = sum(d['opt_lbs'] for d in opt_days)
    tot_drv_lbs = sum(d['drv_lbs'] for d in opt_days)
    tot_opt_miles = sum(d['opt_miles'] for d in opt_days)
    tot_drv_miles = sum(d['drv_miles'] for d in opt_days)
    avg_opt_fill = sum(d['opt_fill_pct'] for d in opt_days) / n_days
    avg_drv_fill = sum(d['drv_fill_pct'] for d in opt_days) / n_days

    print(f"\n  {'═' * 90}")
    print(f"  2-WEEK ROLLING TOTALS ({n_days} delivery days)")
    print(f"  {'═' * 90}")
    print(f"  Total miles:  Optimizer = {tot_opt_miles:>9.0f}   Drivers = {tot_drv_miles:>9.0f}   "
          f"Delta = {tot_opt_miles - tot_drv_miles:>+8.0f} ({(tot_opt_miles/tot_drv_miles-1)*100:>+.1f}%)" if tot_drv_miles else "")
    print(f"  Total lbs:    Optimizer = {tot_opt_lbs:>9,}   Drivers = {tot_drv_lbs:>9,}   "
          f"Delta = {tot_opt_lbs - tot_drv_lbs:>+8,} ({(tot_opt_lbs/tot_drv_lbs-1)*100:>+.1f}%)" if tot_drv_lbs else "")
    print(f"  Total stops:  Optimizer = {tot_opt_stops:>9}   Drivers = {tot_drv_stops:>9}   "
          f"Delta = {tot_opt_stops - tot_drv_stops:>+8}")
    if tot_opt_stops and tot_drv_stops:
        print(f"  Lbs/stop:     Optimizer = {tot_opt_lbs/tot_opt_stops:>9.0f}   Drivers = {tot_drv_lbs/tot_drv_stops:>9.0f}")
    if tot_opt_miles and tot_drv_miles:
        print(f"  Lbs/mile:     Optimizer = {tot_opt_lbs/tot_opt_miles:>9.1f}   Drivers = {tot_drv_lbs/tot_drv_miles:>9.1f}")
        print(f"  Miles/stop:   Optimizer = {tot_opt_miles/tot_opt_stops:>9.1f}   Drivers = {tot_drv_miles/tot_drv_stops:>9.1f}")
    print(f"  Avg fill%:    Optimizer = {avg_opt_fill:>8.0f}%   Drivers = {avg_drv_fill:>8.0f}%")

    if output_file:
        with open(output_file, 'w') as f:
            json.dump(opt_days, f, indent=2)
        print(f"\n  Results saved → {output_file}")

    return opt_days


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='Rolling 2-week simulation vs actual drivers')
    parser.add_argument('--start', default='2026-03-09',
                        help='Plan start date (evening before first delivery day)')
    parser.add_argument('--days', type=int, default=10, help='Number of delivery days')
    parser.add_argument('--solve-sec', type=int, default=8)
    parser.add_argument('--output', default=None)
    args = parser.parse_args()
    run_rolling_comparison(args.start, args.days, args.solve_sec, args.output)
