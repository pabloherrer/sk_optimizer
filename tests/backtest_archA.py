#!/usr/bin/env python3
"""
backtest_archA.py — Rolling backtest using Architecture A (two-phase pipeline).

Phase 1: scheduler.py (K-means clustering + EDF day assignment)
Phase 2: router.py (per-truck TSP)
Orchestrated via rolling_optimizer.py's RollingHorizonOptimizer.

Same rolling simulation as backtest_full.py but using the two-phase solver
instead of the unified solver. Results saved in same format for comparison.
"""

import sys, json, io, re
from pathlib import Path
from contextlib import redirect_stdout, redirect_stderr
from collections import defaultdict

import pandas as pd
import numpy as np

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from config import (
    INPUT_FILE, MATRIX_FILE, DAYS, NUM_DAYS,
    EXCLUDED_CLIENT_IDS, METERS_PER_MILE,
)
from load_data import load_all
from forecast_consumption import estimate_consumption_rates
from inventory import enrich_snapshot
from router import load_matrix
from state import initialise_state_from_snapshot
from schema_loaders import load_time_windows, load_closures, load_depot_config
from scheduler import assign_customers_to_days
from router import _solve_single_truck

_WEEKDAY_SHORT = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
RESULTS_FILE = Path(__file__).parent / 'backtest_archA_results.json'
STATE_FILE = Path(__file__).parent / 'backtest_archA_state.json'


def load_actual_deliveries(input_file):
    import openpyxl
    wb = openpyxl.load_workbook(input_file, data_only=True)
    ws = wb['Delivery_Log']
    ws_cl = wb['Client_List']
    tank_by_name = {}
    for row in ws_cl.iter_rows(min_row=4, values_only=True):
        if row[1]:
            tank_by_name[str(row[1]).strip()] = float(row[9] or 0)
    by_date = defaultdict(list)
    for row in ws.iter_rows(min_row=4, values_only=True):
        dt, cust = row[0], row[1]
        if dt and cust:
            cust_s = str(cust).strip()
            date_str = pd.Timestamp(dt).strftime('%Y-%m-%d')
            nid = extract_numeric_id(cust_s)
            qty = float(row[6] or 0)
            tank = float(row[5] or 0)
            if tank == 0:
                tank = tank_by_name.get(cust_s, 0)
            by_date[date_str].append({'id': nid, 'customer': cust_s, 'qty': qty, 'tank': tank})
    return by_date


def extract_numeric_id(customer_str):
    parts = customer_str.replace('-', ' ').split()
    for p in parts:
        m = re.match(r'^(\d{3,})[A-Za-z]?$', p)
        if m:
            return int(m.group(1))
    return None


def _nn_route(indices, dist_matrix, depot=0):
    if not indices:
        return 0.0
    unvisited = set(indices)
    current = depot
    total_m = 0.0
    while unvisited:
        best_dist = float('inf')
        best_node = None
        for n in unvisited:
            d = dist_matrix[current][n]
            if d < best_dist:
                best_dist = d
                best_node = n
        if best_node is None:
            break
        total_m += best_dist
        unvisited.remove(best_node)
        current = best_node
    total_m += dist_matrix[current][depot]
    return total_m


def estimate_driver_miles_2truck(client_ids, node_index_map, dist_matrix):
    indices = []
    for cid in client_ids:
        for key, idx in node_index_map.items():
            m = re.match(r'^(\d+)', str(key))
            if m and int(m.group(1)) == cid:
                indices.append(idx)
                break
    if not indices:
        return 0.0
    if len(indices) <= 3:
        return _nn_route(indices, dist_matrix) / METERS_PER_MILE
    depot = 0
    sorted_by_dist = sorted(indices, key=lambda i: dist_matrix[depot][i])
    truck_a, truck_b = [], []
    last_a, last_b = depot, depot
    for idx in sorted(indices, key=lambda i: -dist_matrix[depot][i]):
        dist_a = dist_matrix[last_a][idx]
        dist_b = dist_matrix[last_b][idx]
        if dist_a <= dist_b and len(truck_a) <= len(truck_b) + 3:
            truck_a.append(idx); last_a = idx
        elif len(truck_b) <= len(truck_a) + 3:
            truck_b.append(idx); last_b = idx
        else:
            truck_a.append(idx); last_a = idx
    return (_nn_route(truck_a, dist_matrix) + _nn_route(truck_b, dist_matrix)) / METERS_PER_MILE


def run_archA_comparison(start_date: str, end_date: str, batch_size: int = 99,
                         resume: bool = False):
    """Run rolling comparison using Architecture A (two-phase pipeline)."""

    # ── Load static data ────────────────────────────────────────────────
    clients_raw, deliveries_full = load_all(INPUT_FILE)
    dm, tm, nim = load_matrix(MATRIX_FILE)
    actual_by_date = load_actual_deliveries(INPUT_FILE)

    plan_start = pd.Timestamp(start_date).normalize()
    plan_end = pd.Timestamp(end_date).normalize()

    # ── Truncate deliveries ─────────────────────────────────────────────
    cutoff = plan_start + pd.Timedelta(days=1)
    deliveries_cut = deliveries_full[deliveries_full['Date'] < cutoff].copy()
    print(f"  Using {len(deliveries_cut)} deliveries up to {plan_start.date()}")

    clients_df = estimate_consumption_rates(deliveries_cut, clients_raw, today=plan_start)
    state = initialise_state_from_snapshot(clients_df)
    snapshot = enrich_snapshot(clients_df, state)

    # ── Try resume ──────────────────────────────────────────────────────
    saved = None
    if resume and STATE_FILE.exists():
        with open(STATE_FILE) as f:
            saved = json.load(f)
        if saved.get('start_date') == start_date:
            print(f"  Resuming from {saved['last_date']}...")

    # ── Build mutable inventories ───────────────────────────────────────
    opt_inv = {}
    drv_inv = {}
    rates = {}
    tanks = {}
    for _, row in snapshot.iterrows():
        cid = row['ID']
        cur = float(row.get('Current_lbs', 0))
        opt_inv[cid] = cur
        drv_inv[cid] = cur
        rates[cid] = float(row.get('Avg_LbsPerDay', 0))
        tanks[cid] = float(row.get('Tank_lbs', 0))

    # ── Apply saved state if resuming ───────────────────────────────────
    results = []
    current_date = plan_start + pd.Timedelta(days=1)
    if saved and saved.get('start_date') == start_date:
        results = saved['results']
        opt_inv = {k: float(v) for k, v in saved['opt_inventory'].items()}
        drv_inv = {k: float(v) for k, v in saved['drv_inventory'].items()}
        current_date = pd.Timestamp(saved['last_date']) + pd.Timedelta(days=1)

    # ── Override solve time for backtest speed ─────────────────────────
    import config as _cfg
    import router as _router
    _cfg.SOLVE_SEC = 3       # Match unified solver backtest time
    _router.SOLVE_SEC = 3    # Router reads at import time

    from config import TRUCK_NAMES

    workday_set = set(DAYS)
    excluded_nums = {int(x) for x in EXCLUDED_CLIENT_IDS}
    days_done = 0

    print(f"\n{'═' * 110}")
    print(f"  Architecture A (two-phase): {start_date} → {end_date}")
    print(f"{'═' * 110}")
    print(f"  {'Date':<12} {'Day':<4} {'OptStp':>6} {'DrvStp':>6}"
          f" {'OptLbs':>8} {'DrvLbs':>8}"
          f"  {'OptMi':>6} {'DrvMi':>6}"
          f"  {'Opt/s':>5} {'Drv/s':>5}"
          f"  {'OptFl':>5} {'DrvFl':>5}")
    print(f"  {'-' * 100}")

    while current_date <= plan_end and days_done < batch_size:
        day_name = _WEEKDAY_SHORT[current_date.weekday()]
        if day_name not in workday_set:
            for cid in opt_inv:
                opt_inv[cid] = max(0, opt_inv[cid] - rates.get(cid, 0))
            for cid in drv_inv:
                drv_inv[cid] = max(0, drv_inv[cid] - rates.get(cid, 0))
            current_date += pd.Timedelta(days=1)
            continue

        days_done += 1
        date_str = current_date.strftime('%Y-%m-%d')
        day_idx = DAYS.index(day_name)

        # ── Update snapshot with optimizer inventory ──────────────
        snap = snapshot.copy()
        for i, row in snap.iterrows():
            cid = row['ID']
            if cid in opt_inv:
                cur = max(0, opt_inv[cid])
                tank = tanks.get(cid, 1)
                rate = rates.get(cid, 0)
                snap.at[i, 'Current_lbs'] = cur
                snap.at[i, 'Refill_lbs'] = max(0, tank - cur)
                snap.at[i, 'Days_Until_Stockout'] = cur / rate if rate > 0 else 999
                from inventory import urgency_tier
                snap.at[i, 'Urgency'] = urgency_tier(cur / rate if rate > 0 else 999)

        # ── Phase 1: Day assignment ─────────────────────────────────
        opt_stops = 0
        opt_lbs = 0
        opt_miles = 0.0
        opt_fill_ratios = []

        try:
            buf = io.StringIO()
            with redirect_stdout(buf), redirect_stderr(buf):
                enriched = enrich_snapshot(clients_df, opt_inv)
                assignment = assign_customers_to_days(enriched, start_day=day_idx)

            # ── Phase 2: Route only today's assigned clients ────────
            day_clients = assignment[assignment['AssignedDayIndex'] == day_idx].copy()

            if len(day_clients) > 0:
                truck_routes = []
                for truck_name in TRUCK_NAMES:
                    tc = day_clients[day_clients['AssignedTruck'] == truck_name].copy()
                    if len(tc) == 0:
                        continue
                    buf2 = io.StringIO()
                    with redirect_stdout(buf2), redirect_stderr(buf2):
                        rt = _solve_single_truck(
                            tc, day_idx, truck_name, dm, tm, nim
                        )
                    if not rt.empty:
                        truck_routes.append(rt)

                if truck_routes:
                    day0 = pd.concat(truck_routes, ignore_index=True)
                else:
                    day0 = pd.DataFrame()

                if not day0.empty:
                    opt_stops = len(day0)
                    if 'Route_Dist_km' in day0.columns:
                        for truck in day0['Truck'].unique():
                            sub = day0[day0['Truck'] == truck]
                            opt_miles += float(sub['Route_Dist_km'].iloc[0]) / 1.60934
                    for _, stop in day0.iterrows():
                        cid = stop['ID']
                        refill = float(stop.get('Refill_lbs', 0))
                        tank_v = float(stop.get('Tank_lbs', 0)) or tanks.get(cid, 0)
                        opt_lbs += refill
                        if tank_v > 0:
                            opt_fill_ratios.append(refill / tank_v * 100)
                        opt_inv[cid] = min(tanks.get(cid, 10000),
                                           opt_inv.get(cid, 0) + refill)
        except Exception as e:
            print(f"  {date_str}  {day_name}  SOLVER ERROR: {e}")
            import traceback; traceback.print_exc()

        # ── Driver deliveries ───────────────────────────────────────
        actual_all = actual_by_date.get(date_str, [])
        actual = [r for r in actual_all if r['id'] not in excluded_nums]
        drv_stops = len(actual)
        drv_lbs = sum(r['qty'] for r in actual)
        drv_client_ids = [r['id'] for r in actual if r['id'] is not None]
        drv_miles = estimate_driver_miles_2truck(drv_client_ids, nim, dm)
        drv_fill_ratios = [r['qty'] / r['tank'] * 100 for r in actual
                           if r['tank'] > 0 and r['qty'] > 0 and r['qty'] <= r['tank'] * 1.05]

        # Apply driver deliveries to driver inventory
        for rec in actual:
            cid_num = rec['id']
            if cid_num is not None:
                for key in drv_inv:
                    m = re.match(r'^(\d+)', str(key))
                    if m and int(m.group(1)) == cid_num:
                        drv_inv[key] = min(tanks.get(key, 10000),
                                           drv_inv.get(key, 0) + rec['qty'])
                        break

        # ── Deplete ─────────────────────────────────────────────────
        for cid in opt_inv:
            opt_inv[cid] = max(0, opt_inv[cid] - rates.get(cid, 0))
        for cid in drv_inv:
            drv_inv[cid] = max(0, drv_inv[cid] - rates.get(cid, 0))

        opt_per_stop = opt_lbs / opt_stops if opt_stops else 0
        drv_per_stop = drv_lbs / drv_stops if drv_stops else 0
        opt_fill = sum(opt_fill_ratios) / len(opt_fill_ratios) if opt_fill_ratios else 0
        drv_fill = sum(drv_fill_ratios) / len(drv_fill_ratios) if drv_fill_ratios else 0

        day_result = {
            'date': date_str, 'day': day_name,
            'opt_stops': opt_stops, 'drv_stops': drv_stops,
            'opt_lbs': int(opt_lbs), 'drv_lbs': int(drv_lbs),
            'opt_miles': round(opt_miles, 1), 'drv_miles': round(drv_miles, 1),
            'opt_lbs_per_stop': round(opt_per_stop),
            'drv_lbs_per_stop': round(drv_per_stop),
            'opt_fill_pct': round(opt_fill, 1),
            'drv_fill_pct': round(drv_fill, 1),
        }
        results.append(day_result)

        print(f"  {date_str:<12} {day_name:<4} {opt_stops:>6} {drv_stops:>6}"
              f" {int(opt_lbs):>8,} {int(drv_lbs):>8,}"
              f"  {opt_miles:>6.0f} {drv_miles:>6.0f}"
              f"  {opt_per_stop:>5.0f} {drv_per_stop:>5.0f}"
              f"  {opt_fill:>4.0f}% {drv_fill:>4.0f}%")

        # Save state for resume
        try:
            with open(STATE_FILE, 'w') as sf:
                json.dump({
                    'start_date': start_date,
                    'last_date': date_str,
                    'results': results,
                    'opt_inventory': {k: float(v) for k, v in opt_inv.items()},
                    'drv_inventory': {k: float(v) for k, v in drv_inv.items()},
                }, sf)
        except Exception:
            pass

        current_date += pd.Timedelta(days=1)

    # ── Save ────────────────────────────────────────────────────────────
    with open(RESULTS_FILE, 'w') as f:
        json.dump(results, f, indent=2)

    # ── Summary ─────────────────────────────────────────────────────────
    n = len(results)
    if n == 0:
        return results

    print(f"  {'-' * 100}")
    s = lambda k: sum(d[k] for d in results)
    tot_om, tot_dm = s('opt_miles'), s('drv_miles')
    tot_ol, tot_dl = s('opt_lbs'), s('drv_lbs')
    tot_os, tot_ds = s('opt_stops'), s('drv_stops')

    print(f"\n  {'═' * 100}")
    print(f"  ARCH A TOTALS ({n} delivery days)")
    print(f"  {'═' * 100}")
    if tot_dm:
        print(f"  Total miles:  Opt = {tot_om:>8.0f}   Drv = {tot_dm:>8.0f}   Delta = {tot_om-tot_dm:>+8.0f} ({(tot_om/tot_dm-1)*100:>+.1f}%)")
    if tot_dl:
        print(f"  Total lbs:    Opt = {tot_ol:>8,}   Drv = {tot_dl:>8,}   Delta = {tot_ol-tot_dl:>+8,} ({(tot_ol/tot_dl-1)*100:>+.1f}%)")
    print(f"  Total stops:  Opt = {tot_os:>8}   Drv = {tot_ds:>8}   Delta = {tot_os-tot_ds:>+8}")
    if tot_os and tot_ds:
        print(f"  Lbs/stop:     Opt = {tot_ol/tot_os:>8.0f}   Drv = {tot_dl/tot_ds:>8.0f}")
    if tot_om and tot_dm:
        print(f"  Lbs/mile:     Opt = {tot_ol/tot_om:>8.1f}   Drv = {tot_dl/tot_dm:>8.1f}")
    print(f"\n  Results saved → {RESULTS_FILE}")
    return results


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--start', default='2026-03-01')
    parser.add_argument('--end', default='2026-04-21')
    parser.add_argument('--batch', type=int, default=99)
    parser.add_argument('--resume', action='store_true')
    args = parser.parse_args()
    run_archA_comparison(args.start, args.end, args.batch, args.resume)
