#!/usr/bin/env python3
"""
backtest.py — Compare optimizer committed routes vs actual historical deliveries.

For each test date:
  1. Use only delivery data up to that date (simulate "we're planning that afternoon")
  2. Run the optimizer → get committed client IDs for the next working day
  3. Compare to what was actually delivered on that next working day
  4. Report overlap, optimizer-only, and driver-only clients
"""

import sys, json, io
from pathlib import Path
from contextlib import redirect_stdout, redirect_stderr
from collections import defaultdict

import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from config import INPUT_FILE, MATRIX_FILE, SOLVE_SEC_WEEK
from load_data import load_all
from forecast_consumption import estimate_consumption_rates
from inventory import enrich_snapshot
from router import load_matrix
from schema_loaders import load_time_windows, load_closures, load_depot_config
from unified_solver import solve_horizon


# S&K works Tue–Sat
def next_working_day(dt):
    d = dt + pd.Timedelta(days=1)
    while d.weekday() in (0, 6):  # skip Mon=0, Sun=6
        d += pd.Timedelta(days=1)
    return d


def load_actual_deliveries(input_file):
    """Load all deliveries with qty and tank size.
    Returns:
      by_date: date_str -> set of customer strings
      by_date_detail: date_str -> list of {id, customer, qty, tank, product}
      all_records: list of (Timestamp, customer_str)
    """
    import openpyxl
    wb = openpyxl.load_workbook(input_file, data_only=True)
    ws = wb['Delivery_Log']

    # Build tank lookup from Client_List
    ws_cl = wb['Client_List']
    tank_by_name = {}
    for row in ws_cl.iter_rows(min_row=4, values_only=True):
        cust_name = row[1]
        if cust_name:
            tank_by_name[str(cust_name).strip()] = float(row[9] or 0)  # col J = Tank Size

    by_date = defaultdict(set)
    by_date_detail = defaultdict(list)
    all_records = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        dt, cust = row[0], row[1]
        if dt and cust:
            cust_s = str(cust).strip()
            date_str = pd.Timestamp(dt).strftime('%Y-%m-%d')
            by_date[date_str].add(cust_s)
            all_records.append((pd.Timestamp(dt), cust_s))

            nid = extract_numeric_id(cust_s)
            qty = float(row[6] or 0)   # col G = Qty Delivered (lbs)
            tank = float(row[5] or 0)  # col F = Tank (lbs)
            # If tank is 0 from formula, fall back to Client_List
            if tank == 0:
                tank = tank_by_name.get(cust_s, 0)

            by_date_detail[date_str].append({
                'id': nid,
                'customer': cust_s,
                'qty': qty,
                'tank': tank,
            })
    return by_date, by_date_detail, all_records


def extract_numeric_id(customer_str):
    """Extract the numeric ID from a customer string like 'AJO - 1018 - AJO ALS BELL'."""
    import re
    parts = customer_str.replace('-', ' ').split()
    for p in parts:
        # Match numeric IDs, possibly with a trailing letter (e.g. 4015A)
        m = re.match(r'^(\d{3,})[A-Za-z]?$', p)
        if m:
            return int(m.group(1))
    return None


def run_one_backtest(plan_date_str, all_records, by_date, by_date_detail, dm, tm, nim, tw, cl, dc,
                     clients_raw, solve_sec=10):
    """Run optimizer for one date, return comparison dict."""
    plan_date = pd.Timestamp(plan_date_str)
    commit_date = next_working_day(plan_date)
    commit_str = commit_date.strftime('%Y-%m-%d')

    # Filter deliveries to only before plan_date (end of that day)
    cutoff = plan_date + pd.Timedelta(days=1)  # include plan_date's deliveries
    filtered = [(dt, cust) for dt, cust in all_records if dt < cutoff]

    # Build a deliveries DataFrame matching what load_all returns
    if not filtered:
        return None

    deliveries = pd.DataFrame(filtered, columns=['Date', 'Customer'])
    # We need the full deliveries df that estimate_consumption_rates expects
    # Re-load but filter
    clients_raw_copy = clients_raw.copy()

    # estimate_consumption_rates needs: deliveries with Date, Customer, Qty columns at minimum
    # Let's just call load_all and filter the deliveries df by date
    clients_full, deliveries_full = load_all(INPUT_FILE)
    deliveries_cut = deliveries_full[deliveries_full['Date'] < cutoff].copy()

    clients_df = estimate_consumption_rates(deliveries_cut, clients_full, today=plan_date)

    # Initialise state from snapshot (no persisted state in backtest)
    from state import initialise_state_from_snapshot
    state = initialise_state_from_snapshot(clients_df)
    snapshot = enrich_snapshot(clients_df, state)

    # Solve
    buf = io.StringIO()
    with redirect_stdout(buf), redirect_stderr(buf):
        committed, tentative, deferred = solve_horizon(
            clients_df=snapshot,
            dist_matrix=dm,
            time_matrix_min=tm,
            node_index_map=nim,
            today=plan_date,
            horizon_days=5,
            commit_days=1,
            solve_seconds=solve_sec,
            time_windows_df=tw,
            closures_df=cl,
            depot_config=dc,
        )

    # Extract committed client IDs (as ints for consistent comparison)
    optimizer_ids = set()
    optimizer_customers = set()
    for d, df in committed.items():
        if not df.empty:
            for x in df['ID'].tolist():
                    import re as _re
                    m = _re.match(r'^(\d+)', str(x))
                    if m:
                        optimizer_ids.add(int(m.group(1)))
            if 'Customer' in df.columns:
                optimizer_customers.update(df['Customer'].tolist())

    # Actual deliveries on the commit date
    actual_customers = by_date.get(commit_str, set())
    actual_ids = set()
    for c in actual_customers:
        nid = extract_numeric_id(c)
        if nid:
            actual_ids.add(int(nid))

    # Compare
    overlap = optimizer_ids & actual_ids
    optimizer_only = optimizer_ids - actual_ids
    driver_only = actual_ids - optimizer_ids

    # ── Optimizer metrics (from committed routes) ──────────────────────
    opt_total_lbs = 0
    opt_fill_ratios = []  # refill_lbs / tank_lbs per client
    opt_per_client = []   # lbs per client
    for d, df in committed.items():
        if not df.empty:
            for _, row in df.iterrows():
                refill = float(row.get('Refill_lbs', 0))
                tank = float(row.get('Tank_lbs', 0))
                opt_total_lbs += refill
                opt_per_client.append(refill)
                if tank > 0:
                    opt_fill_ratios.append(round(refill / tank * 100, 1))

    # ── Driver metrics (from actual delivery log) ────────────────────
    actual_detail = by_date_detail.get(commit_str, [])
    drv_total_lbs = 0
    drv_fill_ratios = []  # qty / tank per client
    drv_per_client = []
    drv_only_lbs = 0
    drv_only_fill_ratios = []
    drv_only_per_client = []

    for rec in actual_detail:
        qty = rec['qty']
        tank = rec['tank']
        nid = rec['id']
        drv_total_lbs += qty
        drv_per_client.append(qty)
        if tank > 0:
            drv_fill_ratios.append(round(qty / tank * 100, 1))
        # Driver-only subset
        if nid and nid in driver_only:
            drv_only_lbs += qty
            drv_only_per_client.append(qty)
            if tank > 0:
                drv_only_fill_ratios.append(round(qty / tank * 100, 1))

    # ── DTE lookup from snapshot ─────────────────────────────────────
    dte_lookup = {}
    if snapshot is not None and not snapshot.empty:
        import re as _re
        for _, row in snapshot.iterrows():
            _m = _re.match(r'^(\d+)', str(row['ID']))
            if _m:
                cid = int(_m.group(1))
                dte_lookup[cid] = round(float(row.get('Days_Until_Stockout', 999)) - 1, 1)

    def avg_dte(id_set):
        vals = [dte_lookup.get(c, 999) for c in id_set]
        valid = [v for v in vals if v < 900]
        return round(sum(valid) / len(valid), 1) if valid else None

    def safe_avg(lst):
        return round(sum(lst) / len(lst), 1) if lst else 0

    return {
        'plan_date': plan_date_str,
        'commit_date': commit_str,
        # Stops
        'opt_stops': len(optimizer_ids),
        'drv_stops': len(actual_ids),
        'overlap': len(overlap),
        'opt_only_stops': len(optimizer_only),
        'drv_only_stops': len(driver_only),
        'jaccard': round(len(overlap) / len(overlap | optimizer_ids | actual_ids), 3)
            if (overlap | optimizer_ids | actual_ids) else 0,
        # Total lbs
        'opt_total_lbs': int(opt_total_lbs),
        'drv_total_lbs': int(drv_total_lbs),
        # Avg lbs per client
        'opt_avg_lbs': safe_avg(opt_per_client),
        'drv_avg_lbs': safe_avg(drv_per_client),
        'drv_only_avg_lbs': safe_avg(drv_only_per_client),
        # Fill ratio (delivered / tank %) — how full was the delivery relative to tank
        'opt_avg_fill_ratio': safe_avg(opt_fill_ratios),
        'drv_avg_fill_ratio': safe_avg(drv_fill_ratios),
        'drv_only_avg_fill_ratio': safe_avg(drv_only_fill_ratios),
        # DTE
        'opt_avg_dte': avg_dte(optimizer_ids),
        'drv_avg_dte': avg_dte(actual_ids),
        'drv_only_avg_dte': avg_dte(driver_only),
    }


def main(plan_dates, solve_sec=10, output_file=None):
    # Load static data once
    clients_raw, deliveries_full = load_all(INPUT_FILE)
    dm, tm, nim = load_matrix(MATRIX_FILE)
    tw = load_time_windows(INPUT_FILE)
    cl = load_closures(INPUT_FILE)
    dc = load_depot_config(INPUT_FILE)

    by_date, by_date_detail, all_records = load_actual_deliveries(INPUT_FILE)

    results = []
    for pd_str in plan_dates:
        print(f"\n{'─'*60}")
        print(f"  Backtesting: plan={pd_str} → commit={next_working_day(pd.Timestamp(pd_str)).strftime('%Y-%m-%d')}")
        print(f"{'─'*60}")
        try:
            r = run_one_backtest(pd_str, all_records, by_date, by_date_detail,
                                dm, tm, nim, tw, cl, dc, clients_raw, solve_sec=solve_sec)
            if r:
                results.append(r)
                print(f"  Stops: Opt={r['opt_stops']}  Real={r['drv_stops']}  Overlap={r['overlap']}  Jaccard={r['jaccard']:.0%}")
                print(f"  Total lbs:  Opt={r['opt_total_lbs']:,}  Real={r['drv_total_lbs']:,}")
                print(f"  Avg lbs/stop:  Opt={r['opt_avg_lbs']:.0f}  Real={r['drv_avg_lbs']:.0f}  Drv-only={r['drv_only_avg_lbs']:.0f}")
                print(f"  Avg fill ratio (del/tank):  Opt={r['opt_avg_fill_ratio']:.0f}%  Real={r['drv_avg_fill_ratio']:.0f}%  Drv-only={r['drv_only_avg_fill_ratio']:.0f}%")
                print(f"  Avg DTE:  Opt={r['opt_avg_dte']}  Real={r['drv_avg_dte']}  Drv-only={r['drv_only_avg_dte']}")
        except Exception as e:
            print(f"  ERROR: {e}")
            import traceback; traceback.print_exc()

    # Summary
    if results:
        print(f"\n{'═'*110}")
        print(f"  BACKTEST SUMMARY ({len(results)} dates)")
        print(f"{'═'*110}")
        print(f"  {'Plan':<11} {'Commit':<11} {'Stops':>10} {'Ovlp':>4} {'Jacc':>5}"
              f"  {'Opt lbs':>8} {'Drv lbs':>8}"
              f"  {'Opt/stp':>7} {'Drv/stp':>7}"
              f"  {'Opt%tk':>6} {'Drv%tk':>6} {'DO%tk':>5}"
              f"  {'OptDTE':>6} {'DrvDTE':>6}")
        print(f"  {'─'*108}")
        for r in results:
            print(f"  {r['plan_date']:<11} {r['commit_date']:<11} {r['opt_stops']:>4}/{r['drv_stops']:<4} {r['overlap']:>4} {r['jaccard']:>4.0%}"
                  f"  {r['opt_total_lbs']:>8,} {r['drv_total_lbs']:>8,}"
                  f"  {r['opt_avg_lbs']:>7.0f} {r['drv_avg_lbs']:>7.0f}"
                  f"  {r['opt_avg_fill_ratio']:>5.0f}% {r['drv_avg_fill_ratio']:>5.0f}% {r['drv_only_avg_fill_ratio']:>4.0f}%"
                  f"  {r['opt_avg_dte'] or 0:>6.1f} {r['drv_avg_dte'] or 0:>6.1f}")

        n = len(results)
        def a(key): return sum(r[key] for r in results) / n
        def ad(key): return sum(r[key] or 0 for r in results) / n
        print(f"  {'─'*108}")
        print(f"  {'AVERAGE':<26} {a('opt_stops'):>4.0f}/{a('drv_stops'):<4.0f} {a('overlap'):>4.0f} {a('jaccard'):>4.0%}"
              f"  {a('opt_total_lbs'):>8,.0f} {a('drv_total_lbs'):>8,.0f}"
              f"  {a('opt_avg_lbs'):>7.0f} {a('drv_avg_lbs'):>7.0f}"
              f"  {a('opt_avg_fill_ratio'):>5.0f}% {a('drv_avg_fill_ratio'):>5.0f}% {a('drv_only_avg_fill_ratio'):>4.0f}%"
              f"  {ad('opt_avg_dte'):>6.1f} {ad('drv_avg_dte'):>6.1f}")

    if output_file and results:
        with open(output_file, 'w') as f:
            json.dump(results, f, indent=2)
        print(f"\n  Results saved → {output_file}")

    return results


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--dates', nargs='+', required=True)
    parser.add_argument('--solve-sec', type=int, default=10)
    parser.add_argument('--output', default=None)
    args = parser.parse_args()
    main(args.dates, args.solve_sec, args.output)
