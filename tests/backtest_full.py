#!/usr/bin/env python3
"""
backtest_full.py — Extended rolling comparison with stockout tracking.

Runs two parallel simulations from the same starting inventory:
  1. OPTIMIZER: solve each day, carry forward deliveries
  2. DRIVERS:   replay actual historical deliveries, carry forward

Both deplete all tanks daily by consumption rate. Tracks:
  - Miles, lbs, stops per day (both sides)
  - Stockout events: which clients hit 0 lbs and when (both sides)
  - Saves state to disk after each day for resumable execution

Usage:
    python tests/backtest_full.py --start 2026-03-01 --end 2026-04-21 --solve-sec 3
    python tests/backtest_full.py --resume  # resume from saved state
"""

import sys, json, io, re, os
from pathlib import Path
from contextlib import redirect_stdout, redirect_stderr
from collections import defaultdict

import pandas as pd
import numpy as np

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from config import (
    INPUT_FILE, MATRIX_FILE, DAYS, NUM_DAYS,
    EXCLUDED_CLIENT_IDS, METERS_PER_MILE,
)
from load_data import load_all
from forecast_consumption import estimate_consumption_rates
from inventory import enrich_snapshot
from router import load_matrix
from state import initialise_state_from_snapshot
from schema_loaders import load_time_windows, load_closures, load_depot_config
from unified_solver import solve_week
from run_unified import compute_plan_dates

_WEEKDAY_SHORT = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']

STATE_FILE = Path(__file__).parent / 'backtest_state.json'
RESULTS_FILE = Path(__file__).parent / 'backtest_full_results.json'


def load_actual_deliveries(input_file):
    """Load actual deliveries with qty and tank size."""
    import openpyxl
    wb = openpyxl.load_workbook(input_file, data_only=True)
    ws = wb['Delivery_Log']
    ws_cl = wb['Client_List']
    tank_by_name = {}
    for row in ws_cl.iter_rows(min_row=4, values_only=True):
        cust_name = row[1]
        if cust_name:
            tank_by_name[str(cust_name).strip()] = float(row[9] or 0)

    by_date = defaultdict(list)
    for row in ws.iter_rows(min_row=4, values_only=True):
        dt, cust = row[0], row[1]
        if dt and cust:
            cust_s = str(cust).strip()
            date_str = pd.Timestamp(dt).strftime('%Y-%m-%d')
            nid = extract_numeric_id(cust_s)
            qty = float(row[6] or 0)
            tank = float(row[5] or 0)
            if tank == 0:
                tank = tank_by_name.get(cust_s, 0)
            by_date[date_str].append({
                'id': nid,
                'customer': cust_s,
                'qty': qty,
                'tank': tank,
            })
    return by_date


def extract_numeric_id(customer_str):
    parts = customer_str.replace('-', ' ').split()
    for p in parts:
        m = re.match(r'^(\d{3,})[A-Za-z]?$', p)
        if m:
            return int(m.group(1))
    return None


def _nn_route(indices, dist_matrix, depot=0):
    """Nearest-neighbor route through indices, return total meters."""
    if not indices:
        return 0.0
    unvisited = set(indices)
    current = depot
    total_m = 0.0
    while unvisited:
        best_dist = float('inf')
        best_node = None
        for n in unvisited:
            d = dist_matrix[current][n]
            if d < best_dist:
                best_dist = d
                best_node = n
        if best_node is None:
            break
        total_m += best_dist
        unvisited.remove(best_node)
        current = best_node
    total_m += dist_matrix[current][depot]
    return total_m


def _ids_to_indices(client_ids, node_index_map):
    """Map numeric client IDs to distance matrix indices."""
    indices = []
    for cid in client_ids:
        for key, idx in node_index_map.items():
            m = re.match(r'^(\d+)', str(key))
            if m and int(m.group(1)) == cid:
                indices.append(idx)
                break
    return indices


def estimate_driver_miles_2truck(client_ids, node_index_map, dist_matrix):
    """
    Estimate driver miles using 2 trucks (matching real operations).
    Split clients into 2 clusters by distance from depot, then run
    nearest-neighbor on each. Returns total miles for both trucks.

    This is a best-case estimate — real drivers likely drive more.
    """
    indices = _ids_to_indices(client_ids, node_index_map)
    if not indices:
        return 0.0
    if len(indices) <= 3:
        # Too few stops to split — run as one route
        return _nn_route(indices, dist_matrix) / METERS_PER_MILE

    # Split into 2 clusters: sort by distance from depot, alternate assignment
    # (simple farthest-insertion split that balances both distance and count)
    depot = 0
    sorted_by_dist = sorted(indices, key=lambda i: dist_matrix[depot][i])

    # Greedy 2-partition: assign each stop to the truck whose current
    # last-stop is closer (nearest-neighbor per truck in parallel)
    truck_a = []
    truck_b = []
    last_a = depot
    last_b = depot

    for idx in sorted(indices, key=lambda i: -dist_matrix[depot][i]):
        # Assign farthest-first to the truck that's closer to this stop
        dist_a = dist_matrix[last_a][idx]
        dist_b = dist_matrix[last_b][idx]
        if dist_a <= dist_b and len(truck_a) <= len(truck_b) + 3:
            truck_a.append(idx)
            last_a = idx
        elif len(truck_b) <= len(truck_a) + 3:
            truck_b.append(idx)
            last_b = idx
        else:
            truck_a.append(idx)
            last_a = idx

    miles_a = _nn_route(truck_a, dist_matrix) / METERS_PER_MILE
    miles_b = _nn_route(truck_b, dist_matrix) / METERS_PER_MILE
    return miles_a + miles_b


def get_stockouts(inventory, rates, tanks, excluded):
    """Return list of client IDs currently at stockout (0 lbs, positive rate)."""
    out = []
    for cid, lvl in inventory.items():
        if (lvl <= 0 and rates.get(cid, 0) > 0
                and tanks.get(cid, 0) > 0
                and cid not in excluded):
            out.append(cid)
    return out


def deplete_all(inventory, rates):
    """Deplete all tanks by one day of consumption."""
    for cid in inventory:
        inventory[cid] = max(0, inventory[cid] - rates.get(cid, 0))


def apply_deliveries(inventory, tanks, deliveries_list):
    """Apply a list of {id, qty} deliveries to inventory."""
    for d in deliveries_list:
        cid = d['id']
        if cid is not None and cid in inventory:
            inventory[cid] = min(tanks.get(cid, 10000),
                                 inventory.get(cid, 0) + d['qty'])


def save_state(state_dict):
    """Save simulation state for resuming."""
    with open(STATE_FILE, 'w') as f:
        json.dump(state_dict, f)


def load_state():
    """Load saved simulation state."""
    if STATE_FILE.exists():
        with open(STATE_FILE) as f:
            return json.load(f)
    return None


def run_full_comparison(start_date: str, end_date: str, solve_sec: int = 3,
                        resume: bool = False, batch_size: int = 99):
    """
    Run rolling comparison from start_date to end_date with stockout tracking.
    Both optimizer and driver inventories start from the same snapshot.
    """
    # ── Load static data ────────────────────────────────────────────────
    clients_raw, deliveries_full = load_all(INPUT_FILE)
    dm, tm, nim = load_matrix(MATRIX_FILE)
    tw = load_time_windows(INPUT_FILE)
    cl = load_closures(INPUT_FILE)
    dc = load_depot_config(INPUT_FILE)
    actual_by_date = load_actual_deliveries(INPUT_FILE)

    plan_start = pd.Timestamp(start_date).normalize()
    plan_end = pd.Timestamp(end_date).normalize()

    # ── Try to resume ───────────────────────────────────────────────────
    saved = load_state() if resume else None
    if saved and saved.get('start_date') == start_date:
        print(f"  Resuming from {saved['last_date']}...")
        results = saved['results']
        opt_inv = {k: float(v) for k, v in saved['opt_inventory'].items()}
        drv_inv = {k: float(v) for k, v in saved['drv_inventory'].items()}
        rates = {k: float(v) for k, v in saved['rates'].items()}
        tanks = {k: float(v) for k, v in saved['tanks'].items()}
        last_date = pd.Timestamp(saved['last_date'])
        current_date = last_date + pd.Timedelta(days=1)
        # We need snapshot for the solver — rebuild it
        cutoff = plan_start + pd.Timedelta(days=1)
        deliveries_cut = deliveries_full[deliveries_full['Date'] < cutoff].copy()
        clients_df = estimate_consumption_rates(deliveries_cut, clients_raw, today=plan_start)
        state = initialise_state_from_snapshot(clients_df)
        snapshot = enrich_snapshot(clients_df, state)
    else:
        # ── Fresh start ─────────────────────────────────────────────────
        cutoff = plan_start + pd.Timedelta(days=1)
        deliveries_cut = deliveries_full[deliveries_full['Date'] < cutoff].copy()
        print(f"  Using {len(deliveries_cut)} deliveries up to {plan_start.date()}")

        clients_df = estimate_consumption_rates(deliveries_cut, clients_raw, today=plan_start)
        state = initialise_state_from_snapshot(clients_df)
        snapshot = enrich_snapshot(clients_df, state)

        # Both sides start with identical inventory
        opt_inv = {}
        drv_inv = {}
        rates = {}
        tanks = {}
        for _, row in snapshot.iterrows():
            cid = row['ID']
            cur = float(row.get('Current_lbs', 0))
            opt_inv[cid] = cur
            drv_inv[cid] = cur
            rates[cid] = float(row.get('Avg_LbsPerDay', 0))
            tanks[cid] = float(row.get('Tank_lbs', 0))

        results = []
        current_date = plan_start + pd.Timedelta(days=1)

    workday_set = set(DAYS)
    days_done = 0

    print(f"\n{'═' * 110}")
    print(f"  Full comparison: {start_date} → {end_date}")
    print(f"{'═' * 110}")
    print(f"  {'Date':<12} {'Day':<4} {'OptStp':>6} {'DrvStp':>6}"
          f" {'OptLbs':>8} {'DrvLbs':>8}"
          f"  {'OptMi':>6} {'DrvMi':>6}"
          f"  {'Opt/s':>5} {'Drv/s':>5}"
          f"  {'OptSO':>5} {'DrvSO':>5}")
    print(f"  {'-' * 100}")

    while current_date <= plan_end and days_done < batch_size:
        day_name = _WEEKDAY_SHORT[current_date.weekday()]
        if day_name not in workday_set:
            # Non-delivery day: deplete both inventories
            deplete_all(opt_inv, rates)
            deplete_all(drv_inv, rates)
            current_date += pd.Timedelta(days=1)
            continue

        days_done += 1
        date_str = current_date.strftime('%Y-%m-%d')

        # ── Update snapshot with OPTIMIZER inventory ────────────────
        snap = snapshot.copy()
        for i, row in snap.iterrows():
            cid = row['ID']
            if cid in opt_inv:
                cur = max(0, opt_inv[cid])
                tank = tanks.get(cid, 1)
                rate = rates.get(cid, 0)
                snap.at[i, 'Current_lbs'] = cur
                snap.at[i, 'Refill_lbs'] = max(0, tank - cur)
                snap.at[i, 'Days_Until_Stockout'] = cur / rate if rate > 0 else 999
                from inventory import urgency_tier
                snap.at[i, 'Urgency'] = urgency_tier(cur / rate if rate > 0 else 999)

        # ── Run solver ──────────────────────────────────────────────
        plan_evening = current_date - pd.Timedelta(days=1)
        plan_dates = compute_plan_dates(plan_evening)

        opt_stops = 0
        opt_lbs = 0
        opt_miles = 0.0

        try:
            buf = io.StringIO()
            with redirect_stdout(buf), redirect_stderr(buf):
                routes, deferred = solve_week(
                    snap, dm, tm, nim, solve_seconds=solve_sec,
                    time_windows_df=tw, closures_df=cl,
                    today=plan_evening,
                    depot_config=dc, plan_dates=plan_dates,
                )

            day0 = routes.get(0, pd.DataFrame())
            opt_fill_ratios = []
            if not day0.empty:
                opt_stops = len(day0)
                for truck in day0['Truck'].unique():
                    sub = day0[day0['Truck'] == truck]
                    opt_miles += float(sub['Route_Dist_mi'].iloc[0])
                for _, stop in day0.iterrows():
                    cid = stop['ID']
                    planned_refill = float(stop.get('Refill_lbs', 0))
                    tank_cap = float(stop.get('Tank_lbs', 0)) or tanks.get(cid, 0)
                    # Count ACTUAL oil delivered (capped at tank room), not planned.
                    # Planned refill uses projected levels from consumption rates,
                    # which may be inflated. Actual = min(planned, room_in_tank).
                    current_level = opt_inv.get(cid, 0)
                    room = max(0, tank_cap - current_level)
                    actual_refill = min(planned_refill, room)
                    opt_lbs += actual_refill
                    if tank_cap > 0:
                        opt_fill_ratios.append(actual_refill / tank_cap * 100)
                    opt_inv[cid] = min(tank_cap,
                                       current_level + planned_refill)
        except Exception as e:
            print(f"  {date_str}  {day_name}  SOLVER ERROR: {e}")

        # ── Apply actual driver deliveries to driver inventory ──────
        # Filter out Tucson/Flagstaff (excluded from optimizer — unfair to count)
        excluded_nums = {int(x) for x in EXCLUDED_CLIENT_IDS}
        actual_all = actual_by_date.get(date_str, [])
        actual = [r for r in actual_all if r['id'] not in excluded_nums]
        n_excluded = len(actual_all) - len(actual)

        drv_stops = len(actual)
        drv_lbs = sum(r['qty'] for r in actual)
        drv_client_ids = [r['id'] for r in actual if r['id'] is not None]
        drv_miles = estimate_driver_miles_2truck(drv_client_ids, nim, dm)
        drv_fill_ratios = [r['qty'] / r['tank'] * 100 for r in actual
                           if r['tank'] > 0 and r['qty'] > 0 and r['qty'] <= r['tank'] * 1.05]

        # Apply driver deliveries to driver inventory
        for rec in actual:
            cid_num = rec['id']
            if cid_num is not None:
                # Find the string key in drv_inv
                for key in drv_inv:
                    m = re.match(r'^(\d+)', str(key))
                    if m and int(m.group(1)) == cid_num:
                        drv_inv[key] = min(tanks.get(key, 10000),
                                           drv_inv.get(key, 0) + rec['qty'])
                        break

        # ── Deplete both inventories ────────────────────────────────
        deplete_all(opt_inv, rates)
        deplete_all(drv_inv, rates)

        # ── Count stockouts AFTER depletion ─────────────────────────
        opt_stockouts = get_stockouts(opt_inv, rates, tanks, EXCLUDED_CLIENT_IDS)
        drv_stockouts = get_stockouts(drv_inv, rates, tanks, EXCLUDED_CLIENT_IDS)

        opt_per_stop = opt_lbs / opt_stops if opt_stops else 0
        drv_per_stop = drv_lbs / drv_stops if drv_stops else 0
        opt_fill = sum(opt_fill_ratios) / len(opt_fill_ratios) if opt_fill_ratios else 0
        drv_fill = sum(drv_fill_ratios) / len(drv_fill_ratios) if drv_fill_ratios else 0

        day_result = {
            'date': date_str,
            'day': day_name,
            'opt_stops': opt_stops,
            'drv_stops': drv_stops,
            'opt_lbs': int(opt_lbs),
            'drv_lbs': int(drv_lbs),
            'opt_miles': round(opt_miles, 1),
            'drv_miles': round(drv_miles, 1),
            'opt_lbs_per_stop': round(opt_per_stop),
            'drv_lbs_per_stop': round(drv_per_stop),
            'opt_fill_pct': round(opt_fill, 1),
            'drv_fill_pct': round(drv_fill, 1),
            'opt_stockouts': len(opt_stockouts),
            'drv_stockouts': len(drv_stockouts),
            'opt_stockout_ids': opt_stockouts[:5],
            'drv_stockout_ids': drv_stockouts[:5],
        }
        results.append(day_result)

        print(f"  {date_str:<12} {day_name:<4} {opt_stops:>6} {drv_stops:>6}"
              f" {int(opt_lbs):>8,} {int(drv_lbs):>8,}"
              f"  {opt_miles:>6.0f} {drv_miles:>6.0f}"
              f"  {opt_per_stop:>5.0f} {drv_per_stop:>5.0f}"
              f"  {len(opt_stockouts):>5} {len(drv_stockouts):>5}")

        # ── Save state after each day for resume ────────────────────
        save_state({
            'start_date': start_date,
            'end_date': end_date,
            'last_date': date_str,
            'results': results,
            'opt_inventory': {k: float(v) for k, v in opt_inv.items()},
            'drv_inventory': {k: float(v) for k, v in drv_inv.items()},
            'rates': {k: float(v) for k, v in rates.items()},
            'tanks': {k: float(v) for k, v in tanks.items()},
        })

        current_date += pd.Timedelta(days=1)

    # ── Save results ────────────────────────────────────────────────────
    with open(RESULTS_FILE, 'w') as f:
        json.dump(results, f, indent=2)

    # ── Print summary ───────────────────────────────────────────────────
    n = len(results)
    if n == 0:
        print("  No results.")
        return results

    print(f"  {'-' * 100}")

    tot_opt_mi = sum(d['opt_miles'] for d in results)
    tot_drv_mi = sum(d['drv_miles'] for d in results)
    tot_opt_lbs = sum(d['opt_lbs'] for d in results)
    tot_drv_lbs = sum(d['drv_lbs'] for d in results)
    tot_opt_stp = sum(d['opt_stops'] for d in results)
    tot_drv_stp = sum(d['drv_stops'] for d in results)
    max_opt_so = max(d['opt_stockouts'] for d in results)
    max_drv_so = max(d['drv_stockouts'] for d in results)
    avg_opt_so = sum(d['opt_stockouts'] for d in results) / n
    avg_drv_so = sum(d['drv_stockouts'] for d in results) / n

    print(f"\n  {'═' * 100}")
    print(f"  FULL COMPARISON ({n} delivery days, {start_date} → {results[-1]['date']})")
    print(f"  {'═' * 100}")
    if tot_drv_mi:
        print(f"  Total miles:    Opt = {tot_opt_mi:>8.0f}   Drv = {tot_drv_mi:>8.0f}   "
              f"Delta = {tot_opt_mi - tot_drv_mi:>+8.0f} ({(tot_opt_mi/tot_drv_mi-1)*100:>+.1f}%)")
    if tot_drv_lbs:
        print(f"  Total lbs:      Opt = {tot_opt_lbs:>8,}   Drv = {tot_drv_lbs:>8,}   "
              f"Delta = {tot_opt_lbs - tot_drv_lbs:>+8,} ({(tot_opt_lbs/tot_drv_lbs-1)*100:>+.1f}%)")
    print(f"  Total stops:    Opt = {tot_opt_stp:>8}   Drv = {tot_drv_stp:>8}   "
          f"Delta = {tot_opt_stp - tot_drv_stp:>+8}")
    if tot_opt_stp and tot_drv_stp:
        print(f"  Lbs/stop:       Opt = {tot_opt_lbs/tot_opt_stp:>8.0f}   Drv = {tot_drv_lbs/tot_drv_stp:>8.0f}")
    if tot_opt_mi and tot_drv_mi:
        print(f"  Lbs/mile:       Opt = {tot_opt_lbs/tot_opt_mi:>8.1f}   Drv = {tot_drv_lbs/tot_drv_mi:>8.1f}")
        print(f"  Miles/stop:     Opt = {tot_opt_mi/tot_opt_stp:>8.1f}   Drv = {tot_drv_mi/tot_drv_stp:>8.1f}")
    print(f"\n  STOCKOUTS:")
    print(f"  Avg per day:    Opt = {avg_opt_so:>8.1f}   Drv = {avg_drv_so:>8.1f}")
    print(f"  Peak (max day): Opt = {max_opt_so:>8}   Drv = {max_drv_so:>8}")
    print(f"  Days with 0:    Opt = {sum(1 for d in results if d['opt_stockouts']==0):>8}/{n}"
          f"   Drv = {sum(1 for d in results if d['drv_stockouts']==0):>8}/{n}")

    done = current_date > plan_end
    if not done:
        print(f"\n  ⏸  Batch complete ({days_done} days). Run with --resume to continue.")
    else:
        print(f"\n  ✓ Complete!")
        # Clean up state file
        if STATE_FILE.exists():
            STATE_FILE.unlink()

    print(f"  Results saved → {RESULTS_FILE}")
    return results


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='Full rolling comparison with stockout tracking')
    parser.add_argument('--start', default='2026-03-01')
    parser.add_argument('--end', default='2026-04-21')
    parser.add_argument('--solve-sec', type=int, default=3)
    parser.add_argument('--resume', action='store_true')
    parser.add_argument('--batch', type=int, default=99,
                        help='Max delivery days per run (for chunking)')
    args = parser.parse_args()
    run_full_comparison(args.start, args.end, args.solve_sec, args.resume, args.batch)
