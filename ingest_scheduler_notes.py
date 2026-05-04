#!/usr/bin/env python3
"""
ingest_scheduler_notes.py — Pull ANOVA flags, time windows, and closures
from SK's "NEW SCHEDULER" workbook into our SK_Delivery_System.xlsx.

The scheduler file (SmartService export) has free-text notes mixed in
with the delivery log. We parse:

  • ANOVA flag — clients with sensor-monitored tanks
  • DELIVERY time windows — e.g. "DELIVERY 9AM - 10AM"
  • Closures — e.g. "CLOSED TUESDAY", "CLOSED PERMANENTLY 4/20/2026"
  • Misc scheduling notes — e.g. "DO NOT PUT IN ROUTE SCHEDULE"

Outputs:
  • Updates `Client_List` sheet with new ANOVA + Notes columns
  • Adds rows to `Client_Time_Windows` for each parsed time-window
  • Adds rows to `Client_Closures` for permanent closures
  • Stores recurring weekday-closures as time-window rows with a
    closed-all-day flag

Idempotent: re-running with the same scheduler file produces no
duplicate rows.
"""

from __future__ import annotations

import argparse
import re
import sys
from datetime import datetime, time as dtime
from pathlib import Path
import openpyxl

# ─── Paths ─────────────────────────────────────────────────────────────
SK_FILE_DEFAULT = Path('data/SK_Delivery_System.xlsx')

# ─── Regex for client headers ──────────────────────────────────────────
CLIENT_HEADER_RE = re.compile(r'^[A-Z0-9]+\s*-\s*(\d+)\s*-\s*')

# ─── Time-window parsing ───────────────────────────────────────────────
DAY_NAMES = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']


def parse_time(s: str) -> dtime | None:
    """Parse '9AM', '8:30', '9:15 AM', '12' into a time. Defaults AM if
    bare hour < 12 and ambiguous."""
    s = s.strip().upper()
    am_pm = None
    if 'AM' in s:
        am_pm = 'AM'; s = s.replace('AM', '').strip()
    elif 'PM' in s:
        am_pm = 'PM'; s = s.replace('PM', '').strip()
    s = s.rstrip('.')
    try:
        if ':' in s:
            h, m = s.split(':')
            h, m = int(h), int(m)
        else:
            h, m = int(s), 0
    except ValueError:
        return None
    if am_pm == 'PM' and h < 12:
        h += 12
    elif am_pm is None and h < 6:    # ambiguous "1" probably means 13:00
        h += 12
    if not (0 <= h <= 23 and 0 <= m <= 59):
        return None
    return dtime(h, m)


def time_to_min(t: dtime) -> int:
    return t.hour * 60 + t.minute


def min_to_hhmm(mins: int) -> str:
    return f'{mins // 60:02d}:{mins % 60:02d}'


def parse_delivery_window(note: str) -> tuple[int, int] | None:
    """
    Parse a delivery-window note. Returns (open_min, close_min) absolute
    minutes-from-midnight, or None.

    Recognised formats (case-insensitive):
      DELIVERY 9AM - 10AM
      DELIVERY 8:30 - 10 AM
      DELIVERY 9:15 - 9:45
      AFTER 9 AM            → (540, 1020)   [9:00 to 17:00]
      BEFORE 11 AM          → (360, 660)
    """
    n = note.upper().strip()
    # AFTER X AM/PM
    m = re.match(r'^AFTER\s+(\d{1,2}(?::\d{2})?\s*(?:AM|PM)?)', n)
    if m:
        t = parse_time(m.group(1))
        if t:
            return (time_to_min(t), 17 * 60)   # 5 PM closing
    m = re.match(r'^BEFORE\s+(\d{1,2}(?::\d{2})?\s*(?:AM|PM)?)', n)
    if m:
        t = parse_time(m.group(1))
        if t:
            return (6 * 60, time_to_min(t))    # 6 AM start
    # DELIVERY X - Y
    m = re.search(
        r'(?:DELIVERY|TIME)?\s*(\d{1,2}(?::\d{2})?)\s*(AM|PM)?\s*-\s*(\d{1,2}(?::\d{2})?)\s*(AM|PM)?',
        n,
    )
    if m:
        s1 = m.group(1) + ('' if m.group(2) is None else ' ' + m.group(2))
        s2 = m.group(3) + ('' if m.group(4) is None else ' ' + m.group(4))
        # If only the second has AM/PM, propagate it back to the first
        if m.group(2) is None and m.group(4) is not None:
            s1 = m.group(1) + ' ' + m.group(4)
        t1 = parse_time(s1)
        t2 = parse_time(s2)
        if t1 and t2 and time_to_min(t2) > time_to_min(t1):
            # Reject obviously-wrong ranges (e.g. store hours like 12-8 PM)
            duration = time_to_min(t2) - time_to_min(t1)
            if duration <= 4 * 60:    # delivery windows ≤ 4 hours
                return (time_to_min(t1), time_to_min(t2))
    return None


CLOSED_DAY_RE = re.compile(
    r'CLOSED\s+(?:ON\s+|EVERY\s+)?(MON|TUE|WED|THU|FRI|SAT|SUN)(DAY)?S?',
    re.IGNORECASE,
)
CLOSED_PERM_RE = re.compile(
    r'CLOSED\s+PERMANENTLY\s+(\d{1,2}/\d{1,2}/\d{2,4})',
    re.IGNORECASE,
)


def parse_closure(note: str) -> dict | None:
    """
    Recognise:
      'CLOSED TUESDAY'           → recurring weekday
      'CLOSED ON TUESDAYS'       → recurring weekday
      'CLOSED EVERY TUESDAY'     → recurring weekday
      'CLOSED PERMANENTLY 4/20/2026'  → one-shot start date, no end
    Returns None if no closure detected.
    """
    n = note.upper().strip()
    m = CLOSED_PERM_RE.search(n)
    if m:
        try:
            d = datetime.strptime(m.group(1), '%m/%d/%Y')
        except ValueError:
            try:
                d = datetime.strptime(m.group(1), '%m/%d/%y')
            except ValueError:
                return None
        return {'kind': 'permanent', 'start': d, 'end': datetime(2099, 1, 1),
                'reason': 'CLOSED PERMANENTLY'}
    m = CLOSED_DAY_RE.search(n)
    if m:
        # Map MON→Monday etc
        prefix = m.group(1).title()
        day_short = prefix[:3]
        if day_short in DAY_NAMES:
            return {'kind': 'recurring_dow', 'day_short': day_short,
                    'reason': f'CLOSED {prefix.upper()}'}
    return None


# ─── Note extraction from scheduler workbook ────────────────────────────

def extract_notes(scheduler_path: Path) -> dict[str, list[str]]:
    """{client_id: [list of free-text notes]} from the scheduler file."""
    wb = openpyxl.load_workbook(scheduler_path, data_only=True)
    ws = wb['Sales by Product Service Detail']
    out: dict[str, list[str]] = {}
    current_cid: str | None = None
    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if a is None or not isinstance(a, str):
            continue
        a = a.strip()
        if not a:
            continue
        m = CLIENT_HEADER_RE.match(a)
        if m:
            current_cid = m.group(1)
            continue
        if current_cid is None:
            continue
        if a in ('#DIV/0!', '#REF!', ' '):
            continue
        if not re.search(r'[A-Za-z]', a):
            continue
        out.setdefault(current_cid, []).append(a[:200])
    return out


# ─── Update the SK Delivery System workbook ────────────────────────────

def update_sk_workbook(
    sk_path: Path,
    notes_per_client: dict[str, list[str]],
    *,
    dry_run: bool = False,
) -> dict:
    wb = openpyxl.load_workbook(sk_path)

    # 1. Find or create ANOVA / Notes columns in Client_List
    cl = wb['Client_List']
    header_row = 3   # adjust if different
    headers = {cl.cell(header_row, c).value: c for c in range(1, cl.max_column + 1)}
    if 'ANOVA' not in headers:
        new_col = cl.max_column + 1
        cl.cell(header_row, new_col, 'ANOVA')
        headers['ANOVA'] = new_col
    if 'Notes' not in headers:
        new_col = cl.max_column + 1
        cl.cell(header_row, new_col, 'Notes')
        headers['Notes'] = new_col
    if 'Do_Not_Schedule' not in headers:
        new_col = cl.max_column + 1
        cl.cell(header_row, new_col, 'Do_Not_Schedule')
        headers['Do_Not_Schedule'] = new_col
    anova_col = headers['ANOVA']
    notes_col = headers['Notes']
    skip_col  = headers['Do_Not_Schedule']

    # 2. Iterate clients in Client_List
    cid_to_row = {}
    for r in range(header_row + 1, cl.max_row + 1):
        cid_v = cl.cell(r, 1).value
        if cid_v is None:
            continue
        cid_to_row[str(cid_v).strip()] = r

    n_anova_set = 0
    n_notes_set = 0
    n_skip_set = 0
    parsed_time_windows = []   # list of dict for new TW rows
    parsed_closures = []       # list of dict for new closures
    auto_skip_clients: list = []   # clients to mark "Do_Not_Schedule"

    # Patterns that mean "this client should NOT be auto-scheduled"
    # (irregular service — handled manually by SK)
    SKIP_PATTERNS = [
        r'DO\s*NOT\s*(PUT|INCLUDE|SCHEDULE)',
        r'NOT\s*ON\s*ROUTE',
        r'CALL\s*BEFORE\s*DELIV',
        r'IRREGULAR\s*SCHEDULE',
        r'EVERY\s*\d+\s*MONTHS',           # rare visits
        r'BI[-\s]*MONTHLY',
        r'CLOSED\s*PERMANENTLY',
    ]
    skip_re = [re.compile(p, re.IGNORECASE) for p in SKIP_PATTERNS]

    for cid, notes in notes_per_client.items():
        cid_str = str(cid)
        if cid_str not in cid_to_row:
            continue
        r = cid_to_row[cid_str]

        # ANOVA flag
        is_anova = any('ANOVA' in n.upper() for n in notes)
        if is_anova:
            cl.cell(r, anova_col, 'Y')
            n_anova_set += 1

        # Do-Not-Schedule flag
        is_skip = any(any(pat.search(n) for pat in skip_re) for n in notes)
        if is_skip:
            cl.cell(r, skip_col, 'Y')
            auto_skip_clients.append(cid_str)
            n_skip_set += 1

        # Aggregate notes (strip the ANOVA token, that's now its own column;
        # strip throwaway notes like 'THIS IS CORRECT' that have no operational meaning)
        TRIVIAL = {'ANOVA', 'THIS IS CORRECT', 'USAGE IS CORRECT!',
                   'CORRECT', 'CHECK THJS USAGE & DELIVERY'}
        meaningful = [n for n in notes if n.strip().upper() not in TRIVIAL]
        if meaningful:
            joined = ' | '.join(meaningful)[:250]
            cl.cell(r, notes_col, joined)
            n_notes_set += 1

        # Parse time windows
        for n in notes:
            tw = parse_delivery_window(n)
            if tw:
                # Apply to all working days
                open_min, close_min = tw
                for day in DAY_NAMES[1:6]:   # Tue–Sat
                    parsed_time_windows.append({
                        'Client_ID': cid_str,
                        'Day_of_Week': day,
                        'Open_Min': open_min,
                        'Close_Min': close_min,
                    })
                break   # one window per client

        # Parse closures
        for n in notes:
            cl_info = parse_closure(n)
            if cl_info is None:
                continue
            if cl_info['kind'] == 'recurring_dow':
                # Compact representation: one row with Recurring_DOW set,
                # dates left blank. The loader expands at evaluation time.
                parsed_closures.append({
                    'Client_ID': cid_str,
                    'Recurring_DOW': cl_info['day_short'],
                    'Start_Date': None,
                    'End_Date': None,
                    'Reason': cl_info['reason'],
                })
            elif cl_info['kind'] == 'permanent':
                parsed_closures.append({
                    'Client_ID': cid_str,
                    'Start_Date': cl_info['start'],
                    'End_Date': cl_info['end'],
                    'Reason': cl_info['reason'],
                })

    # 3. Write time windows (Tue–Sat only; idempotent — clears existing rows
    #    for the affected clients before writing)
    affected_tw_clients = {tw['Client_ID'] for tw in parsed_time_windows}
    affected_cl_clients = {cl_['Client_ID'] for cl_ in parsed_closures}

    # ── Re-create Client_Time_Windows in COMPACT format ─────────────────
    # New schema (one row per rule):
    #   Col 1: Client_ID
    #   Col 2: Customer (name; for human readability)
    #   Col 3: Day_of_Week — 'Tue' / 'Wed' / 'Thu' / 'Fri' / 'Sat' / 'All'
    #   Col 4: Open_HHMM
    #   Col 5: Close_HHMM
    #   Col 6: Notes
    cust_lookup = {}
    for r in range(header_row + 1, cl.max_row + 1):
        cid_v = cl.cell(r, 1).value
        name_v = cl.cell(r, 2).value
        if cid_v is None:
            continue
        cust_lookup[str(cid_v).strip()] = str(name_v) if name_v else ''

    if 'Client_Time_Windows' in wb.sheetnames:
        del wb['Client_Time_Windows']
    tw_ws = wb.create_sheet('Client_Time_Windows')
    tw_ws.cell(1, 1, 'Client time windows — one row per client. Day_of_Week='
                     '"All" applies to every workday (Tue-Sat).')
    tw_ws.cell(2, 1, 'Edit by hand: add a row per client with restricted hours. '
                     'Empty sheet = no restrictions.')
    headers_tw = ['Client_ID', 'Customer', 'Day_of_Week',
                  'Open_HHMM', 'Close_HHMM', 'Notes']
    for i, h in enumerate(headers_tw, start=1):
        tw_ws.cell(3, i, h)
    tw_ws.cell(4, 1, 'EXAMPLE_C042')
    tw_ws.cell(4, 2, 'EXAMPLE - 0042 - Sample Restaurant')
    tw_ws.cell(4, 3, 'All')
    tw_ws.cell(4, 4, '09:00')
    tw_ws.cell(4, 5, '11:00')
    tw_ws.cell(4, 6, 'morning delivery only')

    # One row per UNIQUE (client_id, window). Aggregate parsed_time_windows
    # back to one row per client when the window is identical across days.
    by_client_window: dict[str, dict] = {}
    for tw in parsed_time_windows:
        if tw.get('_is_closed'):
            continue   # closure rows go to closures sheet
        cid = tw['Client_ID']
        key = (cid, tw['Open_Min'], tw['Close_Min'])
        by_client_window.setdefault(cid, {
            'Open_Min': tw['Open_Min'],
            'Close_Min': tw['Close_Min'],
            'days': set(),
        })['days'].add(tw['Day_of_Week'])

    for cid, info in by_client_window.items():
        days = info['days']
        # If covers all 5 workdays, write 'All'; otherwise list them
        workdays = {'Tue', 'Wed', 'Thu', 'Fri', 'Sat'}
        if days == workdays or 'All' in days:
            day_label = 'All'
        else:
            day_label = ','.join(sorted(days))
        # Find the original note text (for the Notes column)
        original_notes = notes_per_client.get(cid, [])
        note_text = ''
        for n in original_notes:
            if parse_delivery_window(n) == (info['Open_Min'], info['Close_Min']):
                note_text = n
                break
        tw_ws.append([
            cid,
            cust_lookup.get(cid, ''),
            day_label,
            min_to_hhmm(info['Open_Min']),
            min_to_hhmm(info['Close_Min']),
            note_text,
        ])

    # ── Re-create Client_Closures in COMPACT format ────────────────────
    # New schema (one row per rule):
    #   Col 1: Client_ID
    #   Col 2: Customer
    #   Col 3: Recurring_DOW — 'Tue'/'Wed'/etc. or empty for one-off date range
    #   Col 4: Start_Date    (only used if Recurring_DOW empty)
    #   Col 5: End_Date      (only used if Recurring_DOW empty)
    #   Col 6: Reason
    if 'Client_Closures' in wb.sheetnames:
        del wb['Client_Closures']
    cl_ws = wb.create_sheet('Client_Closures')
    cl_ws.cell(1, 1, 'Client closures — one row per rule. For weekly closures '
                     'set Recurring_DOW (e.g. "Tue") and leave dates blank. '
                     'For one-off ranges fill Start/End dates.')
    cl_ws.cell(2, 1, 'Edit by hand: add a row per closure rule. Empty sheet = '
                     'no closures.')
    headers_cl = ['Client_ID', 'Customer', 'Recurring_DOW',
                  'Start_Date', 'End_Date', 'Reason']
    for i, h in enumerate(headers_cl, start=1):
        cl_ws.cell(3, i, h)
    cl_ws.cell(4, 1, 'EXAMPLE_C042')
    cl_ws.cell(4, 2, 'EXAMPLE - 0042 - Sample Restaurant')
    cl_ws.cell(4, 3, 'Tue')
    cl_ws.cell(4, 4, None)
    cl_ws.cell(4, 5, None)
    cl_ws.cell(4, 6, 'closed every Tuesday (example — delete this row)')

    # Compact each parsed closure: dedup by (cid, kind, dow_or_dates)
    seen = set()
    compact_rows = []
    for cl_row in parsed_closures:
        cid = cl_row['Client_ID']
        # Detect recurring vs permanent vs one-off date range based on reason
        # (simple heuristic; ingest_actuals classifies them earlier)
        reason = cl_row.get('Reason', '')
        rec_dow = cl_row.get('Recurring_DOW') or ''
        sd = cl_row.get('Start_Date')
        ed = cl_row.get('End_Date')
        # Bucket by (cid, reason) so we collapse the 26 weekly rows we used
        # to write into ONE row when the reason was a recurring pattern.
        key = (cid, reason)
        if key in seen:
            continue
        seen.add(key)

        # Was this originally a recurring DOW closure?
        # Heuristic: reason contains 'TUE' or 'WED' etc., AND start==end
        rec_match = re.match(r'CLOSED\s+(MON|TUE|WED|THU|FRI|SAT|SUN)', reason or '')
        if rec_match:
            rec_dow = rec_match.group(1).title()
            sd, ed = None, None     # blanks → recurring
        compact_rows.append({
            'cid': cid, 'rec_dow': rec_dow,
            'sd': sd, 'ed': ed, 'reason': reason,
        })

    for row in compact_rows:
        cl_ws.append([
            row['cid'],
            cust_lookup.get(row['cid'], ''),
            row['rec_dow'],
            row['sd'],
            row['ed'],
            row['reason'],
        ])

    if dry_run:
        return {'anova': n_anova_set, 'notes': n_notes_set,
                'time_windows': len(parsed_time_windows),
                'closures': len(parsed_closures), 'saved': False}

    wb.save(sk_path)
    return {'anova': n_anova_set, 'notes': n_notes_set,
            'time_windows': len(parsed_time_windows),
            'closures': len(parsed_closures), 'saved': True}


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument('--scheduler', type=str, required=False,
                   default='/Users/pabloherrera/Downloads/NEW SCHEDULER copy apr 28.xlsx')
    p.add_argument('--sk-file', type=str, default=str(SK_FILE_DEFAULT))
    p.add_argument('--dry-run', action='store_true')
    args = p.parse_args()

    notes = extract_notes(Path(args.scheduler))
    print(f'Found notes for {len(notes)} clients in scheduler file.\n')

    # Show what we found
    print('--- ANOVA clients ---')
    anova = [c for c, ns in notes.items() if any('ANOVA' in n.upper() for n in ns)]
    print(f'  count: {len(anova)}')

    print('\n--- Time windows ---')
    n_tw = 0
    for cid, ns in sorted(notes.items()):
        for n in ns:
            tw = parse_delivery_window(n)
            if tw:
                h1 = f'{tw[0]//60:02d}:{tw[0]%60:02d}'
                h2 = f'{tw[1]//60:02d}:{tw[1]%60:02d}'
                print(f'  {cid:>6}  {h1}-{h2}  ({n})')
                n_tw += 1
                break
    print(f'  parsed: {n_tw}')

    print('\n--- Closures ---')
    n_cl = 0
    for cid, ns in sorted(notes.items()):
        for n in ns:
            cl_info = parse_closure(n)
            if cl_info:
                print(f'  {cid:>6}  {cl_info}')
                n_cl += 1
                break
    print(f'  parsed: {n_cl}')

    print('\nWriting to SK_Delivery_System.xlsx ...')
    stats = update_sk_workbook(Path(args.sk_file), notes, dry_run=args.dry_run)
    print(f'\n  ANOVA flagged:    {stats["anova"]}')
    print(f'  Notes attached:   {stats["notes"]}')
    print(f'  TW rows written:  {stats["time_windows"]}')
    print(f'  Closure rows:     {stats["closures"]}')
    print(f'  Saved:            {stats["saved"]}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
