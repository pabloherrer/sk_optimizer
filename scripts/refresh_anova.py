#!/usr/bin/env python3
"""
Refresh Anova tank data in SK_Delivery_System.xlsx.

Fetches latest readings from the Azure endpoint and writes them
into the Query sheet. All downstream formulas (Anova_Live →
Client_List → Optimizer_Input) update automatically when Excel
recalculates.

Usage:
    python refresh_anova.py              # default file path
    python refresh_anova.py path/to/SK_Delivery_System.xlsx
"""

import json, sys, urllib.request, os
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("ERROR: pip install openpyxl")
    sys.exit(1)

AZURE_URL = "https://skoil-anova-escpe7axeyczdshp.westcentralus-01.azurewebsites.net/api/anova/data"
DEFAULT_FILE = os.path.join(os.path.dirname(__file__), "data", "SK_Delivery_System.xlsx")

HEADERS = ["device_id", "customer", "display_value", "display_unit", "scaled_value",
           "scaled_unit", "product", "timestamp", "received_at"]


def fetch_anova():
    """Fetch latest readings from Azure, return list of dicts."""
    req = urllib.request.Request(AZURE_URL)
    with urllib.request.urlopen(req, timeout=15) as resp:
        data = json.loads(resp.read())
    # Filter to Level channel only
    return [r for r in data if r.get("channel_type", "") == "Level"]


def update_query_sheet(filepath, readings):
    """Write readings into the Query sheet of the workbook."""
    wb = openpyxl.load_workbook(filepath)
    
    if "Query" not in wb.sheetnames:
        print("ERROR: Query sheet not found")
        sys.exit(1)
    
    ws = wb["Query"]
    
    # Write headers in row 1
    for c, hdr in enumerate(HEADERS, 1):
        ws.cell(1, c).value = hdr
    
    # Clear old data (rows 2-100)
    for r in range(2, 101):
        for c in range(1, len(HEADERS) + 1):
            ws.cell(r, c).value = None
    
    # Write new data
    for i, reading in enumerate(readings):
        row = i + 2
        for c, field in enumerate(HEADERS, 1):
            ws.cell(row, c).value = reading.get(field, "")
    
    wb.save(filepath)
    return len(readings)


def main():
    filepath = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_FILE
    
    if not os.path.exists(filepath):
        print(f"ERROR: File not found: {filepath}")
        sys.exit(1)
    
    print(f"Fetching from {AZURE_URL} ...")
    readings = fetch_anova()
    print(f"  Got {len(readings)} Level readings")
    
    for r in readings:
        dev = r.get("device_id", "?")
        val = r.get("display_value", "?")
        ts  = r.get("timestamp", "?")
        print(f"    {dev}: {val} lbs @ {ts}")
    
    print(f"\nUpdating {os.path.basename(filepath)} ...")
    count = update_query_sheet(filepath, readings)
    print(f"  Wrote {count} rows to Query sheet")
    print(f"  Done at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("\nOpen Excel and formulas will auto-calculate with fresh data.")


if __name__ == "__main__":
    main()
