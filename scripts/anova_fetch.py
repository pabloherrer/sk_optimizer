"""
anova_fetch.py — Pull live tank readings from Azure and update local data
=========================================================================

Architecture
------------
Anova readings flow into a dedicated `Anova_Live` sheet inside
SK_Delivery_System.xlsx. The Client_List sheet's Anova columns (R-X) are
XLOOKUP formulas that pull from Anova_Live by client ID. This makes the
data flow visible inside Excel itself — a non-coder can follow it.

The Anova_Live sheet can be refreshed two ways:
  1. Run `python anova_fetch.py` (Python writes to the sheet)
  2. Power Query connected to the Azure GET endpoint (set up in Excel:
     Data → Get Data → From Web → paste AZURE_URL, load to Anova_Live)
Either source produces the same sheet shape; Client_List formulas don't
care which writer populated it.

Entry points:
  fetch_and_update()   — full pipeline: Azure → anova_latest.json, history.csv, Excel
  load_anova_latest()  — read anova_latest.json (no network call)
  update_excel()       — write to Anova_Live sheet + ensure Client_List formulas

Can be run standalone:  python anova_fetch.py
"""

import csv
import json
import os
import sys
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

sys.path.insert(0, str(Path(__file__).parent))
from config import DATA_DIR, INPUT_FILE

AZURE_URL = (
    'https://skoil-anova-escpe7axeyczdshp.westcentralus-01.azurewebsites.net'
    '/api/anova/data'
)
RTU_MAP_FILE = DATA_DIR / 'anova_rtu_map.csv'
LATEST_FILE = DATA_DIR / 'anova_latest.json'
HISTORY_FILE = DATA_DIR / 'anova_history.csv'

ANOVA_LIVE_SHEET = 'Anova_Live'
ANOVA_COL_OFFSET = 17  # Column R (0-indexed) — first Client_List Anova column


# ── RTU mapping ──────────────────────────────────────────────────────────────

def _load_rtu_map() -> dict:
    """Load rtu_id → {client_id, client_name, tank_capacity_lbs, product}."""
    if not RTU_MAP_FILE.exists():
        return {}
    mapping = {}
    with open(RTU_MAP_FILE, newline='', encoding='utf-8') as f:
        for row in csv.DictReader(f):
            mapping[row['rtu_id'].strip()] = {
                'client_id': row['client_id'].strip(),
                'client_name': row['client_name'].strip(),
                'tank_capacity_lbs': float(row['tank_capacity_lbs']),
                'product': row['product'].strip(),
            }
    return mapping


# ── Azure fetch ──────────────────────────────────────────────────────────────

def _fetch_from_azure(timeout: float = 15.0) -> list:
    """GET latest readings from Azure Function App. Returns list of dicts."""
    try:
        resp = requests.get(AZURE_URL, timeout=timeout)
        resp.raise_for_status()
        data = resp.json()
        if isinstance(data, list):
            return data
        return []
    except Exception as e:
        print(f'  ⚠ Azure fetch failed: {e}')
        return []


# ── Map readings to clients ──────────────────────────────────────────────────

def _map_readings(raw_readings: list, rtu_map: dict) -> dict:
    """Map Azure readings to SK client IDs via RTU map. Returns {client_id: reading_dict}."""
    mapped = {}
    unmapped = []

    for r in raw_readings:
        device_id = str(r.get('device_id', '')).strip()
        if not device_id:
            continue

        info = rtu_map.get(device_id)
        if not info:
            unmapped.append(device_id)
            continue

        cid = info['client_id']
        tank_cap = info['tank_capacity_lbs']
        level_lbs = float(r.get('level_lbs') or r.get('display_value') or 0)
        if level_lbs <= 0:
            continue

        ts_str = r.get('timestamp', '')
        try:
            ts = pd.Timestamp(ts_str)
        except Exception:
            continue

        now = pd.Timestamp.now(tz='UTC').tz_localize(None)
        ts_naive = ts.tz_localize(None) if ts.tz else ts
        age_hours = max((now - ts_naive).total_seconds() / 3600.0, 0.0)
        pct_full = round(level_lbs / tank_cap * 100, 1) if tank_cap > 0 else 0.0

        prev = mapped.get(cid)
        if prev and pd.Timestamp(prev['timestamp']) >= ts:
            continue

        mapped[cid] = {
            'client_id': cid,
            'client_name': info['client_name'],
            'rtu_id': device_id,
            'level_lbs': round(level_lbs, 1),
            'tank_capacity_lbs': tank_cap,
            'pct_full': pct_full,
            'product': info['product'],
            'timestamp': str(ts_naive),
            'age_hours': round(age_hours, 1),
            'confidence': 'sensor' if age_hours <= 24.0 else 'stale',
        }

    if unmapped:
        unique = set(unmapped)
        print(f'  ⚠ {len(unique)} unmapped RTU(s): {", ".join(sorted(unique))}')

    return mapped


# ── Local file persistence ───────────────────────────────────────────────────

def _write_latest(mapped: dict) -> None:
    """Write anova_latest.json (overwrite)."""
    payload = {
        'fetched_at': datetime.now(timezone.utc).isoformat(),
        'source': 'azure_push',
        'readings': mapped,
    }
    LATEST_FILE.write_text(json.dumps(payload, indent=2, default=str), encoding='utf-8')


def _append_history(mapped: dict) -> None:
    """Append new readings to anova_history.csv, dedup by (client_id, timestamp)."""
    existing_keys = set()
    if HISTORY_FILE.exists():
        with open(HISTORY_FILE, newline='', encoding='utf-8') as f:
            for row in csv.DictReader(f):
                existing_keys.add((row.get('client_id', ''), row.get('timestamp', '')))

    new_rows = []
    for cid, r in mapped.items():
        key = (cid, r['timestamp'])
        if key not in existing_keys:
            new_rows.append(r)

    if not new_rows:
        return

    write_header = not HISTORY_FILE.exists()
    cols = [
        'timestamp', 'client_id', 'client_name', 'rtu_id',
        'level_lbs', 'pct_full', 'product', 'tank_capacity_lbs',
    ]
    with open(HISTORY_FILE, 'a', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=cols, extrasaction='ignore')
        if write_header:
            w.writeheader()
        w.writerows(new_rows)

    print(f'  ▸ History: {len(new_rows)} new row(s) appended')


# ── Excel update ─────────────────────────────────────────────────────────────

# Anova_Live sheet schema — row 1 has headers, data starts row 2
# (Power-Query-friendly: PQ defaults to writing tables with row-1 headers)
ANOVA_LIVE_HEADERS = [
    'client_id',           # A
    'client_name',         # B
    'rtu_id',              # C
    'level_lbs',           # D
    'pct_full',            # E
    'timestamp',           # F
    'age_hours',           # G
    'source',              # H
    'tank_capacity_lbs',   # I
    'deliverable_lbs',     # J
    'product',             # K
    'fetched_at',          # L  (single cell M2 actually — info row)
]

# Client_List Anova columns (cols R-X = 18-24 1-based)
# Each tuple: (header, Anova_Live source column letter)
CLIENT_LIST_ANOVA_COLS = [
    ('RTU_ID',                 'C'),  # rtu_id
    ('Anova_Level_lbs',        'D'),  # level_lbs
    ('Anova_Pct_Full',         'E'),  # pct_full
    ('Anova_Last_Reading',     'F'),  # timestamp
    ('Anova_Deliverable_lbs',  'J'),  # deliverable_lbs
    ('Anova_Age_Hours',        'G'),  # age_hours
    ('Anova_Source',           'H'),  # source
]


def _write_anova_live_sheet(wb, mapped: dict, fetched_at: str) -> None:
    """Create or refresh the Anova_Live sheet with current readings."""
    if ANOVA_LIVE_SHEET in wb.sheetnames:
        ws = wb[ANOVA_LIVE_SHEET]
        # Clear existing data rows (keep nothing — full refresh)
        ws.delete_rows(1, ws.max_row + 1)
    else:
        ws = wb.create_sheet(ANOVA_LIVE_SHEET)

    # Row 1: headers
    headers = ANOVA_LIVE_HEADERS[:-1]  # exclude 'fetched_at' — it goes in M1/M2
    bold = Font(bold=True)
    fill = PatternFill('solid', fgColor='FFD9E1F2')  # light blue header
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = bold
        c.fill = fill
        c.alignment = Alignment(horizontal='left')

    # Info column M: refresh timestamp
    ws.cell(row=1, column=len(headers) + 1, value='fetched_at').font = bold
    ws.cell(row=2, column=len(headers) + 1, value=fetched_at)

    # Data rows
    for r_idx, (cid, r) in enumerate(sorted(mapped.items()), start=2):
        tank_cap = float(r.get('tank_capacity_lbs', 0))
        level = float(r.get('level_lbs', 0))
        deliverable = max(tank_cap - level, 0) if tank_cap else 0
        row_vals = [
            cid,
            r.get('client_name', ''),
            r.get('rtu_id', ''),
            level,
            r.get('pct_full', 0),
            r.get('timestamp', ''),
            r.get('age_hours', 0),
            r.get('confidence', ''),
            tank_cap,
            round(deliverable, 1),
            r.get('product', ''),
        ]
        for i, v in enumerate(row_vals, start=1):
            ws.cell(row=r_idx, column=i, value=v)

    # Column widths for readability
    widths = [10, 38, 12, 11, 9, 22, 11, 10, 10, 14, 16, 22]
    for i, w in enumerate(widths[:len(headers) + 1], start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    ws.freeze_panes = 'A2'


def _ensure_client_list_formulas(wb) -> int:
    """
    Make sure Client_List columns R-X have:
      - Headers in row 3
      - XLOOKUP formulas in rows 4..N pulling from Anova_Live by client_id

    Idempotent: replaces existing values/formulas every time.
    Returns the number of data rows written.
    """
    ws = wb['Client_List']
    col_start = ANOVA_COL_OFFSET + 1  # 1-based: R = 18

    # Row 3: headers
    bold = Font(bold=True)
    fill = PatternFill('solid', fgColor='FFD9E1F2')
    for i, (header, _src) in enumerate(CLIENT_LIST_ANOVA_COLS):
        c = ws.cell(row=3, column=col_start + i, value=header)
        c.font = bold
        c.fill = fill

    # Find data rows (start row 4, stop where col A is empty)
    last_row = 3
    for r in range(4, ws.max_row + 1):
        if ws.cell(row=r, column=1).value:
            last_row = r

    # Write XLOOKUP formulas
    # Formula: =IFERROR(XLOOKUP($A{row}, Anova_Live!$A:$A, Anova_Live!{src}:{src}, ""), "")
    n = 0
    for r in range(4, last_row + 1):
        if not ws.cell(row=r, column=1).value:
            continue
        for i, (_header, src) in enumerate(CLIENT_LIST_ANOVA_COLS):
            f = (f'=IFERROR(XLOOKUP($A{r},{ANOVA_LIVE_SHEET}!$A:$A,'
                 f'{ANOVA_LIVE_SHEET}!${src}:${src},""),"")')
            ws.cell(row=r, column=col_start + i, value=f)
        n += 1

    # Column widths
    for i, (header, _) in enumerate(CLIENT_LIST_ANOVA_COLS):
        col_letter = ws.cell(row=3, column=col_start + i).column_letter
        ws.column_dimensions[col_letter].width = max(12, len(header) + 2)

    return n


def update_excel(latest_data: dict, excel_path: Path = None) -> int:
    """
    Refresh Anova_Live sheet from `latest_data` and ensure Client_List
    has XLOOKUP formulas in cols R-X.

    Returns count of Anova_Live rows written.
    """
    excel_path = Path(excel_path or INPUT_FILE)
    if not excel_path.exists():
        print(f'  ⚠ Excel not found: {excel_path}')
        return 0

    readings = latest_data.get('readings', {}) if isinstance(latest_data, dict) else {}
    fetched_at = latest_data.get('fetched_at', '') if isinstance(latest_data, dict) else ''

    wb = load_workbook(str(excel_path))
    _write_anova_live_sheet(wb, readings, fetched_at)
    n_clients = _ensure_client_list_formulas(wb)
    wb.save(str(excel_path))

    print(f'  ▸ Excel: Anova_Live sheet refreshed ({len(readings)} sensors), '
          f'Client_List formulas applied to {n_clients} client(s)')
    return len(readings)


# ── Public API ───────────────────────────────────────────────────────────────

def load_anova_latest() -> dict:
    """Read anova_latest.json from disk. No network call."""
    if not LATEST_FILE.exists():
        return {}
    try:
        return json.loads(LATEST_FILE.read_text(encoding='utf-8'))
    except Exception:
        return {}


def fetch_and_update(update_excel_file: bool = True) -> dict:
    """
    Full pipeline: Azure → local JSON/CSV → Excel.
    Returns the mapped readings dict {client_id: reading}.
    """
    print('  Fetching Anova data from Azure...')
    rtu_map = _load_rtu_map()
    if not rtu_map:
        print('  ⚠ No RTU mappings found — skipping Anova fetch')
        return {}

    raw = _fetch_from_azure()
    if not raw:
        print('  ⚠ No readings from Azure (empty response or unreachable)')
        return load_anova_latest().get('readings', {})

    mapped = _map_readings(raw, rtu_map)
    if not mapped:
        print('  ⚠ No readings mapped to SK clients')
        return {}

    _write_latest(mapped)
    _append_history(mapped)
    print(f'  ▸ Anova: {len(mapped)} client(s) with live sensor data')

    if update_excel_file:
        latest = load_anova_latest()
        update_excel(latest)

    return mapped


def load_anova_from_query_sheet(excel_path: Path = None) -> dict:
    """
    Read live Anova readings from the Query sheet in SK_Delivery_System.xlsx.

    The Query sheet is populated by Power Query from the Azure API and contains
    device_id, customer name, display_value (lbs), and timestamp for every
    connected sensor. This is the authoritative source for live tank levels.

    Returns {client_id: reading_dict} keyed by SK client ID (extracted from
    the customer name field, e.g. 'HAR - 8031 - HAROLDS' -> '8031').
    """
    import re
    excel_path = Path(excel_path or INPUT_FILE)
    if not excel_path.exists():
        return {}

    try:
        wb = load_workbook(str(excel_path), data_only=True)
    except Exception as e:
        print(f'  ⚠ Cannot open {excel_path.name}: {e}')
        return {}

    if 'Query' not in wb.sheetnames:
        return {}

    ws = wb['Query']
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if not headers or 'device_id' not in headers:
        return {}

    col_idx = {h: i for i, h in enumerate(headers)}
    readings = {}
    now = pd.Timestamp.now(tz='UTC').tz_localize(None)

    for r in range(2, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if not any(row):
            continue

        device_id = str(row[col_idx.get('device_id', 0)] or '').strip()
        customer = str(row[col_idx.get('customer', 1)] or '').strip()
        level_lbs = row[col_idx.get('display_value', 2)]
        ts_raw = row[col_idx.get('timestamp', 7)]

        if not device_id or not customer or not level_lbs:
            continue
        try:
            level_lbs = float(level_lbs)
        except (ValueError, TypeError):
            continue
        if level_lbs <= 0:
            continue

        # Extract client ID from customer name: 'HAR - 8031 - HAROLDS' -> '8031'
        m = re.search(r'\b(\d{1,6})\b', customer)
        if not m:
            continue
        cid = m.group(1)

        try:
            ts = pd.Timestamp(ts_raw.date() if hasattr(ts_raw, 'date') else ts_raw)
            ts_naive = ts.tz_localize(None) if ts.tz else ts
        except Exception:
            ts_naive = now
        age_hours = max((now - ts_naive).total_seconds() / 3600.0, 0.0)

        readings[cid] = {
            'client_id': cid,
            'client_name': customer,
            'rtu_id': device_id,
            'level_lbs': round(level_lbs, 1),
            'timestamp': str(ts_naive),
            'age_hours': round(age_hours, 1),
            'confidence': 'sensor' if age_hours <= 24.0 else 'stale',
        }

    return readings


# ── CLI entry point ──────────────────────────────────────────────────────────

if __name__ == '__main__':
    print('═' * 50)
    print('  Anova Tank Sensor — Fetch & Update')
    print('═' * 50)
    result = fetch_and_update()
    if result:
        for cid, r in sorted(result.items()):
            pct = r.get('pct_full', 0)
            age = r.get('age_hours', 0)
            indicator = '🟢' if pct >= 40 else ('🟡' if pct >= 20 else '🔴')
            print(f'  {indicator} {cid:>6s}  {r["client_name"][:35]:<35s}  '
                  f'{r["level_lbs"]:>7.0f} lbs  {pct:>5.1f}%  {age:.1f}h ago')
    print('\n✓ Done.')
