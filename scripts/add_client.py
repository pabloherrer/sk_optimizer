"""
add_client.py
=============
Append a new client to the persisted OSRM matrix without rebuilding the
whole thing. Fetches only 2·(N+1) OSRM cells for the new row + column.

Usage
-----
  python add_client.py C999 "New Deli" 33.55 -112.10 \
      --tank 900 --product "CANOLA" --zone 2 --zone-code 2D

  # minimal
  python add_client.py C999 "New Deli" 33.55 -112.10

If the client is also missing from Client_List in the Excel file, pass
--also-write-excel to append it there too (keeps the matrix and the
source-of-truth spreadsheet in sync).
"""

from __future__ import annotations
import argparse
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from config import INPUT_FILE, MATRIX_FILE, NODES_FILE
from scripts.build_matrix import add_client as _add_client_to_matrix


def append_to_client_list(
    client_id: str, customer: str, lat: float, lon: float,
    zone: str, zone_code: str, street: str, city: str,
    tank_lbs: float, product: str,
    input_file: Path = INPUT_FILE,
) -> None:
    """Append a row to the Client_List sheet in the workbook."""
    wb = load_workbook(str(input_file))
    ws = wb['Client_List']

    # Find first empty row starting from row 4
    row = 4
    while ws.cell(row, 1).value:
        row += 1

    ws.cell(row, 1,  client_id)
    ws.cell(row, 2,  customer)
    ws.cell(row, 3,  zone)
    ws.cell(row, 4,  zone_code)
    ws.cell(row, 5,  street)
    ws.cell(row, 6,  city)
    ws.cell(row, 7,  'AZ')
    ws.cell(row, 8,  lat)
    ws.cell(row, 9,  lon)
    ws.cell(row, 10, tank_lbs)
    ws.cell(row, 11, product)

    wb.save(str(input_file))
    print(f"  ✓ Appended {client_id} to Client_List (row {row})")


def main():
    p = argparse.ArgumentParser()
    p.add_argument('client_id',   help="e.g. C999")
    p.add_argument('customer',    help="Customer name")
    p.add_argument('lat', type=float)
    p.add_argument('lon', type=float)
    p.add_argument('--tank',      type=float, default=None, help="tank size lbs")
    p.add_argument('--product',   default='')
    p.add_argument('--zone',      default='')
    p.add_argument('--zone-code', default='')
    p.add_argument('--street',    default='')
    p.add_argument('--city',      default='')
    p.add_argument('--also-write-excel', action='store_true',
                   help='Append to Client_List sheet in SK_Delivery_System.xlsx too')
    args = p.parse_args()

    try:
        _add_client_to_matrix(
            client_id=args.client_id,
            lat=args.lat,
            lon=args.lon,
            customer=args.customer,
            zone=args.zone,
            zone_code=args.zone_code,
            street=args.street,
            city=args.city,
            tank_lbs=args.tank,
            product=args.product,
        )
    except Exception as e:
        print(f"  ✗ Failed to add to matrix: {e}")
        sys.exit(1)

    if args.also_write_excel:
        try:
            append_to_client_list(
                args.client_id, args.customer, args.lat, args.lon,
                args.zone, args.zone_code, args.street, args.city,
                args.tank or '', args.product,
            )
        except Exception as e:
            print(f"  ⚠  Matrix updated but failed to write Excel: {e}")


if __name__ == '__main__':
    main()
