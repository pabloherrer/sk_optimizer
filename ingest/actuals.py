#!/usr/bin/env python3
"""
ingest_actuals.py — Append/update Delivery_Log entries from invoice CSV
========================================================================

Reads an InvoiceItems CSV (QuickBooks-style export) and updates the
Delivery_Log sheet in SK_Delivery_System.xlsx with REAL delivered
quantities. Placeholder rows (Qty=200, the human-entered "I delivered
something but I don't know the amount yet") are replaced when the
invoice arrives with the real number.

Rules applied:
  • Only inventory-items containing 'CAN', 'FRY', or '38500' (palm oil)
    are deliveries. WASTE OIL, FILTERS, SERVICE ITEMS are ignored.
  • If a (date, client_id) row already exists in Delivery_Log:
      - placeholder (Qty=200) → REPLACE with invoice qty
      - real qty                → keep both unless qty differs (warn)
  • If no row exists → APPEND.
  • Customer is matched by the leading "ID" inside the customer name
    (e.g., "POP - 16015 - POPOS BELL" → ID 16015).
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

# Items that count as oil deliveries (substring match on Item column)
OIL_ITEM_PATTERNS = [
    r'30000-CAN',
    r'30000-FRY',
    r'38500',          # palm oil bulk
]

# Items that explicitly are NOT deliveries (rejected even if substring match)
NON_DELIVERY_ITEMS = {
    'WASTE OIL', 'ZECO', 'MISCELLANEOUS', 'ZA-408', 'OIL FINISHED GOODS',
}

PLACEHOLDER_QTY = 200  # rows with this quantity are "I delivered something"


def is_oil_delivery(item: str) -> bool:
    """Return True if the item is a real oil delivery (CAN/FRY/palm bulk)."""
    if not item:
        return False
    upper = item.upper()
    if any(nd in upper for nd in NON_DELIVERY_ITEMS):
        return False
    return any(re.search(pat, upper) for pat in OIL_ITEM_PATTERNS)


def extract_client_id(customer_name: str) -> str | None:
    """
    Pull the numeric client ID out of a customer name like
    'POP - 16015 - POPOS BELL' → '16015'.

    Customer names often have a 3–6 char prefix code (POP, MAN6, etc.)
    followed by ' - <id> - <name>'. ID is the first all-digit token.
    """
    if not customer_name:
        return None
    # Strip any trailing duplicate-customer suffix (rare format e.g. "BLU-...-BLU-...-30239")
    main_part = customer_name.split(' - 30')[0] if ' - 30' in customer_name else customer_name
    tokens = re.split(r'\s+-\s+', main_part)
    for tok in tokens:
        tok = tok.strip()
        if tok.isdigit():
            return tok
    return None


def parse_invoice_date(s: str) -> datetime | None:
    """Parse 'M/D/YYYY' or 'YYYY-MM-DD'."""
    if not s:
        return None
    for fmt in ('%m/%d/%Y', '%Y-%m-%d', '%m/%d/%y'):
        try:
            return datetime.strptime(s.strip(), fmt)
        except ValueError:
            continue
    return None


def _load_active_client_ids(excel_path: Path) -> set[str]:
    """Return the set of client IDs present in the Client_List sheet."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb['Client_List']
    out = set()
    for r in range(4, ws.max_row + 1):
        cid = ws.cell(r, 1).value
        if cid is None:
            continue
        out.add(str(cid).strip())
    return out


def load_csv_deliveries(csv_path: Path, active_ids: set[str] | None = None) -> list[dict]:
    """Parse the InvoiceItems CSV → list of {date, client_id, qty_lbs, item, customer}.
    If active_ids is provided, drops invoice rows for clients not in it
    (e.g. wholesale accounts not part of the route)."""
    out = []
    skipped_inactive = 0
    # encoding='utf-8-sig' strips the BOM that QuickBooks exports prepend
    with open(csv_path, encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            item = (row.get('Item') or '').strip()
            if not is_oil_delivery(item):
                continue
            cust = row.get('Customer Full Name', '').strip()
            cid = extract_client_id(cust)
            if not cid:
                continue
            if active_ids is not None and cid not in active_ids:
                skipped_inactive += 1
                continue
            try:
                qty = float(row.get('Quantity', '0') or '0')
            except ValueError:
                continue
            if qty <= 0:
                continue
            d = parse_invoice_date(row.get('Invoice Date', ''))
            if d is None:
                continue
            out.append({
                'date': d,
                'client_id': cid,
                'qty_lbs': int(round(qty)),
                'item': item,
                'customer': cust,
            })
    if skipped_inactive:
        print(f'  → Skipped {skipped_inactive} invoice line(s) for clients '
              f'not in the active route Client_List (likely wholesale).')
    return out


def find_log_row(
    ws, target_date: datetime, client_id: str,
    *,
    date_window_days: int = 0,
    match_qty: int | None = None,
) -> int | None:
    """
    Find a Delivery_Log row matching (date ± window, client_id).
    If match_qty is given, also requires qty equal to it (for dedup-by-qty).

    The user noted invoice date may be 1–2 days AFTER actual delivery date,
    so we accept a small backward window when matching for dedup.
    """
    from datetime import timedelta
    target_date = target_date.date() if hasattr(target_date, 'date') else target_date
    earliest = target_date - timedelta(days=date_window_days)
    for r in range(4, ws.max_row + 1):
        d = ws.cell(r, 1).value
        cid = ws.cell(r, 2).value
        if d is None or cid is None:
            continue
        d_date = d.date() if hasattr(d, 'date') else d
        if str(cid) != str(client_id):
            continue
        if not (earliest <= d_date <= target_date):
            continue
        if match_qty is not None and ws.cell(r, 4).value != match_qty:
            continue
        return r
    return None


def find_last_data_row(ws) -> int:
    for r in range(ws.max_row, 3, -1):
        if ws.cell(r, 1).value is not None:
            return r
    return 3   # row 3 is header


def write_excel(
    *,
    excel_path: Path,
    csv_deliveries: list[dict],
    pending_visits: list[dict] | None = None,
    dry_run: bool = False,
) -> dict:
    """
    Apply CSV deliveries to the Delivery_Log. Returns a stats dict.

    pending_visits: list of {date, client_id} — adds 200-lb placeholder
    rows for these (e.g., "scheduled but not yet delivered").
    """
    wb = openpyxl.load_workbook(excel_path)
    ws = wb['Delivery_Log']

    stats = {'replaced_placeholder': 0, 'appended_new': 0,
             'kept_existing': 0, 'mismatches': [],
             'pending_appended': 0, 'pending_skipped_existing': 0}

    # Aggregate same-day same-client multiple items (e.g. CROWN PLAZA had 2 palm oil entries)
    agg: dict = {}
    for d in csv_deliveries:
        key = (d['date'].date(), d['client_id'])
        agg.setdefault(key, {'qty': 0, 'items': [], 'customer': d['customer']})
        agg[key]['qty'] += d['qty_lbs']
        agg[key]['items'].append(d['item'])

    for (date_key, cid), info in sorted(agg.items()):
        target_date = datetime.combine(date_key, datetime.min.time())

        # Dedup pass 1: same date exact match (qty agnostic)
        existing_row = find_log_row(ws, target_date, cid, date_window_days=0)

        # Dedup pass 2: invoice date often 1–2 days AFTER actual delivery —
        # search backward for a same-qty match. If found, this is the same
        # delivery, just invoiced later. Skip.
        same_qty_earlier = find_log_row(
            ws, target_date, cid, date_window_days=3, match_qty=info['qty'],
        )

        # Dedup pass 3: same backward window, looking for a placeholder
        # row to replace with the real qty.
        placeholder_earlier = find_log_row(
            ws, target_date, cid, date_window_days=3, match_qty=PLACEHOLDER_QTY,
        )

        if existing_row is None:
            if same_qty_earlier is not None:
                stats['kept_existing'] += 1
                continue   # same delivery, invoiced later — already logged
            if placeholder_earlier is not None:
                ws.cell(placeholder_earlier, 4, info['qty'])
                stats['replaced_placeholder'] += 1
                d_existing = ws.cell(placeholder_earlier, 1).value
                d_str = d_existing.date() if hasattr(d_existing, 'date') else d_existing
                print(f"  ~ {d_str} id={cid:>6}  qty={info['qty']:>6} "
                      f"(replaced placeholder, kept original delivery date)")
                continue
            last = find_last_data_row(ws)
            new_row = last + 1
            ws.cell(new_row, 1, target_date)
            ws.cell(new_row, 2, int(cid))
            ws.cell(new_row, 4, info['qty'])
            stats['appended_new'] += 1
            print(f"  + {date_key} id={cid:>6}  qty={info['qty']:>6} "
                  f"({', '.join(info['items'])[:30]})")
        else:
            existing_qty = ws.cell(existing_row, 4).value
            if existing_qty == PLACEHOLDER_QTY:
                ws.cell(existing_row, 4, info['qty'])
                stats['replaced_placeholder'] += 1
                print(f"  ~ {date_key} id={cid:>6}  qty={info['qty']:>6} "
                      f"(replaced placeholder 200)")
            elif existing_qty == info['qty']:
                stats['kept_existing'] += 1
            else:
                stats['mismatches'].append(
                    (date_key, cid, existing_qty, info['qty'])
                )
                stats['kept_existing'] += 1

    # Pending visits (planned schedule from screenshots) — add 200-lb placeholder
    if pending_visits:
        for v in pending_visits:
            target_date = v['date']
            cid = v['client_id']
            existing_row = find_log_row(ws, target_date, cid)
            if existing_row is not None:
                stats['pending_skipped_existing'] += 1
                continue
            last = find_last_data_row(ws)
            new_row = last + 1
            ws.cell(new_row, 1, target_date)
            ws.cell(new_row, 2, int(cid))
            ws.cell(new_row, 4, PLACEHOLDER_QTY)
            stats['pending_appended'] += 1
            print(f"  ⌛ {target_date.date()} id={cid:>6}  qty={PLACEHOLDER_QTY:>6} "
                  f"(pending — scheduled visit)")

    if dry_run:
        print('\n  DRY RUN — no changes written.')
        return stats

    wb.save(excel_path)
    print(f'\n  Saved → {excel_path}')
    return stats


# ─────────────────────────────────────────────────────────────────────────────
# Pre-built schedules from the screenshots (May 1 + May 2)
# ─────────────────────────────────────────────────────────────────────────────

# May 1, 2026 — JASON B's clients (from screenshot)
SCHEDULE_2026_05_01_JASON_B = [
    '15005',  # ORE - OREGANO GOODYEAR
    '14009',  # NEW - NEW PENNY CAFE
    '16003',  # PET - PETES FISH CENTRAL
    '16005',  # PET - PETES FISH 26TH ST
    '2017',   # BOO - BOOTYS WATSON
    '20091',  # TAI - TAILGATERS BELL SURPRISE
    '19031',  # STA - STATE 48 SURPRISE
    '19080',  # SAR - SARDELLAS LAKE PLEASANT
    '1000',   # ARR - ARRIBAS 79TH
    '20051',  # TIM - TIM FINNEGANS
    '2078',   # BRO - BROTHERS PIZZA INDIAN SCHOOL
]

# May 2, 2026 — JASON B (Flagstaff/far run)
SCHEDULE_2026_05_02_JASON_B = [
    '6113',   # FIL - FILIBERTO CAMP VERDE
    '15004',  # ORE - OREGANO FLAGSTAFF
    '2056',   # B66 - B66 BRUNCH & BREW BBQ
    '11005',  # KAR - KARMA SUSHI
    '19024',  # SOS - SOSOBA FLAGSTAFF
    '12021',  # LOT - LOTUS LOUNGE
    '16039',  # PRO - PROPER MEATS & PROVISIONS
    '15032',  # ORE - OREGANO COUNTRY
    '20010',  # TWI - TWIN ARROWS TRUCK STOP
    '20010',  # TWI - TWIN ARROWS CASINO (same site? two entries in screenshot)
    '18021',  # RRP - R&R PIZZA COTTONWOOD
]

# May 2, 2026 — JOE (metro)
SCHEDULE_2026_05_02_JOE = [
    '15000',  # OHS - OHSO INDIAN SCHOOL
    '6077',   # FOU - FOUR PEAKS BREWING TEMPE
    '16007',  # PET - PETES FISH APACHE TEMPE
    '13094',  # MIS - MISSION BBQ GOODYEAR
    '13012',  # MAN - MANUEL MCDOWELL GOODYEAR
    '13015',  # MAN - MANUEL SOUTHERN TEMPE
    '16010',  # PET - PETES FISH TOLLESON
    '7052',   # GOL - GOLD CANYON RESORT
    '15051',  # ORI - ORIGEN GILBERT
    '15009',  # ORE - OREGANO QUEEN CREEK
    '7002',   # GIL - GILBERT PIZZA
    '15027',  # ORE - OREGANO CHANDLER
    '1011',   # ARR - ARRIBAS CHANDLER
]


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument('--csv', type=str, required=False,
                   default='/Users/pabloherrera/Downloads/InvoiceItems_gridDashboard.csv',
                   help='Invoice CSV path')
    p.add_argument('--excel', type=str, required=False,
                   default='data/SK_Delivery_System.xlsx',
                   help='Delivery_Log Excel path')
    p.add_argument('--add-schedules', action='store_true',
                   help='Also add the May 1 + May 2 scheduled visits as 200-lb placeholders')
    p.add_argument('--dry-run', action='store_true',
                   help="Don't write the Excel file (just preview)")
    args = p.parse_args()

    csv_path = Path(args.csv)
    excel_path = Path(args.excel)
    if not csv_path.exists():
        print(f'CSV not found: {csv_path}', file=sys.stderr); return 1
    if not excel_path.exists():
        print(f'Excel not found: {excel_path}', file=sys.stderr); return 1

    print(f'\n  Reading invoice CSV: {csv_path}')
    active_ids = _load_active_client_ids(excel_path)
    print(f'  → {len(active_ids)} active client IDs in Client_List')
    deliveries = load_csv_deliveries(csv_path, active_ids=active_ids)
    print(f'  → {len(deliveries)} oil-delivery line(s) found '
          f'(WASTE OIL / filters / service items filtered out)')

    pending = []
    if args.add_schedules:
        from datetime import datetime as _dt
        d_may1 = _dt(2026, 5, 1)
        d_may2 = _dt(2026, 5, 2)
        for cid in SCHEDULE_2026_05_01_JASON_B:
            pending.append({'date': d_may1, 'client_id': cid})
        for cid in SCHEDULE_2026_05_02_JASON_B + SCHEDULE_2026_05_02_JOE:
            pending.append({'date': d_may2, 'client_id': cid})
        # Dedupe (same client appearing twice, e.g. TWI ARROWS truck stop + casino)
        seen = set()
        unique = []
        for v in pending:
            k = (v['date'].date(), v['client_id'])
            if k in seen:
                continue
            seen.add(k); unique.append(v)
        pending = unique
        print(f'  → {len(pending)} pending scheduled visit(s) for May 1 + May 2')

    print()
    stats = write_excel(
        excel_path=excel_path,
        csv_deliveries=deliveries,
        pending_visits=pending,
        dry_run=args.dry_run,
    )

    print()
    print(f'  Replaced placeholders:    {stats["replaced_placeholder"]}')
    print(f'  Appended new (CSV):       {stats["appended_new"]}')
    print(f'  Kept existing (matched):  {stats["kept_existing"]}')
    print(f'  Pending appended:         {stats["pending_appended"]}')
    print(f'  Pending skipped (exists): {stats["pending_skipped_existing"]}')
    if stats['mismatches']:
        print(f'\n  ⚠ Quantity mismatches (existing kept; review):')
        for date_key, cid, existing, new in stats['mismatches']:
            print(f'    {date_key} id={cid:>6}  existing={existing}  invoice={new}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
