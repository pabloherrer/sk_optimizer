"""
load_data.py
============
Reads SK_Delivery_System.xlsx and returns clean DataFrames.

Two sheets are used:
  Client_List   — master client records (static / rarely changes)
  Delivery_Log  — historical deliveries (append-only, chronological)

Neither sheet's formula cells are trusted (openpyxl data_only=True cannot
evaluate formulas from files never opened in Excel).  All derived values
are recomputed here in Python.
"""

import warnings
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
from config import INPUT_FILE, PRODUCT_ALIASES, PRODUCTS

warnings.filterwarnings('ignore')


# ── Public API ────────────────────────────────────────────────────────────────

def load_all(input_file: str | Path = INPUT_FILE) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Load and validate both sheets.

    Returns
    -------
    clients_df    : one row per client, ID as string key
    deliveries_df : one row per historical delivery, sorted oldest → newest
    """
    clients_df    = load_clients(input_file)
    deliveries_df = load_deliveries(input_file)
    return clients_df, deliveries_df


def load_clients(input_file: str | Path = INPUT_FILE) -> pd.DataFrame:
    """
    Parse Client_List sheet.

    Expected layout (data starts at row 4):
      Col A : ID         (e.g. 'C001')
      Col B : Customer   (full name string)
      Col C : Zone       (integer zone number, stored as string)
      Col D : Zone_Code  (e.g. '2D')
      Col E : Street     (street portion of address)
      Col F : City
      Col G : State      (always 'AZ' — included for completeness)
      Col H : Latitude
      Col I : Longitude
      Col J : Tank_lbs
      Col K : Product
      Col L : Service_Min (optional per-stop service time override)
      Col M : Access_Notes (optional free text)
      Col N : Phone (optional phone number)
    """
    wb = load_workbook(str(input_file), data_only=True)
    ws = wb['Client_List']

    records = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        # Skip blank rows and the header row (col A = 'ID')
        if not row[0]:
            continue
        if str(row[0]).strip().upper() == 'ID':
            continue
        # Skip rows with no customer name
        if not row[1]:
            continue

        records.append({
            'ID':       str(int(row[0])) if isinstance(row[0], float) else str(row[0]).strip(),
            'Customer': str(row[1]).strip() if row[1] else '',
            'Zone':     str(row[2]).replace('.0', '').strip() if row[2] else '',
            'Zone_Code':str(row[3]).strip() if row[3] else '',
            'Address':  _build_address(row[4], row[5]),
            'Lat':      _to_float(row[7]),
            'Lon':      _to_float(row[8]),
            'Tank_lbs': _to_float(row[9]),
            'Product':  str(row[10]).strip() if row[10] else '',
            'Service_Min': _to_float(row[11]) if len(row) > 11 else None,
            'Access_Notes': str(row[12]).strip() if len(row) > 12 and row[12] else None,
            'Phone': str(row[13]).strip() if len(row) > 13 and row[13] else None,
            # Columns added by ingest_scheduler_notes.py — present if the
            # ingest has been run; otherwise None / empty
            'ANOVA':            str(row[14]).strip() if len(row) > 14 and row[14] else '',
            'Notes':            str(row[15]).strip() if len(row) > 15 and row[15] else '',
            'Do_Not_Schedule':  str(row[16]).strip() if len(row) > 16 and row[16] else '',
        })

    df = pd.DataFrame(records)

    # Remove duplicate IDs — keep first occurrence (sheet order = canonical)
    dupes = df[df.duplicated('ID', keep=False)]
    if len(dupes):
        dup_ids = dupes['ID'].unique().tolist()
        print(f"  ⚠  Duplicate ID(s) in Client_List, keeping first occurrence: {dup_ids}")
        df = df.drop_duplicates(subset='ID', keep='first')

    # Normalise product names to canonical values
    df['Product'] = df['Product'].str.upper().str.strip().map(
        lambda x: PRODUCT_ALIASES.get(x, PRODUCTS[0])
    )

    # Validation: warn on missing GPS or tank size
    no_gps  = df[df['Lat'].isna() | df['Lon'].isna()]
    no_tank = df[df['Tank_lbs'].isna() | (df['Tank_lbs'] <= 0)]
    if len(no_gps):
        print(f"  ⚠  {len(no_gps)} client(s) missing GPS — will be excluded from routing:")
        for _, r in no_gps.iterrows():
            print(f"       {r['ID']}  {r['Customer'][:50]}")
    if len(no_tank):
        print(f"  ⚠  {len(no_tank)} client(s) missing tank size — will be excluded:")
        for _, r in no_tank.iterrows():
            print(f"       {r['ID']}  {r['Customer'][:50]}")

    print(f"  Clients loaded: {len(df)}  "
          f"(routable: {df['Lat'].notna().sum() - len(no_tank)})")
    return df


def load_deliveries(input_file: str | Path = INPUT_FILE) -> pd.DataFrame:
    """
    Parse Delivery_Log sheet.

    Current layout (data starts at row 4):
      Col A (0) : Date delivered    (datetime)
      Col B (1) : Customer ID       (numeric — matches Client_List col A)
      Col C (2) : Customer Name     (formula — may be None in data_only mode)
      Col D (3) : Qty Delivered (lbs)  ← INPUT column
      Col E (4) : Zone              (formula)
      Col F (5) : Zone_Code         (formula)
      Col G (6) : Product           (formula)
      Col H (7) : Tank (lbs)        (formula)
      Col I (8) : Days Since        (formula)
      Col J (9) : Lbs/Day           (formula)
      Col K (10): Running Avg       (formula)

    Rows with missing Date or zero/non-numeric Qty are dropped.
    Customer ID (col B) is resolved to full name via Client_List lookup.
    """
    wb = load_workbook(str(input_file), data_only=True)

    # Build ID → Customer name lookup from Client_List
    cl = wb['Client_List']
    id_to_name = {}
    for r in cl.iter_rows(min_row=4, values_only=True):
        if not r[0] or not r[1]:
            continue
        try:
            cid = str(int(float(r[0])))
        except (ValueError, TypeError):
            import re as _re
            m = _re.match(r'(\d+)', str(r[0]))
            cid = m.group(1) if m else str(r[0]).strip()
        id_to_name[cid] = str(r[1]).strip()

    ws = wb['Delivery_Log']

    records = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[0]:
            continue
        qty = row[3]   # Col D — Qty Delivered
        if not isinstance(qty, (int, float)) or qty <= 0:
            continue

        # Parse date robustly
        raw_date = row[0]
        try:
            d = pd.Timestamp(raw_date.date() if hasattr(raw_date, 'date') else raw_date)
        except Exception:
            continue

        # Resolve customer name from numeric ID in col B
        raw_id = row[1]
        if raw_id is None:
            continue
        try:
            cid_str = str(int(float(raw_id)))
        except (ValueError, TypeError):
            cid_str = str(raw_id).strip()
        customer = id_to_name.get(cid_str, cid_str)   # fall back to raw ID string

        records.append({
            'Date':     d,
            'Customer': customer,
            'Tank_lbs': _to_float(row[7]),   # Col H — Tank
            'Qty_lbs':  float(qty),
            # 200-lb is the human-entered "I delivered something today but
            # I don't know the amount yet" placeholder. The forecaster must
            # exclude it from rate computation or rates collapse to zero.
            'Is_Placeholder': float(qty) == 200.0,
        })

    df = pd.DataFrame(records)
    df = df[df['Date'] >= '2020-01-01'].copy()          # Ignore very old data
    df = df.sort_values(['Customer', 'Date']).reset_index(drop=True)

    print(f"  Deliveries loaded: {len(df)}  "
          f"| Range: {df['Date'].min().date()} → {df['Date'].max().date()}")
    return df


# ── Helpers ───────────────────────────────────────────────────────────────────

def _build_address(street, city) -> str:
    parts = [str(x).strip() for x in [street, city] if x]
    return ', '.join(parts) + ' AZ' if parts else ''


def _to_float(v) -> float | None:
    try:
        return float(v) if v is not None else None
    except (ValueError, TypeError):
        return None
