"""
output.py
=========
Generates the two deliverables drivers and dispatchers actually use:
  1. sk_weekly_schedule.xlsx  — formatted per-truck/per-day route sheets
  2. sk_route_map.html        — interactive Folium map

Both accept a {day_index: routes_DataFrame} dict as input so they can
be regenerated at any point in the rolling horizon.
"""

import math
import time
import urllib.request
import json
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import folium
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from config import (
    DAYS, NUM_DAYS, TRUCK_NAMES, TRUCKS, DEPOT_LAT, DEPOT_LON,
    TRUCK_HEX, TRUCK_MAP_COLORS, URGENCY_FILL_COLORS, SHIFT_MIN, OUTPUT_DIR,
    DATA_DIR,
)

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Persisted route polyline cache (so the map renders with real roads even when
# OSRM is unreachable at view time). Keyed by a hash of the waypoint sequence.
GEOM_CACHE_FILE = DATA_DIR / 'route_geom_cache.json'

# ── Excel ─────────────────────────────────────────────────────────────────────

ROUTE_COLS = [
    'Stop', 'Date', 'Arrival_HHMM', 'Depart_HHMM',
    'Travel_To_Min', 'Dist_To_mi', 'Customer', 'Zone',
    'Days_Until_Stockout', 'DaysToStockoutAtVisit', 'Urgency',
    'Tank_lbs', 'Current_lbs', 'Refill_lbs', 'Fill_Pct',
    'Avg_LbsPerDay', 'Service_Min', 'Product',
    'Route_Dist_mi', 'Route_Time_min', 'Shift_Pct', 'Cap_Pct',
]

# Display labels (shown as header row in driver sheets — more readable than raw column names)
ROUTE_COL_LABELS = {
    'Stop':                  '#',
    'Date':                  'Date',
    'Arrival_HHMM':          'Arrive',
    'Depart_HHMM':           'Depart',
    'Travel_To_Min':         'Drive (min)',
    'Dist_To_mi':            'Dist (mi)',
    'Customer':              'Customer',
    'Zone':                  'Zone',
    'Days_Until_Stockout':   'Days→Empty',
    'DaysToStockoutAtVisit': 'Days@Visit',
    'Urgency':               'Urgency',
    'Tank_lbs':              'Tank (lbs)',
    'Current_lbs':           'Curr (lbs)',
    'Refill_lbs':            'Refill (lbs)',
    'Fill_Pct':              'Fill %',
    'Avg_LbsPerDay':         'Avg lbs/day',
    'Service_Min':           'Svc (min)',
    'Product':               'Product',
    'Route_Dist_mi':         'Route Total (mi)',
    'Route_Time_min':        'Route (min)',
    'Shift_Pct':             'Shift %',
    'Cap_Pct':               'Cap %',
}

# Numeric formats per column (openpyxl number_format strings)
ROUTE_COL_FORMATS = {
    'Travel_To_Min':         '0',
    'Dist_To_mi':            '0.0',
    'Days_Until_Stockout':   '0.0',
    'DaysToStockoutAtVisit': '0.0',
    'Tank_lbs':              '#,##0',
    'Current_lbs':           '#,##0',
    'Refill_lbs':            '#,##0',
    'Fill_Pct':              '0.0"%"',
    'Avg_LbsPerDay':         '0.0',
    'Service_Min':           '0',
    'Route_Dist_mi':         '0.0',
    'Route_Time_min':        '0',
    'Shift_Pct':             '0.0"%"',
    'Cap_Pct':               '0.0"%"',
}


def save_excel_schedule(
    all_routes:  Dict[int, pd.DataFrame],
    deferred_df: Optional[pd.DataFrame] = None,
    filename:    str = 'sk_weekly_schedule.xlsx',
    output_dir:  Path = OUTPUT_DIR,
    plan_dates:  Optional[List[pd.Timestamp]] = None,
    today:       Optional[pd.Timestamp] = None,
    snapshot:    Optional[pd.DataFrame] = None,
) -> Path:
    """Write a styled Excel workbook."""
    filepath = output_dir / filename

    # Combine all route data for the All_Routes sheet
    all_data = [r for r in all_routes.values() if r is not None and not r.empty]
    combined = pd.concat(all_data, ignore_index=True) if all_data else pd.DataFrame()

    # Build summary (manual loop — avoids pandas 3.x groupby.apply breakage
    # where grouping columns are excluded from the group by default)
    if not combined.empty:
        rows = []
        for (truck, day), grp in combined.groupby(['Truck', 'Day']):
            rows.append(_day_summary(grp, truck, day))
        summary = pd.DataFrame(rows)
    else:
        summary = pd.DataFrame()

    with pd.ExcelWriter(str(filepath), engine='openpyxl') as writer:
        # Sheet 0: Plan info (run metadata)
        run_ts = pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')
        plan_info_rows = [{'Field': 'Generated', 'Value': run_ts}]
        if today is not None:
            plan_info_rows.append({'Field': 'Solving as-of', 'Value': today.strftime('%A %b %d, %Y')})
        if plan_dates:
            plan_info_rows.append({'Field': 'Plan start', 'Value': plan_dates[0].strftime('%A %b %d, %Y')})
            plan_info_rows.append({'Field': 'Plan end',   'Value': plan_dates[-1].strftime('%A %b %d, %Y')})
            plan_info_rows.append({'Field': 'Delivery days', 'Value': len(plan_dates)})
            for i, dt in enumerate(plan_dates):
                plan_info_rows.append({'Field': f'  Day {i+1}', 'Value': dt.strftime('%a %b %d, %Y')})
        pd.DataFrame(plan_info_rows).to_excel(writer, sheet_name='0_Plan_Info', index=False)

        # Sheet 1: Week summary
        if not summary.empty:
            summary.to_excel(writer, sheet_name='1_Week_Summary', index=False)

        # Sheet 1b: Stockout Risk (if snapshot provided)
        if snapshot is not None and not snapshot.empty:
            risk_df = _build_stockout_risk(snapshot, combined, today)
            if not risk_df.empty:
                risk_df.to_excel(writer, sheet_name='2_Stockout_Risk', index=False)

        # All routes flat
        if not combined.empty:
            combined.to_excel(writer, sheet_name='All_Routes', index=False)

        # Deferred
        if deferred_df is not None and not deferred_df.empty:
            deferred_df.to_excel(writer, sheet_name='Deferred', index=False)

    # Post-process: apply styling
    _apply_excel_styles(filepath, all_routes)

    # Post-process: add combined printable sheets
    _add_printable_sheets(filepath, all_routes, plan_dates, snapshot, today)

    print(f"  Excel saved: {filepath}")
    return filepath


def _build_stockout_risk(
    snapshot: pd.DataFrame,
    combined_routes: pd.DataFrame,
    today: Optional[pd.Timestamp],
) -> pd.DataFrame:
    """
    Produce a client-by-client stockout risk table.

    Columns:
      ID, Customer, Current_lbs, Tank_lbs, Fill_Pct_Now, Avg_LbsPerDay,
      Days_Until_Stockout, Projected_Stockout_Date, Urgency,
      Scheduled_This_Week, Scheduled_Date, Refill_lbs
    """
    if snapshot is None or snapshot.empty:
        return pd.DataFrame()

    df = snapshot.copy()
    if today is None:
        today = pd.Timestamp.today().normalize()

    # Projected stockout date
    def _proj_date(days):
        try:
            d = float(days)
            if not np.isfinite(d):
                return ''
            return (today + pd.Timedelta(days=max(d, 0))).strftime('%a %b %d')
        except Exception:
            return ''

    df['Projected_Stockout_Date'] = df.get(
        'Days_Until_Stockout', pd.Series([np.nan] * len(df))
    ).apply(_proj_date)

    # Fill % today
    if 'Tank_lbs' in df.columns:
        cur = df.get('Current_lbs', df.get('Est_Current_lbs', 0))
        df['Fill_Pct_Now'] = (cur.astype(float) / df['Tank_lbs'].replace(0, np.nan) * 100).round(1)

    # Scheduling lookup from combined routes
    sched_map: Dict[str, Dict[str, object]] = {}
    if combined_routes is not None and not combined_routes.empty and 'ID' in combined_routes.columns:
        for _, r in combined_routes.iterrows():
            cid = str(r.get('ID', ''))
            if not cid or cid in sched_map:
                continue
            sched_map[cid] = {
                'Scheduled_This_Week': 'Yes',
                'Scheduled_Date':      str(r.get('Date', r.get('Day', ''))),
                'Scheduled_Refill_lbs': int(r.get('Refill_lbs', 0) or 0),
            }

    df['Scheduled_This_Week']   = df['ID'].astype(str).map(lambda i: sched_map.get(i, {}).get('Scheduled_This_Week', 'No'))
    df['Scheduled_Date']        = df['ID'].astype(str).map(lambda i: sched_map.get(i, {}).get('Scheduled_Date', ''))
    df['Scheduled_Refill_lbs']  = df['ID'].astype(str).map(lambda i: sched_map.get(i, {}).get('Scheduled_Refill_lbs', 0))

    # Select & order columns that exist
    cols = [c for c in [
        'ID', 'Customer', 'Urgency',
        'Days_Until_Stockout', 'Projected_Stockout_Date',
        'Current_lbs', 'Tank_lbs', 'Fill_Pct_Now',
        'Avg_LbsPerDay', 'Rate_Source', 'Consumption_Shift',
        'Delivery_Count', 'Product',
        'Scheduled_This_Week', 'Scheduled_Date', 'Scheduled_Refill_lbs',
    ] if c in df.columns]
    out = df[cols].copy()

    # Sort: stockout/critical first, then by days-until-stockout ascending
    urg_rank = {'stockout': 0, 'critical': 1, 'urgent': 2, 'normal': 3}
    if 'Urgency' in out.columns:
        out['_u'] = out['Urgency'].map(lambda u: urg_rank.get(str(u), 9))
        out = out.sort_values(['_u', 'Days_Until_Stockout'], na_position='last').drop(columns=['_u'])
    else:
        out = out.sort_values('Days_Until_Stockout', na_position='last')

    if 'Days_Until_Stockout' in out.columns:
        out['Days_Until_Stockout'] = out['Days_Until_Stockout'].round(1)

    return out.reset_index(drop=True)


def _day_summary(group, truck=None, day=None) -> pd.Series:
    # Grouping columns may be absent in pandas 3.x groupby.apply; accept them
    # as explicit args when available and fall back to group lookup otherwise.
    if truck is None:
        truck = group['Truck'].iloc[0]
    if day is None:
        day = group['Day'].iloc[0]
    first = group.sort_values('Stop').iloc[0]
    last  = group.sort_values('Stop').iloc[-1]
    load  = group['Refill_lbs'].sum()

    comp_a = (f"{first.get('Comp_A_Product','—')}  {int(first.get('Comp_A_lbs',0)):,} lbs"
              if 'Comp_A_Product' in first.index else '—')
    comp_b_lbs = int(first.get('Comp_B_lbs', 0)) if 'Comp_B_lbs' in first.index else 0
    comp_b = (f"{first.get('Comp_B_Product','—')}  {comp_b_lbs:,} lbs"
              if 'Comp_B_Product' in first.index and comp_b_lbs > 0 else '—')

    date_val = group['Date'].iloc[0] if 'Date' in group.columns else ''
    return pd.Series({
        'Truck':         truck,
        'Day':           day,
        'Date':          date_val,
        'Stops':         len(group),
        'Load_lbs':      int(load),
        'Compartment_A': comp_a,
        'Compartment_B': comp_b,
        'Dist_mi':       round(float(first.get('Route_Dist_mi', 0)), 1),
        'Time_min':      first.get('Route_Time_min', 0),
        'Cap_Pct':       round(load / TRUCKS[truck]['capacity_lbs'] * 100, 1),
        'Shift_Pct':     round(first.get('Route_Time_min', 0) / SHIFT_MIN * 100, 1),
        'Avg_Fill_Pct':  round(group['Fill_Pct'].mean(), 1),
    })


def _apply_excel_styles(filepath: Path, all_routes: Dict[int, pd.DataFrame]) -> None:
    wb  = load_workbook(str(filepath))
    bd  = Border(*[Side(style='thin', color='FFBBBBBB')] * 4)

    # Per-truck sheets are no longer generated; style only the summary/info sheets.

    # Style summary sheet (try new name first, fall back to old)
    summary_sn = '1_Week_Summary' if '1_Week_Summary' in wb.sheetnames else '0_Week_Summary'
    if summary_sn in wb.sheetnames:
        ws0 = wb[summary_sn]
        dark = PatternFill('solid', fgColor='FF2C3E50')
        for cell in ws0[1]:
            cell.fill = dark
            cell.font = Font(name='Arial', bold=True, color='FFFFFFFF', size=10)
        for col in ws0.columns:
            ws0.column_dimensions[get_column_letter(col[0].column)].width = 13

    # Style Plan_Info sheet
    if '0_Plan_Info' in wb.sheetnames:
        ws_pi = wb['0_Plan_Info']
        dark = PatternFill('solid', fgColor='FF2C3E50')
        for cell in ws_pi[1]:
            cell.fill = dark
            cell.font = Font(name='Arial', bold=True, color='FFFFFFFF', size=10)
        ws_pi.column_dimensions['A'].width = 18
        ws_pi.column_dimensions['B'].width = 28

    wb.save(str(filepath))


def _add_printable_sheets(
    filepath: Path,
    all_routes: Dict[int, pd.DataFrame],
    plan_dates: Optional[List],
    snapshot: Optional[pd.DataFrame],
    today: Optional[pd.Timestamp],
) -> None:
    """
    Append one combined printable sheet per delivery day to the workbook.

    Each sheet contains:
      • Truck2 block  (banner → manifest → blank → column headers → data rows)
      • Two-row gap
      • Truck9 block  (same structure)
      • Three-row gap + stockout warning section (if any stockouts exist)

    Print settings: landscape, fit-to-width = 1 page, 0.5 in margins.
    Column widths are set to accommodate both the manifest text (product names)
    and the route-table numbers — no hidden or truncated cells.
    """
    wb = load_workbook(str(filepath))
    bd = Border(*[Side(style='thin', color='FFBBBBBB')] * 4)

    # ── Print-friendly column selection (essential info only) ───────────
    PRINT_COLS = [
        'Stop', 'Arrival_HHMM', 'Customer', 'Zone', 'Product',
        'Refill_lbs', 'Days_Until_Stockout', 'Urgency', 'Tank_lbs',
    ]
    PRINT_COL_WIDTHS = {
        'Stop':                6,
        'Arrival_HHMM':        9,
        'Customer':           38,
        'Zone':                8,
        'Product':            18,
        'Refill_lbs':         14,
        'Days_Until_Stockout': 14,
        'Urgency':            12,
        'Tank_lbs':           13,
    }
    # Manifest occupies cols 1-5; enforce minimum widths so product names show.
    # Col 8 = "Status" in the caution table ("Confirmed — 2026-04-21" is ~26 chars).
    MANIFEST_MIN = {1: 12, 2: 24, 3: 26, 4: 16, 5: 10, 8: 28}

    urgency_fills = {
        'stockout': PatternFill('solid', fgColor=URGENCY_FILL_COLORS['stockout']),
        'critical': PatternFill('solid', fgColor=URGENCY_FILL_COLORS['critical']),
        'urgent':   PatternFill('solid', fgColor=URGENCY_FILL_COLORS['urgent']),
        'normal':   PatternFill('solid', fgColor=URGENCY_FILL_COLORS['normal']),
    }
    left_align_cols = {'Customer', 'Product'}

    for d in sorted(all_routes.keys()):
        routes = all_routes[d]
        if routes is None or routes.empty:
            continue

        if plan_dates and d < len(plan_dates):
            dt = plan_dates[d]
            day_short = dt.strftime('%a')
            date_str  = dt.strftime('%b %d')
        elif d < len(DAYS):
            day_short = DAYS[d]
            date_str  = ''
        else:
            day_short = f'Day{d}'
            date_str  = ''

        sn = f'PRINT {day_short} {date_str}'.strip()[:31]
        if sn in wb.sheetnames:
            del wb[sn]
        ws = wb.create_sheet(sn)
        ws.sheet_view.showGridLines = False

        cur = 1  # current writing row (1-based)

        # Determine route columns from first non-empty truck df (print subset)
        route_cols_present: List[str] = []
        for truck in TRUCK_NAMES:
            sub0 = routes[routes['Truck'] == truck]
            if not sub0.empty:
                route_cols_present = [c for c in PRINT_COLS if c in sub0.columns]
                break
        n_cols = max(len(route_cols_present), 5)

        for truck in TRUCK_NAMES:
            sub = routes[routes['Truck'] == truck].sort_values('Stop')
            if sub.empty:
                continue

            load  = int(sub['Refill_lbs'].sum())
            dist  = float(sub['Route_Dist_mi'].iloc[-1]) if 'Route_Dist_mi' in sub.columns else 0.0
            mins  = int(sub['Route_Time_min'].iloc[0]) if 'Route_Time_min' in sub.columns else 0
            hrs, m = divmod(mins, 60)
            cap_p = round(load / TRUCKS[truck]['capacity_lbs'] * 100, 1)

            # ── Banner row ───────────────────────────────────────────────────
            ws.merge_cells(start_row=cur, start_column=1,
                           end_row=cur, end_column=n_cols)
            hdr = ws.cell(cur, 1,
                value=f"{truck}  |  {day_short}  |  {len(sub)} stops  |  "
                      f"{load:,} lbs ({cap_p}%)  |  {dist:.1f} mi  |  {hrs}h {m:02d}m")
            hdr.font      = Font(name='Arial', bold=True, color='FFFFFFFF', size=16)
            hdr.fill      = PatternFill('solid', fgColor=TRUCK_HEX.get(truck, 'FF333333'))
            hdr.alignment = Alignment(horizontal='left', vertical='center')
            ws.row_dimensions[cur].height = 30
            cur += 1

            # ── Manifest header row ──────────────────────────────────────────
            mhdr_fill = PatternFill('solid', fgColor='FFD9DEE3')
            manifest_hdr_labels = ['LOAD', 'Compartment A', 'Compartment B', 'Total Load', 'Stops']
            for ci, label in enumerate(manifest_hdr_labels, 1):
                c = ws.cell(cur, ci, value=label)
                c.fill      = mhdr_fill
                c.font      = Font(name='Arial', bold=True, color='FF2C3E50', size=12)
                c.alignment = Alignment(horizontal='left', vertical='center')
            ws.row_dimensions[cur].height = 22
            cur += 1

            # ── Manifest data row ────────────────────────────────────────────
            comp_a_prod = str(sub['Comp_A_Product'].iloc[0]) if 'Comp_A_Product' in sub.columns else '—'
            comp_a_lbs  = int(sub['Comp_A_lbs'].iloc[0])    if 'Comp_A_lbs'     in sub.columns else 0
            comp_b_prod = str(sub['Comp_B_Product'].iloc[0]) if 'Comp_B_Product' in sub.columns else '—'
            comp_b_lbs  = int(sub['Comp_B_lbs'].iloc[0])    if 'Comp_B_lbs'     in sub.columns else 0

            manifest_vals = [
                truck,
                f'{comp_a_prod}  {comp_a_lbs:,} lbs',
                f'{comp_b_prod}  {comp_b_lbs:,} lbs' if comp_b_lbs > 0 else '—',
                f'{load:,} lbs',
                len(sub),
            ]
            mdata_fill = PatternFill('solid', fgColor='FFF5F7FA')
            for ci, val in enumerate(manifest_vals, 1):
                c = ws.cell(cur, ci, value=val)
                c.fill      = mdata_fill
                c.font      = Font(name='Arial', size=12, bold=True)
                c.alignment = Alignment(
                    horizontal='center' if ci == 5 else 'left',
                    vertical='center',
                )
            ws.row_dimensions[cur].height = 24
            cur += 1

            # ── Blank separator ──────────────────────────────────────────────
            cur += 1

            # ── Route-table column headers (print-friendly subset) ────────────
            rcols = [c for c in PRINT_COLS if c in sub.columns]
            for ci, col_name in enumerate(rcols, 1):
                c = ws.cell(cur, ci)
                c.value     = ROUTE_COL_LABELS.get(col_name, col_name)
                c.fill      = PatternFill('solid', fgColor='FF2C3E50')
                c.font      = Font(name='Arial', bold=True, color='FFFFFFFF', size=12)
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                c.border    = bd
            ws.row_dimensions[cur].height = 30
            cur += 1

            # ── Data rows ────────────────────────────────────────────────────
            for _, row in sub.iterrows():
                tier = str(row.get('Urgency', 'normal')).lower()
                fill = urgency_fills.get(tier, urgency_fills['normal'])
                for ci, col_name in enumerate(rcols, 1):
                    c = ws.cell(cur, ci)
                    c.value     = row.get(col_name)
                    c.fill      = fill
                    c.border    = bd
                    c.font      = Font(name='Arial', size=13)
                    c.alignment = Alignment(
                        horizontal='left' if col_name in left_align_cols else 'center',
                        vertical='center',
                    )
                    if col_name in ROUTE_COL_FORMATS:
                        c.number_format = ROUTE_COL_FORMATS[col_name]
                ws.row_dimensions[cur].height = 24
                cur += 1

            # Gap between trucks
            cur += 2

        # ── Day-specific caution section ─────────────────────────────────────
        # Rule: show clients whose stockout falls ON OR BEFORE this delivery day
        # BUT whose scheduled delivery is AFTER this day (or not scheduled at all).
        # Clients already served on or before today's day are excluded — they're fine.
        if snapshot is not None and not snapshot.empty:
            this_date = plan_dates[d] if (plan_dates and d < len(plan_dates)) else None

            # How many days from "today" to this delivery day
            if today is not None and this_date is not None:
                days_to_this_day = (this_date - today).days
            else:
                days_to_this_day = d + 1  # fallback

            # Build scheduled lookup: client_id → scheduled delivery date (Timestamp)
            all_data = [r for r in all_routes.values() if r is not None and not r.empty]
            combined_all = pd.concat(all_data, ignore_index=True) if all_data else pd.DataFrame()
            sched_date_map: dict = {}   # client_id → pd.Timestamp of delivery
            sched_str_map:  dict = {}   # client_id → display string
            if not combined_all.empty and 'ID' in combined_all.columns:
                for _, r in combined_all.iterrows():
                    cid = str(r.get('ID', ''))
                    if cid not in sched_date_map:
                        raw = r.get('Date', None)
                        try:
                            ts = pd.Timestamp(raw)
                        except Exception:
                            ts = None
                        sched_date_map[cid] = ts
                        sched_str_map[cid]  = str(raw) if raw else '?'

            # Filter step 1: stockout within [days_to_this_day] days
            if 'Days_Until_Stockout' in snapshot.columns:
                at_risk = snapshot[
                    snapshot['Days_Until_Stockout'].apply(
                        lambda x: pd.notna(x) and float(x) <= days_to_this_day
                    )
                ].copy()
            else:
                at_risk = pd.DataFrame()

            # Filter step 2: keep only clients whose delivery is AFTER this day
            # (or not scheduled at all). Clients served on/before this day are fine.
            def _delivered_after(cid):
                ts = sched_date_map.get(str(cid))
                if ts is None:
                    return True   # not scheduled → include
                if this_date is None:
                    return True
                return pd.Timestamp(ts).normalize() > pd.Timestamp(this_date).normalize()

            if not at_risk.empty and 'ID' in at_risk.columns:
                at_risk = at_risk[at_risk['ID'].astype(str).apply(_delivered_after)]

            at_risk = at_risk.sort_values('Days_Until_Stockout', na_position='last') \
                if not at_risk.empty else at_risk

            if not at_risk.empty:
                # Section header
                ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=n_cols)
                c = ws.cell(cur, 1,
                    value=f'⚠  CAUTION — Stocking out by {day_short} {date_str}, '
                          f'delivered AFTER this day')
                c.font      = Font(name='Arial', bold=True, size=15, color='FF993300')
                c.alignment = Alignment(horizontal='left', vertical='center')
                ws.row_dimensions[cur].height = 30
                cur += 1

                ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=n_cols)
                c = ws.cell(cur, 1,
                    value='These clients run out on or before this day but will not receive '
                          'a delivery until later — or are not scheduled at all.')
                c.font      = Font(name='Arial', size=11, color='FF993300', italic=True)
                c.alignment = Alignment(horizontal='left', vertical='center')
                ws.row_dimensions[cur].height = 22
                cur += 2

                # Table header
                caut_cols = ['ID', 'Customer', 'Product', 'Tank (lbs)', 'Curr (lbs)',
                             'Days→Empty', 'Stockout Date', 'Delivery Scheduled']
                caut_hdr_fill = PatternFill('solid', fgColor='FF993300')
                for ci, h in enumerate(caut_cols, 1):
                    c = ws.cell(cur, ci, value=h)
                    c.fill      = caut_hdr_fill
                    c.font      = Font(name='Arial', bold=True, color='FFFFFFFF', size=12)
                    c.alignment = Alignment(horizontal='center', vertical='center')
                    c.border    = bd
                ws.row_dimensions[cur].height = 26
                cur += 1

                for _, row in at_risk.iterrows():
                    cid      = str(row.get('ID', ''))
                    days_val = float(row.get('Days_Until_Stockout', 0) or 0)

                    if today is not None:
                        stk_date = (today + pd.Timedelta(days=max(days_val, 0))).strftime('%a %b %d')
                    else:
                        stk_date = '—'

                    if cid in sched_date_map and sched_date_map[cid] is not None:
                        delivery = f'Scheduled {sched_str_map[cid]}'
                        row_fill = urgency_fills['critical']   # late delivery — highlight
                    else:
                        delivery = 'NOT SCHEDULED ⚠'
                        row_fill = urgency_fills['stockout']

                    vals = [
                        cid,
                        str(row.get('Customer', ''))[:40],
                        str(row.get('Product', '')),
                        row.get('Tank_lbs', ''),
                        row.get('Current_lbs', row.get('Est_Current_lbs', '')),
                        round(days_val, 1),
                        stk_date,
                        delivery,
                    ]
                    for ci, val in enumerate(vals, 1):
                        c = ws.cell(cur, ci, value=val)
                        c.fill      = row_fill
                        c.font      = Font(name='Arial', size=11,
                                          bold=(ci == 8))
                        c.border    = bd
                        c.alignment = Alignment(
                            horizontal='left' if ci in (2, 8) else 'center',
                            vertical='center',
                        )
                    ws.row_dimensions[cur].height = 22
                    cur += 1

        # ── Column widths ────────────────────────────────────────────────────
        print_cols_present = [c for c in PRINT_COLS if c in route_cols_present]
        for ci, col_name in enumerate(print_cols_present, 1):
            w = PRINT_COL_WIDTHS.get(col_name, 12)
            w = max(w, MANIFEST_MIN.get(ci, 0))
            ws.column_dimensions[get_column_letter(ci)].width = w

        # ── Print settings (landscape, fit 1 page wide) ──────────────────────
        ws.page_setup.orientation          = 'landscape'
        ws.page_setup.paperSize            = 1       # Letter
        ws.page_setup.fitToPage            = True
        ws.page_setup.fitToWidth           = 1
        ws.page_setup.fitToHeight          = 0       # auto height pages
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_margins = PageMargins(
            left=0.5, right=0.5, top=0.5, bottom=0.5,
            header=0.2, footer=0.2,
        )
        # Print area covers all written rows
        ws.print_area = f'A1:{get_column_letter(ws.max_column)}{cur}'
        ws.oddHeader.center.text = f'S&K Oil Sales — Route Schedule  |  {day_short} {date_str}'
        ws.oddFooter.right.text  = 'Page &P of &N'

    wb.save(str(filepath))
    print(f'  Printable sheets added: {[s for s in wb.sheetnames if s.startswith("PRINT")]}')


def _style_route_sheet(ws, truck: str, day: str, sub: pd.DataFrame, bd) -> None:
    hfill = PatternFill('solid', fgColor=TRUCK_HEX.get(truck, 'FF333333'))
    load  = int(sub['Refill_lbs'].sum())
    dist  = float(sub['Route_Dist_mi'].iloc[-1]) if len(sub) else 0.0
    mins  = sub['Route_Time_min'].iloc[0] if len(sub) else 0
    hrs, m = divmod(int(mins), 60)
    cap_p  = round(load / TRUCKS[truck]['capacity_lbs'] * 100, 1)

    ws.insert_rows(1)
    ws.merge_cells(
        start_row=1, start_column=1,
        end_row=1, end_column=ws.max_column
    )
    hdr = ws.cell(1, 1,
        value=f"{truck}  |  {day}  |  {len(sub)} stops  |  "
              f"{load:,} lbs ({cap_p}%)  |  {dist:.1f} mi  |  {hrs}h {m:02d}m"
    )
    hdr.font      = Font(name='Arial', bold=True, color='FFFFFFFF', size=12)
    hdr.fill      = hfill
    hdr.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 22

    # Layout after insert_rows(1):
    #   Row 1: banner (just set)
    #   Row 2: manifest column headers
    #   Row 3: manifest data
    #   Row 4: blank
    #   Row 5: route table column headers
    #   Row 6+: route data

    # Manifest header styling (row 2)
    manifest_hdr_fill = PatternFill('solid', fgColor='FFD9DEE3')
    for cell in ws[2]:
        cell.fill      = manifest_hdr_fill
        cell.font      = Font(name='Arial', bold=True, color='FF2C3E50', size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 18

    # Manifest data row styling (row 3)
    for cell in ws[3]:
        cell.font      = Font(name='Arial', size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center')

    # Determine which columns the route table actually wrote (in order)
    route_cols_present = [c for c in ROUTE_COLS if c in sub.columns]
    n_route_cols = len(route_cols_present)

    # Route table column headers (row 5) — apply friendly labels
    for ci, col_name in enumerate(route_cols_present, start=1):
        cell = ws.cell(5, ci)
        cell.value     = ROUTE_COL_LABELS.get(col_name, col_name)
        cell.fill      = PatternFill('solid', fgColor='FF2C3E50')
        cell.font      = Font(name='Arial', bold=True, color='FFFFFFFF', size=9)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = bd
    ws.row_dimensions[5].height = 28

    # Data rows — colour by urgency tier
    urgency_fills = {
        'stockout': PatternFill('solid', fgColor=URGENCY_FILL_COLORS['stockout']),
        'critical': PatternFill('solid', fgColor=URGENCY_FILL_COLORS['critical']),
        'urgent':   PatternFill('solid', fgColor=URGENCY_FILL_COLORS['urgent']),
        'normal':   PatternFill('solid', fgColor=URGENCY_FILL_COLORS['normal']),
    }
    # Columns that should be left-aligned (Customer, Zone, Product — by name)
    left_align_cols = {'Customer', 'Product'}

    for ri, (_, row) in enumerate(sub.iterrows(), start=6):
        tier = str(row.get('Urgency', 'normal')).lower()
        fill = urgency_fills.get(tier, urgency_fills['normal'])
        for ci, col_name in enumerate(route_cols_present, start=1):
            c         = ws.cell(ri, ci)
            c.fill    = fill
            c.border  = bd
            c.font    = Font(name='Arial', size=9)
            c.alignment = Alignment(
                horizontal='left' if col_name in left_align_cols else 'center',
                vertical='center',
            )
            if col_name in ROUTE_COL_FORMATS:
                c.number_format = ROUTE_COL_FORMATS[col_name]

    # Column widths keyed by canonical column name (works regardless of column order)
    col_width_by_name = {
        'Stop':                  6,
        'Date':                  12,
        'Travel_To_Min':         10,
        'Dist_To_mi':            9,
        'Customer':              34,
        'Zone':                  6,
        'Days_Until_Stockout':   11,
        'DaysToStockoutAtVisit': 11,
        'Urgency':               10,
        'Tank_lbs':              10,
        'Current_lbs':           10,
        'Refill_lbs':            11,
        'Fill_Pct':              8,
        'Avg_LbsPerDay':         11,
        'Service_Min':           9,
        'Product':               16,
        'Route_Dist_mi':         12,
        'Route_Time_min':        10,
        'Shift_Pct':             8,
        'Cap_Pct':               8,
    }
    # The manifest row shares columns A–D with the route table.
    # Ensure those columns are wide enough to show compartment product names.
    # e.g. "FRYERS CHOICE  2,633 lbs" (24 chars) needs ≥ 22 units in col C.
    MANIFEST_MIN_WIDTHS = {1: 10, 2: 20, 3: 22, 4: 14}  # col index → min width
    for ci, col_name in enumerate(route_cols_present, start=1):
        w = col_width_by_name.get(col_name, 10)
        w = max(w, MANIFEST_MIN_WIDTHS.get(ci, 0))
        ws.column_dimensions[get_column_letter(ci)].width = w
    # Pad remaining columns if manifest is wider than route table (rare)
    for ci in range(n_route_cols + 1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    # Freeze top area so header stays visible while scrolling
    ws.freeze_panes = 'A6'


# ── Interactive road-geometry map (server-side OSRM fetch) ─────────────────────

import hashlib
import requests as _requests


def _waypoint_cache_key(waypoints: list) -> str:
    """Stable hash of a waypoint sequence. 5-decimal lat/lon so trivial
    rounding doesn't invalidate the cache."""
    s = ';'.join(f'{w["lat"]:.5f},{w["lon"]:.5f}' for w in waypoints)
    return hashlib.sha1(s.encode('utf-8')).hexdigest()


def _load_geom_cache() -> dict:
    if GEOM_CACHE_FILE.exists():
        try:
            return json.loads(GEOM_CACHE_FILE.read_text(encoding='utf-8'))
        except Exception as exc:
            print(f"    ⚠  polyline cache unreadable ({exc}); starting fresh")
    return {}


def _save_geom_cache(cache: dict) -> None:
    try:
        GEOM_CACHE_FILE.parent.mkdir(parents=True, exist_ok=True)
        tmp = GEOM_CACHE_FILE.with_suffix('.tmp.json')
        tmp.write_text(json.dumps(cache), encoding='utf-8')
        tmp.replace(GEOM_CACHE_FILE)
    except Exception as exc:
        print(f"    ⚠  could not persist polyline cache: {exc}")


def _osrm_route_line(coords: list) -> list:
    """Fetch road geometry from OSRM.  coords = [(lat, lon), ...].
    Returns [[lat, lon], ...] for Leaflet.  Falls back to straight lines."""
    if len(coords) < 2:
        return [[la, lo] for la, lo in coords]
    if len(coords) > 98:              # OSRM max-waypoints guard
        coords = coords[:98]
    cs = ';'.join(f'{lo:.5f},{la:.5f}' for la, lo in coords)
    try:
        r = _requests.get(
            f'https://router.project-osrm.org/route/v1/driving/{cs}'
            f'?overview=full&geometries=geojson',
            timeout=20,
            headers={'User-Agent': 'sk-optimizer/1.0'},
        )
        r.raise_for_status()
        d = r.json()
        if d.get('code') == 'Ok':
            return [[p[1], p[0]] for p in d['routes'][0]['geometry']['coordinates']]
    except Exception as exc:
        print(f"    OSRM fallback (straight lines): {exc}")
    return [[la, lo] for la, lo in coords]


def save_route_map(
    all_routes:  Dict[int, pd.DataFrame],
    filename:    str = 'sk_route_map.html',
    output_dir:  Path = OUTPUT_DIR,
    use_osrm:    bool = True,
) -> Path:
    """
    Generate a standalone HTML map with real road-geometry routes.

    Road paths are fetched from OSRM at generation time (server-side) and
    baked into the HTML — the browser does zero async network calls, so the
    map renders instantly.
    """
    filepath = output_dir / filename

    # ── Load persisted polyline cache ─────────────────────────────────────
    geom_cache   = _load_geom_cache()
    cache_hits   = 0
    cache_misses = 0
    cache_fails  = 0

    # ── Collect all routes into a JS-serialisable structure ───────────────
    route_groups = []   # list of dicts, one per (truck, day)

    for d in sorted(all_routes.keys()):
        routes = all_routes[d]
        if routes is None or (hasattr(routes, 'empty') and routes.empty):
            continue
        for truck in TRUCK_NAMES:
            sub = routes[routes['Truck'] == truck].sort_values('Stop')
            if sub.empty:
                continue

            color = TRUCK_MAP_COLORS[truck][d % len(TRUCK_MAP_COLORS[truck])]

            # Waypoints: depot → stops → depot
            waypoints = (
                [{'lat': DEPOT_LAT, 'lon': DEPOT_LON, 'label': 'Depot'}]
                + [
                    {
                        'lat':     float(r['Lat']),
                        'lon':     float(r['Lon']),
                        'stop':    int(r['Stop']),
                        'label':   str(r['Customer'])[:35],
                        'zone':    str(r.get('Zone', '')),
                        'refill':  int(r['Refill_lbs']),
                        'fill_pct': round(float(r.get('Fill_Pct', 0)), 1),
                        'days_left': round(float(r.get('DaysToStockoutAtVisit', 0)), 1),
                        'travel_min': int(r['Travel_To_Min']),
                        'dist_mi':  round(float(r.get('Dist_To_mi', 0)), 1),
                        'urgency':  str(r.get('Urgency', 'normal')).lower(),
                    }
                    for _, r in sub.iterrows()
                ]
                + [{'lat': DEPOT_LAT, 'lon': DEPOT_LON, 'label': 'Depot'}]
            )

            load      = int(sub['Refill_lbs'].sum())
            cap_pct   = round(load / TRUCKS[truck]['capacity_lbs'] * 100, 1)
            dist      = round(float(sub['Route_Dist_mi'].iloc[-1]), 1)
            time_min  = int(sub['Route_Time_min'].iloc[0])
            hrs, mins = divmod(time_min, 60)

            # ── Resolve road-geometry polyline via cache + OSRM fallback ──
            key = _waypoint_cache_key(waypoints)
            polyline = geom_cache.get(key)
            from_cache = polyline is not None
            if polyline is None and use_osrm:
                coords = [(w['lat'], w['lon']) for w in waypoints]
                polyline = _osrm_route_line(coords)
                # Heuristic: straight-line fallback returns exactly len(waypoints)
                # points. Only cache responses that look like real road geometry.
                if polyline and len(polyline) > len(waypoints):
                    geom_cache[key] = polyline
                    cache_misses += 1
                else:
                    cache_fails += 1
            elif from_cache:
                cache_hits += 1
            if polyline is None:
                polyline = [[w['lat'], w['lon']] for w in waypoints]

            # Day label, date, status
            day_label = str(sub.iloc[0].get('Day', f'Day{d}')) if 'Day' in sub.columns else (
                DAYS[d] if d < len(DAYS) else f'Day{d}'
            )
            date_str = str(sub.iloc[0].get('Date', '')) if 'Date' in sub.columns else ''
            status = str(sub.iloc[0].get('Status', 'COMMITTED')) if 'Status' in sub.columns else 'COMMITTED'
            # Unique ID using day index to avoid collisions across weeks
            route_groups.append({
                'id':        f"{truck}_d{d}",
                'truck':     truck,
                'day':       day_label,
                'dayIndex':  d,
                'date':      date_str,
                'status':    status,
                'color':     color,
                'label':     f"{truck} – {day_label}",
                'stops':     len(sub),
                'load':      load,
                'capPct':    cap_pct,
                'dist':      dist,
                'timeMin':   time_min,
                'summary':   f"{len(sub)} stops · {load:,} lbs ({cap_pct}%) · {dist} mi · {hrs}h {mins:02d}m",
                'waypoints': waypoints,
                'polyline':  polyline,
                'from_cache': from_cache,
            })

    # ── Persist updated polyline cache ─────────────────────────────────────
    if cache_misses:
        _save_geom_cache(geom_cache)
    print(f"  Polyline cache: {cache_hits} hit, {cache_misses} fetched, "
          f"{cache_fails} fallback ({len(geom_cache)} total cached)")

    # ── Render HTML ────────────────────────────────────────────────────────
    html = _render_map_html(route_groups)
    filepath.write_text(html, encoding='utf-8')
    print(f"  Map saved:   {filepath}")
    return filepath


def _render_map_html(route_groups: list) -> str:
    import json as _json

    routes_js = _json.dumps(route_groups, ensure_ascii=False)

    # Determine committed day count from data
    committed_days = sorted(set(g['dayIndex'] for g in route_groups if g.get('status') == 'COMMITTED'))
    tentative_days = sorted(set(g['dayIndex'] for g in route_groups if g.get('status') == 'TENTATIVE'))

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>S&amp;K Oil — Route Dispatch</title>
<meta name="viewport" content="width=device-width, initial-scale=1">
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>

<style>
:root {{
  --bg: #ffffff;
  --panel: #f8f9fb;
  --panel-hover: #eef0f4;
  --panel-active: #e4e7ed;
  --accent: #d4542b;
  --blue: #2563eb;
  --green: #059669;
  --text: #1e293b;
  --muted: #64748b;
  --border: #e2e8f0;
  --t2: #1a6faf;
  --t9: #c0392b;
}}
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', 'Inter', sans-serif;
       display: flex; height: 100vh; overflow: hidden; background: var(--bg); color: var(--text); }}

/* ── Sidebar ────────────────────────────── */
#sidebar {{
    width: 340px; min-width: 340px; background: var(--bg);
    display: flex; flex-direction: column; overflow: hidden;
    border-right: 1px solid var(--border);
    box-shadow: 2px 0 12px rgba(0,0,0,.06);
    z-index: 1000;
}}
#sidebar-header {{
    padding: 16px 20px 12px; border-bottom: 1px solid var(--border);
}}
#sidebar-header h1 {{
    font-size: 11px; font-weight: 800; letter-spacing: 2px; text-transform: uppercase;
    color: var(--accent); margin-bottom: 2px;
}}
#sidebar-header h2 {{ font-size: 17px; color: var(--text); font-weight: 700; }}

/* ── View toggle (Committed / Full Plan) ── */
#view-toggle {{
    display: flex; gap: 0; border-bottom: 1px solid var(--border);
}}
.view-btn {{
    flex: 1; padding: 10px 8px; border: none; background: transparent;
    font-size: 12px; font-weight: 700; color: var(--muted); cursor: pointer;
    border-bottom: 2px solid transparent; transition: all .15s; text-align: center;
}}
.view-btn:hover {{ color: var(--text); background: var(--panel); }}
.view-btn.active {{ color: var(--green); border-bottom-color: var(--green); }}
.view-btn.active.tent {{ color: var(--blue); border-bottom-color: var(--blue); }}
.view-btn .vb-count {{
    display: inline-block; background: var(--panel); border-radius: 10px;
    padding: 1px 7px; font-size: 10px; font-weight: 800; margin-left: 4px;
}}

/* ── Day strip ────────────────────────────── */
#day-strip {{
    display: flex; gap: 0; border-bottom: 1px solid var(--border); overflow-x: auto;
    scrollbar-width: none;
}}
#day-strip::-webkit-scrollbar {{ display: none; }}
.ds-day {{
    flex: 0 0 auto; min-width: 62px; padding: 8px 4px; text-align: center;
    cursor: pointer; border-bottom: 3px solid transparent; transition: all .15s;
    position: relative;
}}
.ds-day:hover {{ background: var(--panel); }}
.ds-day.active {{ border-bottom-color: var(--green); background: rgba(5,150,105,.04); }}
.ds-day.active.tent {{ border-bottom-color: var(--blue); background: rgba(37,99,235,.03); }}
.ds-day .ds-weekday {{
    font-size: 10px; font-weight: 700; text-transform: uppercase; letter-spacing: .5px;
    color: var(--muted);
}}
.ds-day.active .ds-weekday {{ color: var(--text); }}
.ds-day .ds-date {{ font-size: 16px; font-weight: 800; color: var(--text); }}
.ds-day .ds-stops {{ font-size: 9px; color: var(--muted); margin-top: 1px; }}
.ds-day .ds-badge {{
    position: absolute; top: 4px; right: 4px; width: 7px; height: 7px;
    border-radius: 50%; background: var(--green);
}}
.ds-day.tent .ds-badge {{ background: var(--blue); }}
.ds-day.tent .ds-weekday {{ opacity: .65; }}
.ds-day.tent .ds-date {{ opacity: .65; }}

/* ── Truck filter ─────────────────────────── */
#truck-filter {{
    display: flex; gap: 6px; padding: 8px 16px; border-bottom: 1px solid var(--border);
}}
.truck-btn {{
    flex: 1; padding: 6px 8px; border: 1.5px solid var(--border); border-radius: 8px;
    background: transparent; color: var(--muted); font-size: 12px; font-weight: 700;
    cursor: pointer; transition: all .15s; text-align: center;
    display: flex; align-items: center; justify-content: center; gap: 6px;
}}
.truck-btn:hover {{ border-color: #94a3b8; color: var(--text); }}
.truck-btn.active[data-truck="Truck2"] {{ background: #eff6ff; border-color: var(--t2); color: var(--t2); }}
.truck-btn.active[data-truck="Truck9"] {{ background: #fef2f2; border-color: var(--t9); color: var(--t9); }}
.truck-btn.active[data-truck="All"] {{ background: var(--panel); border-color: var(--muted); color: var(--text); }}
.truck-dot {{ width: 8px; height: 8px; border-radius: 50%; }}

/* ── Route list ───────────────────────────── */
#route-list {{ overflow-y: auto; flex: 1; }}

.day-section {{ }}
.day-header {{
    padding: 8px 16px 4px; font-size: 11px; font-weight: 800; text-transform: uppercase;
    letter-spacing: .5px; color: var(--muted); background: var(--panel);
    border-bottom: 1px solid var(--border); position: sticky; top: 0; z-index: 10;
    display: flex; align-items: center; gap: 8px;
}}
.day-header .dh-badge {{
    font-size: 9px; font-weight: 700; padding: 1px 6px; border-radius: 4px;
}}
.day-header .dh-committed {{ background: rgba(5,150,105,.1); color: var(--green); }}
.day-header .dh-tentative {{ background: rgba(37,99,235,.08); color: var(--blue); }}

.route-card {{
    padding: 10px 16px; cursor: pointer; border-bottom: 1px solid var(--border);
    border-left: 4px solid transparent; transition: all .15s; opacity: 0.35;
}}
.route-card.visible {{ opacity: 1; }}
.route-card.tent.visible {{ opacity: .7; }}
.route-card:hover {{ background: var(--panel-hover); }}
.route-card.focused {{ background: #eff6ff; }}

.rc-row {{
    display: flex; align-items: center; gap: 10px;
}}
.rc-truck {{
    font-size: 12px; font-weight: 800; white-space: nowrap;
    display: flex; align-items: center; gap: 5px;
}}
.rc-metrics {{
    display: flex; gap: 8px; font-size: 11px; color: var(--muted); flex: 1;
}}
.rc-metrics span {{ white-space: nowrap; }}
.rc-crit {{
    font-size: 9px; font-weight: 800; padding: 2px 6px; border-radius: 4px;
    background: rgba(176,48,48,.1); color: #b03030;
}}

/* ── Stats bar ────────────────────────────── */
#stats-bar {{
    display: flex; gap: 0; border-bottom: 1px solid var(--border);
}}
.stat-cell {{
    flex: 1; text-align: center; padding: 8px 4px;
    border-right: 1px solid var(--border);
}}
.stat-cell:last-child {{ border-right: none; }}
.stat-num {{ font-size: 17px; font-weight: 800; color: var(--text); }}
.stat-label {{ font-size: 8px; text-transform: uppercase; letter-spacing: 0.5px; color: var(--muted); margin-top: 1px; }}

/* ── Legend ────────────────────────────────── */
#legend-bar {{
    display: flex; align-items: center; gap: 10px; padding: 8px 16px;
    border-top: 1px solid var(--border); font-size: 9px; color: var(--muted);
    flex-wrap: wrap; flex-shrink: 0;
}}
.leg-item {{ display: flex; align-items: center; gap: 3px; }}
.leg-swatch {{ width: 7px; height: 7px; border-radius: 50%; }}

/* ── Map ──────────────────────────────────── */
#map {{ flex: 1; }}

/* ── Popup ────────────────────────────────── */
.leaflet-popup-content-wrapper {{
    background: #ffffff !important; color: var(--text) !important;
    border-radius: 10px !important; box-shadow: 0 8px 30px rgba(0,0,0,.12) !important;
    border: 1px solid var(--border) !important;
}}
.leaflet-popup-tip {{ background: #ffffff !important; }}
.leaflet-popup-close-button {{ color: var(--muted) !important; font-size: 18px !important; }}

.sk-popup {{ min-width: 220px; }}
.sk-popup .pop-header {{
    display: flex; align-items: center; gap: 8px; margin-bottom: 8px;
    padding-bottom: 8px; border-bottom: 1px solid var(--border);
}}
.sk-popup .pop-num {{
    width: 28px; height: 28px; border-radius: 50%; display: flex;
    align-items: center; justify-content: center; font-size: 12px;
    font-weight: 800; color: white; flex-shrink: 0;
}}
.sk-popup .pop-name {{ font-size: 12px; font-weight: 700; line-height: 1.3; }}
.sk-popup .pop-route {{ font-size: 10px; color: var(--muted); }}
.sk-popup .pop-grid {{
    display: grid; grid-template-columns: 1fr 1fr; gap: 6px;
}}
.sk-popup .pop-metric {{
    background: #f1f5f9; border-radius: 6px; padding: 6px 8px;
}}
.sk-popup .pop-metric .pm-val {{ font-size: 14px; font-weight: 800; }}
.sk-popup .pop-metric .pm-label {{ font-size: 9px; color: var(--muted); text-transform: uppercase; letter-spacing: 0.3px; }}
.sk-popup .pop-urg {{
    display: inline-block; margin-top: 6px; padding: 2px 8px; border-radius: 4px;
    font-size: 10px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.5px;
}}

/* ── Tooltip ──────────────────────────────── */
.leaflet-tooltip {{
    background: #1e293b !important; color: #ffffff !important;
    border: none !important; border-radius: 6px !important;
    font-size: 11px !important; font-weight: 600 !important; padding: 4px 8px !important;
    box-shadow: 0 4px 12px rgba(0,0,0,.2) !important;
}}
.leaflet-tooltip-top:before {{ border-top-color: #1e293b !important; }}
.leaflet-tooltip-bottom:before {{ border-bottom-color: #1e293b !important; }}
</style>
</head>
<body>

<div id="sidebar">
  <div id="sidebar-header">
    <h1>S&amp;K Oil Sales</h1>
    <h2>Route Dispatch</h2>
  </div>

  <div id="view-toggle"></div>
  <div id="stats-bar"></div>
  <div id="day-strip"></div>
  <div id="truck-filter"></div>
  <div id="route-list"></div>

  <div id="legend-bar">
    <span style="font-weight:700;color:var(--text)">Urgency:</span>
    <div class="leg-item"><div class="leg-swatch" style="background:#b03030"></div> Stockout</div>
    <div class="leg-item"><div class="leg-swatch" style="background:#c47a1a"></div> Critical</div>
    <div class="leg-item"><div class="leg-swatch" style="background:#9e8a1e"></div> Urgent</div>
    <div class="leg-item"><div class="leg-swatch" style="background:#2a8a4a"></div> Normal</div>
  </div>
</div>

<div id="map"></div>

<script>
var ROUTES    = {routes_js};
var DEPOT_LAT = {DEPOT_LAT};
var DEPOT_LON = {DEPOT_LON};

var URG = {{
    stockout: {{ color: '#b03030', bg: 'rgba(176,48,48,.12)', label: 'STOCKOUT' }},
    critical: {{ color: '#c47a1a', bg: 'rgba(196,122,26,.12)', label: 'CRITICAL' }},
    urgent:   {{ color: '#9e8a1e', bg: 'rgba(158,138,30,.12)',  label: 'URGENT' }},
    normal:   {{ color: '#2a8a4a', bg: 'rgba(42,138,74,.12)',  label: 'NORMAL' }},
}};
var TRUCK_COLORS = {{ 'Truck2': '#1a6faf', 'Truck9': '#c0392b' }};

// ── Derived data ───────────────────────────────────
// Build unique day list from route data (preserves order)
var dayMap = {{}};
ROUTES.forEach(function(g) {{
    if (!dayMap[g.dayIndex]) {{
        dayMap[g.dayIndex] = {{
            idx: g.dayIndex, day: g.day, date: g.date, status: g.status,
            stops: 0, lbs: 0, routes: []
        }};
    }}
    dayMap[g.dayIndex].stops += g.stops;
    dayMap[g.dayIndex].lbs += g.load;
    dayMap[g.dayIndex].routes.push(g);
}});
var allDayKeys = Object.keys(dayMap).map(Number).sort(function(a,b){{ return a-b; }});

var committedKeys = allDayKeys.filter(function(k) {{ return dayMap[k].status === 'COMMITTED'; }});
var tentativeKeys = allDayKeys.filter(function(k) {{ return dayMap[k].status === 'TENTATIVE'; }});

// ── Map ─────────────────────────────────────────────
var map = L.map('map', {{ center: [33.48, -112.00], zoom: 11, zoomControl: false }});
L.control.zoom({{ position: 'topright' }}).addTo(map);
L.tileLayer('https://{{s}}.basemaps.cartocdn.com/rastertiles/voyager/{{z}}/{{x}}/{{y}}{{r}}.png', {{
    attribution: '&copy; OpenStreetMap &copy; CARTO', maxZoom: 19, subdomains: 'abcd'
}}).addTo(map);

// Depot marker
L.marker([DEPOT_LAT, DEPOT_LON], {{
    icon: L.divIcon({{
        html: '<div style="width:36px;height:36px;background:#d4542b;border-radius:50%;'
            + 'display:flex;align-items:center;justify-content:center;'
            + 'border:3px solid white;box-shadow:0 0 16px rgba(212,84,43,.35),0 4px 12px rgba(0,0,0,.2);">'
            + '<svg width="16" height="16" viewBox="0 0 24 24" fill="white"><path d="M10 20v-6h4v6h5v-8h3L12 3 2 12h3v8z"/></svg></div>',
        iconSize: [36, 36], iconAnchor: [18, 18], className: ''
    }}), zIndexOffset: 1000
}}).bindTooltip('S&K Depot', {{ direction: 'top', offset: [0, -20] }}).addTo(map);

// ── State ───────────────────────────────────────────
var layerGroups  = {{}};
var focusedRoute = null;
var viewMode     = 'committed';   // 'committed' or 'full'
var selectedDay  = committedKeys.length > 0 ? committedKeys[0] : allDayKeys[0];
var activeTrucks = new Set(['Truck2', 'Truck9']);

// ── Build layers ────────────────────────────────────
function addRoadLine(group) {{
    var poly = group.polyline || group.waypoints.map(function(w) {{ return [w.lat, w.lon]; }});
    var isRoad = poly.length > group.waypoints.length;
    var isTent = group.status === 'TENTATIVE';
    var style = isRoad
        ? {{ color: group.color, weight: isTent ? 3 : 4, opacity: isTent ? 0.5 : 0.85 }}
        : {{ color: group.color, weight: isTent ? 2 : 3, opacity: isTent ? 0.35 : 0.6, dashArray: '8 6' }};
    L.polyline(poly, style).addTo(layerGroups[group.id]);
}}

function buildMarkers(group) {{
    var lg = layerGroups[group.id];
    var stops = group.waypoints.filter(function(w) {{ return w.stop !== undefined; }});
    stops.forEach(function(w) {{
        var u = URG[w.urgency] || URG.normal;
        var sz = w.urgency === 'critical' || w.urgency === 'stockout' ? 28 : 24;
        var fs = sz === 28 ? 11 : 10;
        var glow = w.urgency === 'critical' || w.urgency === 'stockout'
            ? 'box-shadow:0 0 6px ' + u.color + '88,0 2px 4px rgba(0,0,0,.18);'
            : 'box-shadow:0 2px 4px rgba(0,0,0,.12);';
        var icon = L.divIcon({{
            html: '<div style="width:'+sz+'px;height:'+sz+'px;background:'+group.color+';'
                + 'border-radius:50%;display:flex;align-items:center;justify-content:center;'
                + 'font-size:'+fs+'px;font-weight:800;color:white;'
                + 'border:2.5px solid '+u.color+';'+glow+'">' + w.stop + '</div>',
            iconSize: [sz, sz], iconAnchor: [sz/2, sz/2], className: ''
        }});

        // Date label for popup
        var dateLabel = group.date ? new Date(group.date + 'T12:00:00').toLocaleDateString('en-US', {{weekday:'short', month:'short', day:'numeric'}}) : group.day;
        var statusBadge = group.status === 'COMMITTED'
            ? '<span style="color:#059669;font-weight:700">DISPATCH</span>'
            : '<span style="color:#2563eb;font-weight:700">TENTATIVE</span>';

        var popup = '<div class="sk-popup">'
            + '<div class="pop-header">'
            + '<div class="pop-num" style="background:'+group.color+'">'+w.stop+'</div>'
            + '<div><div class="pop-name">'+w.label+'</div>'
            + '<div class="pop-route">'+group.truck+' \\u00b7 '+dateLabel+' \\u00b7 '+statusBadge+'</div></div>'
            + '</div>'
            + '<div class="pop-grid">'
            + '<div class="pop-metric"><div class="pm-val">'+w.refill.toLocaleString()+'</div><div class="pm-label">Lbs Refill</div></div>'
            + '<div class="pop-metric"><div class="pm-val">'+w.fill_pct+'%</div><div class="pm-label">Fill %</div></div>'
            + '<div class="pop-metric"><div class="pm-val">'+w.days_left+'d</div><div class="pm-label">To Stockout</div></div>'
            + '<div class="pop-metric"><div class="pm-val">'+w.dist_mi+' mi</div><div class="pm-label">Drive Dist</div></div>'
            + '</div>'
            + '<div class="pop-urg" style="background:'+u.bg+';color:'+u.color+'">'+u.label+'</div>'
            + '</div>';

        L.marker([w.lat, w.lon], {{ icon: icon, zIndexOffset: w.urgency === 'critical' || w.urgency === 'stockout' ? 500 : 0 }})
         .bindPopup(popup, {{ maxWidth: 280 }})
         .bindTooltip('#'+w.stop+' '+w.label.substring(0,28), {{ direction: 'top', offset: [0, -sz/2 - 4] }})
         .addTo(lg);
    }});
}}

ROUTES.forEach(function(g) {{
    var lg = L.layerGroup();
    layerGroups[g.id] = lg;
    addRoadLine(g);
    buildMarkers(g);
}});

// ── View toggle ─────────────────────────────────────
var cStops = 0, cLbs = 0, tStops = 0, tLbs = 0;
ROUTES.forEach(function(g) {{
    if (g.status === 'COMMITTED') {{ cStops += g.stops; cLbs += g.load; }}
    else {{ tStops += g.stops; tLbs += g.load; }}
}});

var vtBar = document.getElementById('view-toggle');
vtBar.innerHTML =
    '<button class="view-btn active" data-view="committed">'
    + 'Dispatch <span class="vb-count">'+cStops+' stops</span></button>'
    + '<button class="view-btn tent" data-view="full">'
    + 'Full Plan <span class="vb-count">'+(cStops+tStops)+' stops</span></button>';

vtBar.querySelectorAll('.view-btn').forEach(function(btn) {{
    btn.addEventListener('click', function() {{
        viewMode = btn.dataset.view;
        // Reset selected day to first visible
        if (viewMode === 'committed' && tentativeKeys.indexOf(selectedDay) >= 0) {{
            selectedDay = committedKeys[0] || allDayKeys[0];
        }}
        updateAll();
    }});
}});

// ── Stats bar ───────────────────────────────────────
function renderStats() {{
    var visibleRoutes = ROUTES.filter(function(g) {{
        if (viewMode === 'committed' && g.status !== 'COMMITTED') return false;
        return true;
    }});
    var s = 0, lb = 0, mi = 0, r = 0;
    visibleRoutes.forEach(function(g) {{
        s += g.stops; lb += g.load; mi += g.dist; r++;
    }});
    document.getElementById('stats-bar').innerHTML =
        '<div class="stat-cell"><div class="stat-num">'+s+'</div><div class="stat-label">Stops</div></div>'
      + '<div class="stat-cell"><div class="stat-num">'+Math.round(lb/1000)+'k</div><div class="stat-label">Lbs</div></div>'
      + '<div class="stat-cell"><div class="stat-num">'+r+'</div><div class="stat-label">Routes</div></div>'
      + '<div class="stat-cell"><div class="stat-num">'+Math.round(mi)+'</div><div class="stat-label">Miles</div></div>';
}}

// ── Day strip ───────────────────────────────────────
function renderDayStrip() {{
    var strip = document.getElementById('day-strip');
    strip.innerHTML = '';
    var visibleKeys = viewMode === 'committed' ? committedKeys : allDayKeys;

    visibleKeys.forEach(function(k) {{
        var d = dayMap[k];
        var isTent = d.status === 'TENTATIVE';
        var dateObj = d.date ? new Date(d.date + 'T12:00:00') : null;
        var dateNum = dateObj ? dateObj.getDate() : '';

        var el = document.createElement('div');
        el.className = 'ds-day' + (k === selectedDay ? ' active' : '') + (isTent ? ' tent' : '');
        el.dataset.dayIdx = k;
        el.innerHTML =
            '<div class="ds-weekday">' + d.day + '</div>'
          + '<div class="ds-date">' + dateNum + '</div>'
          + '<div class="ds-stops">' + d.stops + ' stops</div>'
          + (k === selectedDay ? '<div class="ds-badge"></div>' : '');
        el.addEventListener('click', function() {{
            selectedDay = k;
            focusedRoute = null;
            updateAll();
            // Zoom to this day's stops
            var coords = [];
            d.routes.forEach(function(g) {{
                g.waypoints.forEach(function(w) {{
                    if (w.stop !== undefined) coords.push([w.lat, w.lon]);
                }});
            }});
            if (coords.length > 1) map.fitBounds(coords, {{ padding: [60, 60] }});
        }});
        strip.appendChild(el);
    }});
}}

// ── Truck filter ────────────────────────────────────
var truckBar = document.getElementById('truck-filter');
['All', 'Truck2', 'Truck9'].forEach(function(t) {{
    var btn = document.createElement('button');
    btn.className = 'truck-btn active';
    btn.dataset.truck = t;
    if (t === 'All') {{
        btn.innerHTML = 'Both Trucks';
    }} else {{
        btn.innerHTML = '<span class="truck-dot" style="background:'+TRUCK_COLORS[t]+'"></span> '
                      + t.replace('ruck', '');
    }}
    btn.addEventListener('click', function() {{
        if (t === 'All') {{
            activeTrucks = new Set(['Truck2', 'Truck9']);
        }} else if (activeTrucks.size === 2) {{
            activeTrucks = new Set([t]);
        }} else if (activeTrucks.has(t) && activeTrucks.size === 1) {{
            activeTrucks = new Set(['Truck2', 'Truck9']);
        }} else {{
            activeTrucks = new Set([t]);
        }}
        updateAll();
    }});
    truckBar.appendChild(btn);
}});

// ── Route list (grouped by day) ─────────────────────
function renderRouteList() {{
    var list = document.getElementById('route-list');
    list.innerHTML = '';
    var visibleKeys = viewMode === 'committed' ? committedKeys : allDayKeys;

    visibleKeys.forEach(function(k) {{
        var d = dayMap[k];
        var isTent = d.status === 'TENTATIVE';
        var dateObj = d.date ? new Date(d.date + 'T12:00:00') : null;
        var dateLabel = dateObj
            ? dateObj.toLocaleDateString('en-US', {{weekday:'short', month:'short', day:'numeric'}})
            : d.day;

        // Day header
        var hdr = document.createElement('div');
        hdr.className = 'day-header';
        hdr.innerHTML = dateLabel
            + ' <span class="dh-badge ' + (isTent ? 'dh-tentative' : 'dh-committed') + '">'
            + (isTent ? 'Tentative' : 'Dispatch') + '</span>'
            + '<span style="margin-left:auto;font-weight:600;font-size:10px;color:var(--muted);">'
            + d.stops + ' stops &middot; ' + Math.round(d.lbs/1000) + 'k lbs</span>';
        list.appendChild(hdr);

        // Route cards for this day
        d.routes.forEach(function(g) {{
            var card = document.createElement('div');
            card.className = 'route-card' + (isTent ? ' tent' : '');
            card.dataset.id = g.id;
            card.dataset.dayIdx = k;
            card.dataset.truck = g.truck;
            card.style.borderLeftColor = g.color;

            var stops = g.waypoints.filter(function(w){{ return w.stop !== undefined; }});
            var critCount = stops.filter(function(w){{ return w.urgency==='critical'||w.urgency==='stockout'; }}).length;

            var hrs = Math.floor(g.timeMin / 60);
            var mins = g.timeMin % 60;

            card.innerHTML =
                '<div class="rc-row">'
                + '<div class="rc-truck"><span class="truck-dot" style="background:'+g.color+'"></span> '
                + g.truck.replace('ruck', '') + '</div>'
                + '<div class="rc-metrics">'
                + '<span>' + g.stops + ' stops</span>'
                + '<span>' + (g.load/1000).toFixed(1) + 'k lbs</span>'
                + '<span>' + g.capPct + '%</span>'
                + '<span>' + g.dist + ' mi</span>'
                + '<span>' + hrs + 'h' + (mins < 10 ? '0' : '') + mins + '</span>'
                + '</div>'
                + (critCount > 0 ? '<span class="rc-crit">' + critCount + '</span>' : '')
                + '</div>';

            card.addEventListener('click', function() {{
                if (focusedRoute === g.id) {{
                    focusedRoute = null;
                    updateAll();
                }} else {{
                    focusedRoute = g.id;
                    selectedDay = k;
                    updateAll();
                    var coords = stops.map(function(w){{ return [w.lat, w.lon]; }});
                    if (coords.length) map.fitBounds(coords, {{ padding: [60, 60] }});
                }}
            }});
            list.appendChild(card);
        }});
    }});
}}

// ── Master update ───────────────────────────────────
function updateAll() {{
    // View toggle
    vtBar.querySelectorAll('.view-btn').forEach(function(b) {{
        b.classList.toggle('active', b.dataset.view === viewMode);
    }});

    renderStats();
    renderDayStrip();
    renderRouteList();

    // Truck buttons
    document.querySelectorAll('.truck-btn').forEach(function(b) {{
        var t = b.dataset.truck;
        if (t === 'All') b.classList.toggle('active', activeTrucks.size === 2);
        else b.classList.toggle('active', activeTrucks.has(t));
    }});

    // Map layers + card visibility
    ROUTES.forEach(function(g) {{
        var dayVisible = (g.dayIndex === selectedDay);
        var truckVisible = activeTrucks.has(g.truck);
        var modeVisible = viewMode === 'full' || g.status === 'COMMITTED';
        var visible = dayVisible && truckVisible && modeVisible;
        if (focusedRoute && focusedRoute !== g.id) visible = false;

        var lg = layerGroups[g.id];
        if (visible) {{
            if (!map.hasLayer(lg)) map.addLayer(lg);
        }} else {{
            if (map.hasLayer(lg)) map.removeLayer(lg);
        }}

        var card = document.querySelector('[data-id="'+g.id+'"]');
        if (card) {{
            card.classList.toggle('visible', dayVisible && truckVisible);
            card.classList.toggle('focused', focusedRoute === g.id);
        }}
    }});
}}

// ── Init ────────────────────────────────────────────
updateAll();
// Zoom to selected day
var initCoords = [];
(dayMap[selectedDay] || {{routes:[]}}).routes.forEach(function(g) {{
    g.waypoints.forEach(function(w) {{
        if (w.stop !== undefined) initCoords.push([w.lat, w.lon]);
    }});
}});
if (initCoords.length > 1) map.fitBounds(initCoords, {{ padding: [60, 60] }});

</script>
</body>
</html>"""
