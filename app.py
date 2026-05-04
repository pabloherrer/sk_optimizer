#!/usr/bin/env python3
"""
S&K Route Optimizer — Web App
==============================
Run:  python app.py
Open: http://localhost:5050
"""

import json
import os
import subprocess
import sys
import threading
import time
import webbrowser
from datetime import datetime
from pathlib import Path

import pandas as pd

# ── Auto-install Flask if missing ─────────────────────────────────────────────
try:
    from flask import Flask, Response, jsonify, request, send_file
except ImportError:
    print("Installing Flask…")
    try:
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'flask', '--quiet'])
    except subprocess.CalledProcessError:
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'flask',
                               '--quiet', '--break-system-packages'])
    from flask import Flask, Response, jsonify, request, send_file

BASE_DIR     = Path(__file__).parent
OUTPUT_DIR   = BASE_DIR / 'output'
SUMMARY_FILE = OUTPUT_DIR / 'last_run_summary.json'

# Input file: env var → local_config.json → default
_default_input = str(BASE_DIR / 'data' / 'SK_Delivery_System.xlsx')
_local_cfg_path = BASE_DIR / 'local_config.json'
_local_input = None
if _local_cfg_path.exists():
    try:
        import json as _j
        _local_input = _j.loads(_local_cfg_path.read_text(encoding='utf-8')).get('input_file')
    except Exception:
        pass
INPUT_FILE = Path(os.environ.get('SK_INPUT_FILE') or _local_input or _default_input)

# Import thresholds from config — never hardcode these
sys.path.insert(0, str(BASE_DIR))
from config import CRITICAL_DAYS, URGENT_DAYS, TRUCK_NAMES, NUM_TRUCKS, SATURDAY_TRUCKS, DATA_DIR

app = Flask(__name__)

_is_running = False
_run_lock   = threading.Lock()


# ── Version (from git) ────────────────────────────────────────────────────────

def _get_version() -> dict:
    """Read git commit info. Falls back gracefully if git is not available."""
    try:
        short = subprocess.check_output(
            ['git', 'rev-parse', '--short', 'HEAD'],
            cwd=str(BASE_DIR), stderr=subprocess.DEVNULL, text=True
        ).strip()
        date = subprocess.check_output(
            ['git', 'log', '-1', '--format=%cd', '--date=format:%b %d, %Y'],
            cwd=str(BASE_DIR), stderr=subprocess.DEVNULL, text=True
        ).strip()
        msg = subprocess.check_output(
            ['git', 'log', '-1', '--format=%s'],
            cwd=str(BASE_DIR), stderr=subprocess.DEVNULL, text=True
        ).strip()
        return {'hash': short, 'date': date, 'msg': msg[:60]}
    except Exception:
        return {'hash': '—', 'date': '—', 'msg': 'git not available'}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _safe_int(v, default=0):
    try:
        if v is None or (isinstance(v, float) and (v != v)):  # NaN check
            return default
        return int(v)
    except (ValueError, TypeError):
        return default


def _safe_float(v, default=0.0):
    try:
        if v is None or (isinstance(v, float) and (v != v)):
            return default
        return float(v)
    except (ValueError, TypeError):
        return default


def _fmt_dt(ts: float) -> str:
    return datetime.fromtimestamp(ts).strftime('%b %d, %Y  %I:%M %p')


def get_data_info() -> dict:
    """Return metadata about the input file — filesystem only, no numpy/pandas."""
    info = {
        'input_path':     str(INPUT_FILE),
        'input_filename': INPUT_FILE.name,
        'input_modified': None,
        'input_mtime_ts': None,
        'last_delivery':  None,
        'client_count':   None,
        'delivery_count': None,
        'stale':          False,
    }
    if INPUT_FILE.exists():
        mtime = INPUT_FILE.stat().st_mtime
        info['input_modified'] = _fmt_dt(mtime)
        info['input_mtime_ts'] = mtime
        info['client_count']   = '—'
        age_days = (time.time() - mtime) / 86400
        info['stale'] = age_days > 2
    return info


def get_data_info_full() -> dict:
    """Richer info via openpyxl — called only from /data-info endpoint."""
    info = get_data_info()
    try:
        from openpyxl import load_workbook
        wb = load_workbook(INPUT_FILE, read_only=True, data_only=True)

        if 'Client_List' in wb.sheetnames:
            ws = wb['Client_List']
            count = sum(
                1 for row in ws.iter_rows(min_row=4, max_col=1, values_only=True)
                if str(row[0] or '').strip().startswith('C')
                and str(row[0] or '').strip()[1:].isdigit()
            )
            info['client_count'] = count

            # Urgency breakdown from Days_Until_Stockout or similar
            all_rows = list(ws.iter_rows(values_only=True))
            if all_rows:
                headers  = [str(c or '').lower() for c in all_rows[0]]
                days_col = next(
                    (i for i, h in enumerate(headers)
                     if ('days' in h and 'stock' in h)
                     or ('days' in h and 'remain' in h)
                     or h in ('days_left', 'days_until_stockout', 'days_remaining')),
                    None
                )
                if days_col is None:
                    days_col = next(
                        (i for i, h in enumerate(headers) if 'days' in h), None
                    )
                if days_col is not None:
                    urg = {'critical': 0, 'urgent': 0, 'normal': 0}
                    for row in all_rows[1:]:
                        v = row[days_col]
                        if isinstance(v, (int, float)):
                            if v <= CRITICAL_DAYS:
                                urg['critical'] += 1
                            elif v <= URGENT_DAYS:
                                urg['urgent'] += 1
                            else:
                                urg['normal'] += 1
                    info['urgency_live'] = urg

        if 'Delivery_Log' in wb.sheetnames:
            ws   = wb['Delivery_Log']
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                # The Delivery_Log sheet starts with a banner row and an
                # instruction row; real column headers live a few rows down.
                # Scan the first 5 rows for one whose first cell is "Date"
                # (case-insensitive) — that's the header row.
                hdr_row_idx = None
                for ri, r in enumerate(rows[:5]):
                    first = str(r[0] or '').strip().lower()
                    if first == 'date':
                        hdr_row_idx = ri
                        break
                if hdr_row_idx is None:
                    hdr_row_idx = 0  # fallback to legacy behavior

                headers  = [str(c or '').lower() for c in rows[hdr_row_idx]]
                date_col = next((i for i, h in enumerate(headers) if 'date' in h), None)
                if date_col is not None:
                    from datetime import date
                    dates = [r[date_col] for r in rows[hdr_row_idx + 1:]
                             if isinstance(r[date_col], (datetime, date))]
                    if dates:
                        last = max(dates)
                        info['last_delivery']  = last.strftime('%b %d, %Y')
                        info['delivery_count'] = len(dates)
        wb.close()
    except Exception as e:
        info['read_error'] = str(e)
    return info


def get_latest_outputs() -> dict:
    """
    Return the most recent Excel + map outputs.
    Distinguishes IRP (sk_irp_*) from legacy unified (sk_unified_*) so the
    dashboard can show which engine produced what — both files coexist
    because there are two entry points (run_irp.py vs run_unified.py).
    """
    result = {
        'excel': None, 'map': None,
        'irp_excel': None, 'irp_map': None, 'irp_csv': None,
        'unified_excel': None, 'unified_map': None,
    }
    if not OUTPUT_DIR.exists():
        return result

    def latest(pattern: str):
        files = sorted(OUTPUT_DIR.glob(pattern),
                       key=lambda f: f.stat().st_mtime, reverse=True)
        return files[0].name if files else None

    result['irp_excel']     = latest('sk_irp_*schedule.xlsx')
    result['irp_map']       = latest('sk_irp_*map.html')
    result['irp_csv']       = latest('sk_irp_*smartservice.csv')
    result['unified_excel'] = latest('sk_unified_*schedule.xlsx')
    result['unified_map']   = latest('sk_unified_*map.html')

    # Backwards-compat: 'excel' / 'map' = the absolute newest, regardless of source
    all_excel = sorted(OUTPUT_DIR.glob('*.xlsx'),
                       key=lambda f: f.stat().st_mtime, reverse=True)
    all_map = sorted(OUTPUT_DIR.glob('*.html'),
                     key=lambda f: f.stat().st_mtime, reverse=True)
    if all_excel:
        result['excel'] = all_excel[0].name
        # Tag which engine produced the latest
        result['latest_source'] = 'irp' if 'sk_irp' in all_excel[0].name else (
            'unified' if 'sk_unified' in all_excel[0].name else 'other'
        )
    if all_map:
        result['map'] = all_map[0].name
    return result


def get_last_run_summary() -> dict | None:
    """Read saved summary from last optimizer run."""
    if SUMMARY_FILE.exists():
        try:
            return json.loads(SUMMARY_FILE.read_text())
        except Exception:
            pass
    return None


def save_run_summary(excel_file: str | None, map_file: str | None, elapsed_sec: float):
    """
    Parse the output Excel to build a run summary and save it as JSON.
    Falls back gracefully if the file can't be read.
    """
    summary = {
        'timestamp':     datetime.now().isoformat(),
        'elapsed_sec':   int(elapsed_sec),
        'excel_file':    excel_file,
        'map_file':      map_file,
        'total_stops':   0,
        'total_miles':   None,
        'routes_count':  0,
        'deferred_count': 0,
        'urgency':       {'critical': 0, 'urgent': 0, 'normal': 0},
    }

    # Parse output Excel for route stats using the Week_Summary sheet
    if excel_file:
        excel_path = OUTPUT_DIR / excel_file
        if excel_path.exists():
            try:
                from openpyxl import load_workbook
                wb = load_workbook(excel_path, read_only=True, data_only=True)

                # 1_Week_Summary has one row per truck-day route
                if '1_Week_Summary' in wb.sheetnames:
                    ws      = wb['1_Week_Summary']
                    rows    = list(ws.iter_rows(values_only=True))
                    if rows:
                        headers   = [str(c or '').lower() for c in rows[0]]
                        data_rows = [r for r in rows[1:] if any(c is not None for c in r)]

                        stops_col = next((i for i, h in enumerate(headers) if h == 'stops'), None)
                        miles_col = next((i for i, h in enumerate(headers) if h == 'dist_mi'), None)

                        summary['routes_count'] = len(data_rows)
                        if stops_col is not None:
                            summary['total_stops'] = sum(
                                r[stops_col] for r in data_rows
                                if isinstance(r[stops_col], (int, float))
                            )
                        if miles_col is not None:
                            summary['total_miles'] = round(sum(
                                r[miles_col] for r in data_rows
                                if isinstance(r[miles_col], (int, float))
                            ), 1)

                # Deferred sheet has one row per unserved client
                if 'Deferred' in wb.sheetnames:
                    ws    = wb['Deferred']
                    drows = list(ws.iter_rows(values_only=True))
                    # subtract 1 for header row
                    summary['deferred_count'] = max(0, len(drows) - 1)

                wb.close()
            except Exception:
                pass

    # Urgency snapshot from input file
    try:
        from openpyxl import load_workbook
        wb = load_workbook(INPUT_FILE, read_only=True, data_only=True)
        if 'Client_List' in wb.sheetnames:
            ws       = wb['Client_List']
            all_rows = list(ws.iter_rows(values_only=True))
            if all_rows:
                headers  = [str(c or '').lower() for c in all_rows[0]]
                days_col = next(
                    (i for i, h in enumerate(headers)
                     if ('days' in h and 'stock' in h)
                     or ('days' in h and 'remain' in h)
                     or h in ('days_left', 'days_until_stockout', 'days_remaining')),
                    None
                )
                if days_col is None:
                    days_col = next(
                        (i for i, h in enumerate(headers) if 'days' in h), None
                    )
                if days_col is not None:
                    for row in all_rows[1:]:
                        v = row[days_col]
                        if isinstance(v, (int, float)):
                            if v <= CRITICAL_DAYS:
                                summary['urgency']['critical'] += 1
                            elif v <= URGENT_DAYS:
                                summary['urgency']['urgent'] += 1
                            else:
                                summary['urgency']['normal'] += 1
        wb.close()
    except Exception:
        pass

    OUTPUT_DIR.mkdir(exist_ok=True)
    SUMMARY_FILE.write_text(json.dumps(summary, indent=2))


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    info    = get_data_info()
    outputs = get_latest_outputs()
    version = _get_version()
    summary = get_last_run_summary()
    return (HTML_TEMPLATE
            .replace('__DATA_INFO__', json.dumps(info))
            .replace('__OUTPUTS__',   json.dumps(outputs))
            .replace('__VERSION__',   json.dumps(version))
            .replace('__SUMMARY__',   json.dumps(summary))
            .replace('__TRUCK_NAMES__', json.dumps(TRUCK_NAMES))
            .replace('__SATURDAY_TRUCKS__', json.dumps(SATURDAY_TRUCKS)))


@app.route('/data-info')
def data_info_route():
    return jsonify(get_data_info_full())


@app.route('/last-run')
def last_run_route():
    return jsonify(get_last_run_summary() or {})


@app.route('/run', methods=['POST'])
def run_optimizer():
    global _is_running

    with _run_lock:
        if _is_running:
            return jsonify({'error': 'Already running'}), 409
        _is_running = True

    data      = request.json or {}
    solve_sec = int(data.get('solve_sec', 600))
    skip_ids  = data.get('skip_ids', [])
    must_ids  = data.get('must_visit_ids', [])
    active_trucks = data.get('active_trucks', TRUCK_NAMES)
    start_time = time.time()

    def generate():
        global _is_running
        try:
            env = {**os.environ, 'PYTHONUNBUFFERED': '1'}
            if skip_ids:
                env['SK_SKIP_IDS'] = ','.join(str(x) for x in skip_ids)
            if must_ids:
                env['SK_MUST_VISIT_IDS'] = ','.join(str(x) for x in must_ids)
            if len(active_trucks) < NUM_TRUCKS:
                env['SK_ACTIVE_TRUCKS'] = ','.join(active_trucks)
            # The dashboard button now runs the new IRP engine by default.
            # State persistence, time windows, closures, multi-visit,
            # SmartService CSV export, planning view all hang off run_irp.py.
            # (The legacy run_unified.py still exists as a CLI fallback —
            # call it directly from the terminal if you ever need it.)
            cmd = [
                sys.executable, '-u',
                str(BASE_DIR / 'run_irp.py'),
                '--solve-sec', str(solve_sec),
                '--input-file', str(INPUT_FILE),
            ]
            proc = subprocess.Popen(
                cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                bufsize=1,
                cwd=str(BASE_DIR),
                env=env,
            )
            for line in proc.stdout:
                yield f"data: {json.dumps({'line': line.rstrip()})}\n\n"
            proc.wait()

            outputs      = get_latest_outputs()
            elapsed_sec  = time.time() - start_time
            success      = proc.returncode == 0

            if success:
                save_run_summary(outputs['excel'], outputs['map'], elapsed_sec)

            result = {
                'done':        True,
                'success':     success,
                'excel':       outputs['excel'],
                'map':         outputs['map'],
                'elapsed_sec': int(elapsed_sec),
            }
            yield f"data: {json.dumps(result)}\n\n"
        finally:
            _is_running = False

    return Response(
        generate(),
        mimetype='text/event-stream',
        headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'},
    )


@app.route('/download/<path:filename>')
def download(filename):
    filepath = (OUTPUT_DIR / filename).resolve()
    if not str(filepath).startswith(str(OUTPUT_DIR.resolve())):
        return 'Forbidden', 403
    if not filepath.exists():
        return 'File not found', 404
    return send_file(filepath, as_attachment=True)


@app.route('/view/<path:filename>')
def view_map(filename):
    filepath = (OUTPUT_DIR / filename).resolve()
    if not str(filepath).startswith(str(OUTPUT_DIR.resolve())):
        return 'Forbidden', 403
    if not filepath.exists():
        return 'File not found', 404
    return send_file(filepath)


@app.route('/status')
def status():
    outputs = get_latest_outputs()
    return jsonify({'running': _is_running, **outputs})


@app.route('/settings', methods=['GET', 'POST'])
def settings_route():
    """Get or save per-machine settings (persists in local_config.json)."""
    from config import LOCAL_CONFIG_FILE, save_local_config
    if request.method == 'GET':
        cfg = {}
        if LOCAL_CONFIG_FILE.exists():
            try:
                cfg = json.loads(LOCAL_CONFIG_FILE.read_text(encoding='utf-8'))
            except Exception:
                pass
        cfg.setdefault('input_file', str(INPUT_FILE))
        return jsonify(cfg)
    else:
        data = request.json or {}
        new_path = data.get('input_file', '').strip()
        if not new_path:
            return jsonify({'error': 'No path provided'}), 400
        p = Path(new_path)
        if not p.exists():
            return jsonify({'error': f'File not found: {new_path}'}), 400
        if not str(p).lower().endswith('.xlsx'):
            return jsonify({'error': 'File must be an .xlsx file'}), 400
        # Save to local_config.json
        cfg = {}
        if LOCAL_CONFIG_FILE.exists():
            try:
                cfg = json.loads(LOCAL_CONFIG_FILE.read_text(encoding='utf-8'))
            except Exception:
                pass
        cfg['input_file'] = str(p)
        save_local_config(cfg)
        return jsonify({'ok': True, 'input_file': str(p),
                        'note': 'Restart the app for the new path to take effect.'})


@app.route('/planning')
def planning_view():
    """Standalone planning dashboard powered by /api/plan-summary,
    /api/anova-status and /api/fill-audit."""
    return PLANNING_HTML


@app.route('/api/anova-status')
def anova_status_route():
    """
    Latest ANOVA tank readings + which SK clients they map to. Used by
    the planning view to show live inventory.
    """
    try:
        from irp_core.anova_integration import (
            load_all_readings, map_asset_to_client,
        )
        from load_data import load_all
        clients_raw, _ = load_all(INPUT_FILE)

        # Look in the standard locations
        anova_csv = Path(__file__).parent.parent / 'anova_data' / 'readings.csv'
        anova_xlsx = Path(__file__).parent.parent / 'anova_live_readings.xlsx'

        readings = load_all_readings(
            csv_path=anova_csv if anova_csv.exists() else None,
            excel_path=anova_xlsx if anova_xlsx.exists() else None,
            fresh_hours=24.0,
        )
        if not readings:
            return jsonify({
                'enabled': False,
                'message': 'No ANOVA readings found. Run anova_pull.py or set up the webhook.',
                'readings': [],
            })

        asset_to_cid = map_asset_to_client(readings, clients_raw)
        cust_lookup = dict(zip(clients_raw['ID'].astype(str), clients_raw['Customer']))
        tank_lookup = dict(zip(clients_raw['ID'].astype(str), clients_raw['Tank_lbs']))

        out = []
        for asset_id, r in readings.items():
            cid = asset_to_cid.get(asset_id)
            tank = tank_lookup.get(cid, 0)
            out.append({
                'asset_id': asset_id,
                'client_id': cid,
                'customer': cust_lookup.get(cid, '<unmapped>'),
                'level_lbs': round(r.level_lbs, 1),
                'tank_lbs': int(tank) if tank else None,
                'pct_full': round(r.level_lbs / tank * 100, 1) if tank else None,
                'product': r.product,
                'timestamp': str(r.timestamp),
                'age_hours': round(r.age_hours, 1),
                'fresh': r.fresh,
            })
        out.sort(key=lambda x: (x['fresh'] is False, x['age_hours']))
        return jsonify({
            'enabled': True,
            'count': len(out),
            'fresh_count': sum(1 for x in out if x['fresh']),
            'readings': out,
        })
    except Exception as e:
        return jsonify({'enabled': False, 'error': str(e)}), 500


@app.route('/api/plan-summary')
def plan_summary_route():
    """
    Read the saved plan.json (output of run_irp.py) and return a
    structured day-by-day summary for the planning view. Includes
    address/zone data so the UI can render "Pending" badges, city,
    and truck-grouped views like SK's existing planning tool.
    """
    try:
        from irp_core.state_manager import load_plan
        plan_path = DATA_DIR / 'plan.json'
        plan = load_plan(plan_path)
        if not plan:
            return jsonify({
                'available': False,
                'message': 'No plan.json found. Run the IRP optimizer (python run_irp.py) to generate one.',
            })

        from load_data import load_all
        clients_raw, _ = load_all(INPUT_FILE)
        cust_lookup  = dict(zip(clients_raw['ID'].astype(str), clients_raw['Customer']))
        zone_lookup  = dict(zip(clients_raw['ID'].astype(str), clients_raw.get('Zone', clients_raw['Customer'])))
        addr_lookup  = dict(zip(clients_raw['ID'].astype(str), clients_raw.get('Address', clients_raw['Customer'])))
        tank_lookup  = dict(zip(clients_raw['ID'].astype(str), clients_raw['Tank_lbs']))

        # Group visits by day
        days: dict = {}
        for v in plan.get('visits', []):
            d = int(v.get('day', 0))
            cid = str(v.get('client_id', ''))
            refill = float(v.get('refill_lbs', 0) or 0)
            tank = float(tank_lookup.get(cid, 0) or 0)
            days.setdefault(d, []).append({
                'client_id': cid,
                'customer': cust_lookup.get(cid, cid),
                'zone': zone_lookup.get(cid, ''),
                'address': addr_lookup.get(cid, ''),
                'truck': v.get('truck'),
                'stop': v.get('stop'),
                'refill_lbs': refill,
                'tank_lbs': tank,
                'fill_pct': round(refill / tank * 100, 0) if tank > 0 else 0,
                'status': v.get('status'),
                'arrival_hhmm': v.get('arrival_hhmm', ''),
                'depart_hhmm':  v.get('depart_hhmm', ''),
                'service_min':  v.get('service_min', 0),
                'travel_min':   v.get('travel_min', 0),
            })
        # Sort each day by stop order
        for d in days:
            days[d].sort(key=lambda x: int(x.get('stop') or 0))

        return jsonify({
            'available': True,
            'solved_at': plan.get('solved_at'),
            'today': plan.get('today'),
            'horizon_days': plan.get('horizon_days'),
            'commit_days': plan.get('commit_days'),
            'plan_dates': plan.get('plan_dates', []),
            'metadata': plan.get('metadata', {}),
            'days': [
                {
                    'day_index': d,
                    'date': plan.get('plan_dates', [None]*100)[d] if d < len(plan.get('plan_dates', [])) else None,
                    'visits': days[d],
                    'n_visits': len(days[d]),
                    'total_lbs': sum(v.get('refill_lbs', 0) for v in days[d]),
                }
                for d in sorted(days.keys())
            ],
            'total_visits': sum(len(v) for v in days.values()),
            'deferred_ids': plan.get('deferred_ids', []),
        })
    except Exception as e:
        return jsonify({'available': False, 'error': str(e)}), 500


@app.route('/api/pending-visits')
def pending_visits_route():
    """
    List all "pending" visits in the Delivery_Log — rows where the qty
    is the 200-lb placeholder. These are visits the driver made (or
    is making) where the actual quantity hasn't been invoiced yet.
    """
    try:
        from load_data import load_all
        clients_raw, deliveries = load_all(INPUT_FILE)
        cust_to_id = dict(zip(clients_raw['Customer'], clients_raw['ID'].astype(str)))
        zone_lookup = dict(zip(clients_raw['ID'].astype(str), clients_raw.get('Zone', '')))
        addr_lookup = dict(zip(clients_raw['ID'].astype(str), clients_raw.get('Address', '')))
        tank_lookup = dict(zip(clients_raw['ID'].astype(str), clients_raw['Tank_lbs']))
        prod_lookup = dict(zip(clients_raw['ID'].astype(str), clients_raw.get('Product', '')))

        pending = deliveries[deliveries['Qty_lbs'] == 200].copy()
        rows = []
        for _, r in pending.iterrows():
            cid = cust_to_id.get(r['Customer'])
            if not cid:
                continue
            rows.append({
                'date': str(r['Date'].date()),
                'client_id': cid,
                'customer': r['Customer'],
                'zone': zone_lookup.get(cid, ''),
                'address': addr_lookup.get(cid, ''),
                'product': prod_lookup.get(cid, ''),
                'tank_lbs': int(tank_lookup.get(cid, 0) or 0),
            })
        rows.sort(key=lambda x: (x['date'], x['client_id']))
        # Group by date
        by_date: dict = {}
        for row in rows:
            by_date.setdefault(row['date'], []).append(row)
        return jsonify({
            'count': len(rows),
            'days': [
                {'date': date, 'visits': v, 'n_visits': len(v)}
                for date, v in sorted(by_date.items())
            ],
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/fill-audit')
def fill_audit_route():
    """
    Historical fill-rate audit per client. Useful for ops to identify
    "we keep visiting this client at 30% empty — is something wrong?"
    """
    try:
        from irp_core.diagnostics import audit_historical_fill_rates
        from load_data import load_all
        clients_raw, deliveries = load_all(INPUT_FILE)
        df = audit_historical_fill_rates(deliveries, clients_raw, horizon_days=180)
        return jsonify({
            'rows': df.to_dict(orient='records'),
            'global_mean_fill_pct': float(df['mean_fill_pct'].mean()) if not df.empty else 0.0,
            'low_fill_clients': int((df['mean_fill_pct'] < 0.50).sum()),
            'horizon_days': 180,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/snapshot')
def snapshot_route():
    """Return pre-solve client estimates: who needs oil, urgency, DTE, refill lbs."""
    try:
        from load_data import load_all
        from forecast_consumption import estimate_consumption_rates
        from inventory import enrich_snapshot
        from state import load_state, initialise_state_from_snapshot
        from config import STATE_FILE as SF, EXCLUDED_CLIENT_IDS

        clients_raw, deliveries = load_all(INPUT_FILE)
        today = pd.Timestamp.today().normalize()
        clients_df = estimate_consumption_rates(deliveries, clients_raw, today=today)
        state = load_state(SF)
        if not state:
            state = initialise_state_from_snapshot(clients_df)
        snapshot = enrich_snapshot(clients_df, state)

        excluded = set(str(x) for x in EXCLUDED_CLIENT_IDS)
        rows = []
        for _, r in snapshot.iterrows():
            cid = str(r.get('ID', ''))
            if cid in excluded:
                continue
            dte = _safe_float(r.get('Days_Until_Stockout'), 999)
            dte = round(dte, 1)
            rows.append({
                'id':       cid,
                'name':     str(r.get('Customer', '')),
                'dte':      dte,
                'urgency':  str(r.get('Urgency', 'normal')),
                'refill':   _safe_int(r.get('Refill_Today_lbs')),
                'tank':     _safe_int(r.get('Tank_lbs')),
                'current':  _safe_int(r.get('Current_lbs')),
                'rate':     round(_safe_float(r.get('Avg_LbsPerDay')), 1),
                'fill_pct': round(_safe_float(r.get('Fill_Pct_Today')) * 100),
            })
        rows.sort(key=lambda x: x['dte'])
        return jsonify({'clients': rows, 'today': str(today.date())})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/clients-all')
def clients_all_route():
    """Return every client in Client_List (name + numeric ID) for the Log Delivery search."""
    try:
        from openpyxl import load_workbook as _lw
        _wb = _lw(str(INPUT_FILE), data_only=True)
        _ws = _wb['Client_List']
        clients = []
        for r in range(4, _ws.max_row + 1):
            cid  = _ws.cell(r, 1).value
            name = _ws.cell(r, 2).value
            prod = _ws.cell(r, 11).value or ''
            if cid is not None and name:
                clients.append({
                    'id':   str(cid),
                    'name': str(name),
                    'prod': str(prod),
                })
        clients.sort(key=lambda x: x['name'])
        return jsonify({'clients': clients})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/log-delivery', methods=['POST'])
def log_delivery_route():
    """Append a delivery to the Delivery_Log sheet."""
    try:
        data = request.get_json(force=True)
        del_date   = data.get('date')       # 'YYYY-MM-DD'
        customer   = data.get('customer')   # numeric client ID (e.g. "1018")
        qty        = data.get('qty')        # integer lbs

        if not del_date or not customer or qty is None:
            return jsonify({'ok': False, 'error': 'Missing date, customer or qty'}), 400

        from openpyxl import load_workbook as _lw
        from datetime import date as _date

        _wb = _lw(str(INPUT_FILE))
        _ws = _wb['Delivery_Log']

        # Find last populated row (scan from bottom for last non-None date),
        # then write to the row immediately after it.
        _last = 3
        for _r in range(4, _ws.max_row + 1):
            if _ws.cell(_r, 1).value is not None:
                _last = _r
        _row = _last + 1

        _dt = _date.fromisoformat(del_date)
        _ws.cell(_row, 1).value = _dt
        _ws.cell(_row, 2).value = int(customer) if str(customer).isdigit() else customer  # B: numeric ID
        _ws.cell(_row, 4).value = int(qty)       # D: Qty Delivered (new column structure)

        _wb.save(str(INPUT_FILE))
        return jsonify({'ok': True, 'row': _row, 'customer': customer, 'qty': int(qty)})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


# ── HTML Template ─────────────────────────────────────────────────────────────

HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>S&amp;K Route Optimizer</title>
<style>
  *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

  body {
    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
    background: #f0f4f8;
    color: #2c3e50;
    min-height: 100vh;
  }

  /* ── Header ── */
  .header {
    background: linear-gradient(135deg, #1a6faf 0%, #155f9a 100%);
    color: white;
    padding: 18px 36px;
    display: flex;
    align-items: center;
    gap: 14px;
    box-shadow: 0 2px 12px rgba(26,111,175,0.25);
  }
  .header-icon { font-size: 30px; line-height: 1; }
  .header-text h1 { font-size: 19px; font-weight: 700; letter-spacing: -0.3px; }
  .header-text p  { font-size: 12.5px; opacity: 0.75; margin-top: 2px; }
  .header-date {
    margin-left: auto;
    font-size: 12.5px;
    opacity: 0.8;
    text-align: right;
    line-height: 1.5;
  }
  .version-badge {
    display: inline-flex;
    align-items: center;
    gap: 5px;
    background: rgba(255,255,255,0.12);
    border: 1px solid rgba(255,255,255,0.2);
    border-radius: 20px;
    padding: 3px 10px;
    font-size: 11px;
    font-weight: 600;
    color: rgba(255,255,255,0.85);
    margin-top: 5px;
    cursor: default;
  }
  .version-dot {
    width: 6px; height: 6px;
    border-radius: 50%;
    background: #4ade80;
  }

  /* ── Layout ── */
  .main {
    max-width: 720px;
    margin: 30px auto;
    padding: 0 20px 60px;
  }

  /* ── Cards ── */
  .card {
    background: white;
    border-radius: 14px;
    padding: 26px 28px;
    box-shadow: 0 1px 5px rgba(0,0,0,0.07), 0 4px 16px rgba(0,0,0,0.04);
    margin-bottom: 18px;
  }
  .card-title {
    font-size: 13px;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.6px;
    color: #1a6faf;
    margin-bottom: 18px;
    display: flex;
    align-items: center;
    gap: 7px;
  }

  /* ── Data Status ── */
  .info-grid {
    display: grid;
    grid-template-columns: repeat(3, 1fr);
    gap: 14px;
  }
  .info-item {
    background: #f7fafc;
    border: 1px solid #e8edf2;
    border-radius: 10px;
    padding: 14px 16px;
  }
  .info-item.stale {
    border-color: #f59e0b;
    background: #fffbeb;
  }
  .stale-banner {
    display: flex;
    align-items: center;
    gap: 7px;
    background: #fef3c7;
    border: 1px solid #f59e0b;
    border-radius: 8px;
    padding: 9px 14px;
    font-size: 12.5px;
    font-weight: 600;
    color: #92400e;
    margin-bottom: 14px;
  }
  .info-item-label {
    font-size: 11px;
    font-weight: 600;
    color: #94a3b8;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    margin-bottom: 5px;
  }
  .info-item-value {
    font-size: 13.5px;
    font-weight: 600;
    color: #2c3e50;
    line-height: 1.35;
  }
  .info-item-sub {
    font-size: 11.5px;
    color: #94a3b8;
    margin-top: 3px;
  }
  .badge {
    display: inline-block;
    padding: 2px 8px;
    border-radius: 20px;
    font-size: 11px;
    font-weight: 600;
  }
  .badge-green  { background: #d1fae5; color: #065f46; }
  .badge-yellow { background: #fef3c7; color: #92400e; }
  .badge-red    { background: #fee2e2; color: #991b1b; }
  .badge-gray   { background: #f1f5f9; color: #64748b; }

  /* ── Plan card ── */
  .plan-window {
    display: flex;
    align-items: center;
    gap: 10px;
    background: #eff6ff;
    border: 1px solid #bfdbfe;
    border-radius: 10px;
    padding: 13px 16px;
    margin-bottom: 20px;
  }
  .plan-window-icon { font-size: 18px; }
  .plan-window-text {
    font-size: 14px;
    font-weight: 700;
    color: #1e40af;
  }
  .plan-window-sub {
    font-size: 12px;
    color: #3b82f6;
    margin-top: 2px;
  }

  /* Advanced collapsible */
  .advanced-details {
    margin-bottom: 20px;
  }
  .advanced-summary {
    font-size: 12px;
    font-weight: 600;
    color: #94a3b8;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    cursor: pointer;
    user-select: none;
    list-style: none;
    display: flex;
    align-items: center;
    gap: 6px;
    padding: 6px 0;
  }
  .advanced-summary::-webkit-details-marker { display: none; }
  .advanced-summary::before {
    content: '▸';
    font-size: 10px;
    transition: transform 0.15s;
  }
  details[open] .advanced-summary::before { transform: rotate(90deg); }
  .advanced-body { padding-top: 14px; }

  /* Speed tiles */
  .speed-tiles { display: flex; gap: 10px; }
  .speed-tile {
    flex: 1;
    padding: 13px 10px;
    border: 2px solid #e2e8f0;
    border-radius: 10px;
    cursor: pointer;
    text-align: center;
    transition: all 0.15s ease;
    user-select: none;
  }
  .speed-tile:hover  { border-color: #93c5fd; }
  .speed-tile.active { border-color: #1a6faf; background: #eff6ff; }
  .speed-tile-name   { font-size: 13.5px; font-weight: 700; color: #2c3e50; }
  .speed-tile-time   { font-size: 12px; color: #94a3b8; margin-top: 3px; }

  /* ── Run button ── */
  .btn-run {
    width: 100%;
    padding: 15px;
    background: linear-gradient(135deg, #e67e22 0%, #d35400 100%);
    color: white;
    border: none;
    border-radius: 10px;
    font-size: 15.5px;
    font-weight: 700;
    cursor: pointer;
    transition: opacity 0.2s, transform 0.1s;
    display: flex;
    align-items: center;
    justify-content: center;
    gap: 10px;
    letter-spacing: 0.2px;
    box-shadow: 0 4px 14px rgba(230,126,34,0.3);
  }
  .btn-run:hover:not(:disabled)  { opacity: 0.93; transform: translateY(-1px); }
  .btn-run:active:not(:disabled) { transform: translateY(0); }
  .btn-run:disabled { background: #cbd5e1; box-shadow: none; cursor: not-allowed; }

  /* ── Log ── */
  .status-row {
    display: flex;
    align-items: center;
    gap: 9px;
    margin-bottom: 10px;
  }
  .dot {
    width: 9px; height: 9px;
    border-radius: 50%;
    background: #cbd5e1;
    flex-shrink: 0;
  }
  .dot.running { background: #e67e22; animation: blink 1.1s ease infinite; }
  .dot.ok      { background: #22c55e; }
  .dot.fail    { background: #ef4444; }
  @keyframes blink { 0%,100%{opacity:1} 50%{opacity:0.35} }
  .status-label { font-size: 13px; font-weight: 600; color: #475569; flex: 1; }
  .status-timer { font-size: 12.5px; font-weight: 600; color: #94a3b8; font-variant-numeric: tabular-nums; }

  .progress-wrap {
    background: #f1f5f9;
    border-radius: 99px;
    height: 6px;
    margin-bottom: 14px;
    overflow: hidden;
  }
  .progress-bar {
    height: 100%;
    border-radius: 99px;
    background: linear-gradient(90deg, #1a6faf, #e67e22);
    width: 0%;
    transition: width 1s linear;
  }
  .progress-bar.done { background: #22c55e; }

  .log-box {
    background: #f8fafc;
    border: 1px solid #e2e8f0;
    border-radius: 9px;
    padding: 14px 16px;
    font-family: 'SF Mono', 'Fira Code', 'Menlo', monospace;
    font-size: 12px;
    line-height: 1.75;
    height: 260px;
    overflow-y: auto;
    color: #334155;
  }
  .ll.ok   { color: #16a34a; font-weight: 700; }
  .ll.err  { color: #dc2626; }
  .ll.head { color: #1a6faf; font-weight: 600; }
  .ll.warn { color: #d97706; }

  /* ── Results ── */
  .dl-row { display: flex; gap: 12px; }
  .dl-btn {
    flex: 1;
    padding: 14px 18px;
    border-radius: 10px;
    font-size: 14px;
    font-weight: 700;
    cursor: pointer;
    text-decoration: none;
    text-align: center;
    display: flex;
    align-items: center;
    justify-content: center;
    gap: 8px;
    transition: all 0.15s;
  }
  .dl-excel {
    background: linear-gradient(135deg, #16a34a, #15803d);
    color: white;
    border: none;
    box-shadow: 0 3px 10px rgba(22,163,74,0.25);
  }
  .dl-excel:hover { opacity: 0.9; }
  .dl-map {
    background: white;
    color: #1a6faf;
    border: 2px solid #1a6faf;
  }
  .dl-map:hover { background: #eff6ff; }

  /* ── Last Run card ── */
  .last-run-meta {
    font-size: 12px;
    color: #94a3b8;
    margin-bottom: 14px;
  }
  .stat-chips {
    display: flex;
    flex-wrap: wrap;
    gap: 10px;
    margin-bottom: 16px;
  }
  .stat-chip {
    background: #f7fafc;
    border: 1px solid #e8edf2;
    border-radius: 10px;
    padding: 10px 14px;
    text-align: center;
    min-width: 80px;
  }
  .stat-chip-val {
    font-size: 20px;
    font-weight: 800;
    color: #1a6faf;
    line-height: 1;
  }
  .stat-chip-label {
    font-size: 11px;
    color: #94a3b8;
    margin-top: 4px;
    font-weight: 600;
  }

  .urgency-row {
    display: flex;
    gap: 10px;
    margin-bottom: 14px;
    flex-wrap: wrap;
  }
  .urgency-pill {
    display: flex;
    align-items: center;
    gap: 6px;
    padding: 7px 13px;
    border-radius: 20px;
    font-size: 13px;
    font-weight: 700;
  }
  .urgency-pill.critical { background: #fee2e2; color: #991b1b; }
  .urgency-pill.urgent   { background: #fef3c7; color: #92400e; }
  .urgency-pill.normal   { background: #d1fae5; color: #065f46; }
  .urgency-dot {
    width: 8px; height: 8px;
    border-radius: 50%;
  }
  .urgency-pill.critical .urgency-dot { background: #dc2626; }
  .urgency-pill.urgent   .urgency-dot { background: #f59e0b; }
  .urgency-pill.normal   .urgency-dot { background: #22c55e; }

  .last-run-links {
    display: flex;
    gap: 10px;
    margin-top: 14px;
    padding-top: 14px;
    border-top: 1px solid #f1f5f9;
    flex-wrap: wrap;
  }
  .prev-link {
    color: #1a6faf;
    text-decoration: none;
    font-weight: 600;
    font-size: 13px;
  }
  .prev-link:hover { text-decoration: underline; }
  .divider { color: #cbd5e1; font-size: 13px; }

  .hidden { display: none !important; }

  /* ── Fleet Control ── */
  .fleet-grid { display: flex; gap: 10px; }
  .truck-tile {
    flex: 1;
    padding: 14px;
    border: 2px solid #22c55e;
    border-radius: 10px;
    cursor: pointer;
    transition: all 0.15s;
    user-select: none;
  }
  .truck-tile.off {
    border-color: #e2e8f0;
    background: #f8fafc;
    opacity: 0.6;
  }
  .truck-tile.backup {
    border: 2px dashed #94a3b8;
    opacity: 0.5;
  }
  .truck-tile.backup.on {
    border: 2px solid #f59e0b;
    opacity: 1;
    background: #fffbeb;
  }
  .truck-tile-name { font-size: 14px; font-weight: 700; }
  .truck-tile-info { font-size: 12px; color: #64748b; margin-top: 3px; }
  .truck-dot {
    width: 10px; height: 10px; border-radius: 50%;
    display: inline-block; margin-right: 6px; vertical-align: middle;
  }
  .truck-dot.on  { background: #22c55e; }
  .truck-dot.off { background: #cbd5e1; }

  /* ── Client Review ── */
  .client-review-header {
    display: flex; align-items: center; gap: 10px;
    margin-bottom: 14px;
  }
  .urgency-counts {
    display: flex; gap: 6px; flex-wrap: wrap;
    margin-bottom: 14px;
  }
  .urg-chip {
    padding: 4px 11px;
    border-radius: 20px;
    font-size: 12px;
    font-weight: 600;
  }
  .urg-chip.stockout  { background: #fee2e2; color: #991b1b; }
  .urg-chip.critical  { background: #ffedd5; color: #9a3412; }
  .urg-chip.urgent    { background: #fef3c7; color: #92400e; }
  .urg-chip.normal    { background: #d1fae5; color: #065f46; }

  .client-list {
    max-height: 420px;
    overflow-y: auto;
    border: 1px solid #e8edf2;
    border-radius: 10px;
    margin-bottom: 14px;
  }
  .client-row {
    display: flex;
    align-items: center;
    gap: 8px;
    padding: 9px 12px;
    border-bottom: 1px solid #f1f5f9;
    font-size: 13px;
    transition: background 0.1s;
  }
  .client-row:last-child { border-bottom: none; }
  .client-row:hover { background: #f8fafc; }
  .client-row.stockout  { background: #fef2f2; }
  .client-row.critical  { background: #fff7ed; }
  .client-row.skipped {
    background: #f8fafc;
    opacity: 0.7;
  }
  .client-row.skipped .client-name {
    text-decoration: line-through;
    color: #94a3b8;
  }
  .client-row.must-visit {
    background: #eff6ff;
    border-left: 3px solid #1a6faf;
  }
  .urg-dot {
    width: 8px; height: 8px;
    border-radius: 50%;
    flex-shrink: 0;
  }
  .urg-dot.stockout { background: #dc2626; }
  .urg-dot.critical { background: #ea580c; }
  .urg-dot.urgent   { background: #d97706; }
  .urg-dot.normal   { background: #22c55e; }

  .client-name { flex: 1; font-weight: 500; }
  .client-meta { font-size: 11.5px; color: #64748b; min-width: 100px; text-align: right; }
  .btn-skip, .btn-must, .btn-undo {
    font-size: 11px;
    padding: 3px 10px;
    border-radius: 14px;
    cursor: pointer;
    border: 1px solid;
    font-weight: 600;
    transition: all 0.1s;
    white-space: nowrap;
  }
  .btn-skip {
    background: #f1f5f9;
    border-color: #e2e8f0;
    color: #64748b;
  }
  .btn-skip:hover { background: #e2e8f0; }
  .btn-must {
    background: #eff6ff;
    border-color: #bfdbfe;
    color: #1d4ed8;
  }
  .btn-must:hover { background: #dbeafe; }
  .btn-undo {
    background: white;
    border-color: #e2e8f0;
    color: #64748b;
  }
  .btn-undo:hover { background: #f8fafc; }

  .review-note {
    font-size: 12px;
    color: #64748b;
    padding: 10px 14px;
    background: #f0f9ff;
    border: 1px solid #bae6fd;
    border-radius: 8px;
    line-height: 1.5;
    margin-bottom: 14px;
  }
  .review-note b { color: #0369a1; }

  .add-client-bar {
    display: flex;
    gap: 8px;
    margin-bottom: 14px;
  }
  .add-client-bar input {
    flex: 1;
    padding: 9px 14px;
    border: 1px solid #e2e8f0;
    border-radius: 8px;
    font-size: 13px;
    outline: none;
  }
  .add-client-bar input:focus { border-color: #93c5fd; box-shadow: 0 0 0 3px rgba(59,130,246,0.1); }
  .add-client-bar button {
    padding: 9px 16px;
    background: #1a6faf;
    color: white;
    border: none;
    border-radius: 8px;
    font-size: 13px;
    font-weight: 600;
    cursor: pointer;
  }
  .add-client-bar button:hover { background: #155f9a; }

  .override-summary {
    font-size: 12.5px;
    padding: 10px 14px;
    background: #fef3c7;
    border: 1px solid #fbbf24;
    border-radius: 8px;
    color: #92400e;
    font-weight: 600;
    margin-bottom: 14px;
  }
  .section-divider {
    font-size: 11px;
    color: #94a3b8;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    padding: 6px 12px;
    background: #f8fafc;
    border-bottom: 1px solid #f1f5f9;
    font-weight: 600;
  }

  .loading-spinner {
    display: inline-block;
    width: 18px; height: 18px;
    border: 2px solid #e2e8f0;
    border-top-color: #1a6faf;
    border-radius: 50%;
    animation: spin 0.7s linear infinite;
  }
  @keyframes spin { to { transform: rotate(360deg); } }

  /* ── Settings Modal ── */
  .settings-btn {
    background: none;
    border: none;
    cursor: pointer;
    font-size: 15px;
    color: #94a3b8;
    padding: 2px 6px;
    border-radius: 6px;
    transition: all 0.15s;
    margin-left: auto;
  }
  .settings-btn:hover { background: #f1f5f9; color: #1a6faf; }
  .modal-overlay {
    position: fixed;
    inset: 0;
    background: rgba(0,0,0,0.35);
    display: flex;
    align-items: center;
    justify-content: center;
    z-index: 1000;
    opacity: 0;
    pointer-events: none;
    transition: opacity 0.2s;
  }
  .modal-overlay.open { opacity: 1; pointer-events: auto; }
  .modal-box {
    background: #fff;
    border-radius: 14px;
    box-shadow: 0 20px 60px rgba(0,0,0,0.18);
    width: 520px;
    max-width: 92vw;
    padding: 28px 30px 24px;
  }
  .modal-title {
    font-size: 15px;
    font-weight: 700;
    color: #1e293b;
    margin-bottom: 6px;
  }
  .modal-sub {
    font-size: 12px;
    color: #94a3b8;
    margin-bottom: 18px;
    line-height: 1.5;
  }
  .modal-label {
    font-size: 11px;
    font-weight: 600;
    color: #64748b;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    margin-bottom: 6px;
  }
  .modal-input {
    width: 100%;
    padding: 10px 12px;
    border: 1.5px solid #e2e8f0;
    border-radius: 8px;
    font-size: 13px;
    font-family: 'SF Mono', 'Consolas', monospace;
    color: #334155;
    box-sizing: border-box;
    transition: border-color 0.15s;
  }
  .modal-input:focus { outline: none; border-color: #1a6faf; }
  .modal-actions {
    display: flex;
    justify-content: flex-end;
    gap: 10px;
    margin-top: 20px;
  }
  .modal-btn {
    padding: 8px 18px;
    border-radius: 8px;
    font-size: 13px;
    font-weight: 600;
    cursor: pointer;
    border: none;
    transition: all 0.15s;
  }
  .modal-btn-cancel {
    background: #f1f5f9;
    color: #64748b;
  }
  .modal-btn-cancel:hover { background: #e2e8f0; }
  .modal-btn-save {
    background: #1a6faf;
    color: #fff;
  }
  .modal-btn-save:hover { background: #15577f; }
  .modal-btn-save:disabled { opacity: 0.5; cursor: not-allowed; }
  .modal-msg {
    font-size: 12px;
    margin-top: 12px;
    padding: 8px 12px;
    border-radius: 7px;
    display: none;
  }
  .modal-msg.ok { display: block; background: #ecfdf5; color: #065f46; border: 1px solid #a7f3d0; }
  .modal-msg.err { display: block; background: #fef2f2; color: #991b1b; border: 1px solid #fecaca; }

  /* ── Log Delivery Panel ── */
  .log-form { display: flex; flex-direction: column; gap: 12px; }
  .log-row   { display: flex; gap: 10px; flex-wrap: wrap; }
  .log-field { display: flex; flex-direction: column; gap: 4px; flex: 1; min-width: 140px; }
  .log-label {
    font-size: 10.5px; font-weight: 700; color: #64748b;
    text-transform: uppercase; letter-spacing: 0.5px;
  }
  .log-input {
    padding: 9px 12px; border-radius: 8px; font-size: 13px;
    border: 1.5px solid #e2e8f0; background: #f8fafc; color: #1e293b;
    transition: border-color 0.15s;
  }
  .log-input:focus { outline: none; border-color: #1a6faf; background: white; }
  .log-search-wrap { position: relative; }
  .log-dropdown {
    position: absolute; top: calc(100% + 4px); left: 0; right: 0;
    background: white; border: 1.5px solid #e2e8f0; border-radius: 10px;
    box-shadow: 0 6px 24px rgba(0,0,0,0.10); z-index: 200;
    max-height: 240px; overflow-y: auto; display: none;
  }
  .log-dropdown.open { display: block; }
  .log-dd-item {
    padding: 9px 14px; cursor: pointer; font-size: 12.5px;
    border-bottom: 1px solid #f1f5f9; line-height: 1.35;
  }
  .log-dd-item:last-child { border-bottom: none; }
  .log-dd-item:hover, .log-dd-item.active { background: #eff6ff; }
  .log-dd-id   { font-size: 10.5px; color: #94a3b8; margin-left: 6px; }
  .log-dd-prod { font-size: 10px; color: #64748b; }
  .log-dd-none { padding: 12px 14px; font-size: 12.5px; color: #94a3b8; }
  .log-submit {
    padding: 10px 20px; border-radius: 9px; font-size: 13px; font-weight: 700;
    background: #1a6faf; color: white; border: none; cursor: pointer;
    transition: opacity 0.15s, transform 0.1s; align-self: flex-end;
  }
  .log-submit:hover:not(:disabled) { opacity: 0.9; transform: translateY(-1px); }
  .log-submit:disabled { opacity: 0.45; cursor: not-allowed; }
  .log-result {
    font-size: 12.5px; padding: 8px 12px; border-radius: 7px;
    display: none; margin-top: 2px;
  }
  .log-result.ok  { display: block; background: #ecfdf5; color: #065f46; border: 1px solid #a7f3d0; }
  .log-result.err { display: block; background: #fef2f2; color: #991b1b; border: 1px solid #fecaca; }
</style>
</head>
<body>

<div class="header">
  <div class="header-icon">🛢️</div>
  <div class="header-text">
    <h1>S&amp;K Route Optimizer</h1>
    <p>S&amp;K Oil Sales — Phoenix, AZ</p>
  </div>
  <div style="margin-left:auto;text-align:right;display:flex;align-items:center;gap:16px;">
    <a href="/planning" style="color:#4a9eff;text-decoration:none;font-size:14px;font-weight:500;padding:6px 12px;border:1px solid #4a9eff;border-radius:4px;">📋 Planning view</a>
    <div>
      <div class="header-date" id="headerDate"></div>
      <div class="version-badge" id="versionBadge" title="Loading...">
        <div class="version-dot"></div>
        <span id="versionText">v—</span>
      </div>
    </div>
  </div>
</div>

<div class="main">

  <!-- ── Data Status ── -->
  <div class="card">
    <div class="card-title">📂 Data Status <button class="settings-btn" onclick="openSettings()" title="Data file settings">⚙️</button></div>
    <div id="staleBanner" class="stale-banner hidden">
      ⚠️ Data file hasn't been updated in over 2 days — routes may not reflect current inventory.
    </div>
    <div class="info-grid">
      <div class="info-item" id="fileCell">
        <div class="info-item-label">Input File</div>
        <div class="info-item-value" id="infoClients">—</div>
        <div class="info-item-sub" id="infoFilename">—</div>
      </div>
      <div class="info-item">
        <div class="info-item-label">Last Delivery</div>
        <div class="info-item-value" id="infoLastDelivery">—</div>
        <div class="info-item-sub" id="infoDeliveryCount">—</div>
      </div>
      <div class="info-item">
        <div class="info-item-label">File Updated</div>
        <div class="info-item-value" id="infoModified">—</div>
        <div class="info-item-sub" id="infoStatus">—</div>
      </div>
    </div>
  </div>

  <!-- ── Plan Week ── -->
  <div class="card">
    <div class="card-title">⚙️ Plan Week</div>

    <div class="plan-window">
      <div class="plan-window-icon">📅</div>
      <div>
        <div class="plan-window-text" id="planWindowText">Loading…</div>
        <div class="plan-window-sub">10-day horizon · 2-day commit · 2 trucks</div>
      </div>
    </div>

    <details class="advanced-details">
      <summary class="advanced-summary">Advanced Options</summary>
      <div class="advanced-body">
        <div style="font-size:12px;font-weight:600;color:#64748b;text-transform:uppercase;letter-spacing:0.5px;margin-bottom:10px;">Solver Speed</div>
        <div class="speed-tiles">
          <div class="speed-tile" data-sec="180" onclick="selectSpeed(this)">
            <div class="speed-tile-name">Quick</div>
            <div class="speed-tile-time">~3 min · rough</div>
          </div>
          <div class="speed-tile active" data-sec="600" onclick="selectSpeed(this)">
            <div class="speed-tile-name">Standard</div>
            <div class="speed-tile-time">~10 min · recommended</div>
          </div>
          <div class="speed-tile" data-sec="1500" onclick="selectSpeed(this)">
            <div class="speed-tile-name">Thorough</div>
            <div class="speed-tile-time">~25 min · best routes</div>
          </div>
        </div>
        <div style="font-size:11px;color:#64748b;margin-top:6px;">
          With 150+ clients, solver needs &gt;5 min for tight geographic
          clustering. Quick mode produces feasible but messy routes.
        </div>
      </div>
    </details>
  </div>

  <!-- ── Fleet Control ── -->
  <div class="card">
    <div class="card-title">🚛 Fleet</div>
    <div class="fleet-grid" id="fleetGrid"></div>
  </div>

  <!-- ── Client Review ── -->
  <div class="card" id="reviewCard">
    <div class="card-title">
      <span>📋 Pre-Solve Review</span>
      <div class="loading-spinner" id="reviewSpinner" style="margin-left:auto;"></div>
    </div>

    <div class="review-note" id="reviewNote">
      Loading inventory estimates… This shows which clients the optimizer will consider scheduling.
      You can <b>skip</b> clients you know don't need oil, or mark clients as <b>must-visit</b> to guarantee delivery.
    </div>

    <div class="urgency-counts" id="urgCounts"></div>

    <div id="overrideSummary" class="override-summary hidden"></div>

    <!-- Search bar — always visible -->
    <div class="add-client-bar">
      <input type="text" id="searchInput" placeholder="Search clients by name or ID…" autocomplete="off">
    </div>

    <div class="client-list" id="clientList">
      <div style="padding:40px;text-align:center;color:#94a3b8;">
        <div class="loading-spinner" style="margin:0 auto 12px;"></div>
        Loading client data…
      </div>
    </div>

    <div style="display:flex;gap:8px;align-items:center;flex-wrap:wrap;">
      <span style="font-size:11.5px;color:#94a3b8;" id="clientCountLabel"></span>
    </div>
  </div>

  <!-- ── Log Delivery ── -->
  <div class="card">
    <div class="card-title">📝 Log Delivery</div>
    <div class="log-form">
      <div class="log-row">
        <!-- Date -->
        <div class="log-field" style="max-width:160px;">
          <label class="log-label" for="ldDate">Date</label>
          <input class="log-input" id="ldDate" type="date">
        </div>
        <!-- Customer search -->
        <div class="log-field" style="flex:2;">
          <label class="log-label" for="ldSearch">Customer (name, number, or code)</label>
          <div class="log-search-wrap">
            <input class="log-input" id="ldSearch" type="text"
                   placeholder="Type name, 1018, or AJO…" autocomplete="off">
            <div class="log-dropdown" id="ldDropdown"></div>
          </div>
        </div>
        <!-- Qty -->
        <div class="log-field" style="max-width:130px;">
          <label class="log-label" for="ldQty">Qty (lbs)</label>
          <input class="log-input" id="ldQty" type="number" min="1" placeholder="e.g. 900">
        </div>
      </div>
      <div style="display:flex;align-items:center;gap:12px;flex-wrap:wrap;">
        <button class="log-submit" id="ldSubmit" onclick="submitDelivery()" disabled>Save Delivery</button>
        <div class="log-result" id="ldResult"></div>
      </div>
    </div>
  </div>

  <!-- ── Generate Routes ── -->
  <div class="card">
    <button class="btn-run" id="runBtn" onclick="runOptimizer()">
      <span>▶</span> Generate Routes
    </button>
  </div>

  <!-- ── Progress log ── -->
  <div class="card hidden" id="logCard">
    <div class="card-title">📋 Progress</div>
    <div class="status-row">
      <div class="dot running" id="statusDot"></div>
      <div class="status-label" id="statusLabel">Running optimizer…</div>
      <div class="status-timer" id="statusTimer"></div>
    </div>
    <div class="progress-wrap">
      <div class="progress-bar" id="progressBar"></div>
    </div>
    <div class="log-box" id="logBox"></div>
  </div>

  <!-- ── Results ── -->
  <div class="card hidden" id="resultsCard">
    <div class="card-title">✅ Routes Ready <span id="sourceLabel" style="font-size:11px;font-weight:500;color:#64748b;margin-left:8px;"></span></div>
    <div id="warmStartBadge" style="font-size:12px;margin-bottom:10px;display:none;"></div>
    <div style="font-size:13px;color:#64748b;margin-bottom:12px;">
      <b>Committed</b> routes (first 2 days) are ready for dispatch.
      <b>Tentative</b> routes (days 3-10) are lookahead — they'll be refined tomorrow.
    </div>
    <div class="dl-row">
      <a class="dl-btn dl-excel" id="excelLink" href="#" download>
        📊 Download Schedule
      </a>
      <a class="dl-btn dl-map" id="mapLink" href="#" target="_blank">
        🗺️ Open Route Map
      </a>
    </div>
    <div id="extraOutputs" style="font-size:12px;color:#64748b;margin-top:10px;display:flex;gap:10px;flex-wrap:wrap;"></div>
  </div>

  <!-- ── Last Run ── -->
  <div class="card hidden" id="lastRunCard">
    <div class="card-title">📁 Last Run</div>
    <div class="last-run-meta" id="lastRunMeta"></div>

    <!-- Urgency snapshot -->
    <div class="urgency-row" id="urgencyRow"></div>

    <!-- Stats chips -->
    <div class="stat-chips" id="statChips"></div>

    <!-- File links -->
    <div class="last-run-links" id="lastRunLinks"></div>
  </div>

</div>

<script>
  // ── Init data ──────────────────────────────────────────────────────────────
  const DATA_INFO = __DATA_INFO__;
  const OUTPUTS   = __OUTPUTS__;
  const VERSION   = __VERSION__;
  const SUMMARY   = __SUMMARY__;

  // On page load: if there's an existing output, show the results card and
  // populate links. Distinguishes IRP vs legacy by filename prefix.
  document.addEventListener('DOMContentLoaded', function() {
    const card = document.getElementById('resultsCard');
    if (!card) return;
    const hasOutput = OUTPUTS.irp_excel || OUTPUTS.unified_excel || OUTPUTS.excel;
    if (!hasOutput) return;
    card.classList.remove('hidden');
    // Prefer IRP if both exist (it's the recommended workflow)
    const excel = OUTPUTS.irp_excel || OUTPUTS.unified_excel || OUTPUTS.excel;
    const map   = OUTPUTS.irp_map   || OUTPUTS.unified_map   || OUTPUTS.map;
    if (excel) document.getElementById('excelLink').href = '/download/' + excel;
    if (map)   document.getElementById('mapLink').href   = '/view/' + map;
    const src = excel && excel.includes('sk_irp') ? 'IRP' :
                excel && excel.includes('sk_unified') ? 'Legacy unified' : '';
    const lbl = document.getElementById('sourceLabel');
    if (lbl) lbl.textContent = src ? `(${src})` : '';
    const extras = document.getElementById('extraOutputs');
    if (extras) {
      const items = [];
      if (OUTPUTS.irp_excel)     items.push(`<a href="/download/${OUTPUTS.irp_excel}" style="color:#1a6faf;">IRP schedule</a>`);
      if (OUTPUTS.irp_map)       items.push(`<a href="/view/${OUTPUTS.irp_map}" target="_blank" style="color:#1a6faf;">IRP map</a>`);
      if (OUTPUTS.irp_csv)       items.push(`<a href="/download/${OUTPUTS.irp_csv}" style="color:#1a6faf;">IRP → SmartService CSV</a>`);
      if (OUTPUTS.unified_excel) items.push(`<a href="/download/${OUTPUTS.unified_excel}" style="color:#888;">Legacy schedule</a>`);
      if (OUTPUTS.unified_map)   items.push(`<a href="/view/${OUTPUTS.unified_map}" target="_blank" style="color:#888;">Legacy map</a>`);
      extras.innerHTML = items.length ? ('All available: ' + items.join(' · ')) : '';
    }
  });

  let selectedSec = 600;

  // ── State ───────────────────────────────────────────────────────────────────
  const TRUCK_NAMES = __TRUCK_NAMES__;
  let activeTrucks = new Set(TRUCK_NAMES);
  let allClients   = [];     // full snapshot from /snapshot
  let skipIds      = new Set();
  let mustVisitIds = new Set();
  let showAllClients = false; // toggle for showing non-urgent clients

  // ── Version badge ───────────────────────────────────────────────────────────
  if (VERSION && VERSION.hash !== '—') {
    document.getElementById('versionText').textContent = VERSION.hash + ' · ' + VERSION.date;
    document.getElementById('versionBadge').title = VERSION.msg || '';
  }

  // ── Header date ────────────────────────────────────────────────────────────
  const now = new Date();
  document.getElementById('headerDate').innerHTML =
    now.toLocaleDateString('en-US', {weekday:'long', month:'long', day:'numeric', year:'numeric'}) +
    '<br>' + now.toLocaleTimeString('en-US', {hour:'2-digit', minute:'2-digit'});

  // ── Compute planning window ────────────────────────────────────────────────
  const WORK_JS_DAYS = new Set([2, 3, 4, 5, 6]);
  const DAY_NAMES    = ['Sun','Mon','Tue','Wed','Thu','Fri','Sat'];
  const today = new Date(now); today.setHours(0,0,0,0);

  const planDates = [];
  const cursor = new Date(today);
  cursor.setDate(cursor.getDate() + 1);
  for (let safety = 0; safety < 30 && planDates.length < 5; safety++) {
    if (WORK_JS_DAYS.has(cursor.getDay())) planDates.push(new Date(cursor));
    cursor.setDate(cursor.getDate() + 1);
  }

  if (planDates.length >= 2) {
    const fmt = d => d.toLocaleDateString('en-US', {weekday:'short', month:'short', day:'numeric'});
    document.getElementById('planWindowText').textContent =
      fmt(planDates[0]) + '  →  ' + fmt(planDates[planDates.length - 1]);
  }

  // ── Populate data info ─────────────────────────────────────────────────────
  function populateDataInfo(d) {
    if (d.stale) {
      document.getElementById('staleBanner').classList.remove('hidden');
      document.getElementById('fileCell').classList.add('stale');
    }
    if (d.client_count && d.client_count !== '—') {
      document.getElementById('infoClients').innerHTML =
        d.client_count + ' clients <span class="badge badge-green">loaded</span>';
    } else if (d.input_modified) {
      document.getElementById('infoClients').innerHTML =
        '<span class="badge badge-green">Connected</span>';
    } else {
      document.getElementById('infoClients').innerHTML =
        '<span class="badge badge-yellow">Not found</span>';
    }
    document.getElementById('infoFilename').textContent =
      d.input_filename || d.input_path || '—';
    if (d.last_delivery) {
      document.getElementById('infoLastDelivery').textContent = d.last_delivery;
      document.getElementById('infoDeliveryCount').textContent =
        (d.delivery_count || '?') + ' deliveries on record';
    } else {
      document.getElementById('infoLastDelivery').textContent = 'No data';
      document.getElementById('infoDeliveryCount').textContent = '—';
    }
    if (d.input_modified) {
      document.getElementById('infoModified').textContent = d.input_modified;
      document.getElementById('infoStatus').textContent =
        d.stale ? '⚠️ Over 2 days old' : 'Up to date';
    } else {
      document.getElementById('infoModified').textContent = 'Not found';
      document.getElementById('infoStatus').textContent = '—';
    }
  }
  populateDataInfo(DATA_INFO);
  fetch('/data-info').then(r => r.json()).then(d => populateDataInfo(d)).catch(() => {});

  // ── Saturday auto-detection ─────────────────────────────────────────────────
  // On Saturdays, one truck runs Tucson/Flagstaff. Auto-restrict metro planning
  // to SATURDAY_TRUCKS (from config). Users can still override manually.
  const SATURDAY_TRUCKS = __SATURDAY_TRUCKS__;
  const isSaturdayPlan = planDates.length > 0 && planDates.some(d => d.getDay() === 6);
  let saturdayMode = false;

  if (isSaturdayPlan && SATURDAY_TRUCKS.length < TRUCK_NAMES.length) {
    saturdayMode = true;
    // Don't auto-remove trucks from activeTrucks — the solver handles per-day
    // restriction internally. But show a notice.
  }

  // ── Fleet Control ──────────────────────────────────────────────────────────
  function renderFleet() {
    const grid = document.getElementById('fleetGrid');
    let html = TRUCK_NAMES.map(name => {
      const isOn = activeTrucks.has(name);
      const isSatOnly = saturdayMode && !SATURDAY_TRUCKS.includes(name);
      const info = isSatOnly ? 'Tucson/Flagstaff on Sat' : (isOn ? 'Active' : 'Out of service');
      return `
        <div class="truck-tile ${isOn ? '' : 'off'}" onclick="toggleTruck('${name}')">
          <div style="display:flex;align-items:center;gap:6px;">
            <div class="truck-dot ${isOn ? 'on' : 'off'}"></div>
            <span class="truck-tile-name">${name}</span>
          </div>
          <div class="truck-tile-info">${info}</div>
          ${isSatOnly && isOn ? '<div style="font-size:11px;color:var(--amber);margin-top:2px;">Sat: 1 truck metro only</div>' : ''}
        </div>`;
    }).join('');
    grid.innerHTML = html;
  }

  function toggleTruck(name) {
    if (activeTrucks.has(name)) {
      if (activeTrucks.size <= 1) return; // need at least 1
      activeTrucks.delete(name);
    } else {
      activeTrucks.add(name);
    }
    renderFleet();
    updateOverrideSummary();
  }

  renderFleet();

  // ── Client Review ──────────────────────────────────────────────────────────
  function loadSnapshot() {
    fetch('/snapshot')
      .then(r => r.json())
      .then(data => {
        if (data.error) throw new Error(data.error);
        allClients = data.clients || [];
        document.getElementById('reviewSpinner').style.display = 'none';
        document.getElementById('reviewNote').innerHTML =
          'Stockout and critical clients are <b>automatically prioritized</b> by the solver. ' +
          'Use <b>Skip</b> if you know a client doesn\\'t need oil yet, or <b>Must</b> to force a normal client onto the schedule.';
        renderClientList();
      })
      .catch(err => {
        document.getElementById('reviewSpinner').style.display = 'none';
        document.getElementById('clientList').innerHTML =
          '<div style="padding:30px;text-align:center;color:#ef4444;">Failed to load: ' + err.message + '</div>';
      });
  }

  function renderClientList() {
    // Determine which clients to show:
    // If searching, show all matches regardless of threshold.
    // Otherwise show urgent/critical/stockout + overrides (DTE ≤ 12),
    // or everything if showAllClients is toggled on.
    const SHOW_THRESHOLD = 12;
    let visible = allClients.filter(c => {
      // Search filter first
      if (_searchQuery) {
        return c.name.toLowerCase().includes(_searchQuery) ||
               c.id.toLowerCase().includes(_searchQuery);
      }
      if (mustVisitIds.has(c.id) || skipIds.has(c.id)) return true;
      if (showAllClients) return true;
      return c.dte <= SHOW_THRESHOLD;
    });

    // Urgency counts (from ALL clients, not just visible)
    const counts = {stockout: 0, critical: 0, urgent: 0, normal: 0};
    allClients.forEach(c => { counts[c.urgency] = (counts[c.urgency] || 0) + 1; });

    document.getElementById('urgCounts').innerHTML = [
      counts.stockout ? `<div class="urg-chip stockout">${counts.stockout} stockout</div>` : '',
      counts.critical ? `<div class="urg-chip critical">${counts.critical} critical</div>` : '',
      counts.urgent   ? `<div class="urg-chip urgent">${counts.urgent} urgent</div>` : '',
      `<div class="urg-chip normal">${counts.normal} on track</div>`,
    ].join('');

    // Group: must-visit first, then by urgency, then skipped at bottom
    const mustFirst = visible.filter(c => mustVisitIds.has(c.id));
    const skipped   = visible.filter(c => skipIds.has(c.id) && !mustVisitIds.has(c.id));
    const regular   = visible.filter(c => !mustVisitIds.has(c.id) && !skipIds.has(c.id));

    const ordered = [...mustFirst, ...regular, ...skipped];

    let html = '';

    if (mustFirst.length) {
      html += '<div class="section-divider">Must Visit</div>';
      html += mustFirst.map(c => clientRowHtml(c)).join('');
    }

    if (regular.length) {
      html += '<div class="section-divider">Estimated Needs (' + regular.length + ')</div>';
      html += regular.map(c => clientRowHtml(c)).join('');
    }

    if (skipped.length) {
      html += '<div class="section-divider">Skipped</div>';
      html += skipped.map(c => clientRowHtml(c)).join('');
    }

    document.getElementById('clientList').innerHTML = html || '<div style="padding:30px;text-align:center;color:#94a3b8;">No clients to show</div>';

    const hiddenCount = allClients.length - visible.length;
    const label = document.getElementById('clientCountLabel');
    if (hiddenCount > 0 && !showAllClients) {
      label.innerHTML = `Showing ${visible.length} of ${allClients.length} clients · <a href="#" onclick="showAll(event)" style="color:#1a6faf;font-weight:600;">Show all</a>`;
    } else if (showAllClients) {
      label.innerHTML = `Showing all ${allClients.length} clients · <a href="#" onclick="showLess(event)" style="color:#1a6faf;font-weight:600;">Show less</a>`;
    } else {
      label.textContent = `${visible.length} clients`;
    }

    updateOverrideSummary();
  }

  function clientRowHtml(c) {
    const isSkipped = skipIds.has(c.id);
    const isMust    = mustVisitIds.has(c.id);
    let cls = 'client-row ' + c.urgency;
    if (isSkipped) cls = 'client-row skipped';
    if (isMust)    cls = 'client-row must-visit';

    const dteStr = c.dte >= 999 ? 'n/a' : c.dte.toFixed(1) + 'd';
    const tankPct = c.tank > 0 ? Math.round((c.current / c.tank) * 100) : 0;

    // Mini tank gauge
    const gaugeColor = c.urgency === 'stockout' ? '#dc2626'
                     : c.urgency === 'critical' ? '#ea580c'
                     : c.urgency === 'urgent'   ? '#d97706'
                     : '#22c55e';

    let buttons = '';
    if (isSkipped) {
      buttons = `<button class="btn-undo" onclick="undoClient('${c.id}')" style="background:#fee2e2;border-color:#fca5a5;color:#dc2626;">✕ Restore</button>`;
    } else if (isMust) {
      buttons = `<button class="btn-undo" onclick="undoClient('${c.id}')">Remove</button>`;
    } else if (c.urgency === 'stockout' || c.urgency === 'critical') {
      // Solver already forces these — only offer Skip override
      buttons = `<button class="btn-skip" onclick="skipClient('${c.id}')">Skip</button>`;
    } else {
      buttons = `<button class="btn-skip" onclick="skipClient('${c.id}')">Skip</button>` +
                `<button class="btn-must" onclick="mustClient('${c.id}')">Must</button>`;
    }

    return `<div class="${cls}" data-id="${c.id}">
      <div class="urg-dot ${c.urgency}"></div>
      <div class="client-name" title="${c.id}">${c.name}</div>
      <div class="client-meta">
        <span title="Days to empty">${dteStr}</span>
        <span style="margin:0 4px;color:#cbd5e1;">·</span>
        <span title="Tank: ${tankPct}% full (${c.current}/${c.tank} lbs)">
          <span style="display:inline-block;width:32px;height:6px;background:#e2e8f0;border-radius:3px;vertical-align:middle;position:relative;overflow:hidden;">
            <span style="position:absolute;left:0;top:0;height:100%;width:${tankPct}%;background:${gaugeColor};border-radius:3px;"></span>
          </span>
          ${tankPct}%
        </span>
        <span style="margin:0 4px;color:#cbd5e1;">·</span>
        <span title="Estimated refill">${c.refill} lbs</span>
      </div>
      <div style="display:flex;gap:4px;flex-shrink:0;">${buttons}</div>
    </div>`;
  }

  function skipClient(id) {
    mustVisitIds.delete(id);
    skipIds.add(id);
    renderClientList();
  }

  function mustClient(id) {
    skipIds.delete(id);
    mustVisitIds.add(id);
    renderClientList();
  }

  function undoClient(id) {
    skipIds.delete(id);
    mustVisitIds.delete(id);
    renderClientList();
  }

  function showAll(e) { e.preventDefault(); showAllClients = true; renderClientList(); }
  function showLess(e) { e.preventDefault(); showAllClients = false; renderClientList(); }

  function updateOverrideSummary() {
    const el = document.getElementById('overrideSummary');
    const parts = [];
    if (skipIds.size) parts.push(skipIds.size + ' skipped');
    if (mustVisitIds.size) parts.push(mustVisitIds.size + ' must-visit');
    if (activeTrucks.size < TRUCK_NAMES.length) parts.push((TRUCK_NAMES.length - activeTrucks.size) + ' truck(s) offline');
    if (parts.length) {
      el.textContent = 'Overrides: ' + parts.join(' · ');
      el.classList.remove('hidden');
    } else {
      el.classList.add('hidden');
    }
  }

  // Inline search — always visible, filters the list in place
  let _searchQuery = '';
  document.getElementById('searchInput')?.addEventListener('input', function(e) {
    _searchQuery = e.target.value.toLowerCase().trim();
    renderClientList();
  });

  // Load snapshot on page load
  loadSnapshot();

  // ── Last Run card ──────────────────────────────────────────────────────────
  function renderLastRun(s) {
    if (!s || Object.keys(s).length === 0) return;
    const card = document.getElementById('lastRunCard');
    card.classList.remove('hidden');

    const ts = s.timestamp ? new Date(s.timestamp) : null;
    const tsStr = ts
      ? ts.toLocaleDateString('en-US', {month:'short', day:'numeric', year:'numeric'}) +
        '  ' + ts.toLocaleTimeString('en-US', {hour:'2-digit', minute:'2-digit'})
      : '';
    const elapsed = s.elapsed_sec
      ? Math.floor(s.elapsed_sec / 60) + ':' + String(s.elapsed_sec % 60).padStart(2,'0') + ' solve time'
      : '';
    document.getElementById('lastRunMeta').textContent =
      [tsStr, elapsed].filter(Boolean).join('  ·  ');

    const urg = s.urgency || {};
    let urgHtml = '';
    if (urg.critical) urgHtml += `<div class="urgency-pill critical"><div class="urgency-dot"></div>${urg.critical} critical</div>`;
    if (urg.urgent)   urgHtml += `<div class="urgency-pill urgent"><div class="urgency-dot"></div>${urg.urgent} urgent</div>`;
    if (urg.normal)   urgHtml += `<div class="urgency-pill normal"><div class="urgency-dot"></div>${urg.normal} on track</div>`;
    document.getElementById('urgencyRow').innerHTML = urgHtml || '<span style="color:#94a3b8;font-size:13px;">No urgency data</span>';

    const chips = [];
    if (s.routes_count)  chips.push({val: s.routes_count,  label: 'Routes'});
    if (s.total_stops)   chips.push({val: s.total_stops,   label: 'Stops'});
    if (s.total_miles != null) chips.push({val: Math.round(s.total_miles) + ' mi', label: 'Est. Miles'});
    if (s.deferred_count) chips.push({val: s.deferred_count, label: 'Deferred'});
    document.getElementById('statChips').innerHTML = chips.map(c =>
      `<div class="stat-chip"><div class="stat-chip-val">${c.val}</div><div class="stat-chip-label">${c.label}</div></div>`
    ).join('');

    let linksHtml = '';
    if (s.excel_file) linksHtml += `<a class="prev-link" href="/download/${s.excel_file}">📊 ${s.excel_file}</a>`;
    if (s.excel_file && s.map_file) linksHtml += `<span class="divider">·</span>`;
    if (s.map_file)   linksHtml += `<a class="prev-link" href="/view/${s.map_file}" target="_blank">🗺️ ${s.map_file}</a>`;
    document.getElementById('lastRunLinks').innerHTML = linksHtml;
  }
  renderLastRun(SUMMARY);

  // ── Speed selection ────────────────────────────────────────────────────────
  function selectSpeed(el) {
    document.querySelectorAll('.speed-tile').forEach(t => t.classList.remove('active'));
    el.classList.add('active');
    selectedSec = parseInt(el.dataset.sec);
  }

  // ── Timer & progress ───────────────────────────────────────────────────────
  let _timerInterval = null;
  let _startTime     = null;

  function startTimer(totalSec) {
    _startTime = Date.now();
    const bar  = document.getElementById('progressBar');
    const tmr  = document.getElementById('statusTimer');
    bar.style.width = '0%';
    bar.classList.remove('done');
    _timerInterval = setInterval(() => {
      const elapsed = (Date.now() - _startTime) / 1000;
      const pct     = Math.min((elapsed / totalSec) * 100, 98);
      bar.style.width = pct + '%';
      const eMin = Math.floor(elapsed / 60);
      const eSec = Math.floor(elapsed % 60).toString().padStart(2,'0');
      const tMin = Math.floor(totalSec / 60);
      const tSec = (totalSec % 60).toString().padStart(2,'0');
      tmr.textContent = `${eMin}:${eSec} / ~${tMin}:${tSec}`;
    }, 1000);
  }

  function stopTimer(success) {
    if (_timerInterval) { clearInterval(_timerInterval); _timerInterval = null; }
    const bar = document.getElementById('progressBar');
    const tmr = document.getElementById('statusTimer');
    bar.style.width = '100%';
    if (success) bar.classList.add('done');
    if (_startTime) {
      const elapsed = (Date.now() - _startTime) / 1000;
      const eMin = Math.floor(elapsed / 60);
      const eSec = Math.floor(elapsed % 60).toString().padStart(2,'0');
      tmr.textContent = `Done in ${eMin}:${eSec}`;
    }
  }

  // ── Run optimizer ──────────────────────────────────────────────────────────
  function runOptimizer() {
    const btn = document.getElementById('runBtn');
    btn.disabled = true;
    btn.innerHTML = '<span>⏳</span> Running…';

    document.getElementById('logCard').classList.remove('hidden');
    document.getElementById('resultsCard').classList.add('hidden');
    const logBox = document.getElementById('logBox');
    logBox.innerHTML = '';

    setStatus('running', 'Running optimizer…');
    startTimer(selectedSec);

    const payload = {
      solve_sec: selectedSec,
      skip_ids: Array.from(skipIds),
      must_visit_ids: Array.from(mustVisitIds),
      active_trucks: Array.from(activeTrucks),
    };

    fetch('/run', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify(payload),
    }).then(resp => {
      const reader  = resp.body.getReader();
      const decoder = new TextDecoder();
      let   buf     = '';

      function pump() {
        reader.read().then(({done, value}) => {
          if (done) return;
          buf += decoder.decode(value, {stream: true});
          const parts = buf.split('\\n\\n');
          buf = parts.pop();
          for (const part of parts) {
            if (!part.startsWith('data: ')) continue;
            try {
              const msg = JSON.parse(part.slice(6));
              if (msg.line !== undefined) addLine(msg.line);
              if (msg.done)              handleDone(msg);
            } catch(e) {}
          }
          pump();
        });
      }
      pump();
    }).catch(err => {
      addLine('Network error: ' + err.message);
      setStatus('fail', 'Connection lost');
      stopTimer(false);
      resetBtn();
    });
  }

  // Captured during the run; surfaced on the results card after completion
  let capturedRunInfo = {};

  function addLine(text) {
    if (!text.trim()) return;
    const logBox = document.getElementById('logBox');
    const d = document.createElement('div');
    d.className = 'll';
    if (/✓|Done|complete/i.test(text))          d.classList.add('ok');
    else if (/ERROR|⛔|failed/i.test(text))      d.classList.add('err');
    else if (/^[=\\[\\d]|^\\/|S&K/.test(text))  d.classList.add('head');
    else if (/warn|⚠/i.test(text))              d.classList.add('warn');
    d.textContent = text;
    logBox.appendChild(d);
    logBox.scrollTop = logBox.scrollHeight;

    // Sniff key signals from the log so we can show them prominently on
    // the results card (warm-start status, plan stability, key counts).
    const m1 = text.match(/Warm start: (\\d+) visits matched.*?(\\d+) missed/i);
    if (m1) capturedRunInfo.warmStart = {matched: +m1[1], missed: +m1[2]};
    if (/Warm start: no usable visits|cold start/i.test(text)) {
      capturedRunInfo.warmStart = {matched: 0, missed: 0, cold: true};
    }
    const m2 = text.match(/plan stability vs yesterday:\\s*([\\d.]+)%/i);
    if (m2) capturedRunInfo.planStability = parseFloat(m2[1]);
    const m3 = text.match(/Prior plan:\\s*(\\S+)\\s*\\((\\d+)\\s*day/i);
    if (m3) capturedRunInfo.priorPlanAge = +m3[2];
  }

  function handleDone(data) {
    stopTimer(data.success);
    if (data.success) {
      setStatus('ok', 'Complete — routes are ready');
      if (data.excel || data.map) {
        const card = document.getElementById('resultsCard');
        card.classList.remove('hidden');
        if (data.excel) document.getElementById('excelLink').href = '/download/' + data.excel;
        if (data.map)   document.getElementById('mapLink').href   = '/view/' + data.map;

        // Warm-start + stability badge
        const wsBadge = document.getElementById('warmStartBadge');
        if (wsBadge) {
          const parts = [];
          const ws = capturedRunInfo.warmStart;
          if (ws) {
            if (ws.cold || (ws.matched === 0 && ws.missed > 0)) {
              parts.push('<span style="color:#d97706;">❄️ Cold start (no usable previous plan)</span>');
            } else if (ws.matched > 0) {
              parts.push(`<span style="color:#16a34a;">↻ Warm-started from yesterday: ${ws.matched} visits carried over</span>`);
            }
          } else {
            parts.push('<span style="color:#64748b;">First run — cold start (next run will warm-start from this plan)</span>');
          }
          if (capturedRunInfo.planStability !== undefined) {
            const s = capturedRunInfo.planStability;
            const color = s >= 50 ? '#16a34a' : (s >= 20 ? '#ca8a04' : '#dc2626');
            parts.push(`<span style="color:${color};margin-left:14px;">📊 Plan stability vs yesterday: <b>${s}%</b></span>`);
          }
          wsBadge.innerHTML = parts.join('');
          wsBadge.style.display = parts.length ? 'block' : 'none';
        }

        // Identify source engine and label it
        const src = (data.excel && data.excel.includes('sk_irp')) ? 'IRP' :
                    (data.excel && data.excel.includes('sk_unified')) ? 'Legacy unified' : '';
        const lbl = document.getElementById('sourceLabel');
        if (lbl) lbl.textContent = src ? `(produced by: ${src})` : '';

        // Refresh full outputs panel to show all available downloads
        fetch('/data-info').then(r => r.json()).then(d => {});  // ensure data-info refresh
        refreshExtraOutputs();
      }
      fetch('/last-run').then(r => r.json()).then(s => renderLastRun(s)).catch(() => {});
    } else {
      setStatus('fail', 'Failed — see log for details');
    }
    resetBtn();
  }

  function refreshExtraOutputs() {
    fetch('/').then(() => fetch('/data-info'));
    // Re-render the secondary outputs row from OUTPUTS (refreshed via window reload would be cleaner)
    const extras = document.getElementById('extraOutputs');
    if (!extras) return;
    const out = OUTPUTS;
    const items = [];
    if (out.irp_excel)     items.push(`<a href="/download/${out.irp_excel}" style="color:#1a6faf;">IRP schedule</a>`);
    if (out.irp_map)       items.push(`<a href="/view/${out.irp_map}" target="_blank" style="color:#1a6faf;">IRP map</a>`);
    if (out.irp_csv)       items.push(`<a href="/download/${out.irp_csv}" style="color:#1a6faf;">IRP → SmartService CSV</a>`);
    if (out.unified_excel) items.push(`<a href="/download/${out.unified_excel}" style="color:#888;">Legacy schedule</a>`);
    if (out.unified_map)   items.push(`<a href="/view/${out.unified_map}" target="_blank" style="color:#888;">Legacy map</a>`);
    extras.innerHTML = items.length ? ('All available: ' + items.join(' · ')) : '';
  }

  function setStatus(state, text) {
    const dot = document.getElementById('statusDot');
    dot.className = 'dot ' + state;
    document.getElementById('statusLabel').textContent = text;
  }

  function resetBtn() {
    const btn = document.getElementById('runBtn');
    btn.disabled = false;
    btn.innerHTML = '<span>▶</span> Generate Routes';
  }
</script>

<!-- ── Log Delivery JS ── -->
<script>
(function () {
  // ── State ──────────────────────────────────────────────────────────────
  var _allClients = [];          // [{id, name, prod}] from /clients-all
  var _selectedClient = null;    // currently chosen client object
  var _ddIdx = -1;               // keyboard-highlighted index in dropdown

  // ── Init ───────────────────────────────────────────────────────────────
  // Set date to today
  var _today = new Date();
  var _mm = String(_today.getMonth()+1).padStart(2,'0');
  var _dd = String(_today.getDate()).padStart(2,'0');
  document.getElementById('ldDate').value = _today.getFullYear() + '-' + _mm + '-' + _dd;

  // Load client list once
  fetch('/clients-all').then(function(r){ return r.json(); }).then(function(data){
    _allClients = data.clients || [];
  }).catch(function(){});

  // ── Search filtering ────────────────────────────────────────────────────
  function _filterClients(q) {
    if (!q) return [];
    var lq = q.toLowerCase();
    return _allClients.filter(function(c) {
      return c.name.toLowerCase().includes(lq) || c.id.includes(lq);
    }).slice(0, 30);
  }

  function _renderDropdown(matches) {
    var dd = document.getElementById('ldDropdown');
    _ddIdx = -1;
    if (!matches.length) {
      dd.innerHTML = '<div class="log-dd-none">No clients found</div>';
      dd.classList.add('open');
      return;
    }
    dd.innerHTML = matches.map(function(c, i) {
      var nid = _numericId(c.name);
      // Show short code + name without the numeric ID in the middle
      var parts = c.name.split(' - ');
      var shortName = parts.length >= 3 ? parts.slice(2).join(' - ') : c.name;
      var code = parts[0] || '';
      return '<div class="log-dd-item" data-idx="' + i + '" onclick="ldPick(' + i + ')">' +
        '<span style="font-weight:600;">' + shortName + '</span>' +
        '<span class="log-dd-id">#' + nid + ' · ' + code + '</span>' +
        (c.prod ? '<div class="log-dd-prod">' + c.prod + '</div>' : '') +
        '</div>';
    }).join('');
    dd.classList.add('open');
    dd._matches = matches;
  }

  document.getElementById('ldSearch').addEventListener('input', function() {
    _selectedClient = null;
    _updateSubmit();
    var matches = _filterClients(this.value.trim());
    if (!this.value.trim()) {
      document.getElementById('ldDropdown').classList.remove('open');
      return;
    }
    _renderDropdown(matches);
  });

  document.getElementById('ldSearch').addEventListener('keydown', function(e) {
    var dd = document.getElementById('ldDropdown');
    var items = dd.querySelectorAll('.log-dd-item');
    if (e.key === 'ArrowDown') {
      e.preventDefault();
      _ddIdx = Math.min(_ddIdx + 1, items.length - 1);
      items.forEach(function(el,i){ el.classList.toggle('active', i===_ddIdx); });
    } else if (e.key === 'ArrowUp') {
      e.preventDefault();
      _ddIdx = Math.max(_ddIdx - 1, 0);
      items.forEach(function(el,i){ el.classList.toggle('active', i===_ddIdx); });
    } else if (e.key === 'Enter' && _ddIdx >= 0) {
      e.preventDefault();
      ldPick(_ddIdx);
    } else if (e.key === 'Escape') {
      dd.classList.remove('open');
    }
  });

  // Close dropdown on outside click
  document.addEventListener('click', function(e) {
    if (!e.target.closest('.log-search-wrap')) {
      document.getElementById('ldDropdown').classList.remove('open');
    }
  });

  // ── Pick a client ───────────────────────────────────────────────────────
  window.ldPick = function(idx) {
    var dd = document.getElementById('ldDropdown');
    var matches = dd._matches || [];
    if (!matches[idx]) return;
    _selectedClient = matches[idx];
    document.getElementById('ldSearch').value = _selectedClient.name;
    dd.classList.remove('open');
    document.getElementById('ldQty').focus();
    _updateSubmit();
  };

  // ── Enable submit only when all fields are filled ──────────────────────
  function _updateSubmit() {
    var btn = document.getElementById('ldSubmit');
    var qty = parseInt(document.getElementById('ldQty').value, 10);
    var dt  = document.getElementById('ldDate').value;
    btn.disabled = !(_selectedClient && qty > 0 && dt);
  }
  document.getElementById('ldQty').addEventListener('input', _updateSubmit);
  document.getElementById('ldDate').addEventListener('input', _updateSubmit);

  // ── Submit ──────────────────────────────────────────────────────────────
  window.submitDelivery = function() {
    var btn = document.getElementById('ldSubmit');
    var res = document.getElementById('ldResult');
    if (!_selectedClient) return;
    var qty = parseInt(document.getElementById('ldQty').value, 10);
    var dt  = document.getElementById('ldDate').value;
    if (!qty || !dt) return;

    btn.disabled = true;
    btn.textContent = 'Saving…';
    res.className = 'log-result';

    fetch('/log-delivery', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({ date: dt, customer: _selectedClient.id, qty: qty })
    }).then(function(r){ return r.json(); }).then(function(data){
      btn.disabled = false;
      btn.textContent = 'Save Delivery';
      if (data.ok) {
        res.className = 'log-result ok';
        res.textContent = '✓ Saved: ' + _selectedClient.name + ' (#' + data.customer + ') — ' + data.qty + ' lbs';
        // Reset form
        _selectedClient = null;
        document.getElementById('ldSearch').value = '';
        document.getElementById('ldQty').value = '';
        _updateSubmit();
      } else {
        res.className = 'log-result err';
        res.textContent = '✗ ' + (data.error || 'Error saving delivery');
      }
    }).catch(function(){
      btn.disabled = false;
      btn.textContent = 'Save Delivery';
      res.className = 'log-result err';
      res.textContent = 'Network error — is the app running?';
    });
  };
})();
</script>

<!-- ── Settings Modal ── -->
<div class="modal-overlay" id="settingsModal">
  <div class="modal-box">
    <div class="modal-title">Data File Settings</div>
    <div class="modal-sub">
      Set the path to your SK_Delivery_System.xlsx file on this computer.
      This is saved locally and won't affect other machines.
    </div>
    <div class="modal-label">File Path</div>
    <input class="modal-input" id="settingsPath" type="text"
           placeholder="C:\\Users\\...\\SK_Delivery_System.xlsx" spellcheck="false">
    <div class="modal-msg" id="settingsMsg"></div>
    <div class="modal-actions">
      <button class="modal-btn modal-btn-cancel" onclick="closeSettings()">Cancel</button>
      <button class="modal-btn modal-btn-save" id="settingsSave" onclick="saveSettings()">Save</button>
    </div>
  </div>
</div>

<script>
  function openSettings() {
    var msg = document.getElementById('settingsMsg');
    msg.className = 'modal-msg'; msg.textContent = '';
    fetch('/settings').then(function(r){ return r.json(); }).then(function(cfg){
      document.getElementById('settingsPath').value = cfg.input_file || '';
    }).catch(function(){});
    document.getElementById('settingsModal').classList.add('open');
  }

  function closeSettings() {
    document.getElementById('settingsModal').classList.remove('open');
  }

  document.getElementById('settingsModal').addEventListener('click', function(e) {
    if (e.target === this) closeSettings();
  });

  function saveSettings() {
    var btn = document.getElementById('settingsSave');
    var msg = document.getElementById('settingsMsg');
    var path = document.getElementById('settingsPath').value.trim();
    if (!path) { msg.className = 'modal-msg err'; msg.textContent = 'Please enter a file path.'; return; }
    btn.disabled = true;
    msg.className = 'modal-msg'; msg.textContent = '';
    fetch('/settings', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({input_file: path})
    }).then(function(r){ return r.json().then(function(d){ return {ok: r.ok, data: d}; }); })
    .then(function(res){
      btn.disabled = false;
      if (res.ok) {
        msg.className = 'modal-msg ok';
        msg.textContent = 'Saved! Restart the app for the new path to take effect.';
      } else {
        msg.className = 'modal-msg err';
        msg.textContent = res.data.error || 'Something went wrong.';
      }
    }).catch(function(){
      btn.disabled = false;
      msg.className = 'modal-msg err';
      msg.textContent = 'Network error — is the app running?';
    });
  }
</script>

</body>
</html>"""


# ── Planning view template ───────────────────────────────────────────────────

PLANNING_HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>S&amp;K Planning · Route Optimizer</title>
<style>
  :root {
    --bg: #f5f7fa;
    --surface: #ffffff;
    --surface-2: #f0f2f5;
    --border: #e1e5eb;
    --text: #1a202c;
    --text-muted: #64748b;
    --accent: #2563eb;
    --green: #16a34a;
    --green-bg: #dcfce7;
    --yellow: #ca8a04;
    --yellow-bg: #fef9c3;
    --orange: #ea580c;
    --orange-bg: #ffedd5;
    --red: #dc2626;
    --red-bg: #fee2e2;
    --pink: #db2777;
    --pink-bg: #fce7f3;
    --truck2: #2563eb;
    --truck9: #dc2626;
  }
  * { box-sizing: border-box; }
  body {
    margin: 0; font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
    background: var(--bg); color: var(--text); line-height: 1.4; font-size: 14px;
  }
  header {
    background: var(--surface); border-bottom: 1px solid var(--border);
    padding: 14px 24px; display: flex; justify-content: space-between; align-items: center;
    position: sticky; top: 0; z-index: 10;
  }
  h1 { margin: 0; font-size: 18px; font-weight: 600; }
  .nav a {
    color: var(--accent); text-decoration: none; margin-left: 16px;
    padding: 6px 12px; border: 1px solid var(--accent); border-radius: 6px;
    font-size: 13px; font-weight: 500;
  }
  .nav a:hover { background: var(--accent); color: white; }
  main { padding: 20px; max-width: 1500px; margin: 0 auto; }
  .grid {
    display: grid; grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
    gap: 12px; margin-bottom: 24px;
  }
  .card {
    background: var(--surface); border: 1px solid var(--border);
    border-radius: 10px; padding: 16px 18px;
  }
  .card h2 {
    margin: 0 0 10px; font-size: 11px; text-transform: uppercase;
    letter-spacing: 0.06em; color: var(--text-muted); font-weight: 600;
  }
  .card.section h2 { font-size: 13px; }
  .big-num { font-size: 26px; font-weight: 700; line-height: 1.1; }
  .label { color: var(--text-muted); font-size: 12px; margin-top: 4px; }
  /* day grid (calendar style) */
  .days-grid {
    display: grid; grid-template-columns: repeat(auto-fill, minmax(540px, 1fr));
    gap: 16px;
  }
  .day-card {
    background: var(--surface); border: 1px solid var(--border);
    border-radius: 10px; padding: 14px 18px;
  }
  .day-header {
    display: flex; justify-content: space-between; align-items: baseline;
    padding-bottom: 10px; margin-bottom: 12px; border-bottom: 1px solid var(--border);
  }
  .day-date { font-weight: 600; font-size: 16px; }
  .day-date small { font-weight: 400; color: var(--text-muted); margin-left: 6px; }
  .day-totals { color: var(--text-muted); font-size: 13px; }
  .day-totals strong { color: var(--text); }
  .truck-cols {
    display: grid; grid-template-columns: 1fr 1fr; gap: 10px;
  }
  .truck-col { min-width: 0; }
  .truck-head {
    font-size: 12px; font-weight: 600; padding: 6px 10px;
    border-radius: 6px; text-align: center; margin-bottom: 6px;
    color: white;
  }
  .truck-2 .truck-head { background: var(--truck2); }
  .truck-9 .truck-head { background: var(--truck9); }
  .stop {
    padding: 6px 10px; margin: 3px 0; border-radius: 6px; font-size: 12.5px;
    border-left: 3px solid var(--text-muted); background: var(--surface-2);
    display: flex; justify-content: space-between; gap: 8px;
  }
  .stop .name { flex: 1; min-width: 0; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
  .stop .qty { color: var(--text-muted); font-variant-numeric: tabular-nums; font-size: 11.5px; flex-shrink: 0; }
  .stop.pending {
    background: var(--yellow-bg); border-left-color: var(--yellow);
    font-weight: 500;
  }
  .stop.pending .qty { color: var(--yellow); font-weight: 600; }
  .stop.committed {
    background: #eef4ff; border-left-color: var(--accent);
  }
  .stop.tentative {
    background: var(--surface-2); border-left-color: #cbd5e1;
  }
  .badge {
    display: inline-block; padding: 1px 7px; border-radius: 999px;
    font-size: 10px; font-weight: 600; text-transform: uppercase;
    letter-spacing: 0.04em;
  }
  .badge-committed { background: #dbeafe; color: var(--accent); }
  .badge-tentative { background: #f1f5f9; color: var(--text-muted); }
  .badge-pending   { background: var(--yellow-bg); color: var(--yellow); }
  .badge-sensor    { background: var(--green-bg); color: var(--green); }
  .badge-stale     { background: var(--orange-bg); color: var(--orange); }
  table { width: 100%; border-collapse: collapse; margin-top: 6px; font-size: 13px; }
  th {
    text-align: left; color: var(--text-muted); font-weight: 500;
    border-bottom: 1px solid var(--border); padding: 6px 8px;
    font-size: 11px; text-transform: uppercase; letter-spacing: 0.04em;
  }
  td { padding: 6px 8px; border-bottom: 1px solid var(--border); }
  tr:hover td { background: var(--surface-2); }
  .empty { color: var(--text-muted); font-style: italic; padding: 16px; text-align: center; font-size: 13px; }
  .progress { width: 60px; height: 6px; background: var(--surface-2); border-radius: 3px; overflow: hidden; display: inline-block; vertical-align: middle; margin-left: 6px; }
  .progress-bar { height: 100%; transition: width 0.3s; }
  .pf-low  { background: var(--red); }
  .pf-mid  { background: var(--yellow); }
  .pf-high { background: var(--green); }
  .age-fresh { color: var(--green); }
  .age-stale { color: var(--orange); }
  .age-old   { color: var(--red); }
  .small { font-size: 12px; color: var(--text-muted); }
  .section-title {
    font-size: 13px; text-transform: uppercase; letter-spacing: 0.06em;
    color: var(--text-muted); font-weight: 600; margin: 0 0 12px;
  }
  code { font-family: ui-monospace, SFMono-Regular, Menlo, monospace; font-size: 12px; color: var(--text-muted); }
</style>
</head>
<body>
<header>
  <h1>📋 S&amp;K Planning</h1>
  <nav class="nav">
    <a href="/">← Dispatch</a>
  </nav>
</header>
<main>

<!-- Top metrics row -->
<div class="grid" id="metrics-row">
  <div class="card"><div class="label">Loading…</div></div>
</div>

<!-- Pending visits (placeholder qty=200) -->
<div class="card section" style="margin-bottom: 16px;">
  <h2>⌛ Pending visits — drivers visited, qty not yet invoiced</h2>
  <div id="pending-content"><div class="empty">Loading…</div></div>
</div>

<!-- Plan day-by-day -->
<div class="card section" style="margin-bottom: 16px;">
  <h2>📅 Current optimizer plan</h2>
  <div id="plan-content"><div class="empty">Loading…</div></div>
</div>

<!-- ANOVA live tank readings (currently SCAFFOLDED but not piped into the IRP) -->
<div class="card section" style="margin-bottom: 16px;">
  <h2>📡 ANOVA live tank readings <span style="color:var(--orange);font-size:11px;">· not in IRP yet</span></h2>
  <div class="small" style="margin-bottom:6px;">
    The receiver code is in place but real-time data isn't flowing into the
    optimizer. When the webhook is live and asset→client mapping is
    confirmed, run with <code>--enable-anova</code> to wire it up.
  </div>
  <div id="anova-content"><div class="empty">Loading…</div></div>
</div>

<!-- Fill-rate audit -->
<div class="card section">
  <h2>📉 Fill-rate audit (last 180 days)</h2>
  <div class="small" style="margin-bottom: 8px;">
    Bottom of list = candidates for less-frequent visits.
  </div>
  <div id="fill-content"><div class="empty">Loading…</div></div>
</div>

</main>

<script>
function pctBar(pct, w=80) {
  const cls = pct < 30 ? 'pf-low' : (pct < 70 ? 'pf-mid' : 'pf-high');
  return `<div class="progress" style="width:${w}px"><div class="progress-bar ${cls}" style="width:${Math.min(pct,100)}%"></div></div>`;
}

function shortName(name) {
  // 'POP - 16015 - POPOS BELL' -> 'POPOS BELL'
  const m = name.match(/^[^-]+-\\s*\\d+\\s*-\\s*(.+)$/);
  return m ? m[1].trim() : name;
}

function dayLabel(dateStr) {
  if (!dateStr) return '';
  const d = new Date(dateStr + 'T00:00:00');
  const today = new Date(); today.setHours(0,0,0,0);
  const tomorrow = new Date(today); tomorrow.setDate(today.getDate()+1);
  const yesterday = new Date(today); yesterday.setDate(today.getDate()-1);
  const fmt = d.toLocaleDateString('en-US', {weekday:'short', month:'short', day:'numeric'});
  if (d.getTime() === today.getTime()) return 'TODAY · ' + fmt;
  if (d.getTime() === tomorrow.getTime()) return 'TOMORROW · ' + fmt;
  if (d.getTime() === yesterday.getTime()) return 'Yesterday · ' + fmt;
  return fmt;
}

function loadAnova() {
  fetch('/api/anova-status').then(r => r.json()).then(d => {
    const el = document.getElementById('anova-content');
    if (!d.enabled) {
      el.innerHTML = `<div class="empty">${d.message || 'ANOVA integration not configured.'}</div>`;
      return;
    }
    const rows = d.readings.map(r => {
      const ageCls = r.age_hours < 6 ? 'age-fresh' : (r.age_hours < 24 ? 'age-stale' : 'age-old');
      const customer = r.customer || '<span style="color:var(--text-muted)">unmapped</span>';
      const pctCell = r.pct_full !== null
        ? `<td>${r.pct_full.toFixed(0)}% ${pctBar(r.pct_full)}</td>`
        : `<td class="small">—</td>`;
      return `<tr>
        <td><span class="badge badge-sensor">live</span> <code>${r.asset_id}</code></td>
        <td>${customer}</td>
        <td>${r.level_lbs.toFixed(0)} lbs</td>
        ${pctCell}
        <td class="${ageCls} small">${r.age_hours.toFixed(1)}h ago</td>
      </tr>`;
    }).join('');
    el.innerHTML = `
      <div class="small" style="margin-bottom: 8px;">
        ${d.fresh_count} of ${d.count} readings fresh (≤24h).
        Fresh readings override the IRP's inventory estimate for those clients.
      </div>
      <table><thead><tr>
        <th>Asset</th><th>Customer</th><th>Level</th><th>% Full</th><th>Age</th>
      </tr></thead><tbody>${rows}</tbody></table>`;
  });
}

function renderTruckColumn(visits, truckClass, truckLabel) {
  if (!visits || !visits.length) {
    return `<div class="truck-col ${truckClass}">
      <div class="truck-head">${truckLabel}</div>
      <div class="empty" style="padding:8px;font-size:11px;">no stops</div>
    </div>`;
  }
  const totalLbs = visits.reduce((a, v) => a + (v.refill_lbs || 0), 0);
  const stops = visits.map(v => {
    const isPending = v.refill_lbs === 200;
    const cls = isPending ? 'pending' : (v.status === 'COMMITTED' ? 'committed' : 'tentative');
    const qtyDisplay = isPending
      ? `<span class="badge badge-pending">pending</span>`
      : `${(v.refill_lbs || 0).toFixed(0)} lbs`;
    const arrivalDisplay = v.arrival_hhmm
      ? `<span style="color:var(--text-muted);font-variant-numeric:tabular-nums;">${v.arrival_hhmm}</span> · `
      : '';
    return `<div class="stop ${cls}" title="${v.customer} — arrives ${v.arrival_hhmm || '?'}, departs ${v.depart_hhmm || '?'}">
      <span class="name">${arrivalDisplay}<strong>${v.client_id}</strong> ${shortName(v.customer)}</span>
      <span class="qty">${qtyDisplay}</span>
    </div>`;
  }).join('');
  return `<div class="truck-col ${truckClass}">
    <div class="truck-head">${truckLabel} · ${visits.length} stops · ${totalLbs.toLocaleString()} lbs</div>
    ${stops}
  </div>`;
}

function loadPending() {
  fetch('/api/pending-visits').then(r => r.json()).then(d => {
    const el = document.getElementById('pending-content');
    if (d.error) {
      el.innerHTML = `<div class="empty">${d.error}</div>`;
      return;
    }
    if (!d.count) {
      el.innerHTML = `<div class="empty">No pending visits — all delivered quantities have been invoiced. ✓</div>`;
      return;
    }
    const dayCards = d.days.map(day => {
      const stops = day.visits.map(v =>
        `<div class="stop pending" title="${v.customer}">
          <span class="name"><strong>${v.client_id}</strong> ${shortName(v.customer)} <small style="color:var(--text-muted)">${v.zone||''}</small></span>
          <span class="qty"><span class="badge badge-pending">pending</span></span>
        </div>`
      ).join('');
      return `<div class="day-card" style="padding:10px 14px;">
        <div class="day-header" style="margin-bottom:8px;padding-bottom:6px;">
          <div class="day-date">${dayLabel(day.date)}</div>
          <div class="day-totals">${day.n_visits} pending</div>
        </div>
        ${stops}
      </div>`;
    }).join('');
    el.innerHTML = `
      <div class="small" style="margin-bottom: 10px;">
        ${d.count} visit(s) waiting for invoiced quantity. Run
        <code>python ingest_actuals.py</code> after the next invoice export
        to fill these in automatically.
      </div>
      <div class="days-grid">${dayCards}</div>`;
  });
}

function loadPlan() {
  fetch('/api/plan-summary').then(r => r.json()).then(d => {
    const el = document.getElementById('plan-content');
    if (!d.available) {
      el.innerHTML = `<div class="empty">${d.message || 'No plan available. Run the optimizer to generate one.'}</div>`;
      document.getElementById('metrics-row').innerHTML = `
        <div class="card"><h2>Plan</h2><div class="big-num">—</div><div class="label">No plan yet</div></div>
      `;
      return;
    }

    // Metrics row
    const totalLbs = d.days.reduce((acc, day) => acc + day.total_lbs, 0);
    const stability = d.metadata.plan_stability_vs_prior;
    document.getElementById('metrics-row').innerHTML = `
      <div class="card">
        <h2>Total visits</h2>
        <div class="big-num">${d.total_visits}</div>
        <div class="label">over ${d.days.length} day(s)</div>
      </div>
      <div class="card">
        <h2>Total lbs</h2>
        <div class="big-num">${totalLbs.toLocaleString()}</div>
        <div class="label">scheduled refills</div>
      </div>
      <div class="card">
        <h2>Solve time</h2>
        <div class="big-num">${(d.metadata.solve_seconds || '?')}<span style="font-size:18px;color:var(--text-muted);"> s</span></div>
        <div class="label">${d.solved_at ? new Date(d.solved_at).toLocaleString() : ''}</div>
      </div>
      <div class="card">
        <h2>Plan stability</h2>
        <div class="big-num">${stability !== undefined && stability !== null
          ? (stability * 100).toFixed(0) + '<span style="font-size:18px;color:var(--text-muted);"> %</span>'
          : '—'}</div>
        <div class="label">vs yesterday's day-1</div>
      </div>
    `;

    if (d.days.length === 0) {
      el.innerHTML = `<div class="empty">Plan has no visits.</div>`;
      return;
    }

    const dayCards = d.days.map((day) => {
      const isCommitted = day.day_index < (d.commit_days || 1);
      const truck2 = day.visits.filter(v => v.truck && v.truck.includes('2'));
      const truck9 = day.visits.filter(v => v.truck && v.truck.includes('9'));
      return `<div class="day-card">
        <div class="day-header">
          <div class="day-date">${dayLabel(day.date)}
            <small><span class="badge ${isCommitted ? 'badge-committed' : 'badge-tentative'}">
              ${isCommitted ? 'committed' : 'tentative'}
            </span></small>
          </div>
          <div class="day-totals">
            <strong>${day.n_visits}</strong> stops · <strong>${day.total_lbs.toLocaleString()}</strong> lbs
          </div>
        </div>
        <div class="truck-cols">
          ${renderTruckColumn(truck2, 'truck-2', 'Truck 2')}
          ${renderTruckColumn(truck9, 'truck-9', 'Truck 9')}
        </div>
      </div>`;
    }).join('');
    el.innerHTML = `<div class="days-grid">${dayCards}</div>`;
  });
}

function loadFill() {
  fetch('/api/fill-audit').then(r => r.json()).then(d => {
    const el = document.getElementById('fill-content');
    if (d.error) {
      el.innerHTML = `<div class="empty">Error: ${d.error}</div>`;
      return;
    }
    const rows = d.rows.slice(0, 15).map(r => {
      const pct = (r.mean_fill_pct * 100).toFixed(0);
      const pctClass = r.mean_fill_pct < 0.5 ? 'pf-low' : (r.mean_fill_pct < 0.75 ? 'pf-mid' : 'pf-high');
      return `<tr>
        <td><code>${r.ID}</code></td>
        <td>${shortName(r.Customer)}</td>
        <td>${r.n_deliveries}</td>
        <td>${pct}% <div class="progress"><div class="progress-bar ${pctClass}" style="width:${Math.min(pct,100)}%"></div></div></td>
        <td>${r.low_fill_count}</td>
      </tr>`;
    }).join('');
    el.innerHTML = `
      <div class="small" style="margin-bottom: 8px;">
        Global mean fill: <strong style="color:var(--text)">${(d.global_mean_fill_pct * 100).toFixed(0)}%</strong>.
        ${d.low_fill_clients} client(s) average below 50% — may be over-served.
      </div>
      <table><thead><tr>
        <th>ID</th><th>Customer</th><th>Visits</th><th>Avg fill</th><th>Low fills</th>
      </tr></thead><tbody>${rows}</tbody></table>`;
  });
}

loadPending(); loadPlan(); loadAnova(); loadFill();
setInterval(loadAnova, 60000);
</script>
</body>
</html>"""


# ── Entrypoint ────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    port = 5050

    def _open_browser():
        time.sleep(1.8)
        webbrowser.open(f'http://localhost:{port}')

    threading.Thread(target=_open_browser, daemon=True).start()

    print(f"\n  S&K Route Optimizer\n  → http://localhost:{port}\n")
    app.run(host='localhost', port=port, debug=False, threaded=True)
