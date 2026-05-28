"""
final.app.server — dashboard for the FINAL route optimizer.

Run:  python -m final.app
Open: http://localhost:5050

Endpoints
---------
GET  /                       dashboard HTML
GET  /api/health             excel mtime, deliveries, anova coverage, etc.
GET  /api/settings           current input file path
POST /api/settings           change input file (writes local_config.json)
POST /api/pick-file          opens native OS file picker, returns path
GET  /api/clients            client list with SOLVER-CORRECT rates + state
GET  /api/last-plan          last archived plan summary (committed/preview split)
GET  /api/overrides          current dashboard Pins/Forbids
POST /api/overrides          add/clear a Pin or Forbid (or all for a client)
GET  /api/truck-availability list of (date, truck_id) unavailable
POST /api/truck-availability replace the unavailability list
POST /api/run                kick off solver subprocess
GET  /api/run/status         polled by frontend; returns status + log tail
GET  /outputs/<filename>     serve generated plan files
"""
from __future__ import annotations

import json
import os
import platform
import subprocess
import sys
import threading
import time
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl
import pandas as pd
from flask import Flask, Response, jsonify, request, send_file, render_template

from final.app.overrides_store import load_user_overrides, save_user_overrides
from final.app.availability_store import load_unavailability, save_unavailability
from final.sk_solver_final import estimate_consumption_recency_weighted
from v2.ingest.excel import load_clients, load_deliveries

# ── Paths ────────────────────────────────────────────────────────────────────
HERE = Path(__file__).resolve().parent
APP_DIR = HERE
REPO = HERE.parent.parent
DATA_DIR = REPO / 'data'
OUTPUT_DIR = REPO / 'final' / 'output'
USER_OVERRIDES = DATA_DIR / 'user_overrides.json'
TRUCK_UNAVAIL = DATA_DIR / 'truck_unavailable.json'
LOCAL_CONFIG = REPO / 'local_config.json'

# ── Flask app ────────────────────────────────────────────────────────────────
app = Flask(__name__,
             template_folder=str(APP_DIR / 'templates'),
             static_folder=str(APP_DIR / 'static'))

# ── Run state ────────────────────────────────────────────────────────────────
_run_lock = threading.Lock()
_current_run: Dict[str, object] = {
    'status': 'idle', 'started_at': None, 'finished_at': None,
    'log': [], 'date': None, 'returncode': None, 'elapsed_s': None,
}


# ════════════════════════════════════════════════════════════════════════════
# SETTINGS  (local_config.json — which Excel file we're reading)
# ════════════════════════════════════════════════════════════════════════════

def _read_input_file() -> Optional[Path]:
    if LOCAL_CONFIG.exists():
        try:
            cfg = json.loads(LOCAL_CONFIG.read_text(encoding='utf-8'))
            p = Path(cfg.get('input_file') or '')
            return p if p.exists() else None
        except Exception:
            return None
    return None


def _write_input_file(path: Path) -> None:
    cfg = {}
    if LOCAL_CONFIG.exists():
        try:
            cfg = json.loads(LOCAL_CONFIG.read_text(encoding='utf-8'))
        except Exception:
            cfg = {}
    cfg['input_file'] = str(path)
    LOCAL_CONFIG.write_text(json.dumps(cfg, indent=2), encoding='utf-8')


# ════════════════════════════════════════════════════════════════════════════
# SAFE CELL PARSING
# ════════════════════════════════════════════════════════════════════════════
# Excel cells may contain error strings ("#VALUE!", "#REF!", "#NAME?",
# "#DIV/0!", "#N/A") when a formula breaks — e.g. while the operator is
# editing the M6 staleness formula. A raw float() on those crashes the
# /api/health endpoint with a 500 (which then renders as HTML and the
# frontend chokes parsing "<!doctype" as JSON). Funnel every cell-to-float
# through this helper so a bad cell becomes None, not an exception.

def _safe_float(v) -> Optional[float]:
    if v is None or v == '':
        return None
    if isinstance(v, bool):
        return float(v)
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s or s.startswith('#'):  # #VALUE! #REF! #NAME? #DIV/0! #N/A
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


# ════════════════════════════════════════════════════════════════════════════
# DATA FRESHNESS GUARDS
# ════════════════════════════════════════════════════════════════════════════

def _data_freshness_warnings(f: Path) -> List[Dict]:
    """Return a list of structured warnings about data freshness/integrity.

    Each warning is {severity, code, message}. Severity is one of:
      'block'  — solver should refuse to run
      'warn'   — display prominently, allow override
      'info'   — informational only

    Checks:
      1. Excel lock file (~$filename.xlsx) exists → file being edited NOW.
         Reading it WILL produce inconsistent data because Excel's in-memory
         state is not on disk. SOLVER MUST NOT RUN.
      2. File mtime ≥ 24h ago → operator may have forgotten to save or sync.
         Could be stale. Warn loudly.
      3. File mtime < 60 seconds ago → was just saved; may be mid-sync.
         Brief info hint.
    """
    warnings: List[Dict] = []
    if f is None or not f.exists():
        return warnings

    # 1) Excel lock file — definitive "file is open and being edited"
    lockfile = f.parent / f'~${f.name}'
    if lockfile.exists():
        warnings.append({
            'severity': 'block',
            'code': 'EXCEL_LOCK',
            'message': (f'Excel has the file open with unsaved changes '
                        f'({lockfile.name}). Save in Excel (Cmd-S) and wait for '
                        f'OneDrive to sync (✓ icon in menu bar) before running.'),
        })

    # 2) File age
    age_hours = (datetime.now() - datetime.fromtimestamp(f.stat().st_mtime)).total_seconds() / 3600.0
    if age_hours > 24:
        sev = 'warn' if age_hours < 72 else 'block'
        warnings.append({
            'severity': sev,
            'code': 'STALE_FILE',
            'message': (f'Data file was last saved {age_hours/24:.1f} days ago. '
                        f'Anova readings and delivery logs may be outdated. '
                        f'Open in Excel, check for unsaved edits, and Save.'),
        })
    elif age_hours < 0.017:   # < 1 minute
        warnings.append({
            'severity': 'info',
            'code': 'JUST_SAVED',
            'message': 'File just saved — give OneDrive a moment to fully sync before running.',
        })
    return warnings


def _stale_client_warnings(f: Path) -> List[Dict]:
    """Find clients whose delivery log looks incomplete — risky to use them
    in scheduling because the est_current may be wildly wrong.

    A delivery missing from the log → spreadsheet thinks customer was last
    served weeks ago → formula clamps current_lbs to 0 → solver schedules
    a full-tank refill → driver arrives with way too much oil → overflow.

    Heuristic: client has active rate (>0) AND est_current ≤ 5% of tank AND
    days_since_last > 14 AND has no fresh Anova reading. These are the
    "likely missing delivery" candidates.
    """
    warnings: List[Dict] = []
    try:
        wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
    except Exception:
        return warnings

    flagged: List[Tuple[str, str, int, int]] = []
    if 'Optimizer_Input' in wb.sheetnames:
        ws = wb['Optimizer_Input']
        # Cols: 2=id 3=name 7=tank 8=rate 11=days_since 12=est_current
        #       19=anova_lvl 20=anova_age
        for row in ws.iter_rows(min_row=6, max_col=20, values_only=True):
            if not row or row[1] is None: continue
            cid = str(row[1])
            name = str(row[2] or '')[:40]
            # Column shifts after Std Dev added at col 9 — everything
            # after col 8 moved +1. AVG (col 8 / idx 7) stayed put.
            # _safe_float() makes us robust to #VALUE!/#REF! cells.
            tank = _safe_float(row[6]) or 0
            rate = _safe_float(row[7]) or 0           # AVG rate
            days_since_raw = row[11] if len(row) > 11 else None
            days_since = _safe_float(days_since_raw)
            est_cur = _safe_float(row[12]) if len(row) > 12 else None
            anova_lvl = row[19] if len(row) > 19 else None
            anova_age = _safe_float(row[20]) if len(row) > 20 else None
            anova_fresh = (anova_lvl is not None and
                            anova_age is not None and
                            abs(anova_age) < 48)

            if rate <= 0 or tank <= 0 or est_cur is None: continue
            if days_since is None: continue
            if anova_fresh: continue   # Anova overrides any concern

            if days_since > 14 and est_cur <= 0.05 * tank:
                flagged.append((cid, name, int(days_since), int(est_cur)))
    wb.close()

    if flagged:
        flagged.sort(key=lambda x: -x[2])  # most-days-stale first
        sample = ', '.join(f'{c[0]}({c[2]}d)' for c in flagged[:6])
        more = f' (+{len(flagged)-6} more)' if len(flagged) > 6 else ''
        warnings.append({
            'severity': 'warn',
            'code': 'STALE_CLIENTS',
            'count': len(flagged),
            'sample': flagged[:10],
            'message': (
                f'{len(flagged)} client(s) show as nearly empty with last delivery '
                f'>14 days ago and no Anova sensor. The estimate may be wrong '
                f'(missing delivery in log). Verify before dispatching. '
                f'Sample: {sample}{more}'
            ),
        })
    return warnings


# ════════════════════════════════════════════════════════════════════════════
# DATA HEALTH
# ════════════════════════════════════════════════════════════════════════════

def _data_health() -> Dict:
    f = _read_input_file()
    if f is None:
        return {'ok': False, 'error': 'No input file selected. Click Browse to pick your SK_Delivery_System.xlsx'}
    stat = f.stat()
    mtime = datetime.fromtimestamp(stat.st_mtime)
    age_hours = (datetime.now() - mtime).total_seconds() / 3600.0

    try:
        wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
    except Exception as e:
        return {'ok': False, 'error': f'Could not open Excel: {e}'}

    out = {
        'ok': True, 'path': str(f),
        'mtime': mtime.isoformat(timespec='seconds'),
        'age_hours': round(age_hours, 1),
        'size_kb': round(stat.st_size / 1024, 0),
    }

    if 'Delivery_Log' in wb.sheetnames:
        ws = wb['Delivery_Log']
        n_deliv = 0
        last_date = None
        for i, row in enumerate(ws.iter_rows(min_row=4, max_col=2, values_only=True)):
            d = row[0] if row else None
            if isinstance(d, datetime):
                n_deliv += 1
                if last_date is None or d > last_date:
                    last_date = d
            if i > 6000:
                break
        out['delivery_count'] = n_deliv
        out['last_delivery'] = last_date.date().isoformat() if last_date else None
        if last_date:
            out['last_delivery_days_ago'] = (datetime.now().date() - last_date.date()).days

    if 'Anova_Live' in wb.sheetnames:
        ws = wb['Anova_Live']
        total = 0; with_reading = 0; max_age = 0.0; min_age = 1e9
        for row in ws.iter_rows(min_row=2, max_col=13, values_only=True):
            cid = row[1] if len(row) > 1 else None
            if cid is None: continue
            total += 1
            lvl = row[4] if len(row) > 4 else None
            age = _safe_float(row[12]) if len(row) > 12 else None
            if lvl is not None and lvl != '':
                with_reading += 1
                a = age if age is not None else 999
                max_age = max(max_age, a); min_age = min(min_age, a)
        out['anova_total'] = total
        out['anova_with_reading'] = with_reading
        out['anova_pct'] = round(100 * with_reading / max(total, 1), 0)
        out['anova_oldest_hours'] = round(max_age, 1) if with_reading > 0 else None
        out['anova_freshest_hours'] = round(min_age, 1) if with_reading > 0 else None

    if 'Client_List' in wb.sheetnames:
        ws = wb['Client_List']
        n_clients = 0; n_dns = 0
        for row in ws.iter_rows(min_row=4, max_col=17, values_only=True):
            cid = row[0] if row else None
            if cid is None: continue
            n_clients += 1
            dns_cell = row[16] if len(row) > 16 else None
            if str(dns_cell or '').strip().upper() in ('Y', 'YES', 'TRUE', '1'):
                n_dns += 1
        out['client_count'] = n_clients
        out['client_dns'] = n_dns

    wb.close()

    # Attach freshness warnings (blocks Run when severe).
    warnings = _data_freshness_warnings(f)
    warnings.extend(_stale_client_warnings(f))
    out['warnings'] = warnings
    out['has_block'] = any(w['severity'] == 'block' for w in warnings)
    return out


# ════════════════════════════════════════════════════════════════════════════
# CLIENT LIST — uses SOLVER's recency-weighted rate (IQR-filtered)
# ════════════════════════════════════════════════════════════════════════════

def _client_list() -> List[Dict]:
    """Build a list of clients with SOLVER-CORRECT rates + current state."""
    f = _read_input_file()
    if f is None:
        return []

    # Rates come from the spreadsheet's AVG column directly — same as the
    # solver uses. No need for our own recency estimator anymore (operator
    # curates AVG + StdDev in the workbook).
    try:
        clients = load_clients(f)
    except Exception as e:
        print(f"WARN: could not load clients: {e}")
        clients = ()

    # Read Optimizer_Input for state (current_lbs, last delivery, rate, std).
    try:
        wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
    except Exception:
        return []

    # Build id → Client lookup for tank cap
    by_id = {c.id: c for c in clients}

    # DNS flag lives in Client_List col Q (17) — read once, pass through.
    dns_flagged: Dict[str, str] = {}   # cid → reason text (if any)
    if 'Client_List' in wb.sheetnames:
        cl = wb['Client_List']
        for crow in cl.iter_rows(min_row=4, max_col=17, values_only=True):
            if not crow or crow[0] is None: continue
            dns_val = crow[16] if len(crow) > 16 else None
            if str(dns_val or '').strip().upper() in ('Y', 'YES', 'TRUE', '1'):
                notes = crow[15] if len(crow) > 15 else None
                dns_flagged[str(crow[0])] = str(notes or 'Do not schedule')

    rows: List[Dict] = []
    ws = wb['Optimizer_Input']
    # Column layout (operator added StdDev at col 9, shifted others +1):
    #   col  8 / idx 7  = AVG rate
    #   col  9 / idx 8  = Std Dev (new)
    #   col 11 / idx 10 = Last Delivery
    #   col 13 / idx 12 = Est. Current
    #   col 20 / idx 19 = Anova Level
    #   col 21 / idx 20 = Anova Age
    for row in ws.iter_rows(min_row=6, max_col=22, values_only=True):
        if not row or row[1] is None:
            continue
        cid = str(row[1])
        # Use _safe_float so #VALUE!/#REF! cells degrade to None instead of crashing.
        spreadsheet_rate = _safe_float(row[7])               # AVG
        spreadsheet_std = _safe_float(row[8]) if len(row) > 8 else None
        last_deliv = row[10] if len(row) > 10 else None
        est_cur = _safe_float(row[12]) if len(row) > 12 else None
        anova_lvl = row[19] if len(row) > 19 else None
        anova_age = _safe_float(row[20]) if len(row) > 20 else None

        # Use the spreadsheet's AVG rate (which the solver also uses now).
        # No more IQR-filtered fallback — operator-curated AVG is canonical.
        rate = round(spreadsheet_rate, 1) if spreadsheet_rate else None
        std = round(spreadsheet_std, 1) if spreadsheet_std else None

        # Tank from clients tuple (master), fallback to Optimizer_Input
        client = by_id.get(cid)
        tank_raw = _safe_float(row[6]) or 0
        tank = float(client.tank_capacity_lbs) if client else tank_raw
        name = client.customer if client else (row[2] or '')

        current_lbs = est_cur or 0.0
        pct = round(100 * current_lbs / tank, 0) if tank > 0 else 0

        # DTE = current / rate. Spreadsheet's DTE column (now col 15 / idx 14)
        # is the same calc; use ours so it stays in sync if Anova adjusts current.
        if rate and rate > 0:
            dte = round(current_lbs / rate, 1)
        else:
            sheet_dte = _safe_float(row[14]) if len(row) > 14 else None
            dte = round(sheet_dte, 1) if sheet_dte is not None else None

        rows.append({
            'id': cid,
            'name': str(name),
            'tank_lbs': tank,
            'rate_lpd': rate,
            'rate_std_dev': std,
            'current_lbs': round(current_lbs, 1),
            'pct_full': pct,
            'dte': dte,
            'last_delivery': str(last_deliv)[:10] if last_deliv else None,
            'has_anova': anova_lvl is not None,
            'anova_age_h': round(anova_age, 1) if anova_age is not None else None,
            'dns': cid in dns_flagged,
            'dns_reason': dns_flagged.get(cid, ''),
        })
    wb.close()
    rows.sort(key=lambda r: r['dte'] if r['dte'] is not None else 999)
    return rows


# ════════════════════════════════════════════════════════════════════════════
# LAST PLAN
# ════════════════════════════════════════════════════════════════════════════

def _last_plan() -> Optional[Dict]:
    archive_dir = OUTPUT_DIR / 'archive'
    if not archive_dir.exists():
        return None
    files = sorted(archive_dir.glob('plan_*.json'), key=lambda p: p.stat().st_mtime)
    if not files: return None
    try:
        plan = json.loads(files[-1].read_text(encoding='utf-8'))
    except Exception:
        return None

    today_iso = plan.get('today', '')
    today = date.fromisoformat(today_iso) if today_iso else date.today()
    horizon = [date.fromisoformat(d) for d in plan.get('horizon_dates', [])]
    commit_days = int(plan.get('commit_days', 2))
    committed_dates = set(horizon[:commit_days])

    # The JSON archive nests each route under a 'route' sub-key (the dict
    # was serialized as [{date, truck_id, route: {…full route…}}, ...] to
    # avoid tuple-keyed dicts). Unwrap to a flat shape so we can read
    # total_load_lbs / stops without surprises.
    raw_routes = plan.get('routes', [])
    flat_routes: List[Dict] = []
    if isinstance(raw_routes, list):
        for r in raw_routes:
            route_data = r.get('route', r) if isinstance(r, dict) else {}
            flat_routes.append({
                'date': r.get('date'),
                'truck_id': r.get('truck_id'),
                **route_data,
            })
    elif isinstance(raw_routes, dict):
        for _k, r in raw_routes.items():
            if 'route' in r and isinstance(r['route'], dict):
                flat_routes.append({**r['route'], 'date': r.get('date'), 'truck_id': r.get('truck_id')})
            else:
                flat_routes.append(r)

    by_date: Dict[str, List[Dict]] = {}
    for r in flat_routes:
        by_date.setdefault(str(r.get('date', '')), []).append(r)

    committed: List[Dict] = []
    preview: List[Dict] = []
    for d_iso in sorted(by_date.keys()):
        day_routes = by_date[d_iso]
        d_obj = date.fromisoformat(d_iso) if len(d_iso) == 10 else None
        is_committed = d_obj in committed_dates if d_obj else False
        summary = {
            'date': d_iso,
            'date_label': d_obj.strftime('%a %b %d') if d_obj else d_iso,
            'committed': is_committed,
            'trucks': [
                {
                    'truck_id': r.get('truck_id'),
                    'stops': len(r.get('stops') or []),
                    'lbs': float(r.get('total_load_lbs') or 0),
                    'cap_pct': float(r.get('cap_pct') or 0),
                    'minutes': int(r.get('total_minutes') or 0),
                    'miles': float(r.get('total_miles') or 0),
                }
                for r in day_routes
            ],
        }
        (committed if is_committed else preview).append(summary)

    # List output files that ACTUALLY exist for this plan, so the frontend
    # doesn't construct guess-URLs that 404.
    files = {}
    if OUTPUT_DIR.exists():
        excel = OUTPUT_DIR / f'plan_{today_iso}.xlsx'
        if excel.exists():
            files['excel'] = excel.name
        map_file = OUTPUT_DIR / f'route_map_{today_iso}.html'
        if map_file.exists():
            files['map'] = map_file.name
        # SmartService CSV is named after first delivery date in plan
        # (typically today+1, but find it by glob to be safe).
        csv_candidates = sorted(OUTPUT_DIR.glob('smartservice_*.csv'),
                                 key=lambda p: p.stat().st_mtime)
        if csv_candidates:
            files['csv'] = csv_candidates[-1].name

    return {
        'today': today_iso,
        'horizon_dates': [d.isoformat() for d in horizon],
        'commit_days': commit_days,
        'horizon_days': len(horizon),
        'committed': committed,
        'preview': preview,
        'total_stops': plan.get('total_stops'),
        'total_lbs': plan.get('total_lbs_delivered'),
        'total_miles': plan.get('total_miles'),
        'avg_fill': plan.get('avg_fill_pct'),
        'objective_dollars': plan.get('objective_cost_dollars'),
        'solve_seconds': plan.get('solve_seconds'),
        'generated_at': plan.get('generated_at'),
        'files': files,
    }


# ════════════════════════════════════════════════════════════════════════════
# WORKDAYS — used by truck-availability widget
# ════════════════════════════════════════════════════════════════════════════

def _upcoming_workdays(start: date, count: int = 10) -> List[date]:
    """Tue–Sat working days starting from `start`."""
    _DOW = ('Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun')
    work_set = {'Tue', 'Wed', 'Thu', 'Fri', 'Sat'}
    out: List[date] = []
    cursor = start
    for _ in range(60):
        if len(out) >= count:
            break
        if _DOW[cursor.weekday()] in work_set:
            out.append(cursor)
        cursor += timedelta(days=1)
    return out


# ════════════════════════════════════════════════════════════════════════════
# ROUTES — UI
# ════════════════════════════════════════════════════════════════════════════

@app.route('/')
def index():
    default_date = (date.today() + timedelta(days=1)).isoformat()
    return render_template('index.html', default_date=default_date)


# ════════════════════════════════════════════════════════════════════════════
# ROUTES — JSON API
# ════════════════════════════════════════════════════════════════════════════

@app.route('/api/health')
def api_health():
    # Wrap in a try/except so any unexpected error still returns JSON.
    # The frontend parses this response with r.json(); an HTML 500 page
    # would throw "Unexpected token '<'" and the strip would show nothing.
    try:
        return jsonify(_data_health())
    except Exception as e:
        return jsonify({'ok': False, 'error': f'health check failed: {e}'}), 200


@app.route('/api/settings', methods=['GET', 'POST'])
def api_settings():
    if request.method == 'GET':
        return jsonify({'input_file': str(_read_input_file()) if _read_input_file() else None})
    body = request.get_json(force=True) or {}
    new_path = body.get('input_file')
    if not new_path:
        return jsonify({'ok': False, 'error': 'missing input_file'}), 400
    p = Path(new_path)
    if not p.exists() or not p.is_file():
        return jsonify({'ok': False, 'error': f'File not found: {new_path}'}), 400
    if p.suffix.lower() != '.xlsx':
        return jsonify({'ok': False, 'error': 'Pick a .xlsx file'}), 400
    _write_input_file(p)
    return jsonify({'ok': True, 'input_file': str(p)})


# ── Solver tuning settings (min_fill_pct etc.) ─────────────────────────────
# Persisted to local_config.json under `solver_settings`, picked up by
# sk_solver_final.py at every run. Defaults match the in-code constants.
_SOLVER_DEFAULTS = {
    'min_fill_pct': 0.50,
}


def _read_solver_settings() -> Dict:
    if not LOCAL_CONFIG.exists():
        return dict(_SOLVER_DEFAULTS)
    try:
        cfg = json.loads(LOCAL_CONFIG.read_text(encoding='utf-8'))
    except Exception:
        return dict(_SOLVER_DEFAULTS)
    user_settings = cfg.get('solver_settings') or {}
    out = dict(_SOLVER_DEFAULTS)
    out.update({k: v for k, v in user_settings.items() if k in _SOLVER_DEFAULTS})
    return out


def _write_solver_settings(updates: Dict) -> Dict:
    cfg = {}
    if LOCAL_CONFIG.exists():
        try:
            cfg = json.loads(LOCAL_CONFIG.read_text(encoding='utf-8'))
        except Exception:
            cfg = {}
    cur = cfg.get('solver_settings') or {}
    cur.update(updates)
    cfg['solver_settings'] = cur
    LOCAL_CONFIG.write_text(json.dumps(cfg, indent=2), encoding='utf-8')
    return _read_solver_settings()


@app.route('/api/settings/solver', methods=['GET', 'POST'])
def api_settings_solver():
    if request.method == 'GET':
        return jsonify(_read_solver_settings())
    body = request.get_json(force=True) or {}
    updates: Dict = {}
    # Validate each field with its own bounds
    if 'min_fill_pct' in body:
        try:
            v = float(body['min_fill_pct'])
        except (TypeError, ValueError):
            return jsonify({'ok': False, 'error': 'min_fill_pct must be a number'}), 400
        if not (0.0 <= v <= 0.95):
            return jsonify({'ok': False, 'error': 'min_fill_pct must be between 0 and 0.95'}), 400
        updates['min_fill_pct'] = v
    if not updates:
        return jsonify({'ok': False, 'error': 'no recognized settings provided'}), 400
    new_state = _write_solver_settings(updates)
    return jsonify({'ok': True, **new_state})


@app.route('/api/pick-file', methods=['POST'])
def api_pick_file():
    """Open native OS file picker. Returns selected path."""
    current = _read_input_file()
    initial_dir = str(current.parent) if current else str(Path.home() / 'Downloads')
    sysname = platform.system()
    try:
        if sysname == 'Darwin':
            script = (
                'POSIX path of (choose file '
                'with prompt "Select your SK_Delivery_System.xlsx file" '
                'of type {"xlsx", "org.openxmlformats.spreadsheetml.sheet"} '
                f'default location POSIX file "{initial_dir}")'
            )
            proc = subprocess.run(
                ['osascript', '-e', script],
                capture_output=True, text=True, timeout=600,
            )
            if proc.returncode == 0 and proc.stdout.strip():
                p = proc.stdout.strip()
                _write_input_file(Path(p))
                return jsonify({'ok': True, 'path': p})
            if 'cancel' in (proc.stderr or '').lower():
                return jsonify({'ok': False, 'cancelled': True})
            return jsonify({'ok': False, 'error': proc.stderr.strip() or 'Picker failed'})

        elif sysname == 'Windows':
            ps = (
                'Add-Type -AssemblyName System.Windows.Forms; '
                '$d = New-Object System.Windows.Forms.OpenFileDialog; '
                '$d.Filter = "Excel files (*.xlsx)|*.xlsx"; '
                f'$d.InitialDirectory = "{initial_dir}"; '
                '$d.Title = "Select your SK_Delivery_System.xlsx file"; '
                'if ($d.ShowDialog() -eq "OK") { Write-Output $d.FileName }'
            )
            proc = subprocess.run(
                ['powershell', '-NoProfile', '-Command', ps],
                capture_output=True, text=True, timeout=600,
            )
            path = (proc.stdout or '').strip()
            if path:
                _write_input_file(Path(path))
                return jsonify({'ok': True, 'path': path})
            return jsonify({'ok': False, 'cancelled': True})

        else:
            for cmd in (
                ['zenity', '--file-selection',
                 '--title=Select your SK_Delivery_System.xlsx file',
                 '--file-filter=Excel files | *.xlsx',
                 f'--filename={initial_dir}/'],
                ['kdialog', '--getopenfilename', initial_dir, '*.xlsx'],
            ):
                try:
                    proc = subprocess.run(cmd, capture_output=True, text=True, timeout=600)
                    if proc.returncode == 0 and proc.stdout.strip():
                        p = proc.stdout.strip()
                        _write_input_file(Path(p))
                        return jsonify({'ok': True, 'path': p})
                except FileNotFoundError:
                    continue
            return jsonify({'ok': False, 'error': 'No GUI file picker available (install zenity or kdialog)'})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/clients')
def api_clients():
    overrides = load_user_overrides(USER_OVERRIDES)
    pin_ids = {p.client_id for p in overrides.pins}
    forbid_ids = {f.client_id for f in overrides.forbids}
    rows = _client_list()
    for r in rows:
        r['pinned'] = r['id'] in pin_ids
        r['forbidden'] = r['id'] in forbid_ids
    return jsonify({'clients': rows, 'count': len(rows)})


@app.route('/api/last-plan')
def api_last_plan():
    plan = _last_plan()
    if plan is None:
        return jsonify({'ok': False, 'error': 'No plan yet — click RUN'})
    plan['ok'] = True
    return jsonify(plan)


# ════════════════════════════════════════════════════════════════════════════
# URGENCY LOCK MATRIX — operator's "did the solver do the right thing?" view
# ════════════════════════════════════════════════════════════════════════════
#
# Joins the latest plan archive's `pre_run_urgency` snapshot (the operator's
# coloring AT RUN TIME — never re-calculates) with the routes (where the
# solver actually scheduled each client). The result is a per-client row:
#   - frozen urgency bucket
#   - the day-index the solver picked (or null = deferred)
#   - which truck, refill lbs
#   - deferred reason if not served
#
# Sorted RED → YEL → GRN → GRY. Within each bucket, by DTE ascending.

@app.route('/api/urgency-matrix')
def api_urgency_matrix():
    archive_dir = OUTPUT_DIR / 'archive'
    if not archive_dir.exists():
        return jsonify({'ok': False, 'error': 'No plan archive yet — click RUN'})
    files = sorted(archive_dir.glob('plan_*.json'), key=lambda p: p.stat().st_mtime)
    if not files:
        return jsonify({'ok': False, 'error': 'No plan archive yet — click RUN'})

    try:
        plan = json.loads(files[-1].read_text(encoding='utf-8'))
    except Exception as e:
        return jsonify({'ok': False, 'error': f'Could not read archive: {e}'})

    snapshot = plan.get('pre_run_urgency') or []
    if not snapshot:
        return jsonify({
            'ok': False,
            'error': ('This plan was produced before the urgency snapshot '
                       'feature was added. Run again to see the matrix.'),
        })

    today_iso = plan.get('today', '')
    horizon_iso = plan.get('horizon_dates', [])
    horizon_dates = [date.fromisoformat(d) for d in horizon_iso]
    commit_days = int(plan.get('commit_days', 2))

    # Build cid → (day_index, truck, refill_lbs, urgency_tier) from routes
    scheduled: Dict[str, Dict] = {}
    for r in plan.get('routes', []):
        rd_str = r.get('date')
        truck_id = r.get('truck_id')
        route_data = r.get('route', r) if isinstance(r, dict) else {}
        try:
            day_idx = horizon_iso.index(rd_str)
        except (ValueError, TypeError):
            day_idx = None
        for stop in (route_data.get('stops') or []):
            cid = str(stop.get('client_id', ''))
            if not cid:
                continue
            scheduled[cid] = {
                'day_index':  day_idx,
                'date':       rd_str,
                'truck':      str(truck_id) if truck_id else None,
                'refill_lbs': float(stop.get('delivery_lbs') or 0),
                'sequence':   int(stop.get('sequence') or 0),
            }

    deferred_map = plan.get('deferred') or {}

    # Bucket ordering for the matrix
    bucket_order = {'RED': 0, 'YEL': 1, 'GRN': 2, 'GRY': 3}

    rows = []
    for s in snapshot:
        cid = str(s.get('client_id', ''))
        bucket = s.get('urgency_bucket', 'GRY')
        dte = s.get('dte_to_reserve')
        sched = scheduled.get(cid)
        deferred_reason = deferred_map.get(cid)
        is_urgent_miss = (
            bucket in ('RED', 'YEL') and sched is None and deferred_reason is not None
        )
        rows.append({
            'client_id':      cid,
            'customer':       s.get('customer'),
            'tank_lbs':       s.get('tank_lbs'),
            'current_lbs':    s.get('current_lbs'),
            'rate_lbs_per_day': s.get('rate_lbs_per_day'),
            'dte_to_reserve': dte,
            'urgency_bucket': bucket,
            'scheduled':      sched,        # None if deferred
            'deferred_reason': deferred_reason,
            'urgent_miss':    is_urgent_miss,
        })

    # Sort: bucket priority, then dte asc (most urgent first within bucket)
    rows.sort(key=lambda r: (
        bucket_order.get(r['urgency_bucket'], 9),
        r['dte_to_reserve'] if r['dte_to_reserve'] is not None else 999,
    ))

    # Tally for header
    counts = {'RED': 0, 'YEL': 0, 'GRN': 0, 'GRY': 0}
    misses = 0
    for r in rows:
        counts[r['urgency_bucket']] = counts.get(r['urgency_bucket'], 0) + 1
        if r['urgent_miss']:
            misses += 1

    return jsonify({
        'ok':           True,
        'today':        today_iso,
        'horizon':      [{'date': d.isoformat(),
                          'label': d.strftime('%a %m/%d'),
                          'committed': i < commit_days}
                          for i, d in enumerate(horizon_dates)],
        'commit_days':  commit_days,
        'reserve_pct':  plan.get('pre_run_reserve_pct', 0.10),
        'counts':       counts,
        'urgent_misses': misses,
        'rows':         rows,
    })


@app.route('/api/overrides', methods=['GET'])
def api_overrides_get():
    if not USER_OVERRIDES.exists():
        return jsonify({'pins': [], 'forbids': []})
    try:
        return jsonify(json.loads(USER_OVERRIDES.read_text(encoding='utf-8')))
    except Exception:
        return jsonify({'pins': [], 'forbids': []})


@app.route('/api/overrides', methods=['POST'])
def api_overrides_post():
    body = request.get_json(force=True) or {}
    action = body.get('action')
    cid = str(body.get('client_id', '')).strip()
    d_str = body.get('date')
    if action not in ('pin', 'skip', 'clear') or not cid:
        return jsonify({'ok': False, 'error': 'bad request'}), 400

    current = json.loads(USER_OVERRIDES.read_text(encoding='utf-8')) \
        if USER_OVERRIDES.exists() else {'pins': [], 'forbids': []}
    pins = [p for p in current.get('pins', []) if p.get('client_id') != cid]
    forbids = [f for f in current.get('forbids', []) if f.get('client_id') != cid]

    if action == 'pin' and d_str:
        # Pin = "must visit on THIS specific date". Single day.
        pins.append({'client_id': cid, 'date': d_str,
                     'reason': 'dashboard pin', 'created_at': datetime.now().isoformat()})
    elif action == 'skip' and d_str:
        # Skip = "don't visit during the firm commit window". Expands to
        # `commit_days` working days starting from d_str. Without this
        # expansion the solver would just push the client to day 1 (the
        # day after) — defeating the operator's intent.
        try:
            start = date.fromisoformat(d_str)
        except Exception:
            return jsonify({'ok': False, 'error': 'bad date'}), 400
        commit_dates = _upcoming_workdays(start, count=2)
        forbids.append({
            'client_id': cid,
            'dates': [d.isoformat() for d in commit_dates],
            'reason': 'dashboard skip (commit window)',
            'created_at': datetime.now().isoformat(),
        })
    save_user_overrides(USER_OVERRIDES, pins, forbids)
    return jsonify({'ok': True, 'pins': pins, 'forbids': forbids})


@app.route('/api/truck-availability', methods=['GET'])
def api_truck_avail_get():
    """Return list of upcoming work-days with each truck's availability status."""
    start = date.today()
    days = _upcoming_workdays(start, count=10)
    unavail = load_unavailability(TRUCK_UNAVAIL)
    trucks = ['Truck2', 'Truck9']  # could read from fleet.yaml later
    rows = []
    _DOW = ('Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun')
    for d in days:
        is_sat = _DOW[d.weekday()] == 'Sat'
        per_truck = {}
        for t in trucks:
            # Saturday rule: Truck9 is automatically unavailable
            sat_blocked = (is_sat and t != 'Truck2')
            user_blocked = ((d, t) in unavail)
            per_truck[t] = {
                'available': not (sat_blocked or user_blocked),
                'reason_sat': sat_blocked,
                'reason_user': user_blocked,
            }
        rows.append({
            'date': d.isoformat(),
            'label': d.strftime('%a %b %d'),
            'is_saturday': is_sat,
            'trucks': per_truck,
        })
    return jsonify({'days': rows, 'trucks': trucks})


@app.route('/api/truck-availability', methods=['POST'])
def api_truck_avail_post():
    """Toggle one (date, truck) entry."""
    body = request.get_json(force=True) or {}
    d_str = body.get('date')
    tid = body.get('truck_id')
    available = bool(body.get('available'))
    if not d_str or not tid:
        return jsonify({'ok': False, 'error': 'missing date or truck_id'}), 400
    try:
        d = date.fromisoformat(d_str)
    except Exception:
        return jsonify({'ok': False, 'error': 'bad date'}), 400

    # Load current
    unavail = load_unavailability(TRUCK_UNAVAIL)
    if available:
        unavail.discard((d, tid))
    else:
        unavail.add((d, tid))
    entries = [
        {'date': d.isoformat(), 'truck_id': t, 'reason': 'dashboard'}
        for (d, t) in sorted(unavail)
    ]
    save_unavailability(TRUCK_UNAVAIL, entries)
    return jsonify({'ok': True, 'unavailable_count': len(entries)})


@app.route('/api/run', methods=['POST'])
def api_run():
    body = request.get_json(force=True) or {}
    plan_date = body.get('date')
    force = bool(body.get('force', False))
    if not plan_date:
        return jsonify({'ok': False, 'error': 'Missing date'}), 400
    try:
        date.fromisoformat(plan_date)
    except Exception:
        return jsonify({'ok': False, 'error': 'Bad date'}), 400

    # Pre-flight: refuse to run on data we suspect is being edited / stale.
    # Operator can override with force=true, but the warning is shown first.
    f = _read_input_file()
    if f is not None:
        warns = _data_freshness_warnings(f)
        blocks = [w for w in warns if w['severity'] == 'block']
        if blocks and not force:
            return jsonify({
                'ok': False,
                'blocked': True,
                'warnings': warns,
                'error': blocks[0]['message'],
            }), 409

    with _run_lock:
        if _current_run['status'] == 'running':
            return jsonify({'ok': False, 'error': 'A run is already in progress'}), 409
        _current_run.update({
            'status': 'running', 'started_at': time.time(),
            'finished_at': None, 'log': [], 'date': plan_date,
            'returncode': None, 'elapsed_s': None,
        })

    solve_seconds = int(body.get('solve_seconds', 120))
    threading.Thread(
        target=_run_solver_thread,
        args=(plan_date, solve_seconds),
        daemon=True,
    ).start()
    return jsonify({'ok': True})


def _append_log(line: str) -> None:
    """Push a line into the live run log (used by both pre-flight & solver)."""
    with _run_lock:
        _current_run['log'].append(line)
        if len(_current_run['log']) > 500:
            _current_run['log'] = _current_run['log'][-500:]


def _ensure_matrix_up_to_date(input_file: Optional[Path]) -> bool:
    """Auto-rebuild the OSRM matrix if any spreadsheet client is missing.

    Returns True on success (matrix is now up-to-date), False on hard failure.
    Streams progress into the run log so the operator sees what happened.
    """
    if input_file is None or not input_file.exists():
        _append_log('   [matrix] no input file configured — skipping pre-flight')
        return True

    matrix_file = DATA_DIR / 'osrm_full_matrix_with_ids.npz'
    try:
        # Fast path: load matrix + clients, diff IDs locally.
        import numpy as np
        clients = load_clients(input_file)
        routable = [c for c in clients
                     if c.lat is not None and c.lon is not None]
        if not matrix_file.exists():
            missing_ids = [c.id for c in routable]
            _append_log(f'   [matrix] file missing — will build {len(missing_ids)} clients')
        else:
            data = np.load(matrix_file, allow_pickle=True)
            ids_in_matrix = set(data['client_ids'])
            missing_ids = [c.id for c in routable if c.id not in ids_in_matrix]

        if not missing_ids:
            _append_log(f'   [matrix] ✓ up to date ({len(routable)} clients)')
            return True

        _append_log(f'   [matrix] ⚠ {len(missing_ids)} client(s) missing: '
                    f'{", ".join(missing_ids[:8])}'
                    + (' …' if len(missing_ids) > 8 else ''))
        _append_log('   [matrix] rebuilding via OSRM (this can take ~30–60 s)…')

        cmd = [sys.executable, '-u', '-m', 'final.build_matrix']
        proc = subprocess.Popen(
            cmd, cwd=str(REPO), env=dict(os.environ),
            stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
            bufsize=1, text=True,
        )
        for line in proc.stdout:  # type: ignore
            _append_log('   [matrix] ' + line.rstrip())
        proc.wait()
        if proc.returncode != 0:
            _append_log(f'   [matrix] ✗ rebuild FAILED (exit {proc.returncode}) — '
                        'aborting solver run')
            return False
        _append_log('   [matrix] ✓ rebuild complete — proceeding to solver')
        return True
    except Exception as e:
        _append_log(f'   [matrix] ✗ pre-flight error: {e} — aborting solver run')
        return False


def _run_solver_thread(plan_date: str, solve_seconds: int):
    # Pre-flight: make sure every spreadsheet client is in the OSRM matrix.
    # New clients (e.g. PYNION, COWBOY COOKIN) would otherwise be silently
    # dropped by the solver with a NOT_IN_MATRIX warning in the Deferred sheet.
    matrix_ok = _ensure_matrix_up_to_date(_read_input_file())
    if not matrix_ok:
        with _run_lock:
            _current_run['status'] = 'error'
            _current_run['finished_at'] = time.time()
            _current_run['elapsed_s'] = _current_run['finished_at'] - _current_run['started_at']
            _current_run['returncode'] = -1
        return

    cmd = [sys.executable, '-u', '-m', 'final.sk_solver_final',
           '--today', plan_date, '--solve-seconds', str(solve_seconds)]
    env = dict(os.environ)
    try:
        proc = subprocess.Popen(
            cmd, cwd=str(REPO), env=env,
            stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
            bufsize=1, text=True,
        )
        for line in proc.stdout:  # type: ignore
            _append_log(line.rstrip())
        proc.wait()
        with _run_lock:
            _current_run['finished_at'] = time.time()
            _current_run['elapsed_s'] = _current_run['finished_at'] - _current_run['started_at']
            _current_run['returncode'] = proc.returncode
            _current_run['status'] = 'done' if proc.returncode == 0 else 'error'
    except Exception as e:
        _append_log(f'ERROR: {e}')
        with _run_lock:
            _current_run['status'] = 'error'
            _current_run['finished_at'] = time.time()
            _current_run['elapsed_s'] = _current_run['finished_at'] - _current_run['started_at']


@app.route('/api/run/status')
def api_run_status():
    with _run_lock:
        snap = dict(_current_run)
        snap['log_tail'] = snap['log'][-80:]
        snap['log_count'] = len(snap['log'])
    elapsed = None
    if snap['started_at']:
        end = snap['finished_at'] or time.time()
        elapsed = round(end - snap['started_at'], 1)
    snap['elapsed_s'] = elapsed
    return jsonify(snap)


@app.route('/outputs/<path:filename>')
def serve_output(filename):
    p = OUTPUT_DIR / filename
    if not p.exists():
        return jsonify({'error': 'not found'}), 404
    # .html (route maps) → render inline in browser.
    # .xlsx and .csv → force download so browser doesn't open a blank tab.
    suffix = p.suffix.lower()
    as_attachment = suffix in ('.xlsx', '.csv', '.json', '.zip')
    return send_file(p, as_attachment=as_attachment, download_name=p.name)


# ════════════════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════════════════

def main(host: str = '127.0.0.1', port: int = 5050):
    print('=' * 70)
    print('  S&K Route Optimizer Dashboard')
    print(f'  Open: http://{host}:{port}')
    print('=' * 70)
    app.run(host=host, port=port, debug=False, threaded=True)


if __name__ == '__main__':
    main()
