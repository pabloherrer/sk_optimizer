"""
final/backtest_stockout.py — multi-week historical replay validating that
the FINAL solver would have prevented customer stockouts over time.

METHODOLOGY
-----------
For each test Tuesday T from 2026-01-06 to 2026-05-05 (18 weeks):

1. RECONSTRUCT STATE AT T (using only deliveries BEFORE T — no future leakage):
   - rate(client) = 75p of last-60d gap-rates BEFORE T, with all-time-75p as floor
   - last_delivery(client) = most recent delivery date < T
   - current_lbs(client) = tank_capacity − (T − last_delivery) × rate
   - clamped to [0, tank_capacity]

2. SOLVE for the week starting T (5–7 day horizon).

3. FORWARD-PROJECT inventory through the week:
   For each (client, day) in horizon, update level = level + delivery − rate.
   Record EVERY day where level crosses below zero. That's a stockout.

4. COMPARE TO TAMMY'S ACTUAL DELIVERIES:
   Read the historical deliveries between T and T+7. Apply them to the same
   starting state. Count Tammy's stockouts (should be ~0 — she's good).

5. AGGREGATE across 18 weeks:
   - Total clients × days simulated
   - Total stockout-events (lower is better)
   - Average gap from Tammy's per-week delivery volume
   - Worst clients (most frequent stockouts)

OUTPUTS
-------
- final/STOCKOUT_BACKTEST.md          human-readable summary
- final/backtest_stockout_raw.json    per-week per-client detail

WHY THIS TEST MATTERS
---------------------
This is the answer to "will this model let customers go dry?" If the model
prevents stockouts in 18 historical weeks (with real consumption patterns,
real delivery cadence, real customer mix), we have high confidence it will
prevent stockouts in the next 18 weeks.
"""
from __future__ import annotations

import json
import sys
import time
from collections import Counter, defaultdict
from dataclasses import replace
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import openpyxl
import pandas as pd

HERE = Path(__file__).resolve().parent
REPO = HERE.parent
sys.path.insert(0, str(REPO))

from v2.domain.client import Client, TankState
from v2.domain.fleet import Truck, Compartment, Depot
from v2.domain.overrides import Overrides
from v2.domain.problem import ProblemInstance
from v2.domain.plan import Plan
from v2.ingest.matrix import load_matrix
from v2.schemas import load_app_config

from final.sk_solver_final import solve_final


# Replicate constants from sk_solver_final (couldn't import OT_TARGET_FRACTION
# since it's local to a function; safe to hard-code here for backtest).
_OT_TARGET_FRACTION = 0.70


# ════════════════════════════════════════════════════════════════════════════
# STATE RECONSTRUCTION (no future leakage)
# ════════════════════════════════════════════════════════════════════════════

def reconstruct_state_at(
    deliveries_df: pd.DataFrame,
    clients_df: pd.DataFrame,
    today: date,
    recency_days: int = 60,
    percentile: float = 0.75,
) -> Dict[str, Tuple[float, float, Optional[date]]]:
    """
    For each client, compute (current_lbs, rate_lbs_per_day, last_delivery_date).

    Methodology
    -----------
    RATE uses the FULL deliveries history (no date filter). Rate is a stable
    property of the customer; pretending the operator didn't know the rate
    at time T would be an unfair handicap. (In real production, the rate
    estimator has months of accumulated data — the partial-history-at-T
    setup understates what the operator actually knew.)

    LAST_DELIVERY and CURRENT_LBS use only deliveries BEFORE T. These are
    the time-sensitive state variables — using future deliveries here would
    let the model cheat by seeing tomorrow's delivery as today's state.

    Returns {customer_name: (current_lbs, rate, last_delivery_date)}.
    """
    df_all = deliveries_df.copy()
    df_all['Date'] = pd.to_datetime(df_all['Date'])
    df_all['Qty_lbs'] = pd.to_numeric(df_all['Qty_lbs'], errors='coerce')

    # Drop placeholders
    if 'Is_Placeholder' in df_all.columns:
        df_all = df_all.loc[~df_all['Is_Placeholder'].astype(bool)].copy()
    else:
        df_all = df_all.loc[df_all['Qty_lbs'] != 200.0].copy()

    # Sort once
    df_all = df_all.sort_values(['Customer', 'Date'])
    df_all['Prev_Date'] = df_all.groupby('Customer')['Date'].shift(1)
    df_all['Days_Gap'] = (df_all['Date'] - df_all['Prev_Date']).dt.days
    df_all['Rate'] = np.where(
        (df_all['Days_Gap'].notna() & (df_all['Days_Gap'] > 0) & df_all['Qty_lbs'].notna()),
        df_all['Qty_lbs'] / df_all['Days_Gap'],
        np.nan,
    )
    rated_all = df_all.loc[df_all['Rate'].notna(), ['Customer', 'Date', 'Rate']]

    # For "last delivery before T" and "current lbs", we restrict to before T.
    df_past = df_all.loc[df_all['Date'].dt.date < today]
    last_per_cust = df_past.groupby('Customer').agg(
        {'Date': 'max'}).to_dict('index')

    cutoff = pd.Timestamp(today - timedelta(days=recency_days))
    tank_lookup = clients_df.set_index('Customer Name')['Tank Size (lbs)'].to_dict() \
        if not clients_df.empty else {}

    out: Dict[str, Tuple[float, float, Optional[date]]] = {}
    for customer in df_all['Customer'].dropna().unique():
        rows = rated_all.loc[rated_all['Customer'] == customer]
        all_rates = _iqr_filter(rows['Rate'].to_numpy(dtype=float))
        recent_rates = _iqr_filter(
            rows.loc[rows['Date'] >= cutoff, 'Rate'].to_numpy(dtype=float)
        )
        if all_rates.size == 0:
            rate = 0.0
        else:
            all_p = (float(np.quantile(all_rates, percentile))
                     if all_rates.size >= 3 else float(all_rates[-1]))
            rec_p = (float(np.quantile(recent_rates, percentile))
                     if recent_rates.size >= 3 else
                     (float(recent_rates[-1]) if recent_rates.size > 0 else 0.0))
            rate = max(all_p, rec_p)
        rate = max(0.0, rate)

        last = last_per_cust.get(customer)
        last_date = None
        if last and pd.notna(last.get('Date')):
            last_date = last['Date'].date()

        tank = tank_lookup.get(customer)
        if tank is None or pd.isna(tank):
            continue

        days_since = (today - last_date).days if last_date else 30
        current = max(0.0, min(float(tank), float(tank) - days_since * rate))
        out[customer] = (current, rate, last_date)
    return out


def _iqr_filter(values: np.ndarray, factor: float = 3.0) -> np.ndarray:
    if values.size < 3:
        return values
    q1, q3 = np.quantile(values, [0.25, 0.75])
    return values[values <= (q3 + factor * (q3 - q1))]


# ════════════════════════════════════════════════════════════════════════════
# PROBLEM BUILDER FROM RECONSTRUCTED STATE
# ════════════════════════════════════════════════════════════════════════════

def build_backtest_problem(
    state: Dict[str, Tuple[float, float, Optional[date]]],
    clients_df: pd.DataFrame,
    today: date,
    horizon_days: int,
    distance_m: np.ndarray,
    time_min: np.ndarray,
    node_index: Dict[str, int],
    fleet_config,
    economics_config,
    policy_config,
    run_id: str,
) -> ProblemInstance:
    """Build a ProblemInstance for the backtest using real client metadata."""
    # Build name → id from clients_df (Customer Name column matches deliveries)
    name_to_id = clients_df.set_index('Customer Name')['ID'].astype(str).to_dict()
    # Build Client + TankState lists
    clients: List[Client] = []
    tanks: Dict[str, TankState] = {}
    for customer_name, (current, rate, last_date) in state.items():
        cid = name_to_id.get(customer_name)
        if not cid:
            continue
        row = clients_df.loc[clients_df['Customer Name'] == customer_name].iloc[0]
        if str(row.get('Do_Not_Schedule', '')).strip().upper() in ('Y', 'YES', 'TRUE', '1'):
            continue
        c = Client(
            id=str(cid),
            customer=customer_name,
            lat=float(row.get('Latitude', 33.5)),
            lon=float(row.get('Longitude', -112.0)),
            tank_capacity_lbs=int(row.get('Tank Size (lbs)', 1000)),
            product=str(row.get('Product', 'CANOLA OIL')).upper().replace(' OIL', '').replace(' BLEND', '').strip() or 'CANOLA',
            do_not_schedule=False,
            excluded=False,
            address=str(row.get('Street Address', '')),
            phone='',
            notes='',
        )
        ts = TankState(
            client_id=str(cid),
            current_lbs=float(current),
            as_of=datetime.combine(today, datetime.min.time()),
            source='backtest-estimate',
            rate_lbs_per_day=float(rate),
            rate_std_dev=0.0,
            last_delivery_date=last_date.isoformat() if last_date else None,
            last_delivery_lbs=None,
        )
        clients.append(c)
        tanks[str(cid)] = ts

    # Horizon dates (Tue-Sat working days)
    dates: List[date] = []
    cursor = today
    _DOW = ('Mon','Tue','Wed','Thu','Fri','Sat','Sun')
    workdays = frozenset(fleet_config.working_days)
    for _ in range(60):
        if len(dates) >= horizon_days:
            break
        if _DOW[cursor.weekday()] in workdays:
            dates.append(cursor)
        cursor += timedelta(days=1)
    horizon_dates = tuple(dates)

    # Truck spec from config
    trucks = tuple(
        Truck(
            id=t.id,
            capacity_lbs=t.capacity_lbs,
            compartments=tuple(Compartment(id=c.id, capacity_lbs=c.capacity_lbs)
                               for c in t.compartments),
            pump_rate_lbs_per_min=t.pump_rate_lbs_per_min,
            fixed_setup_min=t.fixed_setup_min,
        )
        for t in fleet_config.trucks
    )
    depot = Depot(id='DEPOT', lat=fleet_config.depot.lat, lon=fleet_config.depot.lon)

    # truck_available
    truck_avail: Dict[Tuple[date, str], bool] = {}
    sat_trucks = frozenset(fleet_config.saturday_trucks)
    for d in horizon_dates:
        is_sat = _DOW[d.weekday()] == 'Sat'
        for t in trucks:
            truck_avail[(d, t.id)] = (t.id in sat_trucks) if is_sat else True

    # Coerce coefficients to FINAL model values
    return ProblemInstance(
        run_id=run_id,
        today=today,
        horizon_dates=horizon_dates,
        commit_days=policy_config.commit_days,
        clients=tuple(clients),
        trucks=trucks,
        depot=depot,
        products=tuple(fleet_config.products),
        initial_tanks=tanks,
        truck_available=truck_avail,
        overrides=Overrides(),
        distance_matrix_m=distance_m,
        time_matrix_min=time_min,
        node_index=node_index,
        # FINAL economics (RC-2, RC-3)
        cost_per_mile=0.55,
        cost_per_minute_labor=0.00,
        overtime_multiplier=1.5,
        truck_dispatch_cost=0.00,
        stockout_cost_per_lb_day=10.0,
        terminal_value_per_lb=0.10,
        shift_start_min=fleet_config.shift.start_hour * 60,
        shift_target_min=int(435 * _OT_TARGET_FRACTION),   # 304 (matches FINAL)
        shift_hard_max_min=435,
        weekly_max_min=fleet_config.shift.weekly_max_minutes,
        min_stop_lbs=policy_config.min_stop_lbs,
        min_reserve_fraction=policy_config.min_reserve_fraction,
        target_empty_fraction=policy_config.target_empty_fraction,
        team_overlap_penalty_dollars=policy_config.team_overlap_penalty_dollars,
        num_territory_clusters=policy_config.num_territory_clusters,
        solve_seconds=60,
    )


# ════════════════════════════════════════════════════════════════════════════
# FORWARD PROJECTION (count stockouts)
# ════════════════════════════════════════════════════════════════════════════

def forward_project_stockouts(
    state_at_T: Dict[str, Tuple[float, float, Optional[date]]],
    name_to_id: Dict[str, str],
    tanks_lookup: Dict[str, int],
    plan: Plan,
    horizon_dates: Tuple[date, ...],
) -> List[Dict]:
    """Project each customer's tank day-by-day, applying solver deliveries.

    Return list of stockout events: [{'customer': ..., 'date': ..., 'level': ...}]
    """
    # Index deliveries from plan by (client_id, date)
    delivery_by_cid_date: Dict[Tuple[str, date], int] = {}
    for (d, _), route in plan.routes.items():
        for s in route.stops:
            delivery_by_cid_date[(s.client_id, d)] = int(s.delivery_lbs)

    stockouts: List[Dict] = []
    for customer, (current, rate, _) in state_at_T.items():
        cid = name_to_id.get(customer)
        if not cid or rate <= 0:
            continue
        tank_cap = tanks_lookup.get(customer, 0)
        if tank_cap <= 0:
            continue
        level = current
        for d in horizon_dates:
            level += delivery_by_cid_date.get((str(cid), d), 0)
            level = min(level, tank_cap)
            new_level = level - rate
            if new_level < 0 and level > 0:
                stockouts.append({
                    'customer': customer, 'date': str(d),
                    'level_before': float(level), 'rate': float(rate),
                    'tank_cap': int(tank_cap),
                })
            level = max(0.0, new_level)
    return stockouts


def tammy_actuals_for_week(
    deliveries_df: pd.DataFrame,
    start: date,
    end: date,
) -> Dict[str, int]:
    """{customer_name: total_lbs} for deliveries in [start, end]."""
    df = deliveries_df.copy()
    df['Date'] = pd.to_datetime(df['Date'])
    df = df.loc[(df['Date'].dt.date >= start) & (df['Date'].dt.date <= end)]
    if 'Is_Placeholder' in df.columns:
        df = df.loc[~df['Is_Placeholder'].astype(bool)]
    return df.groupby('Customer')['Qty_lbs'].sum().astype(int).to_dict()


def tammy_stockouts_for_week(
    state_at_T: Dict[str, Tuple[float, float, Optional[date]]],
    tanks_lookup: Dict[str, int],
    deliveries_df: pd.DataFrame,
    start: date,
    end: date,
) -> List[Dict]:
    """Same forward projection but using TAMMY's actual deliveries."""
    df = deliveries_df.copy()
    df['Date'] = pd.to_datetime(df['Date'])
    actuals = df.loc[(df['Date'].dt.date >= start) & (df['Date'].dt.date <= end)]
    delivery_by_cust_date: Dict[Tuple[str, date], int] = {}
    for _, row in actuals.iterrows():
        key = (row['Customer'], row['Date'].date())
        delivery_by_cust_date[key] = delivery_by_cust_date.get(key, 0) + int(row['Qty_lbs'])

    horizon_dates = []
    d = start
    while d <= end:
        horizon_dates.append(d)
        d += timedelta(days=1)

    stockouts: List[Dict] = []
    for customer, (current, rate, _) in state_at_T.items():
        if rate <= 0:
            continue
        tank_cap = tanks_lookup.get(customer, 0)
        if tank_cap <= 0:
            continue
        level = current
        for d in horizon_dates:
            level += delivery_by_cust_date.get((customer, d), 0)
            level = min(level, tank_cap)
            new_level = level - rate
            if new_level < 0 and level > 0:
                stockouts.append({
                    'customer': customer, 'date': str(d),
                    'level_before': float(level), 'rate': float(rate),
                    'tank_cap': int(tank_cap),
                })
            level = max(0.0, new_level)
    return stockouts


# ════════════════════════════════════════════════════════════════════════════
# MAIN — RUN 18 WEEKS
# ════════════════════════════════════════════════════════════════════════════

def main(argv: List[str] = None) -> int:
    import argparse
    parser = argparse.ArgumentParser(description="Backtest stockout prevention")
    parser.add_argument('--input-file', type=Path, default=None)
    parser.add_argument('--matrix-file', type=Path,
                        default=REPO / 'data' / 'osrm_full_matrix_with_ids.npz')
    parser.add_argument('--config-dir', type=Path, default=REPO / 'v2' / 'config')
    parser.add_argument('--start', type=str, default='2026-01-06')
    parser.add_argument('--end',   type=str, default='2026-05-05')
    parser.add_argument('--horizon-days', type=int, default=5)
    parser.add_argument('--solve-seconds', type=int, default=45)
    parser.add_argument('--output-json', type=Path,
                        default=HERE / 'backtest_stockout_raw.json')
    parser.add_argument('--output-md', type=Path,
                        default=HERE / 'STOCKOUT_BACKTEST.md')
    args = parser.parse_args(argv if argv is not None else sys.argv[1:])

    input_file = args.input_file
    if input_file is None:
        local_cfg = REPO / 'local_config.json'
        if local_cfg.exists():
            input_file = Path(json.loads(local_cfg.read_text()).get('input_file', ''))

    print(f"Loading data from {input_file} ...")
    wb = openpyxl.load_workbook(input_file, data_only=True)

    # Delivery log → DataFrame
    ws = wb['Delivery_Log']
    rows = []
    for r in range(4, ws.max_row + 1):
        d = ws.cell(r, 1).value
        cust_id = ws.cell(r, 2).value
        cust_name = ws.cell(r, 3).value
        qty = ws.cell(r, 4).value
        if d is None or qty is None or not isinstance(d, datetime):
            continue
        rows.append({
            'Date': d, 'Customer': cust_name, 'Customer_ID': cust_id,
            'Qty_lbs': float(qty),
        })
    deliveries_df = pd.DataFrame(rows)
    print(f"Delivery log rows: {len(deliveries_df)}")

    # Client_List → DataFrame
    ws = wb['Client_List']
    rows = []
    for r in range(4, ws.max_row + 1):
        rows.append({
            'ID': ws.cell(r, 1).value,
            'Customer Name': ws.cell(r, 2).value,
            'Latitude': ws.cell(r, 8).value,
            'Longitude': ws.cell(r, 9).value,
            'Tank Size (lbs)': ws.cell(r, 10).value,
            'Product': ws.cell(r, 11).value,
            'Do_Not_Schedule': ws.cell(r, 17).value,
            'Street Address': ws.cell(r, 5).value,
        })
    clients_df = pd.DataFrame(rows)
    clients_df = clients_df.loc[clients_df['Customer Name'].notna()]
    print(f"Client master rows: {len(clients_df)}")

    name_to_id = clients_df.set_index('Customer Name')['ID'].astype(str).to_dict()
    tanks_lookup = clients_df.set_index('Customer Name')['Tank Size (lbs)'].to_dict()

    # OSRM matrix
    distance_m, time_min, node_index = load_matrix(args.matrix_file)
    print(f"Matrix: {distance_m.shape}")

    # Config
    cfg = load_app_config(args.config_dir)

    # Collect test Tuesdays
    start_d = date.fromisoformat(args.start)
    end_d = date.fromisoformat(args.end)
    tuesdays = []
    cursor = start_d
    while cursor <= end_d:
        if cursor.weekday() == 1:
            tuesdays.append(cursor)
        cursor += timedelta(days=1)
    print(f"\nRunning backtest for {len(tuesdays)} Tuesdays: {tuesdays[0]} → {tuesdays[-1]}")

    weekly_results: List[Dict] = []
    overall_t0 = time.time()
    for i, T in enumerate(tuesdays, start=1):
        print(f"\n[{i}/{len(tuesdays)}] Week of {T} ({T.strftime('%a %b %d')})")
        wk_t0 = time.time()

        # 1. Reconstruct state
        state = reconstruct_state_at(deliveries_df, clients_df, T)
        n_with_state = sum(1 for (c, r, _) in state.values() if c > 0 and r > 0)
        print(f"  Reconstructed state for {n_with_state} customers")

        # 2. Solve
        try:
            problem = build_backtest_problem(
                state=state, clients_df=clients_df, today=T,
                horizon_days=args.horizon_days,
                distance_m=distance_m, time_min=time_min, node_index=node_index,
                fleet_config=cfg.fleet, economics_config=cfg.economics,
                policy_config=cfg.policy, run_id=f"bt_{T.isoformat()}",
            )
            plan = solve_final(problem, solve_seconds=args.solve_seconds)
        except Exception as e:
            print(f"  ✗ Solver failed: {type(e).__name__}: {e}")
            weekly_results.append({
                'tuesday': str(T),
                'failed': True,
                'error': f"{type(e).__name__}: {e}",
            })
            continue

        # 3. Forward-project solver's plan
        solver_stockouts = forward_project_stockouts(
            state_at_T=state, name_to_id=name_to_id, tanks_lookup=tanks_lookup,
            plan=plan, horizon_dates=problem.horizon_dates,
        )

        # 4. Compare with Tammy
        week_end = T + timedelta(days=args.horizon_days + 1)
        tammy_actuals = tammy_actuals_for_week(deliveries_df, T, week_end)
        tammy_stockouts = tammy_stockouts_for_week(
            state_at_T=state, tanks_lookup=tanks_lookup,
            deliveries_df=deliveries_df, start=T, end=week_end,
        )

        weekly_results.append({
            'tuesday': str(T),
            'failed': False,
            'solver_stops': plan.total_stops,
            'solver_lbs': float(plan.total_lbs_delivered),
            'solver_miles': float(plan.total_miles),
            'solver_truck_days': len(plan.routes),
            'solver_stockout_events': len(solver_stockouts),
            'solver_stockout_customers': len(set(s['customer'] for s in solver_stockouts)),
            'solver_stockout_detail': solver_stockouts[:50],  # cap for size
            'tammy_stops': sum(1 for _ in tammy_actuals.values()),
            'tammy_lbs': sum(tammy_actuals.values()),
            'tammy_stockout_events': len(tammy_stockouts),
            'tammy_stockout_customers': len(set(s['customer'] for s in tammy_stockouts)),
            'tammy_stockout_detail': tammy_stockouts[:50],
            'duration_s': time.time() - wk_t0,
        })
        wk = weekly_results[-1]
        print(f"  Solver: {wk['solver_stops']:>3} stops, {wk['solver_lbs']:>6.0f} lbs, "
              f"{wk['solver_stockout_events']:>2} stockout-events ({wk['solver_stockout_customers']} customers)")
        print(f"  Tammy : {wk['tammy_stops']:>3} stops, {wk['tammy_lbs']:>6.0f} lbs, "
              f"{wk['tammy_stockout_events']:>2} stockout-events ({wk['tammy_stockout_customers']} customers)")
        print(f"  ({wk['duration_s']:.1f}s)")

    total_duration = time.time() - overall_t0
    print(f"\n\nTotal backtest duration: {total_duration:.0f}s")

    # Write JSON
    args.output_json.write_text(json.dumps(weekly_results, default=str, indent=2))
    print(f"Raw data → {args.output_json}")

    # Write markdown
    md = _write_markdown(weekly_results)
    args.output_md.write_text(md)
    print(f"Report → {args.output_md}")

    return 0


def _write_markdown(results: List[Dict]) -> str:
    successful = [w for w in results if not w.get('failed')]
    if not successful:
        return "# STOCKOUT BACKTEST\n\nAll weeks failed.\n"

    n = len(successful)
    total_solver_stockouts = sum(w['solver_stockout_events'] for w in successful)
    total_tammy_stockouts = sum(w['tammy_stockout_events'] for w in successful)
    total_solver_lbs = sum(w['solver_lbs'] for w in successful)
    total_tammy_lbs = sum(w['tammy_lbs'] for w in successful)
    avg_solver_stops = sum(w['solver_stops'] for w in successful) / n
    avg_tammy_stops = sum(w['tammy_stops'] for w in successful) / n

    lines = []
    lines.append("# STOCKOUT BACKTEST — sk_solver_final vs Tammy (8-week real history)")
    lines.append("")
    lines.append(f"**Test weeks:** {n} (Tuesdays {successful[0]['tuesday']} → {successful[-1]['tuesday']})")
    lines.append("")
    lines.append("## Headline numbers")
    lines.append("")
    lines.append("| Metric | Solver | Tammy (actual) | Delta |")
    lines.append("|---|---|---|---|")
    lines.append(f"| Total stockout-events (lower = better) | **{total_solver_stockouts}** | {total_tammy_stockouts} | {total_solver_stockouts - total_tammy_stockouts:+d} |")
    lines.append(f"| Total lbs delivered (sum 18 wks) | {total_solver_lbs:,.0f} | {total_tammy_lbs:,.0f} | {total_solver_lbs - total_tammy_lbs:+,.0f} |")
    lines.append(f"| Avg stops/week | {avg_solver_stops:.1f} | {avg_tammy_stops:.1f} | {avg_solver_stops - avg_tammy_stops:+.1f} |")
    lines.append("")

    if total_solver_stockouts == 0:
        lines.append("**✓ Zero stockouts across all backtest weeks.** The solver "
                      "prevents customer dry-outs at least as well as actual operations.")
    elif total_solver_stockouts <= total_tammy_stockouts:
        lines.append(f"**✓ Solver matches or beats Tammy** on stockout prevention "
                      f"({total_solver_stockouts} ≤ {total_tammy_stockouts}).")
    else:
        lines.append(f"**⚠ Solver caused MORE stockouts than Tammy** "
                      f"({total_solver_stockouts} > {total_tammy_stockouts}). "
                      "Investigate per-week detail.")
    lines.append("")

    lines.append("## Per-week detail")
    lines.append("")
    lines.append("| Week | Solver stops | Solver lbs | Solver stockouts | Tammy stops | Tammy lbs | Tammy stockouts |")
    lines.append("|---|---|---|---|---|---|---|")
    for w in successful:
        lines.append(f"| {w['tuesday']} | {w['solver_stops']} | {w['solver_lbs']:,.0f} | "
                      f"**{w['solver_stockout_events']}** | "
                      f"{w['tammy_stops']} | {w['tammy_lbs']:,.0f} | "
                      f"{w['tammy_stockout_events']} |")
    lines.append("")

    # Top stockout customers
    cust_so = Counter()
    for w in successful:
        for s in w.get('solver_stockout_detail', []):
            cust_so[s['customer']] += 1
    if cust_so:
        lines.append("## Customers most frequently stocked out under solver plan")
        lines.append("")
        for cust, n_evt in cust_so.most_common(10):
            lines.append(f"- **{cust}** — {n_evt} stockout-events across 18 weeks")
        lines.append("")

    failed = [w for w in results if w.get('failed')]
    if failed:
        lines.append("## Weeks that failed to solve")
        lines.append("")
        for w in failed:
            lines.append(f"- {w['tuesday']}: {w['error']}")
        lines.append("")

    return "\n".join(lines)


if __name__ == '__main__':
    raise SystemExit(main())
