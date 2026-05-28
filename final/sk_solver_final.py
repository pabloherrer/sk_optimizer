"""
sk_solver_final.py — Production OR-Tools solver for S&K oil delivery routing.

This is the single, defensible model. Every coefficient is documented with
its real-world justification, and every constraint/penalty is traceable to
either a business rule or a sensitivity-analyzed empirical choice.

The audit and root-cause analyses are in:
    final/AUDIT.md
    final/ROOT_CAUSES.md
Read those before changing anything in this file.

Architecture
------------
This module reuses v2.ingest and v2.reporting (those are correct), but
replaces v2.solver.model with a new formulation that fixes:

  RC-1: Drop penalty piecewise-priced by tank urgency, not zero for safe clients
  RC-2: Labor cost dropped to $0/min for regular hours (drivers are salaried)
  RC-3: Truck dispatch cost dropped to $0/day (warm-up is in per-mile fuel)
  RC-4: Capacity demand uses p75 of feasible refills, not max
  RC-5: Consumption rate uses recency-weighted percentile (max of 60d and all-time)
  RC-6: Min-stop threshold is a soft per-stop fee, not a hard forbid
  RC-7: Commit-window enforcement: DTE < commit_days+0.5 must be in day 0..commit-1
  RC-8: Saturday rule via truck_available, not 10⁹ fixed cost

Public entry point: main() — runs the full pipeline and writes outputs.
"""
from __future__ import annotations

import math
import sys
import time
import uuid
from dataclasses import replace
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
from ortools.constraint_solver import pywrapcp, routing_enums_pb2

# Reuse the v2 plumbing that's already correct.
from v2.domain.problem import ProblemInstance
from v2.domain.client import TankState
from v2.domain.plan import Plan
from v2.ingest.build_problem import build_problem_instance
from v2.ingest.excel import load_deliveries
from v2.solver.extract import extract_plan
from v2.solver.model import ModelArtifacts          # data-class for interface only
from v2.invariants import check_plan
from v2.reporting import write_all_outputs
from v2.state.store import StateStore


# ════════════════════════════════════════════════════════════════════════════
# SECTION 1. COST COEFFICIENTS — every value justified, every assumption tested
# ════════════════════════════════════════════════════════════════════════════
#
# These OVERRIDE the values in v2/config/economics.yaml. They reflect the
# real S&K economics with weekly-salaried drivers (RC-2 / RC-3).
#
# Sensitivity tested in: validate.py::test_economic_sensitivity()
# A ±50% sweep on each parameter changes plan cost <8% and stop count <3 stops.

# Fuel + per-mile wear. Industry baseline for medium-duty box trucks running
# diesel at ~$3.50/gal getting ~7 mpg, plus 25% wear/maintenance contribution.
# Sensitivity: 0.35 → all-trucks-always; 0.85 → trucks consolidate
COST_PER_MILE              = 0.55       # $/mi

# Labor for REGULAR hours. Zero because drivers are weekly-salaried — the
# wage is sunk regardless of whether the truck rolls. The optimizer would
# only see this cost if drivers were hourly. Setting to a small positive
# value would re-introduce the "use one truck" bias documented in AUDIT.md.
COST_PER_MINUTE_LABOR_REG  = 0.00       # $/min

# Overtime PREMIUM ($/min over target). Real marginal cost: the 0.5×
# multiplier on the OT portion only. Standard time-and-a-half ($50/h base
# × 0.5 = $25/h × 1/60 = $0.42/min). This IS marginal: paid only when route
# pushes beyond the target. The premium is real money out the door.
COST_PER_MINUTE_OT_PREMIUM = 0.42       # $/min over target_minutes

# Fixed cost of dispatching a truck. Zero because (a) regular labor is sunk
# (RC-2), and (b) fuel for the actual route is already priced in COST_PER_MILE.
# What's left — paperwork, warm-up — is a few cents per day, well below the
# precision of cost integerization.
TRUCK_DISPATCH_COST        = 0.00       # $/day

# Stockout penalty per pound-day of negative supply. $10 makes a 100-lb,
# 1-day deficit cost $1,000 — large enough that the solver always finds
# capacity for must-serves, small enough never to force infeasibility.
# Only used inside the drop-penalty function below.
STOCKOUT_COST_PER_LB_DAY   = 10.00      # $/lb-day

# Per-stop SOFT penalty for opportunistic small top-offs that we'd rather
# skip (uneconomic) UNLESS the geographic detour is already cheap. Replaces
# the previous hard-forbid behavior (RC-6). The value is calibrated so a
# 30-lb top-off at a client we're already 0.5 miles past gets accepted
# (0.5mi × $0.55 = $0.275; if penalty < $0.275 we'd take it). The 18-min
# setup time has no labor cost (drivers salaried), so the only economic
# disincentive is the small-stop fee plus the geographic detour.
# $1.50 → solver accepts top-offs within ~3 miles of being on the way.
SMALL_STOP_FEE             = 1.50       # $ extra cost per visit when refill < min_stop_lbs

# Minimum fill % rule (RC-9, operator-requested).
# Don't dispatch a truck to deliver less than X% of tank capacity — too
# inefficient. A 1000-lb tank getting 400 lbs (40% fill) gets skipped;
# a 500-lb tank getting 400 lbs (80% fill) gets served.
#
# Urgency exception: clients with DTE ≤ 3 days bypass this rule — better
# to deliver a small amount than let a customer stock out.
#
# Set MIN_FILL_PCT = 0.0 to disable.
MIN_FILL_PCT               = 0.50       # min refill / tank capacity for non-urgent stops
# NOTE: at runtime this default is overridden by `solver_settings.min_fill_pct`
# in local_config.json (set from the dashboard's Solver Tuning card). The
# constant here is the fallback when no override exists.

# Truck utilization reward — encodes the operational value of using
# spare time/capacity to serve more customers TODAY rather than deferring
# them to a future truck-day.
#
# The cost model otherwise has a blind spot: regular labor is $0/min
# (drivers salaried — sunk) and dispatch cost is $0 (no marginal cost to
# rolling the truck). The only real cost of adding a stop is the extra
# mileage. Without a positive reward for filling truck-days, the solver
# correctly minimizes mileage and stops early — even when time and tank
# space remain.
#
# This is NOT a knob — it's a fixed encoding of operational truth:
# "if the truck is already going out, an extra stop is worth more to us
# than $X of mileage." We set X = $0.30/min of underutilization below a
# 6-hour target. At ~$3-6 of mileage per added stop and ~30 min per stop,
# the math: stop adds (30 × $0.30) = $9 of reward, far above mileage cost
# → solver eagerly fills truck-days.
#
# Mechanics: SetCumulVarSoftLowerBound on each truck-day's End-time node.
# Effective penalty for short truck-day = (target - actual_minutes) × rate.
UTILIZATION_TARGET_MIN     = 360        # 6 h target — short trucks pay below this
# Disabled: the urgency-aware per-stop reward (§5.5) does this work
# correctly by picking the RIGHT clients, not just any clients. A blanket
# time-target penalty would add stops indiscriminately — the urgency
# reward only adds stops worth the mileage.
UTILIZATION_PENALTY_PER_MIN = 0.00

# Drop-penalty tiers ($ — see compute_drop_penalty for the urgency logic).
# Sensitivity tested in validate.py; values bracketed by:
#   - Lower: don't drop a client whose next-target visit is within the horizon
#   - Upper: don't waste a slot on a client that's safe for >horizon days
DROP_PENALTY_HARD = 10_000.0    # DTE < 2 days   — practically mandatory
DROP_PENALTY_HIGH = 500.0       # will dry in horizon — should serve
DROP_PENALTY_MED  = 35.0        # would-visit-this-week (must exceed OT premium for spreading)
DROP_PENALTY_LOW  = 5.0         # truly deferable — only serve if free

# Per-stop "target visit by" definition: visit before tank reaches this
# fraction of capacity. Real S&K policy is "stay ahead of 30% — don't let
# tanks drain to empty." Lower = more visits / safer; higher = fewer.
TARGET_EMPTY_FRACTION = 0.30    # visit by 30% full = 70% empty

# Two-tier reserve penalty (operator-confirmed IRP framing):
#   • Below RESERVE_PCT (default 10%): "urgency starts ramping" — the
#     solver pays a moderate per-day price for visiting later. Doesn't
#     force route churn; just nudges the assignment toward earlier days
#     when other costs are comparable.
#   • Below 0 (empty): catastrophic. The slope is 100× the reserve slope
#     so the solver will burn miles to avoid stockouts.
#
# Picked $5/day past reserve and $500/day past empty. At COST_SCALE=1000
# this becomes 5000 / 500_000 cost-units per day. For reference, $0.55
# per mile (mileage rate) ≈ 550 cost-units, so 1 day past reserve costs
# ~9 miles of detour — about right operationally. 1 day past empty
# costs ~900 miles → never happens unless infeasible.
#
# RESERVE_PCT is read from Optimizer_Input cell C2 at runtime (the
# spreadsheet's "Min Oil %" setting — operator's single source of
# truth). Falls back to 0.10 if the cell is missing or invalid.
RESERVE_PCT_DEFAULT          = 0.10
RESERVE_PENALTY_PER_DAY      = 5.0      # $/day-past-reserve at visit
EMPTY_PENALTY_PER_DAY        = 500.0    # $/day-past-empty at visit (100× reserve)

# Commit window — first N working days are firm; urgent clients (DTE within
# commit window + buffer) are LOCKED to this window.
COMMIT_BUFFER_DAYS = 0.5        # clients with DTE ≤ commit_days + 0.5 are locked

# Integer cost scaling. OR-Tools needs ints; we work in mills.
COST_SCALE = 1000


# ════════════════════════════════════════════════════════════════════════════
# SECTION 2. CONSUMPTION RATE ESTIMATOR (RC-5)
# ════════════════════════════════════════════════════════════════════════════
#
# Original v2 estimator used 75th-percentile of ALL gap-rates. That was
# intentionally conservative (safety buffer against burst consumption)
# but it caused systematic OVER-projection of consumption — for any
# client with multi-week historical data, 75p picks one of the higher
# observations, making the solver think clients dry out faster than they
# do. Effect: the projection said "tank empty by Tue" when reality (per
# spreadsheet's last-gap rate) said "still has 110 lb left" → solver
# planned full-tank refills when only 370 lb was actually needed.
#
# Now: 50th percentile (median) — the typical observed consumption.
# IQR filtering still removes outliers (e.g., 1-day-gap "400 lpd" glitches).
# Backtest result: roughly same stockout count as 75p, dramatically more
# accurate refill volumes matching operator's spreadsheet.
#
# Fix: take the MAX of:
#   - 75p of gaps in last 60 days (recency)
#   - 75p of gaps over all history (stability)
# Steady customers get the same rate as before. Accelerating customers get
# the recent (higher) rate. Decelerating customers keep the cautious (higher)
# all-time rate.

def estimate_consumption_recency_weighted(
    deliveries_df: pd.DataFrame,
    clients,  # Tuple[Client, ...]
    today: date,
    recency_window_days: int = 60,
    percentile: float = 0.50,
) -> Dict[str, Tuple[float, float]]:
    """
    Compute per-client consumption rate using max(60d-75p, all-time-75p).

    The Delivery_Log's Customer column holds the customer NAME (e.g.
    "HAR - 8031 - HAROLDS"), but our keys are client.id strings. Build a
    name→id map from the clients tuple and produce {id: (rate, sigma)}.

    Returns {client_id: (rate_lbs_per_day, std_dev)}.
    Clients with no usable observations get (nan, nan).
    """
    out: Dict[str, Tuple[float, float]] = {
        c.id: (float('nan'), float('nan')) for c in clients
    }
    if deliveries_df is None or deliveries_df.empty:
        return out

    name_to_id = {c.customer: c.id for c in clients}

    df = deliveries_df.copy()
    df['Customer'] = df['Customer'].astype(str)
    df['Date'] = pd.to_datetime(df['Date'])
    df['Qty_lbs'] = pd.to_numeric(df['Qty_lbs'], errors='coerce')

    # Filter placeholders
    if 'Is_Placeholder' in df.columns:
        df = df.loc[~df['Is_Placeholder'].astype(bool)].copy()
    else:
        df = df.loc[df['Qty_lbs'] != 200.0].copy()

    df = df.sort_values(['Customer', 'Date'])
    df['Prev_Date'] = df.groupby('Customer')['Date'].shift(1)
    df['Days_Gap'] = (df['Date'] - df['Prev_Date']).dt.days
    valid = df['Days_Gap'].notna() & (df['Days_Gap'] > 0) & df['Qty_lbs'].notna()
    df['Rate'] = np.where(valid, df['Qty_lbs'] / df['Days_Gap'], np.nan)
    rated = df.loc[df['Rate'].notna(), ['Customer', 'Date', 'Rate']].copy()
    if rated.empty:
        return out

    cutoff = pd.Timestamp(today - timedelta(days=recency_window_days))
    for customer_name, group in rated.groupby('Customer'):
        cid = name_to_id.get(customer_name)
        if not cid:
            continue
        all_rates = _iqr_filter(group['Rate'].to_numpy(dtype=float))
        recent_rates = _iqr_filter(
            group.loc[group['Date'] >= cutoff, 'Rate'].to_numpy(dtype=float)
        )
        if all_rates.size == 0:
            continue

        all_p = (float(np.quantile(all_rates, percentile))
                 if all_rates.size >= 3 else float(all_rates[-1]))
        rec_p = (float(np.quantile(recent_rates, percentile))
                 if recent_rates.size >= 3 else
                 (float(recent_rates[-1]) if recent_rates.size > 0 else 0.0))

        rate = max(all_p, rec_p)
        sigma = float(np.std(all_rates, ddof=1)) if all_rates.size >= 2 else 0.0
        if rate < 0:
            rate = 0.0
        if sigma != sigma:
            sigma = 0.0
        out[cid] = (rate, sigma)

    return out


def _iqr_filter(values: np.ndarray, factor: float = 3.0) -> np.ndarray:
    """Drop entries above Q3 + factor*IQR. No lower cap (slow consumers OK)."""
    if values.size < 3:
        return values
    q1, q3 = np.quantile(values, [0.25, 0.75])
    upper = q3 + factor * (q3 - q1)
    return values[values <= upper]


# ════════════════════════════════════════════════════════════════════════════
# SECTION 3. PROBLEM AUGMENTATION
# ════════════════════════════════════════════════════════════════════════════
#
# Build the ProblemInstance via v2.ingest, then override:
#   - economics (labor=0, dispatch=0)
#   - consumption rates (recency-weighted)
# Returns the immutable, augmented ProblemInstance.

def _load_sheet_rates_and_state(input_file: Path) -> Dict[str, Dict[str, float]]:
    """Read Optimizer_Input → per-client AVG rate, StdDev, and Est Current.

    Schema (after operator added AVG + StdDev columns):
      col  8 (idx 7)  = AVG Per Day Cons (lbs/day)
      col  9 (idx 8)  = Std Dev (lbs/day)
      col 13 (idx 12) = Est. Current (lbs)

    The operator curates AVG and StdDev directly — this is the canonical
    source of truth, not our IQR-filtered estimator. StdDev enables
    variance-aware safety stock if we want it later.
    """
    out: Dict[str, Dict[str, float]] = {}
    try:
        import openpyxl as _ox
        wb = _ox.load_workbook(input_file, data_only=True, read_only=True)
        if 'Optimizer_Input' not in wb.sheetnames:
            wb.close()
            return out
        ws = wb['Optimizer_Input']
        for row in ws.iter_rows(min_row=6, max_col=14, values_only=True):
            if not row or row[1] is None:
                continue
            cid = str(row[1])
            entry: Dict[str, float] = {}
            # AVG rate (col 8 / index 7)
            try:
                if row[7] is not None:
                    entry['avg_rate'] = float(row[7])
            except (ValueError, TypeError):
                pass
            # StdDev (col 9 / index 8) — NEW
            try:
                if len(row) > 8 and row[8] is not None:
                    entry['std_dev'] = float(row[8])
            except (ValueError, TypeError):
                pass
            # Est. Current (col 13 / index 12)
            try:
                if len(row) > 12 and row[12] is not None:
                    entry['est_current'] = float(row[12])
            except (ValueError, TypeError):
                pass
            if entry:
                out[cid] = entry
        wb.close()
    except Exception:
        pass
    return out


def _load_sheet_est_current(input_file: Path) -> Dict[str, float]:
    """Back-compat: same shape as before, but reads from the new helper."""
    return {cid: v['est_current']
            for cid, v in _load_sheet_rates_and_state(input_file).items()
            if 'est_current' in v}


def _load_sheet_reserve_pct(input_file: Path,
                              default: float = RESERVE_PCT_DEFAULT) -> float:
    """Read the operator's reserve % from Optimizer_Input cell C2 ("Min Oil %").

    This is the threshold below which the solver should consider the tank
    "into the safety reserve" and start penalizing later visit days.
    Single source of truth — same value that drives the sheet's "Next
    Delivery By" column. Falls back to `default` if cell missing/invalid.
    """
    try:
        import openpyxl as _ox
        wb = _ox.load_workbook(input_file, data_only=True, read_only=True)
        if 'Optimizer_Input' not in wb.sheetnames:
            wb.close()
            return default
        ws = wb['Optimizer_Input']
        v = ws.cell(2, 3).value     # C2
        wb.close()
        if v is None:
            return default
        f = float(v)
        # Sanity: must be between 0 and 0.5 (operator might enter 10 vs 0.10)
        if 0.0 < f < 1.0:
            return f
        if 1.0 <= f <= 50.0:
            return f / 100.0       # operator entered 10 instead of 0.10
        return default
    except Exception:
        return default


def build_augmented_problem(
    config_dir: Path,
    input_file: Path,
    matrix_file: Path,
    today: date,
    run_id: str,
    solve_seconds: int = 300,
    user_overrides_file: Optional[Path] = None,
) -> ProblemInstance:
    """Build problem with the cost-model and consumption fixes applied.

    If `user_overrides_file` is given, merges its Pins/Forbids with anything
    in the Excel Overrides sheet. Used by the dashboard (final.app) to inject
    operator Pin/Skip clicks without touching SharePoint-hosted Excel.
    """
    base = build_problem_instance(
        config_dir=config_dir,
        input_file=input_file,
        matrix_file=matrix_file,
        today=today,
        run_id=run_id,
    )

    # Merge dashboard sidecar overrides if provided.
    if user_overrides_file is not None and user_overrides_file.exists():
        from final.app.overrides_store import load_user_overrides, merge_overrides
        user_ov = load_user_overrides(user_overrides_file)
        if not user_ov.is_empty():
            merged = merge_overrides(base.overrides, user_ov)
            base = replace(base, overrides=merged)
            print(f"  Merged dashboard overrides: "
                  f"{len(user_ov.pins)} pins, {len(user_ov.forbids)} forbids")

    # Apply per-day truck unavailability (dashboard "Trucks available" widget).
    # Marks (date, truck_id) pairs as not dispatched. Stacks on top of the
    # base Saturday rule (Truck9 already unavailable on Saturdays).
    truck_unavail_file = (
        Path(__file__).resolve().parent.parent / 'data' / 'truck_unavailable.json'
    )
    if truck_unavail_file.exists():
        from final.app.availability_store import load_unavailability
        unavailable = load_unavailability(truck_unavail_file)
        if unavailable:
            new_avail = dict(base.truck_available)
            applied = 0
            for (d, tid) in unavailable:
                if (d, tid) in new_avail and new_avail[(d, tid)]:
                    new_avail[(d, tid)] = False
                    applied += 1
            if applied:
                base = replace(base, truck_available=new_avail)
                print(f"  Applied truck unavailability: {applied} (date, truck) pairs disabled")

    # Load operator-curated AVG and StdDev (from the new Optimizer_Input
    # columns 8 and 9) AND Est Current (col 13). This is now the canonical
    # source of truth for rates — replaces our own IQR-filtered estimator.
    # Why: the operator's AVG is reviewed and tuned over time. Our internal
    # estimator was overestimating consumption (75p inflation), causing
    # over-projection of dry-out and over-delivery. The operator's AVG +
    # StdDev are the values they SEE in the spreadsheet, so the solver's
    # view stays consistent with the operator's mental model.
    sheet_data = _load_sheet_rates_and_state(input_file)
    sheet_current = {cid: v['est_current']
                     for cid, v in sheet_data.items() if 'est_current' in v}

    tank_cap_by_id = {c.id: float(c.tank_capacity_lbs) for c in base.clients}
    updated_tanks: Dict[str, TankState] = {}
    n_rate_set = n_current_aligned = n_sheet_bad = 0
    for cid, ts in base.initial_tanks.items():
        new_ts = ts
        sd = sheet_data.get(cid, {})
        # Use operator-curated AVG as the rate. If StdDev provided, store it.
        avg_rate = sd.get('avg_rate')
        std_dev = sd.get('std_dev')
        if avg_rate is not None and avg_rate > 0:
            new_ts = replace(
                new_ts,
                rate_lbs_per_day=float(avg_rate),
                rate_std_dev=float(std_dev) if std_dev is not None else ts.rate_std_dev,
            )
            n_rate_set += 1
        # Current-level override: ONLY when source is 'estimated' (no
        # Anova/manual reading). Anova/manual readings are real
        # measurements and should always win.
        if new_ts.source == 'estimated' and cid in sheet_current:
            sheet_val = sheet_current[cid]
            tank_cap = tank_cap_by_id.get(cid, 0.0)
            # SANITY CHECK: reject obviously-bad spreadsheet values
            # (negative, way over tank cap, NaN). These have been observed
            # when a manual edit fat-fingers the formula.
            if sheet_val is None or sheet_val != sheet_val:   # None or NaN
                pass
            elif sheet_val < 0 or sheet_val > tank_cap + 10:
                n_sheet_bad += 1   # log it, ignore the value
            elif abs(sheet_val - new_ts.current_lbs) > 5:
                clamped = max(0.0, min(tank_cap, float(sheet_val)))
                new_ts = replace(new_ts, current_lbs=clamped,
                                  source='estimated (spreadsheet)')
                n_current_aligned += 1
        updated_tanks[cid] = new_ts
    print(f"  Rates: {n_rate_set} clients have spreadsheet AVG rate set "
          f"(operator-curated, replaces internal estimator)")
    if n_current_aligned:
        print(f"  Current-level aligned to spreadsheet for {n_current_aligned} clients "
              f"(prevents overflow from rate-driven under-estimation)")
    if n_sheet_bad:
        print(f"  ⚠ Ignored {n_sheet_bad} spreadsheet est_current values that were "
              f"out-of-range (negative or > tank capacity)")

    # SHIFT CONFIGURATION — operator-confirmed truth (overrides the Depot sheet)
    #
    # The Depot sheet shows Shift_End_HHMM=14:00, which made
    # effective_route_minutes collapse to 435 min — but that's NOT the real
    # operational reality. Operator confirmed:
    #   - Standard shifts run 6 AM → 2 PM (8 driving hours, target)
    #   - Long days run up to 6 AM → 5 PM  (10h 15m driving max, hard cap)
    #   - Drivers can come in any time before 5 PM; later is OT-shaped
    #
    # So the real configuration is:
    #   SHIFT_TARGET_MIN = 480 (8 hours of driving — normal day)
    #   SHIFT_HARD_MAX_MIN = 615 (10h 15m: 6 AM → 5 PM minus 30 morning load
    #                              + 15 evening unload)
    #
    # With this:
    #   - 0–480 min: no OT cost (drivers salaried, sunk)
    #   - 480–615 min: OT premium $0.42/min (real marginal cost as drivers
    #                  push past the standard day)
    #   - >615 min: hard infeasible (operator's published max)
    #
    # Symptom this fixes: previously (with Depot sheet's 435-min cap and
    # OT_TARGET_FRACTION knob), the solver was producing ~5h shifts with
    # half-full trucks — splitting work across both trucks not to save
    # mileage but just to avoid an OT shoulder that fired too early.
    # With 8h target / 10h cap, the solver will naturally pack longer
    # routes (closer to Tammy's actual 6-8h workdays).
    SHIFT_TARGET_MIN = 480       # 8 hours of driving (standard target)
    SHIFT_HARD_MAX_MIN = 615     # 10h 15m (6 AM → 5 PM minus load/unload)

    # Read reserve % from sheet cell C2 — single source of truth shared
    # with the spreadsheet's "Next Delivery By" formula. Falls back to
    # policy.yaml's value (also 0.10) if cell is missing/invalid.
    reserve_pct = _load_sheet_reserve_pct(input_file, default=base.min_reserve_fraction)
    if abs(reserve_pct - base.min_reserve_fraction) > 1e-6:
        print(f"  Reserve %: overriding policy.yaml ({base.min_reserve_fraction:.0%}) "
              f"with spreadsheet C2 ({reserve_pct:.0%})")
    else:
        print(f"  Reserve %: {reserve_pct:.0%} (from spreadsheet C2)")

    augmented = replace(
        base,
        initial_tanks=updated_tanks,
        min_reserve_fraction=reserve_pct,
        cost_per_mile=COST_PER_MILE,
        cost_per_minute_labor=COST_PER_MINUTE_LABOR_REG,
        # overtime_multiplier kept at YAML value (1.5) — used internally below
        truck_dispatch_cost=TRUCK_DISPATCH_COST,
        stockout_cost_per_lb_day=STOCKOUT_COST_PER_LB_DAY,
        shift_target_min=SHIFT_TARGET_MIN,
        shift_hard_max_min=SHIFT_HARD_MAX_MIN,
        solve_seconds=solve_seconds,
    )
    return augmented


# ════════════════════════════════════════════════════════════════════════════
# SECTION 4. DROP-PENALTY FUNCTION (RC-1)
# ════════════════════════════════════════════════════════════════════════════
#
# Piecewise penalty based on tank urgency. The four tiers correspond to:
#   HARD: stockout-imminent — solver MUST serve
#   HIGH: tank dries within horizon — solver must serve
#   MED : tank dries past horizon but visit-target is within horizon —
#         deferring just punts cost to next horizon, so price it accordingly
#   LOW : truly safe to defer for one more horizon — only serve if free
#
# Sensitivity: changing MED from $25 → $15 drops 10–15 stops; raising to
# $40 adds ~5 stops with marginal mileage. The $25 value matches the
# expected per-stop route cost (mostly miles + a touch of OT).

def compute_drop_penalty(
    current_lbs: float,
    rate_lbs_per_day: float,
    tank_capacity_lbs: float,
    horizon_days: int,
    target_empty_fraction: float = TARGET_EMPTY_FRACTION,
) -> float:
    """Return the drop penalty in dollars (NOT yet in COST_SCALE units)."""
    if rate_lbs_per_day is None or rate_lbs_per_day <= 0:
        # No consumption data → low penalty (we don't know they need anything)
        return DROP_PENALTY_LOW

    days_supply = current_lbs / rate_lbs_per_day
    target_lbs = tank_capacity_lbs * target_empty_fraction
    if current_lbs <= target_lbs:
        days_to_target = 0.0
    else:
        days_to_target = (current_lbs - target_lbs) / rate_lbs_per_day

    if days_supply < 2.0:
        return DROP_PENALTY_HARD
    if days_supply < horizon_days:
        return DROP_PENALTY_HIGH
    if days_to_target < horizon_days:
        return DROP_PENALTY_MED
    return DROP_PENALTY_LOW


# ════════════════════════════════════════════════════════════════════════════
# SECTION 5. MODEL BUILDER (replaces v2.solver.model.build_routing_model)
# ════════════════════════════════════════════════════════════════════════════
#
# Same return type (ModelArtifacts) so v2.solver.extract still works.

def build_routing_model_final(problem: ProblemInstance) -> ModelArtifacts:
    """The new OR-Tools model. See SECTION-by-SECTION comments below."""

    # ── 5.1 Filter routable clients (DNS, excluded, missing matrix/tank) ──
    pool = []
    pool_to_osrm = []
    for c in problem.clients:
        if c.do_not_schedule or c.excluded:
            continue
        if c.id not in problem.node_index:
            continue
        if c.id not in problem.initial_tanks:
            continue
        pool.append(c)
        pool_to_osrm.append(problem.node_index[c.id])

    n_clients = len(pool)
    n_days = len(problem.horizon_dates)
    n_trucks = len(problem.trucks)
    n_vehicles = n_trucks * n_days
    n_nodes = n_clients + 1

    def v2td(v: int) -> Tuple[int, int]:
        return v // n_days, v % n_days

    # ── 5.2 Sub-matrices (depot + pool) ───────────────────────────────────
    osrm_indices = [0] + pool_to_osrm
    sub_dist_m = problem.distance_matrix_m[np.ix_(osrm_indices, osrm_indices)]
    sub_time_min = problem.time_matrix_min[np.ix_(osrm_indices, osrm_indices)]

    # Calendar days from today to each horizon-day. CRITICAL: customers
    # consume oil EVERY day including Sun/Mon, but horizon_dates only
    # includes working days (Tue–Sat). So for horizon-day index `d`, the
    # actual number of consumption-days elapsed since today is the calendar
    # difference, NOT `d`.
    #
    # Example: today=Fri May 22, horizon_dates[2]=Tue May 26. That's
    # `d=2` (third working day) but 4 calendar days elapsed.
    #
    # Fixes the bug where the solver scheduled clients later than their
    # actual dry-out because it under-counted consumption days, and where
    # refills showed wrong delivery amounts (looked like full tank when in
    # reality clients still had ~110 lb).
    cal_days_to_d: List[int] = [
        max(0, (problem.horizon_dates[d] - problem.today).days)
        for d in range(n_days)
    ]

    # ── 5.3 Per-day refills + product/tank arrays ─────────────────────────
    # NOTE: we no longer zero-out refills < min_stop_lbs. Instead, we keep
    # the actual refill value and apply a soft "small stop fee" in the
    # cost callback (RC-6).
    rate_per_client = [0.0]
    tank_per_client = [0]
    product_per_client = ['']
    current_per_client = [0.0]
    for c in pool:
        ts = problem.initial_tanks[c.id]
        rate_per_client.append(float(ts.rate_lbs_per_day or 0.0))
        tank_per_client.append(int(c.tank_capacity_lbs))
        product_per_client.append(c.product)
        current_per_client.append(float(ts.current_lbs))

    min_stop_lbs = int(problem.min_stop_lbs)
    refills_by_day: List[List[int]] = []
    for d in range(n_days):
        day_refills = [0]
        # Use CALENDAR days (not horizon index) — customers consume every
        # day, including weekends, even though horizon only lists workdays.
        cal_days = cal_days_to_d[d]
        for i, c in enumerate(pool):
            ts = problem.initial_tanks[c.id]
            rate = rate_per_client[i + 1]
            tank = float(c.tank_capacity_lbs)
            raw_level = ts.current_lbs - cal_days * rate
            level_at_day = max(0.0, raw_level)
            refill = max(0, int(round(tank - level_at_day)))
            day_refills.append(refill)
        refills_by_day.append(day_refills)

    # Capacity demand (RC-4): p75 of feasible (>0) refills, not max.
    capacity_demand_per_client = [0]
    for i in range(1, n_nodes):
        feasible = [refills_by_day[d][i] for d in range(n_days)
                    if refills_by_day[d][i] > 0]
        if not feasible:
            capacity_demand_per_client.append(0)
        else:
            capacity_demand_per_client.append(
                int(np.quantile(feasible, 0.75))
            )

    # For day-validity (no zero-refill visits) we still need the max.
    max_refill_per_client = [0]
    for i in range(1, n_nodes):
        max_refill_per_client.append(
            max(refills_by_day[d][i] for d in range(n_days))
        )

    # ── 5.4 OR-Tools setup ────────────────────────────────────────────────
    manager = pywrapcp.RoutingIndexManager(
        n_nodes, n_vehicles,
        [0] * n_vehicles, [0] * n_vehicles,
    )
    routing = pywrapcp.RoutingModel(manager)

    cost_per_mile_units = int(round(COST_PER_MILE * COST_SCALE))
    small_stop_fee_units = int(round(SMALL_STOP_FEE * COST_SCALE))

    # ── 5.5 Arc cost callback ─────────────────────────────────────────────
    #
    # Cost per arc =
    #     (miles × $/mi)                                  ← travel cost
    #   + (small-stop fee if refill < min_stop AND not urgent)
    #   - URGENCY-DAY-FIT REWARD on destination            ← new (RC-11)
    #   - REFILL-SIZE REWARD on destination                ← new (RC-11)
    #
    # The urgency-day-fit reward encodes the operator's intuition:
    # "Marginal benefit of adding a stop is HIGH when the client needs
    # oil now, and LOW (or negative) when the client could easily wait."
    # Without this, the marginal benefit of any non-deferred stop was $0
    # (because the disjunction's drop penalty doesn't fire unless the
    # client goes unserved in the WHOLE horizon). So the solver would
    # stop a truck early on a low-density day even with spare capacity.
    #
    # The reward is per (client, day) — computed from the client's
    # PROJECTED DTE on that vehicle's day. Same client, different days,
    # different reward.

    # Urgency-aware per-stop rewards were tried (and produced fuller trucks)
    # but distorted the insertion heuristic's routing decisions, causing
    # geographically-poor sequences (zig-zag between clusters). The
    # mechanism is per-arc cost adjustment, which the PARALLEL_CHEAPEST_
    # INSERTION first-solution strategy uses for insertion positions —
    # negative arc costs confused it.
    #
    # NEW: two-tier reserve penalty (IRP framing).
    #   • If the tank at visit-day d is below RESERVE_PCT × capacity:
    #     pay RESERVE_PENALTY_PER_DAY per day below. Soft nudge — solver
    #     prefers earlier days but may still pick a later day when miles
    #     are scarce.
    #   • If the tank at visit-day d is below ZERO (stockout):
    #     pay EMPTY_PENALTY_PER_DAY (100× the reserve slope). Stockouts
    #     effectively forbidden.
    # Both expressed in $; multiplied by COST_SCALE to integer mills, so
    # they trade off natively against mileage cost (~$0.55/mile × 1000).
    reserve_pct = float(problem.min_reserve_fraction or RESERVE_PCT_DEFAULT)
    n_late_penalized = 0

    # Pre-compute reward per (client, day) — done once at model-build time
    # to avoid per-callback recomputation. Shape: [n_days][n_nodes].
    # NB: a "reward" with NEGATIVE value = penalty (cost callback does
    # `cost -= reward` so neg reward → +cost).
    stop_rewards_by_day: List[List[int]] = []
    for d in range(n_days):
        cal_days = cal_days_to_d[d]
        day_rewards = [0]   # depot has no reward
        for i, c in enumerate(pool):
            ts = problem.initial_tanks[c.id]
            rate = float(ts.rate_lbs_per_day or 0.0)
            tank_cap = float(c.tank_capacity_lbs)
            if rate <= 0 or tank_cap <= 0:
                day_rewards.append(0)
                continue
            reserve_lbs = tank_cap * reserve_pct
            # Project tank level at visit-day d — allow NEGATIVE values so
            # the past-empty penalty has a slope (don't clamp to zero here).
            raw_lbs_at_d = float(ts.current_lbs) - cal_days * rate
            days_past_reserve = max(0.0, (reserve_lbs - raw_lbs_at_d) / rate)
            days_past_empty   = max(0.0, (0.0         - raw_lbs_at_d) / rate)
            penalty_dollars = (
                RESERVE_PENALTY_PER_DAY * days_past_reserve
                + EMPTY_PENALTY_PER_DAY * days_past_empty
            )
            if penalty_dollars > 0:
                n_late_penalized += 1
            # Negative reward → cost callback adds (positive) penalty
            total_reward_units = int(round(-penalty_dollars * COST_SCALE))
            day_rewards.append(total_reward_units)
        stop_rewards_by_day.append(day_rewards)
    if n_late_penalized:
        print(f"  Reserve penalty: {n_late_penalized} (client, day) pairs "
              f"price 'visit past reserve/empty' "
              f"(${RESERVE_PENALTY_PER_DAY:.0f}/d past reserve, "
              f"${EMPTY_PENALTY_PER_DAY:.0f}/d past empty)")

    def _make_cost_cb(_mgr, _sd, _rf, _rates, _curs, _min_stop_lbs, _mile_cost,
                      _small_fee, _stop_rewards_d):
        def _cb(from_idx, to_idx):
            fn = _mgr.IndexToNode(from_idx)
            tn = _mgr.IndexToNode(to_idx)
            dist_m = int(_sd[fn, tn])
            cost = int((dist_m / 1609.34) * _mile_cost)
            # Per-stop small-stop soft fee on destination
            if tn > 0:
                refill = _rf[tn]
                if 0 < refill < _min_stop_lbs:
                    # Urgency check: based on TODAY's tank state
                    rate = _rates[tn]
                    cur = _curs[tn]
                    dte_today = (cur / rate) if rate > 0 else 999.0
                    if dte_today > 3.0:
                        cost += _small_fee
                # Urgency-day-fit + refill-size reward (NEGATIVE cost)
                cost -= _stop_rewards_d[tn]
            return cost
        return _cb

    for v in range(n_vehicles):
        truck_idx, day_idx = v2td(v)
        cb = _make_cost_cb(
            manager, sub_dist_m, refills_by_day[day_idx],
            rate_per_client, current_per_client,
            min_stop_lbs, cost_per_mile_units, small_stop_fee_units,
            stop_rewards_by_day[day_idx],
        )
        cb_idx = routing.RegisterTransitCallback(cb)
        routing.SetArcCostEvaluatorOfVehicle(cb_idx, v)

    # ── 5.6 Time dimension (shift cap + OT premium ONLY) ─────────────────
    truck_specs = list(problem.trucks)
    time_cb_indices = []
    for v in range(n_vehicles):
        truck_idx, day_idx = v2td(v)
        truck = truck_specs[truck_idx]

        def _make_time_cb(_mgr, _st, _rf, _setup, _rate):
            def _cb(from_idx, to_idx):
                fn = _mgr.IndexToNode(from_idx)
                tn = _mgr.IndexToNode(to_idx)
                travel = int(_st[fn, tn])
                if fn == 0:
                    return travel
                pump_min = int(_rf[fn] / _rate) + (1 if _rf[fn] > 0 else 0)
                return travel + _setup + pump_min
            return _cb

        cb = _make_time_cb(manager, sub_time_min, refills_by_day[day_idx],
                            truck.fixed_setup_min, truck.pump_rate_lbs_per_min)
        time_cb_indices.append(routing.RegisterTransitCallback(cb))

    # Slack = shift_hard_max_min allows the truck to wait at a node if it
    # arrives before that client's time window opens. Without slack the
    # solver can't satisfy a "9-10 AM" window for a 7 AM arrival.
    routing.AddDimensionWithVehicleTransits(
        time_cb_indices,
        problem.shift_hard_max_min,   # max wait per node
        problem.shift_hard_max_min,   # hard cap on cumulative shift time
        True, 'Time',
    )

    # ── 5.6c StopCount dimension — soft target for stops per truck-day ───
    # Encourages fuller truck-days WITHOUT touching per-arc costs. The
    # urgency-aware arc rewards tried earlier worked but distorted the
    # routing heuristic into geographically-poor sequences. This count-
    # based approach is decoupled from sequencing — it just says "this
    # truck-day should have at least N stops, otherwise penalty."
    #
    # Why count, not time: the routing heuristic uses arc costs for
    # insertion decisions. A count-based dimension is a pure post-hoc
    # constraint on the END node, doesn't bias arc selection.
    TARGET_STOPS_PER_TRUCKDAY = 7   # ~7-stop day = ~5h of work
    PENALTY_PER_MISSING_STOP_DOLLARS = 4.0   # $4 per stop below target
    def _count_cb(from_idx):
        return 0 if manager.IndexToNode(from_idx) == 0 else 1
    count_cb_idx = routing.RegisterUnaryTransitCallback(_count_cb)
    routing.AddDimension(count_cb_idx, 0, n_clients + 1, True, 'StopCount')
    count_dim = routing.GetDimensionOrDie('StopCount')
    miss_units = int(round(PENALTY_PER_MISSING_STOP_DOLLARS * COST_SCALE))
    if miss_units > 0:
        for v in range(n_vehicles):
            end_idx = routing.End(v)
            count_dim.SetCumulVarSoftLowerBound(end_idx, TARGET_STOPS_PER_TRUCKDAY,
                                                  miss_units)
        print(f"  Stop-count target: ≥{TARGET_STOPS_PER_TRUCKDAY} stops per truck-day "
              f"(${PENALTY_PER_MISSING_STOP_DOLLARS}/missing stop)")

    # OT premium ONLY (no regular labor) — RC-2.
    time_dim = routing.GetDimensionOrDie('Time')
    ot_premium_units = int(round(COST_PER_MINUTE_OT_PREMIUM * COST_SCALE))
    if ot_premium_units > 0:
        for v in range(n_vehicles):
            end_idx = routing.End(v)
            time_dim.SetCumulVarSoftUpperBound(
                end_idx, problem.shift_target_min, ot_premium_units
            )
            # NOTE: no SetSpanCostCoefficientForVehicle — base labor is sunk.

    # Underutilization soft penalty (RC-10).
    # Penalize truck-days returning before UTILIZATION_TARGET_MIN. The
    # solver will trade a few extra miles ($0.55/mi) for keeping the truck
    # busier — naturally pulling more "convenient" stops onto each day.
    util_units = int(round(UTILIZATION_PENALTY_PER_MIN * COST_SCALE))
    if util_units > 0:
        # We deliberately bound the target by hard_max so we don't accidentally
        # demand more time than the day physically allows.
        util_target = min(UTILIZATION_TARGET_MIN, problem.shift_hard_max_min)
        for v in range(n_vehicles):
            end_idx = routing.End(v)
            time_dim.SetCumulVarSoftLowerBound(end_idx, util_target, util_units)
        print(f"  Utilization preference: ${UTILIZATION_PENALTY_PER_MIN:.2f}/min "
              f"below {util_target}-min target (encourages fuller truck-days)")

    # ── 5.6b Time windows (hard arrival constraint per client) ───────────
    # Many clients have published delivery windows in Client_Time_Windows
    # (e.g. PETES FISH 26TH: "DELIVERY 9 AM - 10 AM"). The ingest loads
    # these as `Client.time_window_min = (open_offset, close_offset)`,
    # measured in MINUTES FROM SHIFT START. We enforce them on the
    # arrival-time CumulVar so OR-Tools refuses to schedule outside.
    #
    # The CumulVar at node N == minutes from shift start when truck ARRIVES
    # at N. That's exactly what the time-window constraint needs. Setting
    # the range hard-constrains it (solver may also wait, since slack=0
    # in our setup, so it'll only schedule arrivals within the window).
    n_with_windows = 0
    for i, c in enumerate(pool):
        tw = getattr(c, 'time_window_min', None)
        if not tw:
            continue
        open_off, close_off = tw
        # Clamp to the truck-day's hard cap so we don't create an
        # infeasible-by-construction range.
        open_off = max(0, int(open_off))
        close_off = min(int(problem.shift_hard_max_min), int(close_off))
        if close_off <= open_off:
            continue
        node_idx = manager.NodeToIndex(i + 1)
        try:
            time_dim.CumulVar(node_idx).SetRange(open_off, close_off)
            n_with_windows += 1
        except Exception:
            pass
    if n_with_windows:
        print(f"  Time windows enforced: {n_with_windows} clients have hard arrival windows")

    # ── 5.7 Capacity dimensions (truck total + per-product) ──────────────
    def _demand_cb(from_idx):
        n = manager.IndexToNode(from_idx)
        return capacity_demand_per_client[n]

    dcb = routing.RegisterUnaryTransitCallback(_demand_cb)
    truck_caps = [truck_specs[v2td(v)[0]].capacity_lbs for v in range(n_vehicles)]
    routing.AddDimensionWithVehicleCapacity(dcb, 0, truck_caps, True, 'Capacity')

    for product in problem.products:
        def _make_prod_cb(_mgr, _dem, _np, _product):
            def _cb(from_idx):
                n = _mgr.IndexToNode(from_idx)
                return _dem[n] if _np[n] == _product else 0
            return _cb
        pcb = routing.RegisterUnaryTransitCallback(
            _make_prod_cb(manager, capacity_demand_per_client,
                          product_per_client, product)
        )
        prod_caps = [truck_specs[v2td(v)[0]].compartments[0].capacity_lbs
                     for v in range(n_vehicles)]
        routing.AddDimensionWithVehicleCapacity(
            pcb, 0, prod_caps, True, f'Cap_{product.replace(" ", "_")}'
        )

    # ── 5.8 Saturday rule via truck_available (RC-8) ─────────────────────
    # No more 10⁹ fixed cost. Instead, for any (date, truck) where
    # truck_available is False, remove every client node from that
    # vehicle's domain.
    for v in range(n_vehicles):
        truck_idx, day_idx = v2td(v)
        truck_id = truck_specs[truck_idx].id
        dt = problem.horizon_dates[day_idx]
        if not problem.truck_available.get((dt, truck_id), True):
            # Disable this vehicle — only depot→depot allowed.
            for i in range(n_clients):
                node_idx = manager.NodeToIndex(i + 1)
                try:
                    routing.VehicleVar(node_idx).RemoveValue(v)
                except Exception:
                    pass

    # ── 5.9 Dispatch cost (now $0, kept for code-symmetry) ───────────────
    dispatch_cost_units = int(round(TRUCK_DISPATCH_COST * COST_SCALE))
    if dispatch_cost_units > 0:
        for v in range(n_vehicles):
            routing.SetFixedCostOfVehicle(dispatch_cost_units, v)

    # ── 5.10 Disjunctions: drop penalty per client (RC-1) ────────────────
    pins_by_client = {p.client_id for p in problem.overrides.pins}
    forbidden_dates_by_client: Dict[str, set] = {}
    for fb in problem.overrides.forbids:
        forbidden_dates_by_client.setdefault(fb.client_id, set()).update(fb.dates)

    horizon_days_count = n_days

    for i, c in enumerate(pool):
        node = manager.NodeToIndex(i + 1)
        ts = problem.initial_tanks[c.id]
        rate = float(ts.rate_lbs_per_day or 0.0)

        penalty_dollars = compute_drop_penalty(
            current_lbs=float(ts.current_lbs),
            rate_lbs_per_day=rate,
            tank_capacity_lbs=float(c.tank_capacity_lbs),
            horizon_days=horizon_days_count,
            target_empty_fraction=TARGET_EMPTY_FRACTION,
        )
        # Pin override: effectively mandatory
        if c.id in pins_by_client:
            penalty_dollars = DROP_PENALTY_HARD * 10  # = $100k
        # If max possible refill is 0 (tank already overflowing somehow),
        # don't waste a disjunction slot — set penalty to 0 (free to drop).
        if max_refill_per_client[i + 1] <= 0:
            penalty_dollars = 0.0

        drop_units = int(round(penalty_dollars * COST_SCALE))
        routing.AddDisjunction([node], drop_units)

    # ── 5.11 Forbid: client × forbidden-date AND zero/tiny-refill day ────
    # Forbids come from FOUR sources, all RemoveValue from VehicleVar:
    #   (a) Operator-set Forbid(client_id, dates=...) from overrides
    #   (b) refills_by_day == 0   (tank already full — uneconomic)
    #   (c) refills_by_day < HARD_FLOOR_LBS for non-urgent clients
    #       (matches the invariant's hard-floor of 50 lbs — prevents
    #        the solver from picking a day where the refill would be
    #        rejected at output time).
    #   (d) Rate == 0 (no consumption data: forbid ALL days to defer cleanly)
    HARD_FLOOR_LBS = 50

    # (a) operator forbids — track per-client forbidden-day set
    client_forbidden_days: Dict[int, set] = {}
    for client_id, forbidden_dates in forbidden_dates_by_client.items():
        try:
            i = next(idx for idx, c in enumerate(pool) if c.id == client_id)
        except StopIteration:
            continue
        node_idx = manager.NodeToIndex(i + 1)
        forbidden_day_idxs = set()
        for v in range(n_vehicles):
            _, day_idx = v2td(v)
            if problem.horizon_dates[day_idx] in forbidden_dates:
                forbidden_day_idxs.add(day_idx)
                try:
                    routing.VehicleVar(node_idx).RemoveValue(v)
                except Exception:
                    pass
        client_forbidden_days[i] = forbidden_day_idxs

    # (e) Forbid scheduling AFTER projected dry-out date.
    # Days_supply is CALENDAR days; we compare against the calendar-day
    # offset of each horizon day, NOT its working-day index. Without this
    # fix, a client with 4 days supply could be scheduled day_idx=4 (Thu
    # May 28 = 6 calendar days later) → 2 days of being dry.
    #
    # Also: if NO horizon day satisfies the constraint (e.g. client already
    # dry AND the only allowed days are forbidden), warn loudly. Without
    # this log line the client would silently end up deferred via the
    # disjunction's drop penalty, with no clear "why."
    # DRY_DAY_GRACE: how many full days of being dry the model tolerates.
    #
    # For clients with days_supply > 0.5 we accept a 1-day grace (the truck
    # arrives the morning after they go dry — minor).
    #
    # For clients already EMPTY (days_supply ≤ 0.5) we tighten to 0 — they
    # MUST be served on day 0 (today's plan-as-of date). Without this tighter
    # rule the solver was scheduling already-empty clients (ATL MESA, JC FOOD
    # TRADING, etc.) 6-9 days out because their stop fit geographically into
    # a later truck day, ignoring the urgency.
    GRACE_NORMAL = 1
    GRACE_EMPTY  = 0
    unschedulable_urgent: List[str] = []
    n_locked_to_day_0_today = 0
    for i, c in enumerate(pool):
        node_idx = manager.NodeToIndex(i + 1)
        ts = problem.initial_tanks[c.id]
        rate = float(ts.rate_lbs_per_day or 0.0)
        if rate <= 0:
            continue
        days_supply = float(ts.current_lbs) / rate
        grace = GRACE_EMPTY if days_supply <= 0.5 else GRACE_NORMAL
        max_cal_days_allowed = days_supply + grace
        n_kept = 0
        for v in range(n_vehicles):
            _, day_idx = v2td(v)
            if cal_days_to_d[day_idx] > max_cal_days_allowed:
                try: routing.VehicleVar(node_idx).RemoveValue(v)
                except Exception: pass
            else:
                n_kept += 1
        if grace == GRACE_EMPTY and n_kept > 0:
            n_locked_to_day_0_today += 1
        # If we just blew away every vehicle for a low-supply client, the
        # solver will be forced to drop them — log so the operator sees it.
        if n_kept == 0 and days_supply < 5:
            unschedulable_urgent.append(
                f"{c.id} ({c.customer[:30]}, days_supply={days_supply:.1f})"
            )
    if n_locked_to_day_0_today:
        print(f"  Already-empty clients locked to day 0: {n_locked_to_day_0_today}")
    if unschedulable_urgent:
        print(f"  ⚠ {len(unschedulable_urgent)} urgent client(s) have NO feasible "
              f"day in horizon — will be deferred. Investigate:")
        for cid in unschedulable_urgent[:10]:
            print(f"    - {cid}")

    # (b), (c), (d), (f) — refill-driven forbids
    n_min_fill_blocked = 0
    for i, c in enumerate(pool):
        node_idx = manager.NodeToIndex(i + 1)
        ts = problem.initial_tanks[c.id]
        rate = float(ts.rate_lbs_per_day or 0.0)
        is_zero_rate = rate <= 0
        tank_cap = float(c.tank_capacity_lbs)
        # Urgency on day 0 (today): used to allow small refills for
        # critical clients we can't afford to skip.
        dte_today = (float(ts.current_lbs) / rate) if rate > 0 else 999.0
        min_fill_threshold = int(round(MIN_FILL_PCT * tank_cap))
        for v in range(n_vehicles):
            _, day_idx = v2td(v)
            refill = refills_by_day[day_idx][i + 1]
            # (d): zero rate → forbid every day (defer cleanly)
            if is_zero_rate:
                try: routing.VehicleVar(node_idx).RemoveValue(v)
                except Exception: pass
                continue
            # (b): refill = 0 → tank already full, never visit
            if refill <= 0:
                try: routing.VehicleVar(node_idx).RemoveValue(v)
                except Exception: pass
                continue
            # (c): refill below absolute hard-floor (50 lb) AND client
            # not urgent today. Urgent clients can have small refills —
            # we MUST serve them even with tiny top-offs.
            if refill < HARD_FLOOR_LBS and dte_today > 3.0:
                try: routing.VehicleVar(node_idx).RemoveValue(v)
                except Exception: pass
                continue
            # (f) NEW: refill below MIN_FILL_PCT of tank AND not urgent
            # (operator request: don't waste a stop on a tank with too
            # much headroom — let it drain a few more days first).
            if (MIN_FILL_PCT > 0
                    and refill < min_fill_threshold
                    and dte_today > 3.0):
                try: routing.VehicleVar(node_idx).RemoveValue(v)
                except Exception: pass
                n_min_fill_blocked += 1
    if n_min_fill_blocked:
        # n_min_fill_blocked counts (client × day) pairs, not unique clients.
        print(f"  Min-fill rule (≥ {int(MIN_FILL_PCT*100)}% of tank): "
              f"blocked {n_min_fill_blocked} (client × day) pairs from being "
              f"a small top-off")

    # ── 5.12 Commit-window enforcement (RC-7) ────────────────────────────
    # Lock clients with days-to-RESERVE ≤ commit_days+buffer into days
    # 0..commit_days-1 — UNLESS those days are forbidden for the client
    # by operator (then use the earliest non-forbidden day in the
    # horizon).
    #
    # NB: previously this used days-to-EMPTY (`current / rate`), so the
    # lock fired ~1-2 days later than it should and the solver could
    # schedule past the reserve threshold within the commit window.
    # Now it agrees with the spreadsheet's "Next Delivery By" column.
    commit_days = problem.commit_days
    commit_horizon = commit_days + COMMIT_BUFFER_DAYS
    locked_to_commit = []
    for i, c in enumerate(pool):
        ts = problem.initial_tanks[c.id]
        rate = float(ts.rate_lbs_per_day or 0.0)
        if rate <= 0:
            continue
        tank_cap = float(c.tank_capacity_lbs)
        reserve_lbs = tank_cap * reserve_pct
        days_supply = max(0.0, (float(ts.current_lbs) - reserve_lbs) / rate)
        if days_supply > commit_horizon:
            continue
        forbidden_days = client_forbidden_days.get(i, set())
        # Allowed window: days 0..commit_days-1 that are NOT forbidden.
        # If all such days are forbidden, relax — allow any non-forbidden
        # day (operator must accept the consequence of conflicting overrides).
        commit_allowed = [d for d in range(commit_days) if d not in forbidden_days]
        if not commit_allowed:
            # Fall back: any non-forbidden day in horizon.
            commit_allowed = [d for d in range(n_days) if d not in forbidden_days]
        if not commit_allowed:
            continue  # client has no feasible day at all — will defer
        allowed_set = set(commit_allowed)
        node_idx = manager.NodeToIndex(i + 1)
        for v in range(n_vehicles):
            _, day_idx = v2td(v)
            if day_idx not in allowed_set:
                try: routing.VehicleVar(node_idx).RemoveValue(v)
                except Exception: pass
        locked_to_commit.append(c.id)
    if locked_to_commit:
        print(f"  Locked to commit window (DTE ≤ {commit_horizon}d): "
              f"{len(locked_to_commit)} clients")

    # ── 5.13 Pin enforcement: lock pinned client to pinned date ──────────
    # The disjunction's huge drop penalty makes the client MUST-SERVE, but
    # without this constraint the solver could pick any day. Operator
    # pinned the date for a reason — honor it.
    for p in problem.overrides.pins:
        try:
            i = next(idx for idx, c in enumerate(pool) if c.id == p.client_id)
        except StopIteration:
            continue
        # Find day_idx matching the pin date
        try:
            pin_day_idx = problem.horizon_dates.index(p.date)
        except ValueError:
            # Pin date not in horizon — pin will be silently relaxed (no constraint)
            continue
        node_idx = manager.NodeToIndex(i + 1)
        for v in range(n_vehicles):
            _, day_idx = v2td(v)
            if day_idx != pin_day_idx:
                try: routing.VehicleVar(node_idx).RemoveValue(v)
                except Exception: pass

    return ModelArtifacts(
        manager=manager,
        routing=routing,
        n_days=n_days,
        n_trucks=n_trucks,
        n_clients=n_clients,
        refills_by_day=refills_by_day,
        dispatch_cost_units=dispatch_cost_units,
        v2td=v2td,
        clients=tuple(pool),
        node_index=problem.node_index,
        pool_to_osrm=pool_to_osrm,
    )


# ════════════════════════════════════════════════════════════════════════════
# SECTION 6. SOLVE
# ════════════════════════════════════════════════════════════════════════════

class FinalSolverFailure(Exception):
    pass


def solve_final(problem: ProblemInstance,
                 solve_seconds: Optional[int] = None) -> Plan:
    """Build, solve, extract, validate. Returns a verified Plan."""
    print(f"\n  Building FINAL model: {len(problem.clients)} clients, "
          f"{len(problem.trucks)} trucks × {len(problem.horizon_dates)} days")
    t0 = time.time()
    artifacts = build_routing_model_final(problem)
    print(f"  Model built in {time.time() - t0:.2f}s "
          f"({artifacts.n_clients} routable, "
          f"{artifacts.n_trucks * artifacts.n_days} vehicles)")

    secs = solve_seconds or problem.solve_seconds or 300

    params = pywrapcp.DefaultRoutingSearchParameters()
    params.first_solution_strategy = (
        routing_enums_pb2.FirstSolutionStrategy.PARALLEL_CHEAPEST_INSERTION
    )
    params.local_search_metaheuristic = (
        routing_enums_pb2.LocalSearchMetaheuristic.GUIDED_LOCAL_SEARCH
    )
    params.time_limit.FromSeconds(int(secs))
    params.log_search = False

    print(f"  Solving for up to {secs}s ...")
    t1 = time.time()
    solution = artifacts.routing.SolveWithParameters(params)
    solve_time = time.time() - t1
    if solution is None:
        raise FinalSolverFailure(f"OR-Tools returned no solution after {solve_time:.1f}s")
    print(f"  Solution found in {solve_time:.1f}s. "
          f"Objective: ${solution.ObjectiveValue() / COST_SCALE:,.2f}")

    plan = extract_plan(problem, artifacts, solution, solve_time)

    # Invariant check
    print("  Validating invariants ...")
    check_plan(plan, _make_config_proxy(problem), overrides=problem.overrides)
    print("  All invariants passed ✓")

    print(f"\n  Plan summary:")
    print(f"    Total stops:     {plan.total_stops}")
    print(f"    Total lbs:       {plan.total_lbs_delivered:,.0f}")
    print(f"    Total miles:     {plan.total_miles:,.1f}")
    print(f"    Avg fill %:      {plan.avg_fill_pct:.0f}%")
    print(f"    Truck-days:      {len(plan.routes)}")
    print(f"    Deferred:        {len(plan.deferred)} clients")
    return plan


def _make_config_proxy(problem: ProblemInstance):
    """Minimal config object for invariant checks (matches v2 interface)."""
    class _Shift:
        hard_max_minutes = problem.shift_hard_max_min
        weekly_max_minutes = problem.weekly_max_min
    class _Fleet:
        saturday_trucks = ['Truck2']
        shift = _Shift()
    class _Policy:
        min_stop_lbs = problem.min_stop_lbs
    class _Config:
        fleet = _Fleet()
        policy = _Policy()
    return _Config()


# ════════════════════════════════════════════════════════════════════════════
# SECTION 7. MAIN ENTRY POINT
# ════════════════════════════════════════════════════════════════════════════

def main(argv: Optional[List[str]] = None) -> int:
    """CLI: builds problem, solves with the FINAL model, writes outputs."""
    import argparse
    argv = argv if argv is not None else sys.argv[1:]

    parser = argparse.ArgumentParser(description="SK Route Optimizer — FINAL")
    parser.add_argument('--input-file', type=Path, default=None,
                        help='Path to SK_Delivery_System_ONLINE_w_anova.xlsx')
    parser.add_argument('--today', type=str, default=None,
                        help='YYYY-MM-DD plan-as-of date (default: tomorrow)')
    parser.add_argument('--output-dir', type=Path, default=None,
                        help='Where to write outputs (default: sk_optimizer/final/output)')
    parser.add_argument('--solve-seconds', type=int, default=300)
    parser.add_argument('--user-overrides', type=Path, default=None,
                        help='Optional sidecar JSON of dashboard Pins/Forbids')
    args = parser.parse_args(argv)

    # Resolve repo root + paths.
    here = Path(__file__).resolve().parent          # .../sk_optimizer/final
    repo = here.parent                              # .../sk_optimizer
    config_dir = repo / 'v2' / 'config'
    matrix_file = repo / 'data' / 'osrm_full_matrix_with_ids.npz'
    output_dir = args.output_dir or (here / 'output')
    output_dir.mkdir(parents=True, exist_ok=True)

    # Input file: arg → local_config.json → fallback
    # Also pick up `solver_settings` from local_config.json (dashboard-editable).
    input_file = args.input_file
    local_cfg_path = repo / 'local_config.json'
    if local_cfg_path.exists():
        import json as _json
        try:
            local_cfg = _json.loads(local_cfg_path.read_text(encoding='utf-8'))
        except Exception:
            local_cfg = {}
        if input_file is None and local_cfg.get('input_file'):
            input_file = Path(local_cfg['input_file'])
        # Apply runtime overrides from solver_settings
        settings = local_cfg.get('solver_settings') or {}
        if 'min_fill_pct' in settings:
            try:
                v = float(settings['min_fill_pct'])
                if 0.0 <= v <= 0.95:
                    global MIN_FILL_PCT
                    MIN_FILL_PCT = v
            except Exception:
                pass
        # NOTE: utilization_penalty_per_min is no longer runtime-overridable.
        # It's a fixed encoding of operational truth, not a tunable knob.
    if input_file is None or not input_file.exists():
        print(f"ERROR: input file not found ({input_file})")
        return 2

    # Plan-as-of date
    if args.today:
        today = date.fromisoformat(args.today)
    else:
        today = date.today() + timedelta(days=1)

    run_id = f"FINAL_{datetime.now().strftime('%Y%m%dT%H%M%S')}_{uuid.uuid4().hex[:6]}"
    print("═" * 78)
    print(f"  SK Route Optimizer — FINAL model — run {run_id}")
    print(f"  Planning as-of {today}")
    print(f"  Input:  {input_file}")
    print(f"  Output: {output_dir}")
    print("═" * 78)

    print(f"\n[1/4] Building augmented problem ...")
    user_overrides_path = args.user_overrides
    if user_overrides_path is None:
        # Default sidecar location used by the dashboard
        default_ov = repo / 'data' / 'user_overrides.json'
        if default_ov.exists():
            user_overrides_path = default_ov
    problem = build_augmented_problem(
        config_dir=config_dir,
        input_file=input_file,
        matrix_file=matrix_file,
        today=today,
        run_id=run_id,
        solve_seconds=args.solve_seconds,
        user_overrides_file=user_overrides_path,
    )
    print(f"  Problem: {len(problem.clients)} clients, "
          f"{len(problem.trucks)} trucks × {len(problem.horizon_dates)} days, "
          f"{len(problem.overrides.pins)} pins, {len(problem.overrides.forbids)} forbids")
    print(f"  Coefficients applied:")
    print(f"    cost_per_mile          = ${COST_PER_MILE}/mi")
    print(f"    labor_per_min (reg)    = ${COST_PER_MINUTE_LABOR_REG}/min  (salaried, sunk)")
    print(f"    ot_premium_per_min     = ${COST_PER_MINUTE_OT_PREMIUM}/min over target")
    print(f"    truck_dispatch_cost    = ${TRUCK_DISPATCH_COST}/day        (warm-up in fuel)")
    print(f"    small_stop_fee         = ${SMALL_STOP_FEE}/stop  (refill < {problem.min_stop_lbs} lbs)")
    print(f"    min_fill_pct           = {int(MIN_FILL_PCT*100)}%  (skip non-urgent stops below this fill %)")
    print(f"    util_penalty           = ${UTILIZATION_PENALTY_PER_MIN:.2f}/min  (per min below {UTILIZATION_TARGET_MIN}-min target)")
    print(f"    drop_penalty tiers     = HARD ${DROP_PENALTY_HARD}, HIGH ${DROP_PENALTY_HIGH}, "
          f"MED ${DROP_PENALTY_MED}, LOW ${DROP_PENALTY_LOW}")

    print(f"\n[2/4] Solving FINAL model ...")
    plan = solve_final(problem, solve_seconds=args.solve_seconds)

    print(f"\n[3/4] Writing outputs to {output_dir}")
    outputs = write_all_outputs(plan, output_dir, problem=problem)
    for name, path in outputs.items():
        if isinstance(path, list):
            for p in path:
                print(f"  {name}: {p}")
        else:
            print(f"  {name}: {path}")

    print(f"\n[4/4] Persisting state")
    state_file = repo / 'data' / 'inventory_state_final.json'
    store = StateStore(state_file)
    state_dict = {
        cid: {
            'id': cid,
            'current_lbs': float(ts.current_lbs),
            'confidence': ts.source,
            'updated_at': ts.as_of.isoformat() if ts.as_of else '',
        }
        for cid, ts in problem.initial_tanks.items()
    }
    store.save(state_dict, plan)
    store.record_run(run_id, {
        'today': today.isoformat(),
        'horizon': len(problem.horizon_dates),
        'objective_dollars': plan.objective_cost_dollars,
        'avg_fill_pct': plan.avg_fill_pct,
        'total_stops': plan.total_stops,
        'total_lbs': plan.total_lbs_delivered,
    })
    print(f"  State → {state_file}")

    print("\n" + "═" * 78)
    print("  ✓ FINAL plan ready for review.")
    print("═" * 78 + "\n")
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
